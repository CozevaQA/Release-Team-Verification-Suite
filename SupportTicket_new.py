import os
import random
from statistics import mean
import sys
import math
import time
import traceback
import csv
from os import listdir
from os.path import isfile, join
import csv
import time
from datetime import date
from datetime import datetime
import base64
from tkinter import messagebox, ttk

import pandas as pd
from PIL import Image as img
from PIL import ImageTk
import PIL
from selenium import *
from selenium import webdriver
from selenium.webdriver.common import keys
from selenium.webdriver.common.alert import Alert
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import *
from selenium.webdriver.support.wait import WebDriverWait
from openpyxl import Workbook
from openpyxl.styles import PatternFill, NamedStyle, Border
from urllib.parse import urlparse, parse_qs
from openpyxl.styles import *
import support_functions as sf
import setups
import variablestorage as locator
import ExcelProcessor as db
from tkinter import *

global Customer, client_id, user_dict_1, details
ENV = "CERT"




# options = webdriver.ChromeOptions()
# prefs = {"download.default_directory" : "C:\\Users\\sbasu\\Documents\\CSV_Files"}
# options.add_argument("user-data-dir=C:\\Users\\sbasu\\AppData\\Local\\Google\\Chrome\\User Data\\SavedData")
# options.add_argument("--disable-notifications")
# options.add_argument("--disk-cache-size=1")
# options.add_argument("--disable-extensions")
# options.add_argument("--disable-gpu")
# options.add_experimental_option("prefs", prefs)
# driver = webdriver.Chrome(executable_path="C:\\Users\\sbasu\\Documents\\Drivers\\chromedriver.exe", options=options)


global i
i = 1

def SheetColorCoder(sheet, workbook, path, filename):
    rows = sheet.max_row  # Starting of code to color code excel
    cols = sheet.max_column
    for i in range(2, rows + 1):
        for j in range(3, cols + 1):
            if sheet.cell(i, j).value == 'Passed' or sheet.cell(i, j).value == 'Present and Passed':
                sheet.cell(i, j).fill = PatternFill('solid', fgColor='BAD366')
            elif sheet.cell(i, j).value == 'Failed' or sheet.cell(i, j).value == 'Present but Failed':
                sheet.cell(i, j).fill = PatternFill('solid', fgColor='FF707A')
            elif sheet.cell(i, j).value == 'Unexecuted':
                sheet.cell(i, j).fill = PatternFill('solid', fgColor='FCD44D')
            elif sheet.cell(i, j).value == 'Metric registry is empty':
                sheet.cell(i, j).fill = PatternFill('solid', fgColor='FCC0BB')
    workbook.save(path + filename)

def skip_intro(driver):
    time.sleep(2)
    try:
        while driver.find_element(By.CLASS_NAME, "introjs-tooltip"):
            driver.find_element(By.XPATH, "//*[@class='introjs-skipbutton']").click()
            time.sleep(2)
    except NoSuchElementException:
        print("No intro")
    # if len(driver.find_element(By.CLASS_NAME,"introjs-tooltip")) != 0:
    #     driver.find_element(By.XPATH, "//*[@class='introjs-skipbutton']").click()
    # else:
    #     print("no element")

# def ajax_preloader_wait(driver):
#     time.sleep(1)
#     WebDriverWait(driver, 300).until(EC.invisibility_of_element((By.CLASS_NAME, "ajax_preloader")))
#     if len(driver.find_elements(By.CLASS_NAME, "ajax_preloader")) != 0:
#         WebDriverWait(driver, 300).until(EC.invisibility_of_element((By.CLASS_NAME, "ajax_preloader")))
#     WebDriverWait(driver, 300).until(EC.invisibility_of_element((By.CLASS_NAME, "drupal_message_text")))
#     time.sleep(1)


#for masquerading
def login_to_user(Username,ws):
    global i, details
    file = open(r"assets\loginInfo.txt", "r+")
    details = file.readlines()
    try:
        WebDriverWait(driver, 30).until(EC.presence_of_element_located((By.XPATH, "//a[@data-target='table_dropdown_people_list']")))
        driver.find_element(By.XPATH, "//a[@data-target='table_dropdown_people_list']").click()
        WebDriverWait(driver, 30).until(EC.presence_of_element_located((By.XPATH, "//input[@name='search_people']")))
        time.sleep(0.5)
        WebDriverWait(driver, 30).until(EC.presence_of_element_located((By.XPATH, "//input[@name='search_people']")))
        time.sleep(0.5)
        driver.find_element(By.XPATH, "//input[@name='search_people']").clear()
        driver.find_element(By.XPATH, "//input[@name='search_people']").send_keys(Username)
        time.sleep(0.5)
        driver.find_element(By.LINK_TEXT, 'Apply').click()
        time.sleep(0.5)
        sf.ajax_preloader_wait(driver)
        WebDriverWait(driver, 30).until(EC.presence_of_element_located((By.XPATH, "(//input[@class='filled-in selector'])[1]")))
        checkbox = driver.find_element(By.XPATH, "(//input[@class='filled-in selector'])[1]")
        driver.execute_script("arguments[0].click();", checkbox)
        time.sleep(1)
        WebDriverWait(driver, 30).until(EC.presence_of_element_located((By.XPATH, "//a[@data-tooltip='Actions']")))
        driver.find_element(By.XPATH, "//a[@data-tooltip='Actions']").click()
        WebDriverWait(driver, 30).until(EC.presence_of_element_located((By.XPATH, "//a[@id='masquerade_user']")))
        driver.find_element(By.XPATH, "//a[@id='masquerade_user']").click()
        time.sleep(0.5)
        WebDriverWait(driver, 30).until(EC.presence_of_element_located((By.XPATH, "//input[@id='edit-masquerade-reason-field']")))
        driver.find_element(By.XPATH, "//input[@id='edit-masquerade-reason-field']").clear()
        driver.find_element(By.XPATH, "//input[@id='edit-masquerade-reason-field']").send_keys(details[4].strip())
        WebDriverWait(driver, 30).until(EC.presence_of_element_located((By.XPATH, "//input[@id='edit-drsign']")))
        driver.find_element(By.XPATH, "//input[@id='edit-drsign']").clear()
        driver.find_element(By.XPATH, "//input[@id='edit-drsign']").send_keys(details[3].strip())
        time.sleep(3)
        driver.find_element(By.XPATH, "//button[@id='edit-submit']").click()
        time.sleep(1)
        sf.ajax_preloader_wait(driver)
        try:
            if driver.find_element(By.XPATH, "//*[@id='edit-later']"):
                driver.find_element(By.XPATH, "//*[@id='edit-later']").click()
                sf.ajax_preloader_wait(driver)
            else:
                sf.ajax_preloader_wait(driver)
        except NoSuchElementException as e:
            print("TOS page not present")
        print("Masqueraded to " + Username)
        if driver.title != "Registries | Cozeva":
            print("Switching to registries")
            driver.find_element(By.XPATH, "//*[@data-target='app_dropdown']").click()
            driver.find_element(By.XPATH, "//*[@class='no-hover app_registries']").click()
            driver.switch_to.window(driver.window_handles[1])
        sf.ajax_preloader_wait(driver)
        ws.append(["TC" + str(i), "Masqueraded to selected user successfully", "Passed", Username])
        i += 1
        return True
    except (NoSuchElementException, ElementNotInteractableException, ElementClickInterceptedException, ElementNotVisibleException, TimeoutException, ElementNotSelectableException) as e:
        print(e)
        traceback.print_exc()
        ws.append(["TC" + str(i), "Masquerade to the user failed", "Failed", user])
        i += 1
        return False


def OpenSupportTicket(user,ws):
# SupportTicket list
    global i
    try:
        driver.find_element(By.XPATH, "//*[@data-target='help_menu_dropdown']").click()
        time.sleep(2)
        driver.find_element(By.XPATH, "//dt[text()='Support Tickets']").click()
        WebDriverWait(driver, 300).until(EC.visibility_of_element_located((By.CLASS_NAME, "dataTables_info")))
        i += 1
        ID = ""
        try:
            driver.find_element(By.XPATH, "//*[@class='dataTables_empty']")
            ws.append(["TC" + str(i), "Navigated to Support ticket page successfully", "Passed",
                       "No Data available"])
            i += 1
        except NoSuchElementException:
            row = driver.find_element(By.XPATH, "//*[@id='support_data']").find_element(By.TAG_NAME, 'tbody').find_elements(
                By.TAG_NAME, 'tr')[0]
            print(row)
            subject = row.find_elements(By.TAG_NAME, 'td')[2].text
            print(subject)
            ID = row.find_elements(By.TAG_NAME, 'td')[1].text
            ID = ID.replace('#', '')
            print(ID)
            ws.append(["TC" + str(i), "Navigated to Support ticket page successfully", "Passed", "Last submitted ticket ID:" + str(ID)])
            i+=1
    except (NoSuchElementException, ElementNotInteractableException, ElementClickInterceptedException,
                ElementNotVisibleException, TimeoutException, ElementNotSelectableException) as e:
        print(e)
        ws.append(["TC" + str(i), "Navigation to Support ticket page failed", "Failed", "-"])
        i += 1
    # Screenshot
    try:
        screenshot_path = os.path.join(folder_path, user+"_Support Page_Before Submission.png")
        driver.save_screenshot(screenshot_path)
        print(f"Screenshot saved to {screenshot_path}")
        ws.append(["TC" + str(i), "Support Page Screenshot taken pre submission", "Passed", "_"])
        i += 1
    except (NoSuchElementException, ElementNotInteractableException, ElementClickInterceptedException,
        ElementNotVisibleException, TimeoutException, ElementNotSelectableException) as e:
        print(e)
        ws.append(["TC" + str(i), "Screenshot failed", "Failed", "-"])
        i += 1


def SubmitSupportTicket(user,ws, flag):
    global i
    ID = 0
# SupportTicket Submission
    sf.ajax_preloader_wait(driver)
    plus_xpath = "//a[@class='btn-floating btn-large red waves-effect waves-light new_support_activity_btn']"
    WebDriverWait(driver, 60).until(EC.visibility_of_element_located((By.XPATH, plus_xpath)))
    driver.find_element(By.XPATH, "//i[contains(text(),'add')]/..").click()
    time.sleep(2)
    try:
        WebDriverWait(driver, 300).until(EC.visibility_of_element_located((By.ID, "modal_dialog_support_activity_form")))
        WebDriverWait(driver, 300).until(EC.invisibility_of_element((By.XPATH, "//*[@id='modal_dialog_support_activity_form']//*[@class='ajax_preloader']")))
        ws.append(["TC" + str(i), "Support ticket submission modal opened successfully", "Passed", "-"])
        i+=1
    except (NoSuchElementException, ElementNotInteractableException, ElementClickInterceptedException,
        ElementNotVisibleException, TimeoutException, ElementNotSelectableException) as e:
        ws.append(["TC" + str(i), "Support ticket submission modal did not open", "Failed", "-"])
        i += 1
    try:
        sf.ajax_preloader_wait(driver)
        driver.find_element(By.XPATH, "//*[@name = 'subject']").send_keys("Submitting support ticket for testing purpose_CozevaQA")
        time.sleep(.2)
        # Module drop down
        driver.find_element(By.XPATH,
                            "//input[@id='select-support-activity-module-materialize-dropdown-input']").click()
        time.sleep(.2)
        module = WebDriverWait(driver, 10).until(
            EC.presence_of_all_elements_located((By.XPATH, "//ul[contains(@id,'select-options')]/li")))
        valid_options = [opt for opt in module if
                         opt.text.strip() != "" and opt.text.strip().lower() != "select" and "disabled" not in opt.get_attribute(
                             "class")]
        # Choose a random option
        random_module = random.choice(valid_options)
        print("Selected Option:", random_module.text)
        random_module.click()

        # Support Type drop down
        support_type_input = WebDriverWait(driver, 10).until(
            EC.element_to_be_clickable((By.XPATH, "//input[starts-with(@id, 'edit-support-type')]")))
        support_type_input.click()
        time.sleep(.2)
        support_type = WebDriverWait(driver, 10).until(
            EC.presence_of_all_elements_located((By.XPATH, "//ul[contains(@id,'select-options')]/li")))
        valid_options = [opt for opt in support_type if
                         opt.text.strip() != "" and opt.text.strip().lower() != "select" and "disabled" not in opt.get_attribute(
                             "class")]
        random_support_type = random.choice(valid_options)
        print("Selected Support Type:", random_support_type.text)
        random_support_type.click()
        time.sleep(2)
        driver.find_element(By.XPATH, "//*[@name = 'problem_cozeva_id']").send_keys("NA")
        time.sleep(.2)
        driver.find_element(By.XPATH, "//*[@name = 'recreate_issue_note']").send_keys("NA")
        time.sleep(1)
        if flag == 0:
            driver.find_element(By.XPATH, "(//i[@class='tiny material-icons ac-icon ac-clear'])[2]").click()
            time.sleep(3)
            driver.find_element(By.XPATH, '(//input[@name="assignee"])').send_keys("Aritra")
            #time.sleep(20)


            WebDriverWait(driver, 20).until(
                EC.element_to_be_clickable(
                    (By.XPATH, "//ul[@class='dropdown-content mat-ac-dropdown ']//li[@tabindex='0'][1]"))).click()
            time.sleep(2)
            screenshot_path = os.path.join(folder_path, user + "_AfterSearch.png")
            driver.save_screenshot(screenshot_path)
            time.sleep(1)
        # to add note
        driver.find_element(By.XPATH, "//*[@name = 'note']").send_keys("Note for testing purpose_CozevaQA")
        time.sleep(1)
        driver.find_element(By.XPATH, "//*[contains(text(),'What is the urgency?')]/../div").click()
        time.sleep(.2)
        driver.find_element(By.XPATH, "//*[contains(text(),'What is the urgency?')]/../div/ul//*[contains(text(),'Normal')]").click()
        WebDriverWait(driver, 300).until(EC.invisibility_of_element((By.XPATH, "//*[@class='throbber']")))
        time.sleep(1.5)
        # to add attachment
        try:
            file_paths = [
                "assets\\test_files\\Doc_CSV.csv",
                "assets\\test_files\\Doc_DOC.docx",
                "assets\\test_files\\Doc_JPEG.jpg",
                "assets\\test_files\\Doc_PDF.pdf",
                "assets\\test_files\\Doc_PNG.png",
                "assets\\test_files\\Doc_PPT.pptx",
                "assets\\test_files\\Doc_TIF.tif",
                "assets\\test_files\\Doc_XLSX.xlsx",
                "assets\\test_files\\Doc_ZIP.zip"
            ]
            for iterator, path in enumerate(file_paths):
                file_paths[iterator] = os.path.join(os.getcwd(), path)
            selected_file = random.choice(file_paths)
            file_input = driver.find_element(By.XPATH, "//*[@name = 'files[file_upload]']")
            driver.execute_script("arguments[0].scrollIntoView();", file_input)
            file_input.send_keys(selected_file)
            selected_file_name = os.path.basename(selected_file)
            WebDriverWait(driver, 300).until(EC.invisibility_of_element((By.XPATH, "//*[contains(text(),'*File upload in progress')]")))
            WebDriverWait(driver, 300).until(EC.visibility_of_element_located((By.XPATH, "//*[@data-tooltip='Delete']")))
            element2 = driver.find_element(By.XPATH, "//*[@id='edit-submit-down']")
            driver.execute_script("arguments[0].scrollIntoView();", element2)
            ws.append(["TC" + str(i), "File attached successfully", "Passed", selected_file_name])
        except Exception as e:
            print(e)
            print('File problem')
            ws.append(["TC" + str(i), "File attachment error", "Failed", "-"])
        element2 = driver.find_element(By.XPATH, "//*[@id='edit-submit-down']")
        driver.execute_script("arguments[0].scrollIntoView();", element2)
        WebDriverWait(driver, 300).until(EC.invisibility_of_element((By.XPATH, "//*[@class='fthumb uploading']")))
        WebDriverWait(driver, 300).until(EC.visibility_of_all_elements_located((By.XPATH, "//*[@class='fthumb']")))
        #WebDriverWait(driver, 300).until(EC.element_to_be_clickable((By.XPATH, "//*[@class = 'messaging_submit_btn support_ticket_phi_submit button js-form-submit form-submit btn-default btn']")))
        time.sleep(3)
        element2.click()

        WebDriverWait(driver, 300).until(EC.invisibility_of_element((By.XPATH, "//*[@class='throbber']")))

        time.sleep(1.5)
        sf.ajax_preloader_wait(driver)
        print("Submitted")
        ws.append(["TC" + str(i), "Support ticket submitted successfully", "Passed", "-"])
        i += 1
    except(NoSuchElementException, ElementNotInteractableException, ElementClickInterceptedException,
        ElementNotVisibleException, TimeoutException, ElementNotSelectableException) as e:
        traceback.print_exc()
        ws.append(["TC" + str(i), "Support ticket submission failed", "Failed", "-"])
        i += 1
# Screenshot
    try:
        print("Waiting for modal_dialog_support_activity_form")
        WebDriverWait(driver, 300).until(
            EC.invisibility_of_element_located((By.ID, "modal_dialog_support_activity_form")))
        screenshot_path2 = os.path.join(folder_path, user + "_Support Page_After Submission.png")
        driver.save_screenshot(screenshot_path2)
        print(f"Screenshot saved to {screenshot_path2}")
        WebDriverWait(driver, 300).until(EC.visibility_of_element_located((By.CLASS_NAME, "dataTables_info")))
        print("Waiting for dataTables_info")
        list_row = driver.find_element(By.XPATH, "//*[@id='support_data']").find_element(By.TAG_NAME, 'tbody').find_elements(By.TAG_NAME, 'tr')
        check_flag = 0
        while ID == 0:
            row = list_row[check_flag]
            print(row)
            subject = row.find_elements(By.TAG_NAME, 'td')[2].text
            print(subject)
            if subject == "Submitting support ticket for testing purpose_CozevaQA":
                ID = row.find_elements(By.TAG_NAME, 'td')[1].text
                ID = ID.replace('#', '')
                print(ID)
                break
            else:
                check_flag +=1
                ID = 0
        ws.append(["TC" + str(i), "Screenshot taken post submission", "Passed", "ID :"+str(ID)])
        i += 1
        return ID
    except (NoSuchElementException, ElementNotInteractableException, ElementClickInterceptedException,
        ElementNotVisibleException, TimeoutException, ElementNotSelectableException) as e:
        ws.append(["TC" + str(i), "Screenshot failed post submission", "Failed", "-"])
        i+=1
        print(e)

def EditResubmit(user,ws,ID,flag):
    global i
    #ticket_xpath = f"//a[@class='ticket_no ' and contains(text(), '#{ID}')]"
    ticket_xpath = f"//a[contains(@class, 'ticket_no') and contains(., '#{ID}')]"

    sf.ajax_preloader_wait(driver)
    first_ticket_link = driver.find_element(By.XPATH, ticket_xpath)
    original_window = driver.current_window_handle
    before_windows = driver.window_handles
    first_ticket_link.click()
    WebDriverWait(driver, 10).until(lambda d: len(d.window_handles) > len(before_windows))
    for handle in driver.window_handles:
        if handle != original_window:
            driver.switch_to.window(handle)
            break
    driver.find_element(By.XPATH, '//*[@id="edit-submit-down"]').click()
    WebDriverWait(driver, 300).until(EC.invisibility_of_element((By.XPATH, "//*[@class='throbber']")))
    time.sleep(5)
    driver.find_element(By.XPATH, "(//i[@class='tiny material-icons ac-icon ac-clear'])[2]").click()
    time.sleep(2)
    if flag == 0:
        driver.find_element(By.XPATH, "//*[@name='assignee']").send_keys("Kaushik")
        time.sleep(20)
    else:
        driver.find_element(By.XPATH, "//*[@name='assignee']").send_keys("SimulatedCustomer2_GlobalCustomerSupport")
        time.sleep(5)
    screenshot_path4 = os.path.join(folder_path, user + "_SearchWhileEdit.png")
    driver.save_screenshot(screenshot_path4)
    time.sleep(1)
    # driver.find_element(By.XPATH, "//ul[@class='dropdown-content mat-ac-dropdown ']//li[@tabindex='0']").click()
    WebDriverWait(driver, 20).until(
        EC.element_to_be_clickable(
            (By.XPATH, "//ul[@class='dropdown-content mat-ac-dropdown ']//li[@tabindex='0'][1]"))).click()

    time.sleep(2)
    save_xpath = "//*[contains(text(),'Save')]" #"//button[contains(@class, 'messaging_submit_btn') and text()='Save']"
    save_button = driver.find_element(By.XPATH, save_xpath)
    save_button.click()
    ws.append(["TC" + str(i), "Support ticket ID:" + str(ID) + " edited successfully", "Passed", "-"])
    i += 1
    driver.close()
    driver.switch_to.window(driver.window_handles[0])


def SupportTicketReply(user,ws,ID):
    global i
    #ticket_xpath = f"//a[@class='ticket_no ' and contains(text(), '#{ID}')]"
    ticket_xpath = f"//a[contains(@class, 'ticket_no') and contains(., '#{ID}')]"
    first_ticket_link = driver.find_element(By.XPATH, ticket_xpath)
    original_window = driver.current_window_handle
    before_windows = driver.window_handles
    first_ticket_link.click()
    WebDriverWait(driver, 10).until(lambda d: len(d.window_handles) > len(before_windows))
    for handle in driver.window_handles:
        if handle != original_window:
            driver.switch_to.window(handle)
            break
    driver.find_element(By.XPATH, '//*[@id="edit-reply-form"]').click()
    ReplyTextXpath = "//textarea[contains(@class, 'reply-text') and contains(@class, 'format_text_new') and contains(@class, 'form-textarea') and contains(@class, 'materialize-textarea') and contains(@class, 'required')]"
    try:
        time.sleep(5)
        reply_box = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, ReplyTextXpath)))
        print("Reply text area found. Proceeding to the next step...")
        reply_box.click()
        reply_box.send_keys("Support ticket reply for testing purpose CozevaQA")
        # to add attachment
        try:
            file_paths = [
                "assets\\test_files\\Doc_CSV.csv"
            ]
            for iterator, path in enumerate(file_paths):
                file_paths[iterator] = os.path.join(os.getcwd(), path)
            selected_file = random.choice(file_paths)
            file_input = driver.find_element(By.XPATH, "//*[@name = 'files[file_upload]']")
            driver.execute_script("arguments[0].scrollIntoView();", file_input)
            file_input.send_keys(selected_file)
            selected_file_name = os.path.basename(selected_file)
            WebDriverWait(driver, 300).until(
                EC.invisibility_of_element((By.XPATH, "//*[contains(text(),'*File upload in progress')]")))

            WebDriverWait(driver, 300).until(EC.visibility_of_element_located((By.XPATH, "//*[@data-tooltip='Delete']")))
            ReplySendButtonXpath = "//button[contains(@class, 'use-ajax-submit') and contains(@class, 'support_ticket_phi_submit') and @type='submit' and text()='Send']"
            send_button = driver.find_element(By.XPATH, ReplySendButtonXpath)
            send_button.click()
            time.sleep(5)
            ws.append(["TC" + str(i), "File attached for reply successfully", "Passed", selected_file_name])
            i += 1
        except Exception as e:
            print(e)
            print('File problem')
            ws.append(["TC" + str(i), "File attachment error", "Failed", "-"])
            i += 1
        ws.append(["TC" + str(i), "Support ticket ID:" + str(ID) + " replied successfully", "Passed", "-"])
        i += 1
    except NoSuchElementException:
        print("Reply text area not found.")
    screenshot_path3 = os.path.join(folder_path, user + "_Support Ticket_After Reply.png")
    driver.save_screenshot(screenshot_path3)
    print(f"Screenshot saved to {screenshot_path3}")
    time.sleep(2)
    driver.close()
    driver.switch_to.window(driver.window_handles[0])


#for switching back
def switch_back():
    try:
        WebDriverWait(driver, 60).until(
            EC.presence_of_element_located((By.XPATH, '//*[@id="nav"]/div/ul/li[2]/a')))
        driver.find_element(By.XPATH, '//*[@id="nav"]/div/ul/li[2]/a').click()
        WebDriverWait(driver, 30).until(
            EC.presence_of_element_located((By.XPATH, '//*[@id="quick_switch_links"]/div/ul/li/a')))
        driver.find_element(By.XPATH, '//*[@id="quick_switch_links"]/div/ul/li/a').click()
        WebDriverWait(driver, 300).until(EC.invisibility_of_element((By.CLASS_NAME, "drupal_message_text")))
    except Exception as e:
        print("error : Unable to switch back - " + e)

#This set of input will come from RTVS
#Customer = "Altamed"

#user_dict_1 = {"Customer Support" : "abc","Regional Support" : "def", "Office Admin" : "ghi", "Provider" : "8ji"}
#user_dict_1 = {"Cozeva Support" : "sbasu.cs", "Customer Support" : "AltaMed_AlUtria", "Regional Support" : "alena.meza", "Provider" : "mpm_GeTabi"}
#user_dict_1 = {"Cozeva Support" : "sbasu.cs"}


def fetch_client_name():
    root = Tk()

    root.configure(background='white')
    style = ttk.Style()
    style.theme_use('alt')
    style.configure('My.TButton', font=('Helvetica', 13, 'bold'), foreground='Black', background='#5a9c32',
                    padding=15, highlightthickness=0, height=1, width=25)
    style.configure('Configs.TButton', font=('Helvetica', 8, 'bold'), foreground='Black', background='#5a9c32',
                    highlightthickness=0)
    style.configure('Next.TButton', font=('Helvetica', 10, 'bold'), foreground='Black', background='#5a9c32',
                    highlightthickness=0)
    style.configure('CheckbuttonStyle.TCheckbutton', font=('Helvetica', 13, 'bold'), foreground='Black',
                    background='white')
    style.configure('TCombobox', fieldbackground=('readonly', 'green'), background=('readonly', 'white'),
                    foreground=('readonly', 'black'))

    style.map('My.TButton', background=[('active', '#72B132')])


    def on_start():
        global client_id, Customer, user_dict_1, ENV, details
        file = open(r"assets\loginInfo.txt", "r+")
        details_in_ui = file.readlines()
        client_id = db.fetchCustomerID(selected_cust.get())
        Customer = db.fetchCustomerName(client_id)
        fetched_user_list = db.getDefaultUserNames(Customer)
        if "Cozeva support" not in fetched_user_list:
            user_dict_1 = {"Cozeva Support": details_in_ui[0].strip()}
            user_dict_1.update(fetched_user_list)
        else:
            user_dict_1 = fetched_user_list






        ENV = selected_env.get()

        root.destroy()

    print(PIL.__version__)

    #image_small = img.open("assets/images/cozeva_logo.png").resize((25, 25))
    cozeva_logo_image = ImageTk.PhotoImage(img.open("assets/images/cozeva_logo.png").resize((280, 60)))
    logo_label = Label(root, image=cozeva_logo_image, background="white")
    logo_label.grid(row=1, column=0, padx=25, columnspan=4)

    please_select_label = Label(root, text="Support Activity Validation", background="white",
                                font=("Times New Roman", 15))
    please_select_label.grid(row=2, column=0, columnspan=4)

    #customer_label = Label(root, text="Select customer", font=("Nunito Sans", 10))
    selected_cust = StringVar()
    selected_cust.set("Select Customer")
    customer_list = db.getCustomerList()  # vs.customer_list
    customer_drop = ttk.Combobox(root, textvariable=selected_cust, values=customer_list, state='readonly',
                                 style='TCombobox', width=35)
    customer_drop.set('Select Customer')


    #customer_label.grid(row=0, column=0)
    customer_drop.grid(row=4, column=0, columnspan=4)

    envlist = ["PROD", "CERT", "Stage"]
    selected_env = StringVar()
    selected_env.set("PROD")
    env_dropdown = ttk.Combobox(root, textvariable=selected_env, values=envlist, state='readonly',
                                 style='TCombobox', width=35)

    env_dropdown.grid(row=5, column=0, columnspan=4)


    ttk.Button(root, text="Start Test", command=on_start,style='Next.TButton').grid(row=6, column=0, columnspan=5)





    root.title("Support Activity Validation")
    root.iconbitmap("assets/icon.ico")
    # root.geometry("400x400")
    root.mainloop()


fetch_client_name()

user_list = user_dict_1.values()
user_type = user_dict_1.keys()
masq_flag = 0
print("hello")

for username in user_list:
    print(username)
    if username == "99999":
        masq_flag = 1
        messagebox.showinfo("Alert!!", "Unable to add/delete/view onshore support tickets")
        sys.exit(0)



for user in user_list:
    print(user)

# FolderPath Creation
now = datetime.now()
date_time_str = now.strftime("%Y-%m-%d")
folder_path = f'C://VerificationReports//SupportTicketVerification/'+date_time_str
os.makedirs(folder_path, exist_ok=True)
folder_path = os.path.join(folder_path, Customer)
os.makedirs(folder_path, exist_ok=True)

# path1 = "C:\\Users\\sbasu\\Documents\\Report\\"
# name_date = datetime.datetime.now()
header = NamedStyle(name="header")
header.font = Font(bold=True)
header.border = Border(bottom=Side(border_style="thin"))
header.alignment = Alignment(horizontal="center", vertical="center")
red_background = PatternFill(patternType='solid', fgColor="00FF0000")
green_background = PatternFill(patternType='solid', fgColor="50C878")
gray_background = PatternFill(patternType='solid', fgColor="5F9EA0")

filename = "//Support Ticket Verification.xlsx"
wb = Workbook()
ws = wb.active
sheet_name = Customer
ws.title = sheet_name
ws.append(["ID", "Scenario", "Status", "Comments"])

header_row = ws[1]
for cell in header_row:
    cell.style = header

# logout_url = "https://cert.cozeva.com/user/logout"
# login_url = "https://cert.cozeva.com/user/login"
# base_url = "https://www.cozeva.com/"
# User = os.environ.get('CS2_User')
# Pass = os.environ.get('CS2_Password')
#
# driver.get(logout_url)
# driver.get(login_url)
# driver.maximize_window()
# print("Initial window= " + driver.title)
# driver.find_element(By.ID, "edit-name").send_keys(User)
# driver.find_element(By.ID, "edit-pass").send_keys(Pass)
# driver.find_element(By.ID, "edit-submit").click()
# time.sleep(1)
#
# try:
#     WebDriverWait(driver, 120).until(EC.presence_of_element_located((By.ID, "reason_textbox")))
#     driver.find_element(By.ID, "reason_textbox").send_keys("https://redmine2.cozeva.com/issues/18008")
#     driver.find_element(By.ID, "edit-submit").click()
# except NoSuchElementException:
#     traceback.print_exc()
#     driver.quit()
driver = setups.driver_setup()

if ENV == 'CERT':
    setups.login_to_cozeva_cert(client_id)
elif ENV == 'STAGE':
    setups.login_to_cozeva_stage()
elif ENV == "PROD":
    setups.login_to_cozeva(client_id)
else:
    print("ENV INVALID")
    exit(3)








print("Landing page= " + driver.title)
sf.ajax_preloader_wait(driver)
sf.skip_intro(driver)

if driver.title == "Registries | Cozeva":
    driver.find_element(By.XPATH, "//a[@data-target='user_menu_dropdown']").click()
    WebDriverWait(driver, 30).until(
        EC.presence_of_element_located((By.XPATH, "//a[@class='not_for_mobile'][contains(@href,'/users_list')]")))
    time.sleep(0.5)
    WebDriverWait(driver, 120).until(
        EC.element_to_be_clickable((By.XPATH, "//a[@class='not_for_mobile'][contains(@href,'/users_list')]")))
    driver.find_element(By.XPATH, "//a[@class='not_for_mobile'][contains(@href,'/users_list')]").click()
    sf.ajax_preloader_wait(driver)
    print("Users list opened")
    user_list_url = driver.current_url

ticket_detail_list = []
temp_id  = ""
for (type, user) in user_dict_1.items():
    if type == "Cozeva Support":
        print(user)
        OpenSupportTicket(user, ws)
        ID = SubmitSupportTicket(user, ws, masq_flag)  # SubmitSupportTicket() (Done by Neha)
        temp_id = ID
        print(temp_id)
        EditResubmit(user, ws, ID, masq_flag)
        SupportTicketReply(user, ws, ID)
        ticket_detail_list.append(user)
        ticket_detail_list.append(ID)
        ticket_detail_list.append(masq_flag)
        wb.save(folder_path + filename)
        driver.get(user_list_url)
    else:
        print(user)
        success = True
        success = login_to_user(user, ws)
        if not success:
            continue
        skip_intro(driver)
        masq_flag = 1
        try:
            driver.find_element(By.ID, "announcement_list")
            driver.find_element(By.XPATH, "//*[@id='delete_all_banners']").click()
        except (NoSuchElementException, ElementNotInteractableException) as e:
            print("No announcement")
            traceback.print_exc()
        if type == "Customer Support":
            OpenSupportTicket(user, ws)
            SupportTicketReply(user, ws, temp_id)
            ID = SubmitSupportTicket(user, ws, masq_flag)
            if Customer == "Simulated Customer 2 (Deid)":
                EditResubmit(user, ws, ID, masq_flag)
            else:
                print("Not Simulated Customer skipping edit resubmit from CU")
            ticket_detail_list.append(user)
            ticket_detail_list.append(ID)
            ticket_detail_list.append(masq_flag)
            switch_back()
            wb.save(folder_path + filename)
        else:
            OpenSupportTicket(user, ws)
            ID = SubmitSupportTicket(user, ws, masq_flag)  # SubmitSupportTicket() (Done by Neha)
            ticket_detail_list.append(user)
            ticket_detail_list.append(ID)
            ticket_detail_list.append(masq_flag)
            switch_back()
            wb.save(folder_path + filename)


#for user in user_list:
#    if masq_flag != 1:
#        print(user)
#        login_to_user(user, ws)
#        sf.skip_intro(driver)
#        try:
#            driver.find_element(By.ID, "announcement_list")
#           driver.find_element(By.XPATH, "//*[@id='delete_all_banners']").click()
#       except Exception as e:
#           print("No announcement")
#       OpenSupportTicket(user, ws)
#       ID = SubmitSupportTicket(user, ws) #SubmitSupportTicket() (Done by Neha)


#       ticket_detail_list.append(user)
#       ticket_detail_list.append(ID)
#       switch_back()
#       wb.save(folder_path + filename)
#   else:
#       OpenSupportTicket(user, ws)
#       ID = SubmitSupportTicket(user, ws) #SubmitSupportTicket() (Done by Neha)
#       ticket_detail_list.append(user)
#       ticket_detail_list.append(ID)
#       wb.save(folder_path + filename)

wb.save(folder_path+filename)
sf.ajax_preloader_wait(driver)
driver.find_element(By.XPATH, "//*[@data-target='app_dropdown']").click()
time.sleep(1)
driver.find_element(By.XPATH, "//*[@class='no-hover app_ut']").click()
driver.switch_to.window(driver.window_handles[1])
sf.ajax_preloader_wait(driver)
print(ticket_detail_list)
sf.ajax_preloader_wait(driver)
for k in range(0, len(ticket_detail_list)-1,3):
    try:
        driver.find_element(By.XPATH, "//a[@data-target='sidenav_slide_out']").click()
        time.sleep(0.5)
        driver.find_element(By.XPATH, "//span[text()='Delete Testing Data']/../..").click()
        time.sleep(2)
        if ticket_detail_list[k+2] == 1:
            driver.find_element(By.XPATH, "//*[@id='masquaraded_to']/../span").click()
            driver.find_element(By.XPATH, "//*[@id='logged_or_masquaraded_user_name']").send_keys(ticket_detail_list[k])
            time.sleep(3)
            driver.find_element(By.XPATH, "//ul[@id='ac-dropdown-logged_or_masquaraded_user_name']").click()
        driver.find_element(By.XPATH, "//*[text()='Select To Clean *']/..").click()
        time.sleep(1)
        driver.find_element(By.XPATH, "//span[text()='Support Activity ID']/..").click()
        time.sleep(1)
        driver.find_element(By.XPATH, "//*[@id='delete_type_id']").send_keys(ticket_detail_list[k+1])
        time.sleep(2)
        screenshot_path = os.path.join(folder_path, "Deleting modal "+str(ticket_detail_list[k]+".png"))
        driver.save_screenshot(screenshot_path)
        print(f"Screenshot saved to {screenshot_path}")
        driver.find_element(By.XPATH, "//*[@data-index='confirm']").click()
        sf.ajax_preloader_wait(driver)
        WebDriverWait(driver, 300).until(EC.presence_of_element_located((By.XPATH, "//*[contains(text(),'Deleted Successfully')]")))
        ws.append(["TC" + str(i), "Support ticket ID:"+str(ticket_detail_list[k+1])+" deleted successfully", "Passed", "User: "+str(ticket_detail_list[k])])
        i+=1
    except(NoSuchElementException, ElementNotInteractableException, ElementClickInterceptedException,
        ElementNotVisibleException, TimeoutException, ElementNotSelectableException) as e:
        ws.append(
            ["TC" + str(i), "Support ticket ID:" + str(ticket_detail_list[k + 1]) + " deletion failed", "Failed",
             "-"])
        i += 1

driver.close()
wb.save(folder_path+filename)
driver.switch_to.window(driver.window_handles[0])
sf.ajax_preloader_wait(driver)
driver.find_element(By.XPATH, "//*[@data-target='help_menu_dropdown']").click()
driver.find_element(By.XPATH, "//*[contains(text(), 'Support Tickets')]").click()
sf.ajax_preloader_wait(driver)
screenshot_path = os.path.join(folder_path, "Support Page after deletion From CS Context.png")
driver.save_screenshot(screenshot_path)
print(f"Screenshot saved to {screenshot_path}")
SheetColorCoder(ws, wb, folder_path, filename)
driver.quit()




