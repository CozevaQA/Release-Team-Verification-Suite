import base64
import os
import traceback
from datetime import date

from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import context_functions
import guiwindow
import variablestorage as locator
import support_functions as sf
import time
import logging
from threading import Timer

import summary_sheet as ss


driver = ""
details = ""


def driver_setup():
    options = webdriver.ChromeOptions()
    options.add_argument("--disable-notifications")
    options.add_argument("--start-maximized")
    options.add_argument(locator.chrome_profile_path)  # Path to your chrome profile
    if guiwindow.headlessmode == 1:
        options.add_argument("--headless")

    options.add_argument('--disable-gpu')
    # options.add_argument("--window-size=1920,1080")
    # options.add_argument("--start-maximized")
    options.add_argument("--no-sandbox");
    options.add_argument("--dns-prefetch-disable");
    global driver
    driver = webdriver.Chrome(executable_path=locator.chrome_driver_path, options=options)
    print(guiwindow.Window_location)
    if guiwindow.Window_location == 1:
        driver.set_window_position(-1000, 0)
    elif guiwindow.Window_location == 0:
        driver.set_window_position(1000, 0)
    driver.maximize_window()
    driver.implicitly_wait(0.75)
    return driver


def create_folders(role):
    dateandtime = sf.date_time()
    datenow = date.today()
    path0 = os.path.join(locator.parent_dir, str(datenow))
    isdir = os.path.isdir(path0)
    if not isdir:
        os.mkdir(path0)

    CustID_Name_string = str(guiwindow.verification_specs[1])+"_"+guiwindow.verification_specs[0]
    path1 = os.path.join(path0, CustID_Name_string)
    isdir = os.path.isdir(path1)
    if not isdir:
        os.mkdir(path1)

    path = os.path.join(path1, str(role)+"_"+str(guiwindow.verification_specs[3][role])+"_"+dateandtime)
    os.mkdir(path)
    return str(path)


def login_to_cozeva():
    driver.get(locator.logout_link)
    driver.get(locator.login_link)
    driver.maximize_window()
    file = open(r"assets\loginInfo.txt", "r+")
    global details
    details = file.readlines()
    driver.find_element_by_id("edit-name").send_keys(details[0].strip())
    driver.find_element_by_id("edit-pass").send_keys(details[1].strip())
    file.seek(0)
    file.close()
    driver.find_element_by_id("edit-submit").click()
    time.sleep(4)
    otpurl = driver.current_url
    sub_str = "/twostepAuthSettings"
    if otpurl.find(sub_str) != -1:
        # print("Need to enter OTP for login. Please paste the OTP here")
        # wait_time = 60
        # start_time = time.perf_counter()
        # otp = ""
        # while (time.perf_counter() - start_time) < wait_time:
        #     otp = input()
        #     if len(otp) > 0:
        #         print("OTP Recieved")
        # if len(otp) == 0:
        #     print("You did not enter an OTP!!")
        #     exit(999)

        timeout = 60
        t = Timer(timeout, print, ['You did not enter the OTP'])
        t.start()
        prompt = "You have %d seconds to enter the OTP here\n" % timeout
        otp = input(prompt)
        t.cancel()

        driver.find_element_by_id("edit-twostep-code").send_keys(otp)
        time.sleep(1)

        driver.find_element_by_id("edit-twostep").click()
    WebDriverWait(driver, 120).until(EC.presence_of_element_located((By.ID, "reason_textbox")))
    driver.find_element_by_id("reason_textbox").send_keys(details[4].strip())
    time.sleep(0.5)
    driver.find_element_by_id("edit-submit").click()
    sf.ajax_preloader_wait(driver)
    WebDriverWait(driver, 30).until(
        EC.presence_of_element_located((By.XPATH, locator.xpath_filter_measure_list)))
    print("Logged in to Cozeva!")


def login_to_cozeva_cert():
    driver.get(locator.logout_link_cert)
    driver.get(locator.login_link_cert)
    driver.maximize_window()
    file = open(r"assets\loginInfo.txt", "r+")
    global details
    details = file.readlines()
    driver.find_element_by_id("edit-name").send_keys(details[0].strip())
    driver.find_element_by_id("edit-pass").send_keys(details[1].strip())
    file.seek(0)
    file.close()
    driver.find_element_by_id("edit-submit").click()
    time.sleep(4)
    otpurl = driver.current_url
    sub_str = "/twostepAuthSettings"
    if otpurl.find(sub_str) != -1:
        # print("Need to enter OTP for login. Please paste the OTP here")
        # wait_time = 60
        # start_time = time.perf_counter()
        # otp = ""
        # while (time.perf_counter() - start_time) < wait_time:
        #     otp = input()
        #     if len(otp) > 0:
        #         print("OTP Recieved")
        # if len(otp) == 0:
        #     print("You did not enter an OTP!!")
        #     exit(999)

        timeout = 60
        t = Timer(timeout, print, ['You did not enter the OTP'])
        t.start()
        prompt = "You have %d seconds to enter the OTP here\n" % timeout
        otp = input(prompt)
        t.cancel()

        driver.find_element_by_id("edit-twostep-code").send_keys(otp)
        time.sleep(1)

        driver.find_element_by_id("edit-twostep").click()
    WebDriverWait(driver, 90).until(EC.presence_of_element_located((By.ID, "reason_textbox")))
    driver.find_element_by_id("reason_textbox").send_keys(details[4].strip())
    time.sleep(0.5)
    driver.find_element_by_id("edit-submit").click()
    sf.ajax_preloader_wait(driver)
    WebDriverWait(driver, 30).until(
        EC.presence_of_element_located((By.XPATH, locator.xpath_filter_measure_list)))
    print("Logged in to Cozeva!")


def login_to_cozeva_stage():
    driver.get(locator.logout_link_stage)
    driver.get(locator.login_link_stage)
    driver.maximize_window()
    file = open(r"assets\loginInfo.txt", "r+")
    global details
    details = file.readlines()
    driver.find_element_by_id("edit-name").send_keys(details[0].strip())
    driver.find_element_by_id("edit-pass").send_keys(details[1].strip())
    file.seek(0)
    file.close()
    driver.find_element_by_id("edit-submit").click()
    time.sleep(4)
    otpurl = driver.current_url
    sub_str = "/twostepAuthSettings"
    if otpurl.find(sub_str) != -1:
        # print("Need to enter OTP for login. Please paste the OTP here")
        # wait_time = 60
        # start_time = time.perf_counter()
        # otp = ""
        # while (time.perf_counter() - start_time) < wait_time:
        #     otp = input()
        #     if len(otp) > 0:
        #         print("OTP Recieved")
        # if len(otp) == 0:
        #     print("You did not enter an OTP!!")
        #     exit(999)

        timeout = 60
        t = Timer(timeout, print, ['You did not enter the OTP'])
        t.start()
        prompt = "You have %d seconds to enter the OTP here\n" % timeout
        otp = input(prompt)
        t.cancel()

        driver.find_element_by_id("edit-twostep-code").send_keys(otp)
        time.sleep(1)

        driver.find_element_by_id("edit-twostep").click()


    WebDriverWait(driver, 90).until(EC.presence_of_element_located((By.ID, "reason_textbox")))
    driver.find_element_by_id("reason_textbox").send_keys(details[4].strip())
    time.sleep(0.5)
    driver.find_element_by_id("edit-submit").click()
    sf.ajax_preloader_wait(driver)
    WebDriverWait(driver, 30).until(
        EC.presence_of_element_located((By.XPATH, locator.xpath_filter_measure_list)))
    print("Logged in to Cozeva!")


def switch_customer_context(cusID):
    try:
        sm_customer_id = (str(cusID)).strip()
        #print(sm_customer_id)
        session_var = 'app_id=registries&custId=' + str(sm_customer_id) + '&payerId=' + str(
            sm_customer_id) + '&orgId=' + str(sm_customer_id)
        encoded_string = base64.b64encode(session_var.encode('utf-8'))
        driver.get("https://www.cozeva.com/registries?session=" + encoded_string.decode('utf-8'))
        sf.ajax_preloader_wait(driver)
        print("Navigated to customer's context")

    except Exception as e:
        print(e)
        #logger.exception("Exception occurred in OpenRegistryPageforCS2 block!")
        raise


def switch_customer_context_cert(cusID):
    try:
        sm_customer_id = (str(cusID)).strip()
        #print(sm_customer_id)
        session_var = 'app_id=registries&custId=' + str(sm_customer_id) + '&payerId=' + str(
            sm_customer_id) + '&orgId=' + str(sm_customer_id)
        encoded_string = base64.b64encode(session_var.encode('utf-8'))
        driver.get("https://cert.cozeva.com/registries?session=" + encoded_string.decode('utf-8'))
        sf.ajax_preloader_wait(driver)
        print("Navigated to customer's context")
    except Exception as e:
        print(e)
        #logger.exception("Exception occurred in OpenRegistryPageforCS2 block!")
        raise


def login_to_user(Username):
    try:
        sf.ajax_preloader_wait(driver)
        driver.find_element_by_xpath("//a[@data-target='user_menu_dropdown']").click()
        WebDriverWait(driver, 30).until(
            EC.presence_of_element_located((By.XPATH, "//a[@class='not_for_mobile'][contains(@href,'/users_list')]")))
        time.sleep(0.5)
        WebDriverWait(driver, 120).until(
            EC.element_to_be_clickable((By.XPATH, "//a[@class='not_for_mobile'][contains(@href,'/users_list')]")))

        driver.find_element_by_xpath("//a[@class='not_for_mobile'][contains(@href,'/users_list')]").click()
        sf.ajax_preloader_wait(driver)
        #logger.info("Users list opened")
        print("Users list opened")
        WebDriverWait(driver, 30).until(
            EC.presence_of_element_located((By.XPATH, "//a[@data-target='table_dropdown_people_list']")))
        driver.find_element_by_xpath("//a[@data-target='table_dropdown_people_list']").click()
        WebDriverWait(driver, 30).until(EC.presence_of_element_located((By.XPATH, "//input[@name='search_people']")))
        time.sleep(0.5)
        driver.find_element_by_xpath("//input[@name='search_people']").clear()
        driver.find_element_by_xpath("//input[@name='search_people']").send_keys(Username)
        time.sleep(0.5)
        driver.find_element_by_link_text('Apply').click()
        time.sleep(0.5)
        sf.ajax_preloader_wait(driver)
        WebDriverWait(driver, 30).until(
            EC.presence_of_element_located((By.XPATH, "(//input[@class='filled-in selector'])[1]")))
        checkbox = driver.find_element_by_xpath("(//input[@class='filled-in selector'])[1]")
        driver.execute_script("arguments[0].click();", checkbox)
        #logger.info("User to be masqueraded is selected.")
        time.sleep(1)
        WebDriverWait(driver, 30).until(EC.presence_of_element_located((By.XPATH, "//a[@data-tooltip='Actions']")))
        driver.find_element_by_xpath("//a[@data-tooltip='Actions']").click()
        WebDriverWait(driver, 30).until(EC.presence_of_element_located((By.XPATH, "//a[@id='masquerade_user']")))
        driver.find_element_by_xpath("//a[@id='masquerade_user']").click()
        time.sleep(0.5)
        WebDriverWait(driver, 30).until(
            EC.presence_of_element_located((By.XPATH, "//input[@id='edit-masquerade-reason-field']")))
        driver.find_element_by_xpath("//input[@id='edit-masquerade-reason-field']").clear()
        driver.find_element_by_xpath("//input[@id='edit-masquerade-reason-field']").send_keys(details[4].strip())
        WebDriverWait(driver, 30).until(EC.presence_of_element_located((By.XPATH, "//input[@id='edit-drsign']")))
        driver.find_element_by_xpath("//input[@id='edit-drsign']").clear()
        driver.find_element_by_xpath("//input[@id='edit-drsign']").send_keys(details[3].strip())
        time.sleep(3)
        driver.find_element_by_xpath("//button[@id='edit-submit']").click()
        time.sleep(1)
        sf.ajax_preloader_wait(driver)
        if len(driver.find_elements_by_xpath(locator.xpath_skip_button)) != 0:
            driver.find_element_by_xpath(locator.xpath_skip_button).click()
            sf.ajax_preloader_wait(driver)
        print("Masqueraded to user's context")
        #logger.info("Masqueraded to user's context")

    except Exception as e:
        print(e)
        #logger.exception("Exception occurred in Masquerade block!")
        raise


def switch_back():
    try:
        WebDriverWait(driver, 60).until(
            EC.presence_of_element_located((By.XPATH, locator.xpath_switch_back)))
        driver.find_element_by_xpath(locator.xpath_switch_back).click()
        WebDriverWait(driver, 30).until(
            EC.presence_of_element_located((By.XPATH, locator.xpath_switch_back2)))
        driver.find_element_by_xpath(locator.xpath_switch_back2).click()
    except:
        print("error : Unable to switch back")

def create_reporting_workbook(path):
    workbook = Workbook()
    workbook.save(path+"\\Report.xlsx")
    return workbook

def logger_setup(path):
    LOG_FORMAT = "%(levelname)s %(asctime)s - %(message)s"
    logging.basicConfig(filename=path + "\\Info.log", level=logging.INFO, format=LOG_FORMAT, filemode='w')
    logger = logging.getLogger()
    logger.setLevel(logging.INFO)
    return logger

def switch_to_registries():
    context_url = driver.current_url
    sub_str1 = "/case_management?"
    sub_str2 = "/ehr"
    if context_url.find(sub_str1) > 0 or context_url.find(sub_str2) > 0 :
        try:
            print("Not in registries!")
            WebDriverWait(driver, 30).until(
                EC.invisibility_of_element((By.ID, "toast-container")))
            driver.find_element_by_xpath("//a[@data-target='app_dropdown']").click()
            print("App_dropdown clicked")
            driver.find_element_by_xpath("//a[@title='Registries']").click()
            print("Registries clicked")
            time.sleep(1)
            driver.close()
            driver.switch_to.window(driver.window_handles[0])
            sf.ajax_preloader_wait(driver)
            WebDriverWait(driver, 45).until(
                EC.presence_of_element_located((By.XPATH, locator.xpath_filter_measure_list)))

        except Exception as e:
            traceback.print_exc()
            print(e)
            return

    else:
        print("in registries")
        return

def new_launch(environment):
    print("Entered New Launch")
    report_folder = create_folders("Cozeva Support")
    workbook = create_reporting_workbook(report_folder)
    logger = logger_setup(report_folder)
    if environment == "PROD":
        switch_customer_context(guiwindow.verification_specs[1])
    elif environment == "CERT":
        switch_customer_context_cert(guiwindow.verification_specs[1])
    ws = None
    run_from = "Cozeva Support"
    checklist = guiwindow.verification_specs[4]
    context_functions.init_global_search()
    if checklist[16] == 1:
        context_functions.market_sheet(driver, workbook,logger, run_from)
        workbook.save(report_folder + "\\Report.xlsx")
    if checklist[0] == 1:
        context_functions.support_menubar(driver,workbook, ws, logger, run_from)
        workbook.save(report_folder + "\\Report.xlsx")
    if checklist[1] == 1:
        context_functions.practice_menubar(driver, workbook, logger, run_from)
        workbook.save(report_folder + "\\Report.xlsx")
    if checklist[2] == 1:
        context_functions.provider_menubar(driver, workbook, logger, run_from)
        workbook.save(report_folder + "\\Report.xlsx")
    if checklist[4] == 1:
        context_functions.provider_registry(driver, workbook, logger, run_from)
        workbook.save(report_folder + "\\Report.xlsx")
    if checklist[5] == 1:
        context_functions.practice_registry(driver, workbook, logger, run_from)
        workbook.save(report_folder + "\\Report.xlsx")
    if checklist[24] == 1:
        context_functions.sticket_validation(driver, workbook, logger, report_folder, run_from, guiwindow.verification_specs[1])
        workbook.save(report_folder + "\\Report.xlsx")
    if checklist[28] == 1:
        context_functions.group_menubar(driver, workbook, logger, report_folder, run_from)
        workbook.save(report_folder + "\\Report.xlsx")
    if checklist[15] == 1:
        print("Practice tab ss section")
        context_functions.practice_tab_ss(driver, workbook, logger, report_folder, run_from)
        workbook.save(report_folder + "\\Report.xlsx")
    if checklist[17] == 1:
        context_functions.patient_medication(driver,workbook, logger, report_folder, run_from)
    if checklist[19] == 1:
        context_functions.apptray_access_check(driver, workbook, logger, report_folder, run_from)
        workbook.save(report_folder + "\\Report.xlsx")
    if checklist[20] == 1:
        context_functions.training_resources(driver, workbook, logger, report_folder, run_from)
    if checklist[7] == 1:
        context_functions.global_search(driver, workbook, logger, run_from)
        workbook.save(report_folder + "\\Report.xlsx")
    if checklist[25] == 1:
        context_functions.patient_dashboard(driver, workbook, logger, run_from)
        workbook.save(report_folder + "\\Report.xlsx")
    if checklist[12] == 1:
        context_functions.SupportpageAccordionValidation(driver, workbook, logger, run_from)
        workbook.save(report_folder + "\\Report.xlsx")
    if checklist[18] == 1:
        context_functions.map_codingtool(driver, workbook, logger, run_from, guiwindow.verification_specs[1])
        workbook.save(report_folder + "\\Report.xlsx")

    if checklist[23] == 1:
        context_functions.cetoggle(driver, workbook, logger, report_folder, run_from)


    workbook.save(report_folder + "\\Report.xlsx")

def cozeva_support(environment):
    report_folder = create_folders("Cozeva Support")
    workbook = create_reporting_workbook(report_folder)
    logger = logger_setup(report_folder)
    if environment == "PROD":
        switch_customer_context(guiwindow.verification_specs[1])
    elif environment == "CERT":
        switch_customer_context_cert(guiwindow.verification_specs[1])

    ws = None
    run_from = "Cozeva Support"
    checklist = guiwindow.verification_specs[4]
    context_functions.init_global_search()
    time.sleep(3)
    if checklist[0] == 1:
        context_functions.support_menubar(driver, workbook, ws, logger, run_from)
        workbook.save(report_folder + "\\Report.xlsx")
    if checklist[1] == 1:
        context_functions.practice_menubar(driver, workbook, logger, run_from)
        workbook.save(report_folder + "\\Report.xlsx")
    if checklist[2] == 1:
        context_functions.provider_menubar(driver, workbook, logger, run_from)
        workbook.save(report_folder + "\\Report.xlsx")
    if checklist[3] == 1:
        context_functions.patient_dashboard(driver, workbook, logger, run_from)
        workbook.save(report_folder + "\\Report.xlsx")
    if checklist[4] == 1:
        context_functions.provider_registry(driver, workbook, logger, run_from)
        workbook.save(report_folder + "\\Report.xlsx")
    if checklist[5] == 1:
        context_functions.practice_registry(driver, workbook, logger, run_from)
        workbook.save(report_folder + "\\Report.xlsx")
    if checklist[6] == 1:
        context_functions.support_level(driver, workbook, logger, run_from)
        workbook.save(report_folder + "\\Report.xlsx")
    if checklist[7] == 1:
        context_functions.global_search(driver, workbook, logger, run_from)
        workbook.save(report_folder + "\\Report.xlsx")
    if checklist[8] == 1:
        context_functions.provider_mspl(driver, workbook, logger, run_from)
        workbook.save(report_folder + "\\Report.xlsx")
    if checklist[9] == 1:
        context_functions.time_capsule(driver, workbook, logger, run_from)
        workbook.save(report_folder + "\\Report.xlsx")
    if checklist[10] == 1:
        context_functions.secure_messaging(driver, workbook, logger, run_from)
        workbook.save(report_folder + "\\Report.xlsx")
    if checklist[11] == 1:
        context_functions.analytics(driver, workbook, logger, run_from)
        workbook.save(report_folder + "\\Report.xlsx")
    if checklist[12] == 1:
        context_functions.SupportpageAccordionValidation(driver, workbook, logger, run_from)
        workbook.save(report_folder + "\\Report.xlsx")

    workbook.save(report_folder+"\\Report.xlsx")
    ss.summarize_report(workbook, report_folder)
    workbook.save(report_folder + "\\Report.xlsx")


def limited_cozeva_support(username):
    report_folder = create_folders("Limited Cozeva Support")
    workbook = create_reporting_workbook(report_folder)
    logger = logger_setup(report_folder)
    login_to_user(username)
    switch_to_registries()
    run_from = "Limited Cozeva Support"
    ws = None
    context_functions.init_global_search()
    checklist = guiwindow.verification_specs[4]
    if checklist[0] == 1:
        context_functions.support_menubar(driver, workbook, ws, logger, run_from)
        workbook.save(report_folder + "\\Report.xlsx")
    if checklist[1] == 1:
        context_functions.practice_menubar(driver, workbook, logger, run_from)
        workbook.save(report_folder + "\\Report.xlsx")
    if checklist[2] == 1:
        context_functions.provider_menubar(driver, workbook, logger, run_from)
        workbook.save(report_folder + "\\Report.xlsx")
    if checklist[3] == 1:
        context_functions.patient_dashboard(driver, workbook, logger, run_from)
        workbook.save(report_folder + "\\Report.xlsx")
    if checklist[4] == 1:
        context_functions.provider_registry(driver, workbook, logger, run_from)
        workbook.save(report_folder + "\\Report.xlsx")
    if checklist[5] == 1:
        context_functions.practice_registry(driver, workbook, logger, run_from)
        workbook.save(report_folder + "\\Report.xlsx")
    if checklist[6] == 1:
        context_functions.support_level(driver, workbook, logger, run_from)
        workbook.save(report_folder + "\\Report.xlsx")
    if checklist[7] == 1:
        context_functions.global_search(driver, workbook, logger, run_from)
        workbook.save(report_folder + "\\Report.xlsx")
    if checklist[8] == 1:
        context_functions.provider_mspl(driver, workbook, logger, run_from)
        workbook.save(report_folder + "\\Report.xlsx")
    if checklist[9] == 1:
        context_functions.time_capsule(driver, workbook, logger, run_from)
        workbook.save(report_folder + "\\Report.xlsx")
    if checklist[10] == 1:
        context_functions.secure_messaging(driver, workbook, logger, run_from)
        workbook.save(report_folder + "\\Report.xlsx")
    if checklist[12] == 1:
        context_functions.SupportpageAccordionValidation(driver, workbook, logger, run_from)
        workbook.save(report_folder + "\\Report.xlsx")
    time.sleep(5)
    workbook.save(report_folder + "\\Report.xlsx")

    switch_back()
    ss.summarize_report(workbook, report_folder)
    workbook.save(report_folder + "\\Report.xlsx")

def customer_support(username):
    report_folder = create_folders("Customer Support")
    workbook = create_reporting_workbook(report_folder)
    logger = logger_setup(report_folder)
    login_to_user(username)
    switch_to_registries()
    run_from = "Customer Support"
    ws = None
    context_functions.init_global_search()
    checklist = guiwindow.verification_specs[4]
    if checklist[0] == 1:
        context_functions.support_menubar(driver, workbook, ws, logger, run_from)
        workbook.save(report_folder + "\\Report.xlsx")
    if checklist[1] == 1:
        context_functions.practice_menubar(driver, workbook, logger, run_from)
        workbook.save(report_folder + "\\Report.xlsx")
    if checklist[2] == 1:
        context_functions.provider_menubar(driver, workbook, logger, run_from)
        workbook.save(report_folder + "\\Report.xlsx")
    if checklist[3] == 1:
        context_functions.patient_dashboard(driver, workbook, logger, run_from)
        workbook.save(report_folder + "\\Report.xlsx")
    if checklist[4] == 1:
        context_functions.provider_registry(driver, workbook, logger, run_from)
        workbook.save(report_folder + "\\Report.xlsx")
    if checklist[5] == 1:
        context_functions.practice_registry(driver, workbook, logger, run_from)
        workbook.save(report_folder + "\\Report.xlsx")
    if checklist[6] == 1:
        context_functions.support_level(driver, workbook, logger, run_from)
        workbook.save(report_folder + "\\Report.xlsx")
    if checklist[7] == 1:
        context_functions.global_search(driver, workbook, logger, run_from)
        workbook.save(report_folder + "\\Report.xlsx")
    if checklist[8] == 1:
        context_functions.provider_mspl(driver, workbook, logger, run_from)
        workbook.save(report_folder + "\\Report.xlsx")
    if checklist[9] == 1:
        context_functions.time_capsule(driver, workbook, logger, run_from)
        workbook.save(report_folder + "\\Report.xlsx")
    if checklist[10] == 1:
        context_functions.secure_messaging(driver, workbook, logger, run_from)
        workbook.save(report_folder + "\\Report.xlsx")
    if checklist[12] == 1:
        context_functions.SupportpageAccordionValidation(driver, workbook, logger, run_from)
        workbook.save(report_folder + "\\Report.xlsx")
    time.sleep(5)
    workbook.save(report_folder + "\\Report.xlsx")
    switch_back()
    ss.summarize_report(workbook, report_folder)
    workbook.save(report_folder + "\\Report.xlsx")

def regional_suport(username):
    report_folder = create_folders("Regional Support")
    workbook = create_reporting_workbook(report_folder)
    logger = logger_setup(report_folder)
    login_to_user(username)
    switch_to_registries()
    run_from = "Regional Support"
    ws = None
    context_functions.init_global_search()
    checklist = guiwindow.verification_specs[4]
    if checklist[0] == 1:
        context_functions.support_menubar(driver, workbook, ws, logger, run_from)
        workbook.save(report_folder + "\\Report.xlsx")
    if checklist[1] == 1:
        context_functions.practice_menubar(driver, workbook, logger, run_from)
        workbook.save(report_folder + "\\Report.xlsx")
    if checklist[2] == 1:
        context_functions.provider_menubar(driver, workbook, logger, run_from)
        workbook.save(report_folder + "\\Report.xlsx")
    if checklist[3] == 1:
        context_functions.patient_dashboard(driver, workbook, logger, run_from)
        workbook.save(report_folder + "\\Report.xlsx")
    if checklist[4] == 1:
        context_functions.provider_registry(driver, workbook, logger, run_from)
        workbook.save(report_folder + "\\Report.xlsx")
    if checklist[5] == 1:
        context_functions.practice_registry(driver, workbook, logger, run_from)
        workbook.save(report_folder + "\\Report.xlsx")
    if checklist[6] == 1:
        context_functions.support_level(driver, workbook, logger, run_from)
        workbook.save(report_folder + "\\Report.xlsx")
    if checklist[7] == 1:
        context_functions.global_search(driver, workbook, logger, run_from)
        workbook.save(report_folder + "\\Report.xlsx")
    if checklist[8] == 1:
        context_functions.provider_mspl(driver, workbook, logger, run_from)
        workbook.save(report_folder + "\\Report.xlsx")
    if checklist[10] == 1:
        context_functions.secure_messaging(driver, workbook, logger, run_from)
        workbook.save(report_folder + "\\Report.xlsx")
    if checklist[12] == 1:
        context_functions.SupportpageAccordionValidation(driver, workbook, logger, run_from)
        workbook.save(report_folder + "\\Report.xlsx")
    time.sleep(5)
    workbook.save(report_folder + "\\Report.xlsx")
    switch_back()
    ss.summarize_report(workbook, report_folder)
    workbook.save(report_folder + "\\Report.xlsx")

def office_admin_Prac(username):
    report_folder = create_folders("Office Admin Practice Delegate")
    workbook = create_reporting_workbook(report_folder)
    logger = logger_setup(report_folder)
    login_to_user(username)
    switch_to_registries()
    run_from = "Office Admin Practice Delegate"
    ws = None
    context_functions.init_global_search()
    checklist = guiwindow.verification_specs[4]
    if checklist[1] == 1:
        context_functions.practice_menubar(driver, workbook, logger, run_from)
        workbook.save(report_folder + "\\Report.xlsx")
    if checklist[2] == 1:
        context_functions.provider_menubar(driver, workbook, logger, run_from)
        workbook.save(report_folder + "\\Report.xlsx")
    if checklist[3] == 1:
        context_functions.patient_dashboard(driver, workbook, logger, run_from)
        workbook.save(report_folder + "\\Report.xlsx")
    if checklist[4] == 1:
        context_functions.provider_registry(driver, workbook, logger, run_from)
        workbook.save(report_folder + "\\Report.xlsx")
    if checklist[5] == 1:
        context_functions.practice_registry(driver, workbook, logger, run_from)
        workbook.save(report_folder + "\\Report.xlsx")
    if checklist[7] == 1:
        context_functions.global_search(driver, workbook, logger, run_from)
        workbook.save(report_folder + "\\Report.xlsx")
    if checklist[8] == 1:
        context_functions.provider_mspl(driver, workbook, logger, run_from)
        workbook.save(report_folder + "\\Report.xlsx")
    if checklist[10] == 1:
        context_functions.secure_messaging(driver, workbook, logger, run_from)
        workbook.save(report_folder + "\\Report.xlsx")
    time.sleep(5)
    workbook.save(report_folder + "\\Report.xlsx")
    switch_back()
    ss.summarize_report(workbook, report_folder)
    workbook.save(report_folder + "\\Report.xlsx")

def office_admin_prov(username):
    report_folder = create_folders("Office Admin Provider Delegate")
    workbook = create_reporting_workbook(report_folder)
    logger = logger_setup(report_folder)
    login_to_user(username)
    switch_to_registries()
    run_from = "Office Admin Provider Delegate"
    ws = None
    context_functions.init_global_search()
    checklist = guiwindow.verification_specs[4]
    if checklist[2] == 1:
        context_functions.provider_menubar(driver, workbook, logger, run_from)
        workbook.save(report_folder + "\\Report.xlsx")
    if checklist[3] == 1:
        context_functions.patient_dashboard(driver, workbook, logger, run_from)
        workbook.save(report_folder + "\\Report.xlsx")
    if checklist[4] == 1:
        context_functions.provider_registry(driver, workbook, logger, run_from)
        workbook.save(report_folder + "\\Report.xlsx")
    if checklist[7] == 1:
        context_functions.global_search(driver, workbook, logger, run_from)
        workbook.save(report_folder + "\\Report.xlsx")
    if checklist[8] == 1:
        context_functions.provider_mspl(driver, workbook, logger, run_from)
        workbook.save(report_folder + "\\Report.xlsx")
    if checklist[10] == 1:
        context_functions.secure_messaging(driver, workbook, logger, run_from)
        workbook.save(report_folder + "\\Report.xlsx")
    time.sleep(5)
    workbook.save(report_folder + "\\Report.xlsx")
    switch_back()
    ss.summarize_report(workbook, report_folder)
    workbook.save(report_folder + "\\Report.xlsx")

def prov(username):
    report_folder = create_folders("Provider")
    workbook = create_reporting_workbook(report_folder)
    logger = logger_setup(report_folder)
    login_to_user(username)
    switch_to_registries()
    run_from = "Provider"
    ws = None
    context_functions.init_global_search()
    checklist = guiwindow.verification_specs[4]
    if checklist[2] == 1:
        context_functions.provider_menubar(driver, workbook, logger, run_from)
        workbook.save(report_folder + "\\Report.xlsx")
    if checklist[3] == 1:
        context_functions.patient_dashboard(driver, workbook, logger, run_from)
        workbook.save(report_folder + "\\Report.xlsx")
    if checklist[4] == 1:
        context_functions.provider_registry(driver, workbook, logger, run_from)
        workbook.save(report_folder + "\\Report.xlsx")
    if checklist[7] == 1:
        context_functions.global_search(driver, workbook, logger, run_from)
        workbook.save(report_folder + "\\Report.xlsx")
    if checklist[8] == 1:
        context_functions.provider_mspl(driver, workbook, logger, run_from)
        workbook.save(report_folder + "\\Report.xlsx")
    if checklist[10] == 1:
        context_functions.secure_messaging(driver, workbook, logger, run_from)
        workbook.save(report_folder + "\\Report.xlsx")
    time.sleep(5)
    workbook.save(report_folder + "\\Report.xlsx")
    switch_back()
    ss.summarize_report(workbook, report_folder)
    workbook.save(report_folder + "\\Report.xlsx")

