#!/usr/bin/env python
# coding: utf-8

# In[105]:
import pickle
import variablestorage as vs
# In[105]:

import setups

import os
import pickle
import random
from statistics import mean

import pandas as pd
import sys
import math
import time
import traceback
import csv
from os import listdir
from os.path import isfile, join

import datetime
import tkinter as tk
from tkinter import ttk

from selenium import *
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import *
from selenium.webdriver.support.wait import WebDriverWait
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from openpyxl.formatting import Rule
from openpyxl.styles import *
from openpyxl.styles.differential import DifferentialStyle
from selenium.webdriver.common.keys import Keys
from PIL import ImageTk, Image
from tkinter import messagebox



# In[71]:

# def SelectedChecklist_GUI():
#     def select_all():
#         for checkbox in checkboxes.values():
#             checkbox.set(True)
#
#     def deselect_all():
#         for checkbox in checkboxes.values():
#             checkbox.set(False)
#
#     def get_selected_items():
#         selected_ids = [item_id for item_id, var in checkboxes.items() if var.get() == 1]
#         selected_items = [items_with_ids[item_id] for item_id in selected_ids]
#         print("Selected IDs:", selected_ids)
#         print("Selected measures:", selected_items)
#         root.destroy()  # Close the GUI after getting selected items
#         return selected_ids
#
#     # Sample list of items
#     items_with_ids = {
#         400: "Review of Chronic Conditions (Blended) [NA for USRC]",
#         551: "Review of Chronic Conditions (Risk Adjustment Version 24)",
#         552: "Review of Suspect conditions",
#         553: "HCC Score(Blended)",
#         554: "Review of ACA Chronic Conditions",
#         555: "Review of ACA Suspect Conditions",
#         556: "ACA HCC Score",
#         557: "ID: 557"
#     }
#
#     root = tk.Tk()
#     root.title("Please select the measures")
#
#     # Create a frame to hold the checkboxes
#     checkbox_frame = ttk.Frame(root)
#     checkbox_frame.pack(pady=10)
#
#     # Create IntVar variables for each checkbox
#     checkboxes = {item_id: tk.IntVar() for item_id in items_with_ids.keys()}
#
#     # Create Checkbuttons for each item
#     for i, (item_id, item) in enumerate(items_with_ids.items()):
#         checkbox = ttk.Checkbutton(checkbox_frame, text=item, variable=checkboxes[item_id])
#         checkbox.grid(row=i // 2, column=i % 2, padx=5, pady=2, sticky=tk.W)
#
#     # Create a frame to hold the buttons
#     button_frame = ttk.Frame(root)
#     button_frame.pack(pady=5)
#
#     # Create buttons
#     select_all_button = ttk.Button(button_frame, text="Select All", command=select_all)
#     select_all_button.grid(row=0, column=0, padx=5, pady=5)
#
#     deselect_all_button = ttk.Button(button_frame, text="Deselect All", command=deselect_all)
#     deselect_all_button.grid(row=0, column=1, padx=5, pady=5)
#
#     get_selected_button = ttk.Button(root, text="Run for selected measures", command=get_selected_items)
#     get_selected_button.pack(pady=10)
#
#     # Center the window on the screen
#     root.update_idletasks()
#     window_width = root.winfo_width()
#     window_height = root.winfo_height()
#     screen_width = root.winfo_screenwidth()
#     screen_height = root.winfo_screenheight()
#     x_coordinate = (screen_width - window_width) // 2
#     y_coordinate = (screen_height - window_height) // 2
#     root.geometry(f"{window_width}x{window_height}+{x_coordinate}+{y_coordinate}")
#
#     root.mainloop()
#


def ajax_preloader_wait(driver):
    time.sleep(2)
    WebDriverWait(driver, 300).until(EC.invisibility_of_element((By.CLASS_NAME, "ajax_preloader")))
    if len(driver.find_elements(By.CLASS_NAME, "ajax_preloader")) != 0:
        WebDriverWait(driver, 300).until(EC.invisibility_of_element((By.CLASS_NAME, "ajax_preloader")))
    WebDriverWait(driver, 300).until(EC.invisibility_of_element((By.CLASS_NAME, "drupal_message_text")))
    time.sleep(2)


def SheetColorCoder(sheet, workbook, path, filename):
    rows = sheet.max_row  # Starting of code to color code excel
    cols = sheet.max_column
    for i in range(2, rows + 1):
        for j in range(3, cols + 1):
            if sheet.cell(i, j).value == 'Passed' or sheet.cell(i, j).value == 'Present and Passed' or "is matching" in sheet.cell(i, j).value or "Successfully" in sheet.cell(i, j).value or "are within" in sheet.cell(i, j).value or "is present" in sheet.cell(i, j).value:
                sheet.cell(i, j).fill = PatternFill('solid', fgColor='BAD366')
            elif sheet.cell(i, j).value == 'Failed' or sheet.cell(i, j).value == 'Present but Failed' or "is not matching" in sheet.cell(i, j).value or "are not within" in sheet.cell(i, j).value or "Issue" in sheet.cell(i, j).value:
                sheet.cell(i, j).fill = PatternFill('solid', fgColor='FF707A')
            elif sheet.cell(i, j).value == 'Unexecuted' or "Warning" in sheet.cell(i, j).value or sheet.cell(i, j).value == 'Skipped':
                sheet.cell(i, j).fill = PatternFill('solid', fgColor='FCD44D')
            elif sheet.cell(i, j).value == 'Present but not calculated' or sheet.cell(i, j).value == 'WAD':
                sheet.cell(i, j).fill = PatternFill('solid', fgColor='49B99C')
            elif sheet.cell(i, j).value == 'Not present':
                sheet.cell(i, j).fill = PatternFill('solid', fgColor='FCC0BB')
    workbook.save(path + filename)


# In[73]:


def PatientDashboardScoreCheck(driver, patient_count, ws, wb, LOB_Name, Measure):
    pro_count = 1
    name = driver.find_element(By.XPATH, "//span[@class='specific_most']").text

    while pro_count != 0:
        demographic_comment = "Unexecuted"
        actual_raf_comment = "Unexecuted"
        potential_raf_comment = "Unexecuted"
        coded_hcc_comment = "Unexecuted"
        potential_hcc_comment = "Unexecuted"
        patient_url = '-'
        pat_count = patient_count
        pro_count = pro_count - 1
        WebDriverWait(driver, 6000).until(EC.presence_of_element_located((By.XPATH, "//*[@id='quality_registry_list']/tbody")))
        try:
            table = driver.find_element(By.ID, "quality_registry_list").find_element(By.TAG_NAME, "tbody").find_elements(By.TAG_NAME, 'tr')
            if len(table) < pat_count:
                pat_count = len(table)
                print("Not enough patients in MSPL")
            else:
                pat_count = patient_count
            try:
                if len(table) == 1 and "No data" in table[0].text:
                    print("No patients for this provider")
                    ws.append([name, LOB_Name, Measure, "Unexecuted", "Unexecuted", "Unexecuted", "Unexecuted", "Unexecuted", "MSPL is blank for this provider", driver.current_url])
                    break
                else:
                    while pat_count != 0:
                        mspl_actual_raf = 0.000
                        mspl_clinical_raf = 0.000
                        mspl_potential_raf = 0.000
                        demographic_calculated = 0.000
                        potential_hcc_calculated = 0.000
                        dashboard_demographic = 0.000
                        dashboard_actual_raf = 0.000
                        dashboard_potential_raf = 0.000
                        dashboard_coded_hcc = 0.000
                        dashboard_potential_hcc = 0.000
                        pat_count = pat_count - 1
                        selected_patient = table[random.randint(0, len(table) - 1)]
                        driver.execute_script("arguments[0].scrollIntoView();", selected_patient)
                        if selected_patient.find_element(By.CLASS_NAME, "risk_score_gap ").text == "NA":
                            patient_url = selected_patient.find_element(By.TAG_NAME, 'a').get_attribute(
                                'href')
                            ws.append([name, LOB_Name, Measure, "Unexecuted", "Unexecuted", "Unexecuted", "Unexecuted", "Unexecuted", "MSPL HCC value is NA so no risk score should not be present", patient_url])
                            continue
                        else:
                            mspl_actual_raf = float(
                                selected_patient.find_element(By.CLASS_NAME, "actual_raf ").text)
                            mspl_clinical_raf = float(
                                selected_patient.find_element(By.CLASS_NAME, "clinical_raf ").text)
                            mspl_potential_raf = float(
                                selected_patient.find_element(By.CLASS_NAME, "potential_raf ").text)
                            demographic_calculated = mspl_actual_raf - mspl_clinical_raf
                            demographic_calculated = round(demographic_calculated, 3)
                            potential_hcc_calculated = mspl_potential_raf - demographic_calculated
                            potential_hcc_calculated = round(potential_hcc_calculated, 3)
                            time.sleep(1)
                            patient_link = selected_patient.find_element(By.TAG_NAME, 'a')
                            driver.execute_script("arguments[0].click();", patient_link)
                            time.sleep(1)
                            driver.switch_to.window(driver.window_handles[1])
                            ajax_preloader_wait(driver)
                            WebDriverWait(driver, 6000).until(
                                EC.presence_of_element_located((By.CLASS_NAME, "patient_header_wrapper")))
                            patient_url = driver.current_url
                            print("Patient Dashboard URL: " + driver.current_url)
                            score = driver.find_element(By.XPATH, "//*[@data-tooltip='Demographic RAF']").text
                            if " -" not in score:
                                dashboard_demographic = float(
                                    driver.find_element(By.XPATH, "//*[@data-tooltip='Demographic RAF']").text)
                                dashboard_actual_raf = float(
                                    driver.find_element(By.XPATH, "//*[@data-tooltip='Actual RAF']").text)
                                dashboard_potential_raf = float((driver.find_element(By.XPATH,
                                                                                     "//*[@data-tooltip='Potential RAF']").text.replace(
                                    "/ ", "")))
                                dashboard_coded_hcc = float(
                                    driver.find_element(By.XPATH, "//*[@data-tooltip='Coded HCC']").text)
                                dashboard_potential_hcc = float((driver.find_element(By.XPATH,
                                                                                     "//*[@data-tooltip='Potential HCC']").text.replace(
                                    "/ ", "")))
                                if demographic_calculated == dashboard_demographic:
                                    demographic_comment = "Passed"
                                else:
                                    demographic_comment = "Failed"
                                if mspl_actual_raf == dashboard_actual_raf:
                                    actual_raf_comment = "Passed"
                                else:
                                    actual_raf_comment = "Failed"
                                if mspl_potential_raf == dashboard_potential_raf:
                                    potential_raf_comment = "Passed"
                                else:
                                    potential_raf_comment = "Failed"
                                if mspl_clinical_raf == dashboard_coded_hcc:
                                    coded_hcc_comment = "Passed"
                                else:
                                    coded_hcc_comment = "Failed"
                                if potential_hcc_calculated == dashboard_potential_hcc:
                                    potential_hcc_comment = "Passed"
                                else:
                                    potential_hcc_comment = "Failed"
                                print("Patient Dashboard vs MSPL details")
                                print("MSPL actual raf: " + str(mspl_actual_raf))
                                print("MSPL clinical raf: " + str(mspl_clinical_raf))
                                print("MSPL potential raf: " + str(mspl_potential_raf))
                                print("Calculated demographic score: " + str(demographic_calculated))
                                print("Calculated potential HCC: " + str(demographic_calculated))
                                print("Dashboard actual raf: " + str(dashboard_actual_raf))
                                print("Dashboard potential raf: " + str(dashboard_potential_raf))
                                print("Dashboard coded HCC: " + str(dashboard_coded_hcc))
                                print("Dashboard potential HCC: " + str(dashboard_potential_hcc))
                                print("Dashboard demographic score: " + str(dashboard_demographic))
                                ws.append([name, LOB_Name, Measure, actual_raf_comment, potential_raf_comment, coded_hcc_comment, potential_hcc_comment, demographic_comment, "Patient count adjusted accordingly", patient_url])
                                SheetColorCoder(ws, wb, path1, filename)
                                driver.close()
                                driver.switch_to.window(driver.window_handles[0])
                                time.sleep(2)
                                table = driver.find_element(By.ID, "quality_registry_list").find_element(
                                    By.TAG_NAME, "tbody").find_elements(By.TAG_NAME, 'tr')
                            else:
                                ws.append([name, LOB_Name, Measure, "Skipped", "Skipped", "Skipped", "Skipped", "Skipped", 'Skipped because NULL HCC value', patient_url])
                                SheetColorCoder(ws, wb, path1, filename)
                                driver.close()
                                driver.switch_to.window(driver.window_handles[0])
                                time.sleep(2)
                                table = driver.find_element(By.ID, "quality_registry_list").find_element(By.TAG_NAME, "tbody").find_elements(By.TAG_NAME, 'tr')
            except (NoSuchElementException, ElementNotInteractableException, ElementClickInterceptedException, ElementNotVisibleException, TimeoutException, ElementNotSelectableException) as e:
                print(e)
                ws.append([name, LOB_Name, Measure, actual_raf_comment, potential_raf_comment, coded_hcc_comment, potential_hcc_comment, demographic_comment, "Error encountered", patient_url])
        except (NoSuchElementException, ElementNotInteractableException, ElementClickInterceptedException, ElementNotVisibleException, TimeoutException, ElementNotSelectableException) as e:
            ws.append([name, LOB_Name, Measure, '-', "Unexecuted", "Unexecuted", "Unexecuted", "Unexecuted", "Unexecuted", "No Measure", '-'])


def addition(total, row, switch, arr):
    try:
        for i in range(len(arr[1])):
            if switch == 1:
                if str(arr[1][i]).endswith("Gaps"):
                    total = total + int(arr[row][i])
            if switch == 2:
                if str(arr[1][i]).endswith("Conditions"):
                    total = total + int(arr[row][i])
            if switch == 3:
                if str(arr[1][i]).endswith("Disconfirms"):
                    total = total + int(arr[row][i])
            if switch == 4:
                if str(arr[1][i]).endswith("Clinical RAF"):
                    total = total + float(arr[row][i])
            if switch == 5:
                if str(arr[1][i]).endswith("Potential RAF"):
                    total = total + float(arr[row][i])
            if switch == 6:
                if str(arr[1][i]).endswith("Coded RAF"):
                    total = total + float(arr[row][i])
            if switch == 7:
                if str(arr[1][i]).endswith("Numerator"):
                    total = total + float(arr[row][i])
            if switch == 8:
                if str(arr[1][i]).endswith("Denominator"):
                    total = total + float(arr[row][i])
        return total
    except ValueError:
        print("No number detected in " + arr[1][i] + " .For row no = " + str(row + 1))
        return total

def NumDenAddition(path):
    import csv
    with open(path, newline='') as csvfile:
        rows = csv.reader(csvfile, delimiter=',')
        rows = list(rows)
    total_num = 0
    total_den = 0
    for i in range(2, len(rows)):
        total_num = addition(total_num, i, 7, rows)
        total_den = addition(total_den, i, 8, rows)
    return total_num, total_den

def csvAddition(filepath):
    import csv
    with open(filepath, newline='') as csvfile:
        rows = csv.reader(csvfile, delimiter=',')
        rows = list(rows)
    Gaps = 0
    Conditions = 0
    Disconfirms = 0
    Clinical = 0
    Potential = 0
    Coded = 0
    for ind in range(2, len(rows)):
        Gaps = addition(Gaps, ind, 1, rows)
        Conditions = addition(Conditions, ind, 2, rows)
        Disconfirms = addition(Disconfirms, ind, 3, rows)
        Clinical = addition(Clinical, ind, 4, rows)
        Potential = addition(Potential, ind, 5, rows)
        Coded = addition(Coded, ind, 6, rows)
    return Gaps, Conditions, Disconfirms, Clinical, Potential, Coded, (len(rows) - 2)

def MSPLScoreCheck(driver,ws, wb, LOB_Name, Measure):
    suspect_list = []
    clinical_score_list = []
    tag_count_list = [0, 0, 0]
    name = driver.find_element(By.XPATH, "//span[@class='specific_most']").text
    present_url = driver.current_url
    list1 = ["HCC Score", "ACA HCC Score", "RAF Score", "RAF Score (Version 28)", "RAF Score (ESRD Model)"]
    list2 = ["One-Year ACA HCC Recapture Rate", "Review of ACA Chronic Condition", "One-Year ACA HCC Suspect Rate", "Review of ACA Suspect Condition"]
    search_var = Measure.split(' | ')[1]

    Domain_name_MSPL = driver.find_element(By.XPATH, "//*[@class='ch metric_specific_patient_list_title']").text
    print("Domain Name MSPL page: " + Domain_name_MSPL)
    if Measure == Domain_name_MSPL:
        Domain_comment = "Passed"
    else:
        Domain_comment = "Failed"
    DataToBeValidated = driver.find_element(By.XPATH, "//*[@class='tab']").find_elements(By.TAG_NAME, 'span')
    Provider_Specific_url = driver.current_url
    print("Provider URL: " + Provider_Specific_url)
    DataToBeValidated_num = DataToBeValidated[0].text
    DataToBeValidated_num = DataToBeValidated_num.replace(',', '')
    print("MSPL Numerator: " + DataToBeValidated_num)
    DataToBeValidated_denum = DataToBeValidated[1].text
    DataToBeValidated_denum = DataToBeValidated_denum.replace(',', '')
    print("MSPL Denominator: " + DataToBeValidated_denum)
    driver.find_element(By.XPATH,
                        "//*[@data-target='datatable_bulk_filter_0_quality_registry_list']").click()
    driver.find_element(By.XPATH, "//*[contains(text(),'Export all to CSV')]").click()
    time.sleep(10)
    ajax_preloader_wait(driver)
    try:
        driver.find_element(By.XPATH, "//*[@class='tabs']").find_elements(By.TAG_NAME, 'li')[1].click()
        ajax_preloader_wait(driver)
        Performance_percentage_UI = driver.find_element(By.XPATH, "//*[@class='performance_value']").text
        Performance_percentage_UI = Performance_percentage_UI.replace('%', '')
        Performance_num_UI = driver.find_element(By.XPATH, "//*[@class='numerator']").text
        Performance_num_UI = Performance_num_UI.replace('Numerator: ', '')
        Performance_denum_UI = driver.find_element(By.XPATH, "//*[@class='denominator']").text
        Performance_denum_UI = Performance_denum_UI.replace('Denominator: ', '')
        Performance_percentage_calculated = round((float(Performance_num_UI) / float(Performance_denum_UI)) * 100, 4)
        print("Performance Tab Numerator: " + Performance_num_UI)
        print("Performance Tab Denominator: " + Performance_denum_UI)
        print("Performance Tab Percentage: " + Performance_percentage_UI)
        print("Performance calculated: " + str(Performance_percentage_calculated))
        if (float(Performance_percentage_UI) - float(Performance_percentage_calculated)) < 0.02:
            Performance_comment = "Passed"
        else:
            Performance_comment = "Failed"
    except (NoSuchElementException, ElementNotInteractableException, ElementClickInterceptedException, ElementNotVisibleException, TimeoutException, ElementNotSelectableException):
        Performance_comment = "The Performance tab is not clickable"
    try:
        driver.find_element(By.XPATH, "//*[@class='tabs']").find_elements(By.TAG_NAME, 'li')[2].click()
        ajax_preloader_wait(driver)
        if EC.presence_of_element_located((By.XPATH, "//*[@id='network_comparison_chart']")):
            Network_comment = "Passed"
        else:
            Network_comment = "Failed"
    except (NoSuchElementException, ElementNotInteractableException, ElementClickInterceptedException, ElementNotVisibleException, TimeoutException, ElementNotSelectableException) as e:
        Network_comment = "The Network tab is not clickable"
        traceback.print_exc()
        print(e)
    onlyfiles = [f for f in listdir("assets\\Download\\CSV_Files_1") if isfile(join("assets\\Download\\CSV_Files_1", f))]
    path = "assets\\Download\\CSV_Files_1\\" + onlyfiles[0]
    result = csvAddition(path)
    print("Total Gap Count: " + str(result[0]))
    print("Total Condition Count: " + str(result[1]))
    print("Total Disconfirm Count: " + str(result[2]))
    print("Total Clinical RAF Score: " + str(result[3]))
    print("Total Potential Score: " + str(result[4]))
    print("Total Coded Score: " + str(result[5]))
    print("Total Patient Count: " + str(result[6]))
    os.remove(path)
    # comments = "Suspects HCCs: %s, Recaptures HCCs: %s, New HCCs: %s" % (str(tag_count_list[0]), str(tag_count_list[1]), str(tag_count_list[2]))
    comments = "-"

    if search_var in list1:
        DataToBeValidated_num = float(DataToBeValidated_denum) - float(DataToBeValidated_num)
        DataToBeValidated_num = round(DataToBeValidated_num, 3)
        DataToBeValidated_denum = round(float(DataToBeValidated_denum), 3)
        num = float(result[3] / result[6])
        num = round(num, 3)
        temp = float((result[5] - result[3]) / result[6])
        denum = float(result[4] / result[6])
        denum = denum - temp
        denum = round(denum, 3)
        if abs(float(DataToBeValidated_num) - num) < 0.015 and abs(float(DataToBeValidated_denum) - denum) < 0.015:
            ws.append([name, LOB_Name, Measure, Domain_comment, Performance_comment, Network_comment, 'Passed', comments, Provider_Specific_url])
        else:
            ws.append([name, LOB_Name, Measure, Domain_comment, Performance_comment, Network_comment, 'Failed', comments, Provider_Specific_url])
    elif search_var in list2:
        if int(DataToBeValidated_num) == int(DataToBeValidated_denum) - int(result[1]) - int(result[2]):
            ws.append([name, LOB_Name, Measure, Domain_comment, Performance_comment, Network_comment, 'Passed', comments, Provider_Specific_url])
        else:
            ws.append([name, LOB_Name, Measure, Domain_comment, Performance_comment, Network_comment, 'Failed', comments, Provider_Specific_url])
    else:
        if float(DataToBeValidated_num) == float(result[0]) and float(DataToBeValidated_denum) == float(
                result[0] + result[1] + result[2]):
            ws.append([name, LOB_Name, Measure, Domain_comment, Performance_comment, Network_comment, 'Passed', comments, Provider_Specific_url])
        else:
            ws.append([name, LOB_Name, Measure, Domain_comment, Performance_comment, Network_comment, 'Failed', comments, Provider_Specific_url])

    driver.get(present_url)

# In[95]:


def NameURLextractor(driver, LOB, checklist, URL_name_list, num_den_list, customer):
    HCC_measure_checklist = checklist
    for i in HCC_measure_checklist:
        try:
            Measure_Specific_url = driver.find_element(By.XPATH, "//*[@id=" + str(i) + "]//a").get_attribute('href')
            Measure = driver.find_element(By.XPATH, "//*[@id=" + str(i) + "]//*[@class='met-name top']").text
            Domain_name_registry = driver.find_element(By.XPATH, "//*[@id=" + str(i) + "]/..//*[@class='group-name-wrapper']").text
            Measure = Domain_name_registry + " | " + Measure
            URL_name_list.append(Measure_Specific_url)
            URL_name_list.append(Measure)
            URL_name_list.append(LOB)
        except Exception as e:
                print("No measure of id "+ str(i))
    return URL_name_list, num_den_list


# In[86]:


def TabNavigator(driver, Tabs, ws, wb, measure, LOB, num_den_list, provider_count, id_list):
    name = driver.find_element(By.XPATH, "//span[@class='specific_most']").text
    Tabs = driver.find_element(By.XPATH, "//*[@id='qt-mt-support-ls']").find_elements(By.TAG_NAME, 'li')
    practice_link = ""
    provider_link = []
    pro_count = provider_count

    for i in range(0, len(Tabs) - 1):
        comments = ""
        try:
            Tabs[i].click()
            ajax_preloader_wait(driver)
            WebDriverWait(driver, 300).until(EC.presence_of_element_located((By.CLASS_NAME, 'tab')))
            field = driver.find_element(By.XPATH, "//*[@class='handler active']").text
            print(field)
            if field == "Practices":
                try:
                    ListRow = driver.find_element(By.XPATH, "//*[@id='metric-support-prac-ls']").find_element(By.TAG_NAME, "tbody").find_elements(By.TAG_NAME, 'tr')
                    if "No data available" in ListRow[0].text and len(ListRow) == 1:
                        comments = "No data available in practice tab"
                    else:
                        # Row = ListRow[random.randint(0, len(ListRow) - 1)]
                        Row = ListRow[0]
                        practice_link = Row.find_elements(By.TAG_NAME, 'a')[1].get_attribute('href')
                        comments = driver.find_element(By.XPATH, "//*[@class='dataTables_info']").text
                except NoSuchElementException as e:
                    comments = "Tab faced an error while opening"
            if field == "Providers":
                try:
                    ListRow = driver.find_element(By.XPATH, "//*[@id='metric-support-prov-ls']").find_element(By.TAG_NAME,
                                                                                                              "tbody").find_elements(
                        By.TAG_NAME, 'tr')
                    if "No data available" in ListRow[0].text and len(ListRow) == 1:
                        comments = "No data available in provider tab"
                    else:
                        if len(ListRow) < pro_count:
                            pro_count = len(ListRow)
                            print("Has less number of providers than specified")
                        while pro_count > 0:
                            Row = ListRow[pro_count-1]
                            Row = ListRow[random.randint(0, (len(ListRow) - pro_count))]
                            link = Row.find_elements(By.TAG_NAME, 'a')[1].get_attribute('href')
                            provider_link.append(link)
                            pro_count -= 1
                        comments = driver.find_element(By.XPATH, "//*[@class='dataTables_info']").text
                except NoSuchElementException as e:
                    comments = "Tab faced an error while opening"
                try:
                    if measure in id_list:
                        driver.find_element(By.XPATH, "//*[@data-target='datatable_bulk_filter_0_metric-support-prov-ls']").click()
                        driver.find_element(By.XPATH, "//*[contains(text(),'Export all to CSV')]").click()
                        time.sleep(2)
                        ajax_preloader_wait(driver)
                        ws.append([name, LOB, measure, "Successfully downloaded export file", "-", driver.current_url])
                        onlyfiles = [f for f in listdir("assets\\Download\\CSV_Files_1") if isfile(join("assets\\Download\\CSV_Files_1", f))]
                        path = "assets\\Download\\CSV_Files_1\\" + onlyfiles[0]
                        [total_num, total_den] = NumDenAddition(path)
                        total_num = int(total_num)
                        total_den = int(total_den)
                        print(total_num)
                        print(total_den)
                        for j in range(0, len(num_den_list)-1, 4):
                            if LOB in num_den_list[j+2] and measure in num_den_list[j+3]:
                                if total_num == num_den_list[j] and total_den == num_den_list[j+1]:
                                    ws.append([name, LOB, measure, "The registry num/denum count is matching with export", "Registry:" + str(num_den_list[j])+"/"+str(num_den_list[j+1]) + "|| Export:" + str(total_num)+"/"+str(total_den), driver.current_url])
                                else:
                                    ws.append([name, LOB, measure, "The registry num/denum count is not matching with export",
                                               "Registry:" + str(num_den_list[j]) + "/" + str(
                                                   num_den_list[j + 1]) + "|| Export:" + str(total_num) + "/" + str(
                                                       total_den), driver.current_url])
                        os.remove(path)
                except NoSuchElementException as e:
                    comments = "Tab faced an error while opening"

            if field == "Patients":
                try:
                    ListRow = driver.find_element(By.XPATH, "//*[@id='metric-support-pat-ls']").find_element(By.TAG_NAME, "tbody").find_elements(By.TAG_NAME, 'tr')
                    if "No data available" in ListRow[0].text and len(ListRow) == 1:
                        comments = "No data available in patients tab"
                    else:
                        comments = driver.find_element(By.XPATH, "//*[@class='dataTables_info']").text
                except NoSuchElementException as e:
                    comments = "Tab faced an error while opening"
            if field == "Performance Statistics":
                try:
                    Performance_percentage_UI = driver.find_element(By.XPATH, "//*[@class='performance_value']").text
                    Performance_percentage_UI = Performance_percentage_UI.replace('%', '')
                    Performance_num_UI = driver.find_element(By.XPATH, "//*[@class='numerator']").text
                    Performance_num_UI = Performance_num_UI.replace('Numerator: ', '')
                    Performance_denum_UI = driver.find_element(By.XPATH, "//*[@class='denominator']").text
                    Performance_denum_UI = Performance_denum_UI.replace('Denominator: ', '')
                    try:
                        Performance_percentage_calculated = round((float(Performance_num_UI) / float(Performance_denum_UI)) * 100, 4)
                        if (float(Performance_percentage_UI) - float(Performance_percentage_calculated)) < 0.02:
                            comments = "The performance percentage matches with num/denum value"
                        else:
                            comments = "Performance percentage does not match with num/denum value"
                    except Exception as e:
                        print(e)
                        comments = "Denominator value is zero"
                except NoSuchElementException as e:
                    comments = "Tab faced an error while opening"
            ws.append([name, LOB, measure, "Successfully navigated to " + field, comments, driver.current_url])
            Tabs = driver.find_element(By.ID, 'qt-mt-support-ls').find_elements(By.TAG_NAME, 'li')
        except (NoSuchElementException, ElementNotInteractableException, ElementClickInterceptedException, ElementNotVisibleException, TimeoutException, ElementNotSelectableException) as e:
            ws.append([name, LOB, measure, "Issue: Tab encountered error for tab no.: " + str(i+1), "NA", driver.current_url])
    return practice_link, provider_link

# In[75]:

with open("assets\\hcc_data.pkl", 'rb') as hcc_data:
    hcc_data_list = pickle.load(hcc_data)


#login_url = "https://cert.cozeva.com/user/login"  #CERT environment
#logout_url = "https://cert.cozeva.com/user/logout"
login_url = "https://www.cozeva.com/user/login"  #PROD environment
logout_url = "https://www.cozeva.com/user/logout"
Customer_value = hcc_data_list[0]  # this is to select customer
year = str(hcc_data_list[1])       #select MY
selected_LOB = ""
provider_count = hcc_data_list[2] #Number of providers [set as default value with current code format]
patient_count = 2 #number of patients
User = os.environ.get('CS2_User')
Pass = os.environ.get('CS2_Password')
reason_for_login = "https://redmine2.cozeva.com/issues/34088"
Global_checklist = [33, 400, 551, 439, 553, 554, 555, 556, 557, 516, 526, 558]
Selected_checklist = []
Selected_LOB = ["Medicare", "ALL"]
id_list = ["Review of Chronic Conditions", "Review of Chronic Conditions (Risk Adjustment Version 24)", "One-Year Recapture Rate", "One-Year Recapture Rate (Risk Adjustment Version 24)"]
customer_list = ["U.S. Renal Care", "L.A. Care", "Molina Healthcare"]
patient_flag = 0

# In[76]:

def select_all():
    for checkbox in checkboxes.values():
        checkbox.set(True)


def deselect_all():
    for checkbox in checkboxes.values():
        checkbox.set(False)

def get_selected_items():
    selected_ids = [item_id for item_id, var in checkboxes.items() if var.get() == 1]
    selected_items = [items_with_ids[item_id] for item_id in selected_ids]
    global Selected_checklist
    Selected_checklist = selected_ids
    print("Selected IDs:", selected_ids)
    print("Selected measures:", selected_items)
    root.destroy()  # Close the GUI after getting selected items

def set_patient_dashboard_flag():
    global PatientDashboardFlag
    PatientDashboardFlag = patient_dashboard_var.get()

# Function to execute both set_patient_dashboard_flag and get_selected_items
def run_selected_measures():
    set_patient_dashboard_flag()
    get_selected_items()

def display_information():
    information_text = "Please read through this information box before selecting measures:\n"\
                       "\n" \
                       "- Review of Chronic Condition measure is basically RCCV28 measure.\n"\
                        "\n"\
                        "- HCC Score(Blended) measure will be available as HCC Measure for USRC, Healthnet, and LA Care.\n" \
                       "\n" \
                       "- HCC Efficiency measure will only be available in HPMG.\n" \
                       "\n" \
                       "- Review of Medicaid Chronic Conditions and Review of Medicaid Suspect conditions will only be present for Molina Healthcare.\n" \
                       "\n" \
                       "- For some customers, the Clinical factor and Suspect score may not be present (as per design).\n" \
                       "\n" \
                       "- For onshore customers, RAF scores are not calculated, so HCC Ribbon will be null.\n" \
                       "\n" \
                       "- Use 'Select All' or 'Deselect All' to toggle the checkboxes.\n" \
                       "\n" \
                       "- Click 'Run for Selected Items' to retrieve the selected items."
    tk.messagebox.showinfo("Information", information_text)

# Sample list of items
items_with_ids = {
    400: "Review of Chronic Conditions",
    551: "Review of Suspect conditions",
    553: "HCC Score",
    554: "Review of ACA Chronic Conditions",
    555: "Review of ACA Suspect Conditions",
    556: "ACA HCC Score",
    552: "HCC Recapture",
    516: "Review of Medicaid Chronic Conditions",
    526: "Review of Medicaid Suspect Conditions",
    558: "RAF Score",
    557: "HCC Opportunity"
}

root = tk.Tk()
root.title("Please select the measures (V28 Only)")
# Create a frame to hold the checkboxes
checkbox_frame = ttk.Frame(root)
checkbox_frame.pack(pady=10)

# Create IntVar variables for each checkbox
checkboxes = {item_id: tk.IntVar() for item_id in items_with_ids.keys()}

# Create Checkbuttons for each item
for i, (item_id, item) in enumerate(items_with_ids.items()):
    checkbox = ttk.Checkbutton(checkbox_frame, text=item, variable=checkboxes[item_id])
    checkbox.grid(row=i // 2, column=i % 2, padx=5, pady=2, sticky=tk.W)

# Create a frame to hold the buttons
button_frame = ttk.Frame(root)
button_frame.pack(pady=5)

# Create buttons
select_all_button = ttk.Button(button_frame, text="Select All", command=select_all)
select_all_button.grid(row=0, column=0, padx=5, pady=5)

deselect_all_button = ttk.Button(button_frame, text="Deselect All", command=deselect_all)
deselect_all_button.grid(row=0, column=1, padx=5, pady=5)

get_selected_button = ttk.Button(root, text="Run for selected measures", command=run_selected_measures)
get_selected_button.pack(pady=10)

# Create information button
information_button = ttk.Button(root, text="Read Me !", command=display_information)
information_button.pack(pady=5)

# Create a frame to hold the radio button
radio_frame = ttk.Frame(root)
radio_frame.pack(pady=10)

# Create a label for the radio button
radio_label = ttk.Label(radio_frame, text="Patient Dashboard:")
radio_label.grid(row=0, column=0, padx=5, pady=2, sticky=tk.W)

# Create a variable to store the selected value (Yes or No)
patient_dashboard_var = tk.StringVar(value="No")

# Create radio buttons for Yes and No
yes_radio = ttk.Radiobutton(radio_frame, text="Yes", variable=patient_dashboard_var, value="Yes")
no_radio = ttk.Radiobutton(radio_frame, text="No", variable=patient_dashboard_var, value="No")

yes_radio.grid(row=0, column=1, padx=5, pady=2, sticky=tk.W)
no_radio.grid(row=0, column=2, padx=5, pady=2, sticky=tk.W)


# Create a button to exit the application
exit_button = ttk.Button(root, text="Exit", command=get_selected_items)
exit_button.pack(pady=10)

# Center the window on the screen
root.update_idletasks()
window_width = root.winfo_width()
window_height = root.winfo_height()
screen_width = root.winfo_screenwidth()
screen_height = root.winfo_screenheight()
x_coordinate = (screen_width - window_width) // 2
y_coordinate = (screen_height - window_height) // 2
root.geometry(f"{window_width}x{window_height}+{x_coordinate}+{y_coordinate}")


root.mainloop()


header = NamedStyle(name="header")
header.font = Font(bold=True)
header.border = Border(bottom=Side(border_style="thin"))
header.alignment = Alignment(horizontal="center", vertical="center")
red_background = PatternFill(patternType='solid', fgColor="00FF0000")
green_background = PatternFill(patternType='solid', fgColor="50C878")
gray_background = PatternFill(patternType='solid', fgColor="5F9EA0")

blended_dl_path = os.path.join(os.getcwd(), "assets\\Download\\CSV_Files_1")
isdir = os.path.isdir(blended_dl_path)
if not isdir:
    os.mkdir(blended_dl_path)

driver = setups.driver_setup_with_download(blended_dl_path)

# options = webdriver.ChromeOptions()
# prefs = {"download.default_directory" : "C:\\Users\\sbasu\\Documents\\CSV_Files_1"}
# options.add_argument("user-data-dir=C:\\Users\\sbasu\\AppData\\Local\\Google\\Chrome\\User Data\\DefaultAutomation99")
# options.add_argument("--disable-notifications")
# options.add_experimental_option("prefs", prefs)
# driver = webdriver.Chrome(executable_path="C:\\Users\\sbasu\\Documents\\Drivers\\chromedriver.exe", options=options)
# Main


# In[77]:

print(Selected_checklist)

if len(Selected_checklist) == 0:
    print("No HCC Measures are selected. Stopping the whole code!")
    driver.close()
    sys.exit()
else:
    if PatientDashboardFlag == "Yes":
        patient_flag = 1
    else:
        patient_flag = 0


setups.login_to_cozeva(Customer_value)
# driver.get(logout_url)
# driver.get(login_url)
# driver.maximize_window()
# print("Initial window= " + driver.title)
# driver.find_element(By.ID, "edit-name").send_keys(User)
# driver.find_element(By.ID, "edit-pass").send_keys(Pass)
# driver.find_element(By.ID, "edit-submit").click()
# time.sleep(1)

# try:
#     WebDriverWait(driver, 300).until(EC.presence_of_element_located((By.ID, "reason_textbox")))
#     CustomersList = driver.find_element(By.XPATH, "//*[@id='select-customer']").find_elements(By.TAG_NAME, "option")
#     for Customer in CustomersList:
#         if Customer_value == Customer.get_attribute('value'):
#             Customer.click()
#     driver.find_element(By.ID, "reason_textbox").send_keys(reason_for_login)
#     driver.find_element(By.ID, "edit-submit").click()
# except NoSuchElementException:
#     traceback.print_exc()
#     driver.quit()
# ajax_preloader_wait(driver)
# Registry_URL = driver.current_url
customer_name = driver.find_element(By.CLASS_NAME, "specific_most").text


# In[14]:


wb = Workbook()
ws = wb.active
sheet_name = customer_name + " Support Registry " + str(year)
ws.title = sheet_name
ws.append(["Customer Name", "LOB", "Measure", "Status", "Comments", "URL"])
header_row = ws[1]
for cell in header_row:
    cell.style = header
wb.create_sheet("Practice Registry")
ws = wb["Practice Registry"]
ws.append(["Practice Name", "LOB", "Measure", "Status", "Comments", "URL"])
header_row = ws[1]
for cell in header_row:
    cell.style = header
wb.create_sheet("Provider MSPL")
ws = wb["Provider MSPL"]
ws.append(["Provider Name", "LOB", "Measure", 'Domain Name Check', 'Performance Statistics Check', 'Network Comparison Check', 'Risk Score Check', 'Comments', 'Provider URL'])
header_row = ws[1]
for cell in header_row:
    cell.style = header
wb.create_sheet("Patient Dashboard")
ws = wb["Patient Dashboard"]
ws.append(["Provider name", "LOB", "Measure", "Actual RAF Check", "Potential RAF Check", "Coded HCC Check", "Potential HCC Check", "Demographic Score Check", "Comment", "URL"])
header_row = ws[1]
for cell in header_row:
    cell.style = header


# Activated this if provider score check needed
# wb.create_sheet("Provider Registry score check")
# ws = wb["Provider Registry score check"]
# ws.append(["LOB", "Measure name", "Provider Name", "Registry Score of ID: 400","Registry Score of ID: 33", "Registry Score of ID: 439", "Comments", "Status","URL"])
# header_row = ws[1]
# for cell in header_row:
#     cell.style = header

# Activate this if practice score check needed
# wb.create_sheet("Practice Registry score check")
# ws = wb["Practice Registry score check"]
# ws.append(["LOB", "Measure name", "Practice Name", "Registry Score of ID: 400","Registry Score of ID: 33", "Registry Score of ID: 439", "Comments", "Status","URL"])
# header_row = ws[1]
# for cell in header_row:
#     cell.style = header

path1 = os.path.join(vs.parent_dir,"HCC Multi Validation Reports")
isdir = os.path.isdir(path1)
if not isdir:
    os.mkdir(path1)

path1 = path1 + "\\"
name_date = datetime.datetime.now()
formatted_date = name_date.strftime("%m-%d %H-%M")
filename = customer_name + "_HCC new code validation MY"+year+"_" + str(formatted_date) + ".xlsx"


# In[97]:


driver.find_element(By.XPATH, "//*[@id='qt-filter-label']").click()
Quarter_list = driver.find_element(By.XPATH, "//*[@id='filter-quarter']").find_elements(By.TAG_NAME, "li")
for quarter in Quarter_list:
    if quarter.text == year:
        quarter.click()
        break
LOB_list = driver.find_element(By.XPATH, "//*[@id='filter-lob']").find_elements(By.TAG_NAME, 'li')
URL_name_registry_list = []
num_den_value_list = []
num_den_prac_list = []
flag = 0

for i in range(0, len(LOB_list)):
    flag += 1
    LOB_Name = LOB_list[i].text
    print("LOB Name for " + year + ": " + LOB_Name)
    try:
        if selected_LOB != "":
            if LOB_Name == selected_LOB:
                LOB_list[i].click()
            else:
                continue
        else:
            LOB_list[i].click()
    except ElementNotInteractableException as e:
        continue
    driver.find_element(By.ID, "reg-filter-apply").click()
    registry_url = driver.current_url
    ajax_preloader_wait(driver)
    if driver.find_element(By.XPATH, "//*[@id='conti_enroll']").is_selected():
        driver.find_element(By.XPATH, "//*[@class='cont_disc_toggle']").click()
    print("Current support registry URL: " + registry_url)
    [URL_name_registry_list, num_den_value_list] = NameURLextractor(driver, LOB_Name, Selected_checklist, URL_name_registry_list, num_den_value_list, customer_name)
    while flag < len(LOB_list):
        driver.find_element(By.XPATH, "//*[@id='qt-filter-label']").click()
        LOB_list = driver.find_element(By.XPATH, "//*[@id='filter-lob']").find_elements(By.TAG_NAME, 'li')
        break
    print(URL_name_registry_list)
    print(num_den_value_list)

ws = wb[sheet_name]
# In[113]:

practice_link = ""
provider_link_supports = []
provider_link_practice = []
i = 0
previous_measure = ""
print(URL_name_registry_list[i])
_ids = {
    400: "Review of Chronic Conditions Version 28",
    551: "Review of Suspect conditions",
    553: "HCC Score",
    554: "Review of ACA Chronic Conditions",
    555: "Review of ACA Suspect Conditions",
    556: "ACA HCC Score",
    557: "HCC Efficiency",
    516: "Review of Medicaid Chronic Conditions",
    526: "Review of Medicaid Suspect Conditions",
    558: "RAF Score"
}
for i in range(0, len(URL_name_registry_list), 3):
    url = URL_name_registry_list[i]
    print(url)
    measure = URL_name_registry_list[i+1]
    LOB = URL_name_registry_list[i+2]
    driver.get(url)
    ajax_preloader_wait(driver)
    time.sleep(3)
    ws = wb[sheet_name]
    MSPL_name = driver.find_element(By.XPATH, "//*[@class='ch metric_specific_patient_list_title valign-wrapper']").text
    print(MSPL_name)
    if URL_name_registry_list[i+1] in MSPL_name:
        ws.append([driver.find_element(By.XPATH, "//span[@class='specific_most']").text, LOB, measure, "Domain name successful validation", "Checked for Registry vs MSPL", url])
    else:
        ws.append([driver.find_element(By.XPATH, "//span[@class='specific_most']").text, LOB, measure, "Domain name mismatch", "Not matching for Registry vs MSPL", url])
    WebDriverWait(driver, 300).until(EC.presence_of_element_located((By.ID, "qt-mt-support-ls")))
    WebDriverWait(driver, 300).until(EC.presence_of_element_located((By.CLASS_NAME, 'tab')))
    Tabs = driver.find_element(By.ID, 'qt-mt-support-ls').find_elements(By.TAG_NAME, 'li')
    [practice_link, provider_link_supports] = TabNavigator(driver, Tabs, ws, wb, measure, LOB, num_den_value_list, provider_count, id_list)
    SheetColorCoder(ws, wb, path1, filename)
    ws = wb["Practice Registry"]
    try:
        driver.get(practice_link)
        ajax_preloader_wait(driver)
        WebDriverWait(driver, 300).until(EC.presence_of_element_located(((By.XPATH, "//*[@data-target='qt-reg-nav-filters']"))))
        #driver.get(link)
        search_var = measure.split(' | ')[1]
        print(search_var)
        time.sleep(2)
        driver.find_element(By.XPATH, "//*[contains(@class, 'met-name top') and contains(text(), '" + search_var + "')] /.. /.. /.. /..").click()
        ajax_preloader_wait(driver)
        Tabs = driver.find_element(By.ID, 'qt-mt-support-ls').find_elements(By.TAG_NAME, 'li')
        [null_link, provider_link_practice] = TabNavigator(driver, Tabs, ws, wb, measure, LOB, num_den_prac_list, provider_count, id_list)
        SheetColorCoder(ws, wb, path1, filename)
        print("Provider links")
        print(provider_link_supports)
        for provider_link_support in provider_link_supports:
            driver.get(provider_link_support)
            ajax_preloader_wait(driver)
            ws = wb["Provider MSPL"]
            MSPLScoreCheck(driver, ws, wb, LOB, measure)
            SheetColorCoder(ws, wb, path1, filename)
            ajax_preloader_wait(driver)
            ws = wb["Patient Dashboard"]
            print("Patient flag: " + str(patient_flag))
            if patient_flag != 0:
                PatientDashboardScoreCheck(driver, patient_count, ws, wb, LOB, measure)
            else:
                print("Patient Dashboard Score check skipped")
            SheetColorCoder(ws, wb, path1, filename)
    except Exception as e:
        print(e)
        ws.append([customer_name, LOB, measure, "Issue: Blank MSPL encountered","NA",driver.current_url])
        SheetColorCoder(ws, wb, path1, filename)


driver.quit()
# In[ ]:




