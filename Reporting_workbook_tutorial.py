from openpyxl import Workbook, load_workbook


wb = Workbook()  # This will create a workbook object.
ws = wb.active   # This will set the currently active sheet
ws.title = "Report"

""" Data can be added to a worksheet by making a list.
Treat ws(the worksheet) as an array, and start appending rows to it
for example i want to fill a sheet with 4 rows and 4 columns with numbers """

ws.append([1,2,3,4])
ws.append([5,6,7,8])
ws.append([9,10,11,12])
ws.append([13,14,15,16])

""" once we had added any data to the worksheet, it is a good idea to save the data to a physical file.
To save this workbook, we will need a filename and a path. For now lets store this in a folder in c:/
"""
filename = "Reporting_workbook.xlsx"
path = "C:\\VerificationReports\\"  # edit your own path here. Do not score it directly in c:/

""" To save, we use the original workbook object. """

wb.save(path+filename)  # append the chosen path with the filename. This will create a .xlsx file in the path
                        # with numbers 1-16 as added previously




"""The way workbooks work is, they are a three dimensional array. The three dimensions are 
Sheet name
Row
Columns
New sheets can be created and switched to using the original workbook object. 
I will create a sheet called Data2
"""

wb.create_sheet("Data2")  # A new sheet has been created, but the ws object is still pointing to the first sheet
ws = wb["Data2"]  #switched the worksheet object to a new sheet in the workbook

#lets add data to the second sheet

ws.append([17,18,19,20])
ws.append([21,22,23,24])
ws.append([25,26,27,28])

#save this workbook now. This time, the workbook object has 2 sheets. And they both have data in it.

wb.save(path+filename)

#Generally in a reporting worksheet, the first row is the header. you would add it like



