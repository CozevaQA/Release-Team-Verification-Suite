import time
import traceback
from selenium import webdriver
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import NoSuchElementException, ElementNotInteractableException, \
    ElementClickInterceptedException, TimeoutException, UnexpectedAlertPresentException
from openpyxl.styles import PatternFill, Font
from selenium.webdriver.common.by import By
from selenium.webdriver.support.wait import WebDriverWait


import setups
import variablestorage as locator
import support_functions as sf


def time_capsule(driver, workbook, logger, run_from):
    workbook.create_sheet('Time Capsule')
    ws = workbook['Time Capsule']

    ws.append(['ID', 'Context', 'Scenario', 'Status', 'Comments'])
    header_font = Font(color='FFFFFF', bold=False, size=12)
    header_cell_color = PatternFill('solid', fgColor='030303')
    ws['A1'].font = header_font
    ws['A1'].fill = header_cell_color
    ws['B1'].font = header_font
    ws['B1'].fill = header_cell_color
    ws['C1'].font = header_font
    ws['C1'].fill = header_cell_color
    ws['D1'].font = header_font
    ws['D1'].fill = header_cell_color
    ws['E1'].font = header_font
    ws['E1'].fill = header_cell_color

    ws.name = "Arial"
    test_case_id = 1
    try:
        last_url = driver.current_url
        window_switched = 0
        driver.find_element_by_xpath(locator.xpath_app_Tray_Link).click()
        driver.find_element_by_xpath(locator.xpath_app_Time_Capsule).click()
        driver.switch_to.window(driver.window_handles[1])
        window_switched = 1
        sf.ajax_preloader_wait(driver)
        try:
            sf.ajax_preloader_wait(driver)
            current_url = driver.current_url
            access_message = sf.CheckAccessDenied(current_url)

            if access_message == 1:
                print("Access Denied found!")
                # logger.critical("Access Denied found!")
                test_case_id+= 1
                ws.append(
                    (test_case_id, 'Time Capsule', 'Access check for Time Capsule', 'Failed', 'Access Denied', driver.current_url))

            else:
                print("Access Check done!")
                # logger.info("Access Check done!")
                error_message = sf.CheckErrorMessage(driver)

                if error_message == 1:
                    print("Error toast message is displayed")
                    # logger.critical("ERROR TOAST MESSAGE IS DISPLAYED!")
                    test_case_id += 1
                    ws.append((test_case_id, 'Time Capsule', 'Navigation to Time Capsule without error message',
                                'Failed', 'Error toast message is displayed', driver.current_url))

                else:
                    ws.append((test_case_id, 'Time Capsule', 'Time Capsule page loading',
                                    'Passed'))
                    if len(driver.find_elements_by_xpath(locator.xpath_latest_Card_Title)) != 0:
                        latest_computation_dete = driver.find_element_by_xpath(
                            locator.xpath_latest_Card_Title).text
                        test_case_id += 1
                        ws.append((test_case_id, 'Time Capsule', 'Computation Card',
                                    'Passed', 'Latest Computation date: ' + latest_computation_dete))
                    else:
                        test_case_id += 1
                        ws.append((test_case_id, 'Time Capsule', 'Computation Card', 'Failed',
                                    'Computation card details is not found!', driver.current_url))

                    # time capsule cards comparision, Making variables
                    measure_performance = []
                    claim_count = []
                    claim_provider_count = []
                    claim_member_count = []
                    pmc_count = []
                    lab_count = []
                    lab_members_count = []

                    cards = driver.find_element_by_class_name("card_section_wrapper").find_elements_by_class_name("card_wrapper")
                    cards = cards[0:4]
                    for i,card in enumerate(cards):
                        measure_performance.append(str(card.find_element_by_class_name("card_body_value").text))
                        card.click()
                        time.sleep(2)
                        sf.ajax_preloader_wait(driver)
                        WebDriverWait(driver, 60).until(
                            EC.presence_of_element_located((By.CLASS_NAME, 'card_details_body')))
                        card_details = driver.find_element_by_id('card_details_modal')

                        claim_count.append(str(
                            card_details.find_element_by_class_name('card_claims').find_elements_by_tag_name('span')[
                                1].text))
                        claim_provider_count.append(str(
                            card_details.find_element_by_class_name('card_providers').find_elements_by_tag_name('span')[
                                1].text))
                        claim_member_count.append(str(
                            card_details.find_element_by_class_name('card_members').find_elements_by_tag_name('span')[
                                1].text))
                        pmc_count.append(str(
                            card_details.find_element_by_class_name('card_pmc').find_elements_by_tag_name('span')[
                                1].text))
                        lab_count.append(str(
                            card_details.find_element_by_class_name('card_labs').find_elements_by_tag_name('span')[
                                1].text))
                        lab_members_count.append(str(
                            card_details.find_element_by_class_name('card_lab_members').find_elements_by_tag_name('span')[
                                1].text))
                        #Add extra checks for latest card
                        if i == 0:
                            # verify if any count is zero
                            current_claim_count = claim_count[0]
                            current_provider_count = claim_provider_count[0]
                            current_claim_member_count = claim_member_count[0]
                            current_pmc_count = pmc_count[0]
                            current_lab_count = lab_count[0]
                            current_lab_members_count = lab_members_count[0]

                            current_claim_count = claim_count[0].replace(',', '')
                            current_provider_count = claim_provider_count[0].replace(',', '')
                            current_claim_member_count = claim_member_count[0].replace(',', '')
                            current_pmc_count = pmc_count[0].replace(',', '')
                            current_lab_count = lab_count[0].replace(',', '')
                            current_lab_members_count = lab_members_count[0].replace(',', '')


                            # Claim count verification
                            if int(current_claim_count) == 0:
                                ws.append((
                                    test_case_id, 'Time Capsule', 'Claim count in latest card', 'Failed',
                                    'Claim count is zero, please check',
                                    driver.current_url
                                ))
                            else:
                                ws.append((
                                    test_case_id, 'Time Capsule', 'Claim count displayed correctly', 'Passed',
                                    f'Claim count is displayed correctly: {current_claim_count}',
                                    driver.current_url
                                ))

                            # Provider count verification
                            if int(current_provider_count) == 0:
                                ws.append((
                                    test_case_id, 'Time Capsule', 'Provider count in latest card', 'Failed',
                                    'Provider count is zero, please check',
                                    driver.current_url
                                ))
                            else:
                                ws.append((
                                    test_case_id, 'Time Capsule', 'Provider count displayed correctly', 'Passed',
                                    f'Provider count is displayed correctly: {current_provider_count}',
                                    driver.current_url
                                ))

                            # Claim member count verification
                            if int(current_claim_member_count) == 0:
                                ws.append((
                                    test_case_id, 'Time Capsule', 'Claim member count in latest card', 'Failed',
                                    'Claim member count is zero, please check',
                                    driver.current_url
                                ))
                            else:
                                ws.append((
                                    test_case_id, 'Time Capsule', 'Claim member count displayed correctly', 'Passed',
                                    f'Claim member count is displayed correctly: {current_claim_member_count}',
                                    driver.current_url
                                ))

                            # PMC count verification
                            if int(current_pmc_count) == 0:
                                ws.append((
                                    test_case_id, 'Time Capsule', 'PMC count in latest card', 'Failed',
                                    'PMC count is zero, please check',
                                    driver.current_url
                                ))
                            else:
                                ws.append((
                                    test_case_id, 'Time Capsule', 'PMC count displayed correctly', 'Passed',
                                    f'PMC count is displayed correctly: {current_pmc_count}',
                                    driver.current_url
                                ))

                            # Lab count verification
                            if int(current_lab_count) == 0:
                                ws.append((
                                    test_case_id, 'Time Capsule', 'Lab count in latest card', 'Failed',
                                    'Lab count is zero, please check',
                                    driver.current_url
                                ))
                            else:
                                ws.append((
                                    test_case_id, 'Time Capsule', 'Lab count displayed correctly', 'Passed',
                                    f'Lab count is displayed correctly: {current_lab_count}',
                                    driver.current_url
                                ))

                            # Lab member count verification
                            if int(current_lab_members_count) == 0:
                                ws.append((
                                    test_case_id, 'Time Capsule', 'Lab member count in latest card', 'Failed',
                                    'Lab member count is zero, please check',
                                    driver.current_url
                                ))
                            else:
                                ws.append((
                                    test_case_id, 'Time Capsule', 'Lab member count displayed correctly', 'Passed',
                                    f'Lab member count is displayed correctly: {current_lab_members_count}',
                                    driver.current_url
                                ))



                            # Verify if LOB section is present
                            lob_section_xpath="//div[@class='card_details_body']//child::div[@class='col s12 m3 l2 lfloat lob_section_wrapper']"
                            lob_count=len(driver.find_elements_by_xpath(lob_section_xpath)) - 1
                            if int(lob_count)==0:
                                ws.append((test_case_id, 'Time Capsule', 'LOB Section appearing correctly ', 'Failed',
                                           'LOB Details is not displayed ,Please check',
                                           driver.current_url))
                            else:
                                ws.append((test_case_id, 'Time Capsule', 'LOB Section appearing correctly ', 'Passed',
                                           f'LOB Details is displayed for {lob_count},Please check',
                                           driver.current_url))



                        # Verify if summary of metric score change has valid table

                        measure_performance_copy = [*set(measure_performance)]
                        claim_count_copy = [*set(claim_count)]
                        claim_member_count_copy = [*set(claim_member_count)]
                        claim_provider_count_copy = [*set(claim_provider_count)]
                        pmc_count_copy = [*set(pmc_count)]
                        lab_count_copy = [*set(lab_count)]
                        lab_members_count_copy = [*set(lab_members_count)]

                        #add appends
                    print(measure_performance)
                    print(claim_count)
                    print(pmc_count)
                    if float(measure_performance[1].replace("%", "")) - float(measure_performance[0].replace("%", "")) < 2:
                        ws.append((test_case_id, 'Time Capsule', 'Measure performance increase', 'Passed',
                                   'Measure performance has increased or is the same', driver.current_url))
                    elif float(measure_performance[1].replace("%", "")) - float(measure_performance[0].replace("%", "")) > 2:
                        ws.append((test_case_id, 'Time Capsule', 'Measure performance increase', 'Failed',
                                   'Measure performance has decreased since previous computation', driver.current_url))
                    if len(measure_performance_copy) == len(measure_performance):
                        ws.append((test_case_id, 'Time Capsule', 'Measure Performance Duplicates', 'Passed',
                                   'All cards have different values', driver.current_url))
                    else:
                        ws.append((test_case_id, 'Time Capsule', 'Measure Performance Duplicates', 'Failed',
                                   'Unique values: \"'+'\",\"'.join([str(elem) for elem in measure_performance_copy])+'\"', driver.current_url))
                    test_case_id+=1
                    if len(claim_count_copy) == len(claim_count):
                        ws.append((test_case_id, 'Time Capsule', 'Claim count Duplicates', 'Passed',
                                   'All cards have different values', driver.current_url))
                    else:
                        ws.append((test_case_id, 'Time Capsule', 'Claim count Duplicates', 'Failed',
                                   'Unique values: \"'+'\",\"'.join([str(elem) for elem in claim_count_copy])+'\"', driver.current_url))
                    test_case_id+=1
                    if len(claim_provider_count_copy) == len(claim_provider_count):
                        ws.append((test_case_id, 'Time Capsule', 'Claim Provider Count Duplicates', 'Passed',
                                   'All cards have different values', driver.current_url))
                    else:
                        ws.append((test_case_id, 'Time Capsule', 'Claim Provider Count Duplicates', 'Failed',
                                   'Unique values: \"'+'\",\"'.join([str(elem) for elem in claim_provider_count_copy])+'\"', driver.current_url))
                    test_case_id+=1
                    if len(claim_member_count_copy) == len(claim_member_count):
                        ws.append((test_case_id, 'Time Capsule', 'Claim Member Count Duplicates', 'Passed',
                                   'All cards have different values', driver.current_url))
                    else:
                        ws.append((test_case_id, 'Time Capsule', 'Claim Member Count Duplicates', 'Failed',
                                   'Unique values: \"'+'\",\"'.join([str(elem) for elem in claim_member_count_copy])+'\"', driver.current_url))
                    test_case_id+=1
                    if len(pmc_count_copy) == len(pmc_count):
                        ws.append((test_case_id, 'Time Capsule', 'PMC Count Duplicates', 'Passed',
                                   'All cards have different values', driver.current_url))
                    else:
                        ws.append((test_case_id, 'Time Capsule', 'PMC Count Duplicates', 'Failed',
                                   'Unique values: \"'+'\",\"'.join([str(elem) for elem in pmc_count_copy])+'\"', driver.current_url))
                    test_case_id+=1
                    if len(lab_count_copy) == len(lab_count):
                        ws.append((test_case_id, 'Time Capsule', 'Lab Count Duplicates', 'Passed',
                                   'All cards have different values', driver.current_url))
                    elif "0" in lab_count_copy or "1" in lab_count_copy:
                        ws.append((test_case_id, 'Time Capsule', 'Lab Count Duplicates', 'Passed',
                                   'Card values are 0 or 1', driver.current_url))
                    else:
                        ws.append((test_case_id, 'Time Capsule', 'Lab Count Duplicates', 'Failed',
                                   'Unique values: \"'+'\",\"'.join([str(elem) for elem in lab_count_copy])+'\"', driver.current_url))
                    test_case_id+=1
                    if len(lab_members_count_copy) == len(lab_members_count):
                        ws.append((test_case_id, 'Time Capsule', 'Lab Members Count Duplicates', 'Passed',
                                   'All cards have different values', driver.current_url))
                    elif "0" in lab_members_count_copy or "1" in lab_members_count_copy:
                        ws.append((test_case_id, 'Time Capsule', 'Lab Members Count Duplicates', 'Passed',
                                   'Card values are 0 or 1', driver.current_url))
                    else:
                        ws.append((test_case_id, 'Time Capsule', 'Lab Members Count Duplicates', 'Failed',
                                   'Unique values: \"'+'\",\"'.join([str(elem) for elem in lab_members_count_copy])+'\"', driver.current_url))
                    test_case_id+=1

                    #measure performance check for lower performance in nearest week







                    #populate these counts to the 4th capsule card and then make copies of the list with _copy
                    #remove duplicates and compare the list lengths. If length is different, fail the test case


        except Exception as e:
            print(e)
            test_case_id += 1
            ws.append(
                (test_case_id, 'Time Capsule', 'Navigation to Time Capsule', 'Failed', 'Exception occurred!', driver.current_url))
            traceback.print_exc()
        finally:

            time.sleep(1)
            if window_switched == 1:
                driver.close()
                driver.switch_to.window(driver.window_handles[0])
                sf.ajax_preloader_wait(driver)
                window_switched = 0


    except Exception as e:
        print(e)
        test_case_id += 1
        ws.append(
            (test_case_id, 'Time Capsule', 'Navigation to Time Capsule', 'Failed', 'Exception occurred!', driver.current_url))
        driver.get(last_url)
        sf.ajax_preloader_wait(driver)
        WebDriverWait(driver, 30).until(
            EC.presence_of_element_located((By.XPATH, locator.xpath_app_Tray_Link)))

    rows = ws.max_row
    cols = ws.max_column
    for i in range(2, rows + 1):
        for j in range(3, cols + 1):
            if ws.cell(i, j).value == 'Passed':
                ws.cell(i, j).fill = PatternFill('solid', fgColor='0FC404')
            elif ws.cell(i, j).value == 'Failed':
                ws.cell(i, j).fill = PatternFill('solid', fgColor='FC0E03')
            elif ws.cell(i, j).value == 'Data table is empty':
                ws.cell(i, j).fill = PatternFill('solid', fgColor='FCC0BB')


driver=setups.driver_setup()
setups.login_to_cozeva("1950")
workbook=setups.create_reporting_workbook(locator.parent_dir)
logger=setups.logger_setup(locator.parent_dir)
time_capsule(driver,workbook,logger,"Cozeva Support")
workbook.save(locator.parent_dir + "\\Report.xlsx")
