import os
import time
import traceback

from openpyxl.styles import Font, PatternFill
from selenium.webdriver.common.by import By
from selenium.webdriver.support.wait import WebDriverWait
from tkinter import *
import setups
import logging
import ExcelProcessor as db
import context_functions as cf
import support_functions as sf
import setups as st
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
import variablestorage as locator
from openpyxl import Workbook, load_workbook
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import NoSuchElementException, ElementNotInteractableException, \
    ElementClickInterceptedException, TimeoutException

ENV = 'PROD'
Client_list = db.getCustomerList()
print(Client_list)
id_list = []
for customer in Client_list:
    id_list.append(str(db.fetchCustomerID(customer).strip()))
print(id_list)
id_list = id_list[1:len(id_list) - 1]
print(id_list)
print(len(id_list))

report_folder = os.path.join(locator.parent_dir,"Page specific access checks")
isdir = os.path.isdir(report_folder)
if not isdir:
    os.mkdir(report_folder)

filename = "Hospital Activity Access_"+sf.date_time()+".xlsx"

wb = Workbook()
ws = wb.active
ws.title = "Hospital Activity_"+ENV

wb.save(report_folder+"\\"+filename)

ws.append(["Client", "Hospital Activity", "Render time"])

driver = setups.driver_setup()
setups.login_to_cozeva("1500")

for client_id in id_list:
    setups.switch_customer_context(client_id)
    sf.ajax_preloader_wait(driver)
    WebDriverWait(driver, 30).until(
        EC.presence_of_element_located((By.XPATH, locator.xpath_filter_measure_list)))

    driver.find_element_by_xpath(locator.xpath_side_nav_SlideOut).click()
    time.sleep(0.5)
    start_time = time.perf_counter()
    driver.find_element(By.ID, "hospital_activity_tab").click()
    sf.ajax_preloader_wait(driver)
    total_time = time.perf_counter() - start_time - 2
    if sf.CheckErrorMessage(driver) == 0:
        if len(driver.find_elements_by_xpath(locator.xpath_data_Table_Info)) != 0:
            time.sleep(0.5)
            datatable_info = driver.find_element_by_xpath(locator.xpath_data_Table_Info).text
        ws.append([db.fetchCustomerName(client_id), "Passed", str(total_time), datatable_info])
    elif sf.CheckErrorMessage(driver) == 1:
        if len(driver.find_elements_by_xpath(locator.xpath_data_Table_Info)) != 0:
            time.sleep(0.5)
            datatable_info = driver.find_element_by_xpath(locator.xpath_data_Table_Info).text
        ws.append([db.fetchCustomerName(client_id), "Failed", str(total_time), "-"])
    wb.save(report_folder + "\\" + filename)

rows = ws.max_row
cols = ws.max_column
for i in range(2, rows + 1):
    for j in range(3, cols + 1):
        if ws.cell(i, j).value == 'Passed':
            ws.cell(i, j).fill = PatternFill('solid', fgColor='0FC404')
        elif ws.cell(i, j).value == 'Failed':
            ws.cell(i, j).fill = PatternFill('solid', fgColor='FC0E03')
        elif ws.cell(i, j).value == 'Showing 0 to 0':
            ws.cell(i, j).fill = PatternFill('solid', fgColor='FCC0BB')
















