import os
import time
import traceback
from tkinter import ttk, messagebox

from openpyxl.styles import Font, PatternFill
from selenium.webdriver.common.by import By
from selenium.webdriver.support.wait import WebDriverWait
from tkinter import *
import setups
import pickle
import logging
import ExcelProcessor as db
import context_functions as cf
import support_functions as sf
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
import variablestorage as locator
from openpyxl import Workbook, load_workbook
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import NoSuchElementException, ElementNotInteractableException, \
    ElementClickInterceptedException, TimeoutException

ENV = 'PROD'
URL = ""
client_list = []
Selected_checklist = []
provider_count = 1
measurement_year = "2022"


def get_ID_list():
    root = Tk()
    customer_list = db.getCustomerList()
    #print(customer_list)
    Checkbox_variables = [None]*len(customer_list)
    for i in range(0, len(Checkbox_variables)):
        Checkbox_variables[i] = IntVar()
    print(len(Checkbox_variables))
    Checkbox_widgets = []



    def on_submitbutton():
        global provider_count, measurement_year
        provider_count = int(provider_count_entrybox.get().strip())
        measurement_year = str(measurement_year_entrybox.get().strip())

        for i in range(0,len(Checkbox_variables)):
            if Checkbox_variables[i].get() == 1:
                client_list.append(db.fetchCustomerID(customer_list[i]))
        print(client_list)
        root.destroy()

    def on_new():
        global provider_count, measurement_year, run_config
        provider_count = int(provider_count_entrybox.get().strip())
        measurement_year = str(measurement_year_entrybox.get().strip())

        for i in range(0, len(Checkbox_variables)):
            if Checkbox_variables[i].get() == 1:
                client_list.append(db.fetchCustomerID(customer_list[i]))
        print(client_list)
        run_config = "New"
        root.destroy()


    #GenerateCheckboxesForall
    for i in range(0,len(customer_list)):
        Checkbox_widgets.append(Checkbutton(root, text=customer_list[i], variable=Checkbox_variables[i], font=("Nunito Sans", 10)))
    submit_button = Button(root, text="Submit", command=on_submitbutton, font=("Nunito Sans", 10))
    new_button = Button(root, text="New code", command=on_new, font=("Nunito Sans", 10))
    provider_count_entrybox = Entry(root)
    provider_measure_label = Label(root, text="Enter number of providers and MY", font=("Nunito Sans", 10))
    measurement_year_entrybox = Entry(root)
    #add all checkboxes to a grid
    #practice_sidemenu_checkbox.grid(row=3, column=0, columnspan=5, sticky="w")
    for i in range(1, len(Checkbox_widgets)):
        if i <= 20:
            Checkbox_widgets[i].grid(row=i, column=0, sticky="w")
        elif 20 < i <= 40:
            Checkbox_widgets[i].grid(row=i-20, column=1, sticky="w")
        elif 40 < i <= 60:
            Checkbox_widgets[i].grid(row=i-40, column=2, sticky="w")
        elif 60 < i <= 80:
            Checkbox_widgets[i].grid(row=i-60, column=3, sticky="w")
        submit_button.grid(row=0, column=3, sticky="e")
        new_button.grid(row=0, column=4, sticky="w")
        provider_measure_label.grid(row=0, column=0, sticky="w")
        provider_count_entrybox.grid(row=0, column=1, sticky="w")
        measurement_year_entrybox.grid(row=0, column=2, sticky="w")
    root.title("Multi HCC Validation")
    root.iconbitmap("assets/icon.ico")
    root.mainloop()


def display_new_gui():
    # Sample list of items
    global items_with_ids
    items_with_ids = {
        400: "Review of Chronic Conditions (Blended)",
        551: "Review of Chronic Conditions (Risk Adjustment Version 24)",
        552: "Review of Suspect conditions",
        553: "HCC Score(Blended)",
        554: "Review of ACA Chronic Conditions",
        555: "Review of ACA Suspect Conditions",
        556: "ACA HCC Score",
        557: "HCC Efficiency"
    }

    def get_selected_items():
        selected_ids = [item_id for item_id, var in checkboxes.items() if var.get() == 1]
        selected_items = [items_with_ids[item_id] for item_id in selected_ids]
        global Selected_checklist
        Selected_checklist = selected_ids
        print("Selected IDs:", selected_ids)
        print("Selected measures:", selected_items)
        root.destroy()  # Close the GUI after getting selected items

    def select_all():
        for checkbox in checkboxes.values():
            checkbox.set(True)

    def deselect_all():
        for checkbox in checkboxes.values():
            checkbox.set(False)

    def display_information():
        information_text = "Please read through this information box before selecting measures:\n" \
                           "\n" \
                           "- RCC(Blended) measure is not available for USRC, Healthnet, and LA Care.\n" \
                           "\n" \
                           "- HCC Score(Blended) measure will be available as HCC Measure for USRC, Healthnet, and LA Care.\n" \
                           "\n" \
                           "- HCC Efficiency measure will only be available in HPMG.\n" \
                           "\n" \
                           "- For some customers, the Clinical factor and Suspect score may not be present (as per design).\n" \
                           "\n" \
                           "- For onshore customers, RAF scores are not calculated, so HCC Ribbon will be null.\n" \
                           "\n" \
                           "- RCC(V28) measure is hidden from the registry for MY2023.\n" \
                           "\n" \
                           "- Use 'Select All' or 'Deselect All' to toggle the checkboxes.\n" \
                           "\n" \
                           "- Click 'Get Selected Items' to retrieve the selected items."
        messagebox.showinfo("Information", information_text)

    root = Tk()
    root.title("Please select the measures")
    # Create a frame to hold the checkboxes
    checkbox_frame = ttk.Frame(root)
    checkbox_frame.pack(pady=10)

    # Create IntVar variables for each checkbox
    checkboxes = {item_id: IntVar() for item_id in items_with_ids.keys()}

    # Create Checkbuttons for each item
    for i, (item_id, item) in enumerate(items_with_ids.items()):
        checkbox = ttk.Checkbutton(checkbox_frame, text=item, variable=checkboxes[item_id])
        checkbox.grid(row=i // 2, column=i % 2, padx=5, pady=2, sticky='W')

    # Create a frame to hold the buttons
    button_frame = ttk.Frame(root)
    button_frame.pack(pady=5)

    # Create buttons
    select_all_button = ttk.Button(button_frame, text="Select All", command=select_all)
    select_all_button.grid(row=0, column=0, padx=5, pady=5)

    deselect_all_button = ttk.Button(button_frame, text="Deselect All", command=deselect_all)
    deselect_all_button.grid(row=0, column=1, padx=5, pady=5)

    get_selected_button = ttk.Button(root, text="Run for selected measures", command=get_selected_items)
    get_selected_button.pack(pady=10)

    # Create information button
    information_button = ttk.Button(root, text="Read Me !", command=display_information)
    information_button.pack(pady=5)

    # Center the window on the screen
    root.update_idletasks()
    window_width = root.winfo_width()
    window_height = root.winfo_height()
    screen_width = root.winfo_screenwidth()
    screen_height = root.winfo_screenheight()
    x_coordinate = (screen_width - window_width) // 2
    y_coordinate = (screen_height - window_height) // 2
    root.geometry(f"{window_width}x{window_height}+{x_coordinate}+{y_coordinate}")

    root.title("HCC v28 Validation")
    root.iconbitmap("assets/icon.ico")
    root.mainloop()

if __name__ == '__main__':
    x=0

print("Hello World")
run_config = "Old"
get_ID_list()

if run_config == "New":
    display_new_gui()

report_folder = os.path.join(locator.parent_dir,"HCC Multi Validation Reports")
isdir = os.path.isdir(report_folder)
if not isdir:
    os.mkdir(report_folder)
workbook_title = "HCC Multi Validation_"+sf.date_time()+".xlsx"

wb = Workbook()
ws = wb.active
ws.title = 'HCC_Validation ' + ENV
for ID in client_list:
    driver = setups.driver_setup()
    if ENV == 'CERT':
        setups.login_to_cozeva_cert(ID)
    elif ENV == 'STAGE':
        setups.login_to_cozeva_stage()
    elif ENV == "PROD":
        setups.login_to_cozeva(ID)
    else:
        print("ENV INVALID")
        exit(3)
    print("Run HCC Validation for "+str(ID)+ ",For providers: "+str(provider_count))
    if run_config == "Old":
        cf.hccvalidation_multi(driver, ID, measurement_year, wb, provider_count, report_folder, "Cozeva Support", workbook_title)
        wb.save(report_folder + "\\" + workbook_title)
        driver.quit()
    elif run_config == "New":
        with open("assets\\hcc_data.pkl", 'wb') as hcc_file:
            pickle.dump([ID, provider_count, measurement_year, Selected_checklist], hcc_file)
        with open("HCC_v28.py") as hcc_code:
            exec(hcc_code.read())
        #import HCC_v28
        driver.quit()




