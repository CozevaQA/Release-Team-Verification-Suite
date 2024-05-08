import base64
import traceback
import tkinter as tk
from os.path import isfile, join
from tkinter import filedialog

import pytz
from openpyxl import Workbook, load_workbook
import os
import sys
from csv import DictReader
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
import json
import openpyxl as xlrd
from selenium.common.exceptions import TimeoutException, StaleElementReferenceException, \
    ElementNotInteractableException, ElementClickInterceptedException
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.select import Select
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium import webdriver
from selenium.common.exceptions import NoSuchElementException
import xlwt
from xlwt import Workbook
import time
import csv
from csv import DictReader
from colorama import Fore, Back, Style
import logging
import os
import shutil
import datetime as x
from datetime import date, datetime,timedelta
import configparser
import openpyxl
import setups
import variablestorage as locator
import guiwindow
import support_functions as sf
from threading import Timer

selected_month = ""

# test comment
config = configparser.RawConfigParser()
config.read("locator-config.properties")

#generate separate report for errors- if string is not formed completely then throw error and store it in error log


# wb = Workbook
# codingsheet = wb.add_sheet("CodingSheet")

begin = time.time()


def date_time():
    today = date.today()
    # print("Today's date:", today)
    tz_In = pytz.timezone('Asia/Kolkata')
    datetime_In = datetime.now(tz_In)
    # print("IN time:", datetime_In.strftime("%I;%M;%S %p"))
    time = datetime_In.strftime("[%I-%M-%S %p]")
    now = str(today) + time
    print(now)
    # logger.info("Date and Time captured!")
    return (now)


# sys.stdout=open("codingsheet.txt","w")

def makedir(foldername):
    path1 = str(foldername)
    if not os.path.exists(path1):
        try:
            os.mkdir(path1)
            return path1
        except OSError as error:
            print(error)
            return False
    else:
        try:
            shutil.rmtree(path1)
            os.mkdir(path1)
            return path1
        except OSError as error:
            print(error)
            return False
def driver_setup(path):
    options = webdriver.ChromeOptions()
    options.add_argument("--disable-notifications")
    options.add_argument("--start-maximized")
    options.add_argument(locator.chrome_profile_path)  # Path to your chrome profile
    if guiwindow.headlessmode == 1:
        options.add_argument("--headless")

    options.add_argument('--disable-gpu')
    # options.add_argument("--window-size=1920,1080")
    # options.add_argument("--start-maximized")
    options.add_argument("--no-sandbox")
    options.add_argument("--dns-prefetch-disable")
    preferences = {
        "download.default_directory": "c:\\VerificationReports\\TaskIngestion\\AuditLogDownload"}
    options.add_experimental_option("prefs", preferences)
    global driver
    driver = webdriver.Chrome(executable_path=path, options=options)
    print(guiwindow.Window_location)
    if guiwindow.Window_location == 1:
        driver.set_window_position(-1000, 0)
    elif guiwindow.Window_location == 0:
        driver.set_window_position(1000, 0)
    driver.maximize_window()
    driver.implicitly_wait(0.75)
    return driver


def setup(val):
    # #val.lower() == "firefox":
    # driver = webdriver.Firefox(executable_path=r"C:\\Users\\ssrivastava\\PycharmProjects\\Python_Practise\\driver2\\geckodriver.exe")
    # driver.implicitly_wait(10)
    # title = driver.get("https://stage.cozeva.com/user/login")
    # driver.maximize_window()
    if val.lower() == "chrome":
        options = Options()
        options.add_argument("--disable-notifications")
        # change for user
        options.add_argument("user-data-dir=C:\\Users\\ssrivastava\\PycharmProjects\\codingtooldata1")
        # options.add_experimental_option("detach", True)
        # self.driver = webdriver.Chrome(executable_path="../chromedriver.exe", chrome_options=options)
        # change for user
        preferences = {
            "download.default_directory": locator.download_dir}
        options.add_experimental_option("prefs", preferences)
        # change for user
        driverpath = "C:\\Users\\ssrivastava\\PycharmProjects\\Python_Practise\\driver2\\chromedriver.exe"
        driver = webdriver.Chrome(driverpath, options=options)
        driver.get("https://www.cozeva.com/user/logout")
        title = driver.get("https://www.cozeva.com/user/login")
        driver.maximize_window()
    return driver

def login_to_cozeva(CusID, path):
    driver.get(locator.logout_link)
    driver.get(locator.login_link)
    driver.maximize_window()
    file = open(path, "r+")
    global details
    details = file.readlines()
    driver.find_element_by_id("edit-name").send_keys(details[0].strip())
    driver.find_element_by_id("edit-pass").send_keys(details[1].strip())
    file.seek(0)
    file.close()
    driver.find_element_by_id("edit-submit").click()
    time.sleep(4)
    otpurl = driver.current_url
    sub_str = "/twostepAuthSettings"
    if otpurl.find(sub_str) != -1:
        # print("Need to enter OTP for login. Please paste the OTP here")
        # wait_time = 60
        # start_time = time.perf_counter()
        # otp = ""
        # while (time.perf_counter() - start_time) < wait_time:
        #     otp = input()
        #     if len(otp) > 0:
        #         print("OTP Recieved")
        # if len(otp) == 0:
        #     print("You did not enter an OTP!!")
        #     exit(999)

        timeout = 60
        t = Timer(timeout, print, ['You did not enter the OTP'])
        t.start()
        prompt = "You have %d seconds to enter the OTP here\n" % timeout
        otp = input(prompt)
        t.cancel()

        driver.find_element_by_id("edit-twostep-code").send_keys(otp)
        time.sleep(1)

        driver.find_element_by_id("edit-twostep").click()

    WebDriverWait(driver, 120).until(EC.presence_of_element_located((By.ID, "reason_textbox")))
    driver.find_element_by_id("reason_textbox").send_keys(details[4].strip())
    time.sleep(0.5)

    global cust_switched
    cust_switched = 0
    try:
        # trying to switch customercontext before registries load
        dropdown_element = Select(driver.find_element(By.ID, "select-customer"))
        time.sleep(1)
        dropdown_element.select_by_value(CusID)
        # dropdown_element.select_by_visible_text("OPTUM")
        time.sleep(0.5)
        cust_switched = 1
    except Exception as e:
        cust_switched = 0
        traceback.print_exc()

    driver.find_element_by_id("edit-submit").click()
    sf.ajax_preloader_wait(driver)
    WebDriverWait(driver, 30).until(
        EC.presence_of_element_located((By.XPATH, locator.xpath_filter_measure_list)))
    print("Logged in to Cozeva!")

def wait_to_load(element_xpath):
    try:
        WebDriverWait(driver, 100).until(EC.presence_of_element_located((By.XPATH, element_xpath)))

    except TimeoutException:
        print("Time out loaded ")
        pass

def wait_to_load_service_year(element_xpath):
    try:
        WebDriverWait(driver, 100).until(EC.visibility_of_element_located((By.XPATH, element_xpath)))
    except TimeoutException:
        print("Time out loaded ")
        pass

def action_click(driver,element):
    try:
        element.click()
    except (ElementNotInteractableException, ElementClickInterceptedException):
        driver.execute_script("arguments[0].click();", element)
    except StaleElementReferenceException:
        driver.refresh()
        WebDriverWait(driver, 300).until(EC.invisibility_of_element((By.XPATH, "//div[@class=\"ajax_preloader\"]")))
        check_xpath="//td[@class=' chart_chase_id_select']"
        WebDriverWait(driver, 300).until(EC.element_to_be_clickable((By.XPATH, check_xpath)))
        driver.execute_script("arguments[0].click();", element)



def check_exists_by_xpath(driver, xpath):
    try:
        driver.find_element_by_xpath(xpath)
    except NoSuchElementException:
        return False
    return True


def extractchoiceinfo(taskid, metricid):
    choice_option_xpath = "//tr[@class=\"metric-tr qc_enabled\" and @data-subtask-id={}]//following-sibling::tr[@metric-id=\"{}\"]//child::td[@class=\"choice_option\"]//div[@class=\"select-wrapper qc_enabled tooltipped\"]".format(
        taskid, metricid)
    choice_options = driver.find_elements_by_xpath(choice_option_xpath)
    choice_value = []
    for choice_option in choice_options:
        choice_value.append(choice_option.get_attribute("data-tooltip"))
    return choice_value


def extractcodeinfo(taskid, metricid,code_entered_xpath):
    metric_name_xpath = "//tr[@class=\"metric-tr qc_enabled\" and @data-subtask-id={}]//following::div[@class=\"met-name mrb grey-text text-darken-4\"]/span[1]".format(
        taskid)
    code_xpath = "//tr[@class=\"metric-tr qc_enabled\" and @data-subtask-id={}]//following-sibling::tr[@metric-id={}]//child::td[@class=\"codes\"]//div[@class=\"tooltipped qc_enabled\"]//input".format(
        taskid, metricid)
    bp_code_xpath = "//tr[@class=\"metric-tr qc_enabled\" and @data-subtask-id={}]//following-sibling::tr[@metric-id={}]//child::td[@class=\"codes\"]//div[@class=\"codebox_inbox\"]//input".format(
        taskid, metricid)
    code_value = []
    metric_name = driver.find_element_by_xpath(metric_name_xpath).text
    if "BP Control" in metric_name or "Body Mass" in metric_name:
        codes = driver.find_elements_by_xpath(bp_code_xpath)
        for mcode in codes:
            code_value.append(mcode.get_attribute('value'))
    else:
        codes = driver.find_elements_by_xpath(code_entered_xpath)
        for mcode in codes:  # mcode is  the code for list of codes obtained
            code_value.append(mcode.get_attribute("data-tooltip"))

    return code_value

def validateDate(date):
    format = "%m/%d/%Y"
    res = []
    for dat in date:

        try:
            res.append(bool(datetime.strptime(dat, format)))
        except ValueError:
            res.append(False)

    if False in res:
        return False
    else:
        return True


def extractdateinfo(taskid, metricid):
    diff = 0
    date_xpath = "//tr[@class=\"metric-tr qc_enabled\" and @data-subtask-id={}]//following-sibling::tr[@metric-id={}]//child::td[@class=\"service_date\"]//child::input".format(
        taskid, metricid)
    date_elements = driver.find_elements_by_xpath(date_xpath)
    if len(date_elements) == 0:
        diff = 1
        date_xpath = "//tr[@class='metric-tr qc_enabled' and @data-subtask-id={}]//following-sibling::tr[@metric-id={}]//child::td[@class='service_date custom-disabled']//child::input".format(
            taskid, metricid)
        date_elements = driver.find_elements_by_xpath(date_xpath)
    dates = []
    for date_element in date_elements:
        dates.append(date_element.get_attribute("value"))
    return dates


def extractrenderingproviderinfo(taskid, metricid):
    rprov_xpath = "//tr[@class=\"metric-tr qc_enabled\" and @data-subtask-id={}]//following-sibling::tr[@metric-id={}]//child::td[@class=\"rendering_provider\"]//child::input".format(
        taskid, metricid)
    rprov_elements = driver.find_elements_by_xpath(rprov_xpath)
    if (len(rprov_elements) == 0):
        rprov_xpath = "//tr[@class='metric-tr qc_enabled' and @data-subtask-id={}]//following-sibling::tr[@metric-id={}]//child::td[@class='rendering_provider custom-disabled']//child::input".format(
            taskid, metricid)
        rprov_elements = driver.find_elements_by_xpath(rprov_xpath)
    rendering_provider = []
    for rprov_element in rprov_elements:
        rendering_provider.append(rprov_element.get_attribute("value"))
    return rendering_provider


def extractchoiceinfo_old():
    choice_xpath = "//tr[@class=\"choice_tr saved_data  custom-disabled\"]//child::td[@class=\"choice_option\"]//input"
    choice_options = driver.find_elements_by_xpath(choice_xpath)
    choice_value = []
    for choice_option in choice_options:
        choice_value.append(choice_option.get_attribute("value"))
    return choice_value
def code_to_rows(t, measure_abbreviation, dates, choicevalues, codevalues, ren_provider):
    coded_string = []
    number_of_rows = len(choicevalues)
    list1=[]
    list2=[]
    lst=[list1]
    count=0
    if ("BP" in measure_abbreviation):
        end = int(len(codevalues) / 4)
        # print(end)
        p = 0
        for j in range(0, end):
            lab_value = codevalues[p] + "/" + codevalues[p + 1]
            # print(lab_value)
            office_value = codevalues[p + 2]
            # print(office_value)
            if len(dates) == 0:
                dategiven = "None"
            else:
                try:
                    dategiven = dates[j]
                except IndexError:
                    dategiven = dates[0]
            if len(ren_provider) == 0:
                ren_providergiven = "None"
            else:
                try:
                    ren_providergiven = ren_provider[j]
                except IndexError:
                    ren_providergiven = ren_provider[0]

            s1 = "{}`{}`{}`{}`{}".format(t, measure_abbreviation, dategiven, lab_value, ren_providergiven)
            coded_string.append(s1)
            list1.append(t)
            list1.append(measure_abbreviation)
            list1.append(dategiven)
            list1.append(choicevalues[0])
            list1.append(lab_value)
            list1.append(ren_provider[0])
            s2 = "{}`{}`{}`{}`{}".format(t, measure_abbreviation, dategiven, office_value, ren_providergiven)
            coded_string.append(s2)
            list2.append(t)
            list2.append(measure_abbreviation)
            list2.append(dategiven)
            list2.append(choicevalues[0])
            list2.append(office_value)
            list2.append(ren_provider[0])
            p = p + 4
            lst=[list1,list2]
            count=2
    else:
        for i in range(0, number_of_rows):
            try:
                code = codevalues[i] + "(" + codevalues[i + 1] + ")"
            except:
                code = codevalues[i]

            if len(dates) == 0:
                dategiven = "None"
            else:
                try:
                    dategiven = dates[i]
                except IndexError:
                    dategiven = dates[0]

            if len(ren_provider) == 0:
                ren_providergiven = "None"
            else:
                try:
                    ren_providergiven = ren_provider[i]
                except IndexError:
                    ren_providergiven = ren_provider[0]

            s = "{}`{}`{}`{}`{}".format(t, measure_abbreviation, dategiven, code, ren_providergiven)
            list1.append(t)
            list1.append(measure_abbreviation)
            list1.append(dategiven)
            list1.append(choicevalues)
            list1.append(code)
            list1.append(ren_provider)
            lst=[list1]
            count=1
    return count,lst

def code_to_string(t, measure_abbreviation, dates, choicevalues, codevalues, ren_provider):
    coded_string = []
    number_of_rows = len(choicevalues)
    list1=[]
    list2=[]
    if ("BP" in measure_abbreviation):
        end = int(len(codevalues) / 4)
        # print(end)
        p = 0
        for j in range(0, end):
            lab_value = codevalues[p] + "/" + codevalues[p + 1]
            # print(lab_value)
            office_value = codevalues[p + 2]
            # print(office_value)
            if len(dates) == 0:
                dategiven = "None"
            else:
                try:
                    dategiven = dates[j]
                except IndexError:
                    dategiven = dates[0]
            if len(ren_provider) == 0:
                ren_providergiven = "None"
            else:
                try:
                    ren_providergiven = ren_provider[j]
                except IndexError:
                    ren_providergiven = ren_provider[0]

            s1 = "{}`{}`{}`{}`{}".format(t, measure_abbreviation, dategiven, lab_value, ren_providergiven)
            coded_string.append(s1)
            s2 = "{}`{}`{}`{}`{}".format(t, measure_abbreviation, dategiven, office_value, ren_providergiven)
            coded_string.append(s2)
            p = p + 4
    else:
        for i in range(0, number_of_rows):
            try:
                code = codevalues[i] + "(" + codevalues[i + 1] + ")"
            except:
                code = codevalues[i]

            if len(dates) == 0:
                dategiven = "None"
            else:
                try:
                    dategiven = dates[i]
                except IndexError:
                    dategiven = dates[0]

            if len(ren_provider) == 0:
                ren_providergiven = "None"
            else:
                try:
                    ren_providergiven = ren_provider[i]
                except IndexError:
                    ren_providergiven = ren_provider[0]

            s = "{}`{}`{}`{}`{}".format(t, measure_abbreviation, dategiven, code, ren_providergiven)
            coded_string.append(s)
    return coded_string


def extractdateinfo_old():
    diff = 0
    date_xpath = "//tr[@class=\"choice_tr saved_data  custom-disabled\"]//child::td[@class=\"service_date\"]//descendant::input"

    date_elements = driver.find_elements_by_xpath(date_xpath)
    if len(date_elements) == 0:
        diff = 1
        date_xpath = "//tr[@class='choice_tr saved_data  custom-disabled']//child::td[@class='service_date custom-disabled']//descendant::input"
        date_elements = driver.find_elements_by_xpath(date_xpath)
    dates = []
    for date_element in date_elements:
        dates.append(date_element.get_attribute("value"))
    return dates


def extractrenderingproviderinfo_old():
    rprov_xpath = "//tr[@class=\"choice_tr saved_data  custom-disabled\"]//child::td[@class=\"rendering_provider\"]//descendant::input"
    rprov_elements = driver.find_elements_by_xpath(rprov_xpath)
    if (len(rprov_elements) == 0):
        rprov_xpath = "//tr[@class='choice_tr saved_data  custom-disabled']//child::td[@class='rendering_provider custom-disabled']//descendant::input"
        rprov_elements = driver.find_elements_by_xpath(rprov_xpath)
    rendering_provider = []
    for rprov_element in rprov_elements:
        rendering_provider.append(rprov_element.get_attribute("value"))
    return rendering_provider

def check_exists_by_xpath(driver,xpath):
    try:
        driver.find_element_by_xpath(xpath)
    except NoSuchElementException:
        return False
    return True

#
# def validate_date_displayed(date)
#     date_string = time_displayed_list[0]
#     format = "%m/%d/%Y"
#     date_status=""
#     try:
#         x.datetime.strptime(date_string, format)
#         date_status="PASS"
#         print("This is the correct date string format.")
#         logger.info("This is the correct date string format.")
#     except ValueError:
#         date_status="FAIL"
#         print("This is the incorrect date string format. It should be MM-DD-YYYY")
#         logger.error("This is the incorrect date string format. It should be MM-DD-YYYY")

def extractcodeinfo_old():
    code_xpath = "//tr[@class=\"choice_tr saved_data  custom-disabled\"]//child::td[@class=\"codes\"]//descendant::input"
    code_values = driver.find_elements_by_xpath(code_xpath)
    code_list = []
    for code_value in code_values:
        code_list.append(code_value.get_attribute("value"))
    return code_list


def report_write_func(report,row,task_id,measure_abb,service_date,choice,code,rendering_provider):
    report.write(row, 0, task_id)
    report.write(row, 1, measure_abb)
    report.write(row, 2, service_date)
    report.write(row, 3, choice)
    report.write(row, 4, code)
    report.write(row,5,rendering_provider)


def verify_codingsheetHCC(driver,workbook,logger,run_from, path_task):

    try:
        workbook.create_sheet("CodingToolHCC")
        ws = wb.active  # This will set the currently active sheet
        ws.title = "Report"
        sh1 = workbook["CodingToolHCC"]
        if (run_from == "CozevaSupport"):
            # initialize report
            sh1['A1'].value = "Task"
            sh1['A1'].font = Font(bold=True, size=13)
            sh1['B1'].value = "Service Date Match"  # if missing , test case is fail
            sh1['B1'].font = Font(bold=True, size=13)
            sh1['C1'].value = "Month = "+str(selected_month)  # if missing , test case is fail
            sh1['C1'].font = Font(bold=True, size=13)
            sh1['D1'].value = "Rendering Provider Match" # if missing , test case is fail
            sh1['D1'].font = Font(bold=True, size=13)
            sh1['E1'].value = "Measure/Condition Match"  #if missing , ignore
            sh1['E1'].font = Font(bold=True, size=13)
            sh1['F1'].value = "Code & Value match "  #if missing , ignore
            sh1['F1'].font = Font(bold=True, size=13)
            sh1['G1'].value = "Review 1 Present"
            sh1['G1'].font = Font(bold=True, size=13)
            sh1['H1'].value = "Audit Log"
            sh1['H1'].font = Font(bold=True, size=13)
            sh1['I1'].value = "No Annotation/Notes"
            sh1['I1'].font = Font(bold=True, size=13)
            sh1['J1'].value = "URL"
            sh1['J1'].font = Font(bold=True, size=13)

    except Exception as e:
        print("Report Sheet not created ")
    # column_xpath = "//*[@id=\"chart_chase\"]/thead/tr/th[2]"
    # wait_to_load(column_xpath)
    # wait_to_load(column_xpath)
    hamburger_xpath = "//i[@class=\"material-icons\" and text()=\"menu\"]"
    hamburger = driver.find_element_by_xpath(hamburger_xpath)
    action_click(driver,hamburger)

    side_bar_nav_xpath = "//ul[@id=\"sidenav_slide_out\"]"
    wait_to_load(side_bar_nav_xpath)
    choice=config.get("HCC","choice")
    print(choice)
    if(choice=="HCC"):
        supp_data_xpath = "//i[@class=\"material-icons sidenav_main\"]//following-sibling::span[text()=\"HCC Chart List\"]"
    else:
        supp_data_xpath = "//i[@class=\"material-icons sidenav_main\"]//following-sibling::span[text()=\"AWV Chart List\"]"
    print(supp_data_xpath)
    supp_data = driver.find_element_by_xpath(supp_data_xpath)
    action_click(driver,supp_data)
    sf.ajax_preloader_wait(driver)

    #wait_to_load(column_xpath)
    # remove other tag

    # other_tag_close_xpath='//div[@class="dt_tag_wrapper"]//child::span[text()="close"]'
    # other_tag_close=driver.find_element_by_xpath(other_tag_close_xpath)
    # action_click(other_tag_close)
    # WebDriverWait(driver, 30).until(EC.invisibility_of_element((By.XPATH, "//div[@class=\"ajax_preloader\"]")))
    # wait_to_load(column_xpath)
    # click on task -> open in new tab ->print element -> Close the tab ->Click on next task
    # random_page_xpath = "//button[text()=\"4\"]"
    # random_page = driver.find_element_by_xpath(random_page_xpath)
    # action_click(random_page)
    filter_list_xpath = "//i[text()=\"filter_list\"]"
    filter_list = driver.find_element_by_xpath(filter_list_xpath)
    filter_list.click()

    #take date from locator config
    new_creation_date_filter_from_xpath = "//input[@name='chart_chase_uploaded_from']"
    new_creation_date_filter_to_xpath = "//input[@name='chart_chase_uploaded_to']"

    date_filter_from = driver.find_element_by_xpath(new_creation_date_filter_from_xpath)
    date_filter_from.clear()
    date_filter_from_input=config.get("HCC","date_from")
    date_filter_from.send_keys(date_filter_from_input)  # start date

    date_filter_to = driver.find_element_by_xpath(new_creation_date_filter_to_xpath)
    date_filter_to.clear()
    date_filter_to_input = config.get("HCC", "date_to")
    date_filter_to.send_keys(date_filter_to_input)  # end date
    apply_xpath = "//a[text()=\"Apply\"]"
    apply = driver.find_element_by_xpath(apply_xpath)
    apply.click()
    WebDriverWait(driver, 300).until(EC.invisibility_of_element((By.XPATH, "//div[@class=\"ajax_preloader\"]")))
    #wait_to_load(column_xpath)


    #read file

    file1 = open(path_task, 'r')
    Lines = file1.readlines()

    count = 0
    for line in Lines:
        workbook.save("Report_HCCCoding.xlsx")
        count += 1
        task_id=line.strip()
        print("Task{}: {}".format(count, line.strip()))
        WebDriverWait(driver, 300).until(EC.invisibility_of_element((By.XPATH, "//div[@class=\"ajax_preloader\"]")))

        #click on filter
        wait_to_load(filter_list_xpath)
        filter_list=driver.find_element_by_xpath(filter_list_xpath)
        action_click(driver,filter_list)
        #enter task input
        task_input_xpath="//input[@title='Task #']"
        task_input=driver.find_element_by_xpath(task_input_xpath)
        task_input.location_once_scrolled_into_view
        task_input.clear()
        task_input.send_keys(task_id)
        #click on apply
        apply_xpath = "//a[text()=\"Apply\"]"
        apply = driver.find_element_by_xpath(apply_xpath)
        apply.location_once_scrolled_into_view
        action_click(driver,apply)
        time.sleep(1)
        #wait for page to load
        #WebDriverWait(driver, 300).until(EC.invisibility_of_element((By.XPATH, "//div[@class=\"ajax_preloader\"]")))
        sf.ajax_preloader_wait(driver)
        task_xpath="//th[@class='chart_chase_task sorting']"
        wait_to_load(task_xpath)

        #verify if task is there

        task_xpath = "//td[@class=' chart_chase_task']//child::a"
        task_found=False
        try:
            task_link = driver.find_element_by_xpath(task_xpath)
            task_found=True
        except NoSuchElementException as e:

            sh1.append((task_id,"Task Not Found"))
            workbook.save("Report_HCCCoding.xlsx")
            continue
        if(task_found==True):
            # extract value service date
            service_date_claim_link_xpath = "//div[@class='col s3 enc_service_date '] //child::a"
            service_date_xpath = "//div[@class='col s3 enc_service_date ']"

            service_date = []
            if (check_exists_by_xpath(driver, service_date_xpath)):
                wait_to_load_service_year(service_date_xpath)
                service_date_data = driver.find_elements_by_xpath(service_date_xpath)
                print(len(service_date_data))
                for q in range(0, len(service_date_data)):
                    check_claim_link_xpath = "(" + service_date_xpath + ")" + "[" + str(q + 1) + "]" + "//child::a"
                    if (check_exists_by_xpath(driver, check_claim_link_xpath)):
                        print("Claim link exists")
                        service_date_value = driver.find_element_by_xpath(check_claim_link_xpath)
                        service_date.append(service_date_value.get_attribute("innerHTML"))
                    elif (len(service_date_data) == 1):
                        print("Service date 1 found")
                        service_date.append(service_date_data[0].get_attribute('innerHTML'))
                    else:
                        service_date_value_xpath = "(" + service_date_xpath + ")" + "[" + str(q + 1) + "]"
                        service_date_value = driver.find_element_by_xpath(service_date_value_xpath)
                        service_date.append(service_date_value.get_attribute("innerHTML"))

            #     if(len(service_date_data)==1):
            #         print("Service date 1 found")
            #         service_date.append(service_date_data[0].get_attribute('innerHTML'))
            #     else:
            #
            #
            # if (check_exists_by_xpath(driver, service_date_claim_link_xpath)):
            #     print("Service date claim link")
            #     # number of service year is one
            #     service_date_data_claim = driver.find_elements_by_xpath(service_date_claim_link_xpath)
            #     if (len(service_date_data_claim) == 1):
            #
            #         service_date.append(service_date_data_claim[0].get_attribute('innerHTML'))
            #     else:
            #         for q in range(0, len(service_date_data_claim)):
            #             service_date.append(service_date_data_claim[q].get_attribute("innerHTML"))

            print(service_date)
            #check if task is in selected month

            month_match = ""
            for taskdate in service_date:
                month_number = taskdate.split('/')[0]

                months = ['January', 'February', 'March', 'April', 'May', 'June', 'July',
                          'August', 'September', 'October', 'November', 'December']

                if int(months.index(str(selected_month)))+1 == int(month_number):
                    month_match = "PASS"
                else:
                    month_match = months[int(month_number)-1]+" : FAIL"
                    break






            # extract rendering provider

            rendering_provider_xpath = "//div[@class='col s3 enc_prov td_vert_flex tooltipped']//child::span[1]"

            rendering_provider = []
            if (check_exists_by_xpath(driver, rendering_provider_xpath)):
                print("Rendering PRovider found")
                # number of service year is one
                rendering_provider_data = driver.find_elements_by_xpath(rendering_provider_xpath)
                if (len(rendering_provider_data) == 1):
                    print("1 provider")
                    rendering_provider.append(rendering_provider_data[0].get_attribute('innerHTML'))
                else:
                    for q in range(0, len(rendering_provider_data)):
                        rendering_provider.append(rendering_provider_data[q].get_attribute("innerHTML"))

            print(rendering_provider)

            # extract condition

            condition_xpath = "//div[@class='col s3 enc_meas tooltipped']"

            condition = []
            if (check_exists_by_xpath(driver, condition_xpath)):
                print("Condition found")
                # number of service year is one
                condition_data = driver.find_elements_by_xpath(condition_xpath)
                if (len(condition_data) == 1):
                    print("1 condition")
                    condition.append(condition_data[0].get_attribute('innerHTML'))
                else:
                    for q in range(0, len(condition_data)):
                        condition.append(condition_data[q].get_attribute("innerHTML"))

            print(condition)

            # extract code

            code_xpath = "//div[@class='col s6 enc_code tooltipped']//child::div"

            code = []
            if (check_exists_by_xpath(driver, code_xpath)):
                print("Code found")
                # number of service year is one
                code_data = driver.find_elements_by_xpath(code_xpath)
                if (len(code_data) == 1):
                    print("1 code")
                    code.append(code_data[0].get_attribute('innerHTML'))
                else:
                    for q in range(0, len(code_data)):
                        code.append(code_data[q].get_attribute("innerHTML"))

            print(code)

            #Check Review 1 status
            review1_status = driver.find_element_by_id("chart_chase").find_element_by_tag_name("tbody").find_element_by_class_name("chart_chase_review_6").text

            if "Completed" in review1_status:
                review_status = "PASS"
            else:
                review_status = "FAIL"

            logger.info(str(task_id) + str(service_date) + str(rendering_provider) + "Condition " + str(
                condition) + " Code " + str(code))

            # click on task
            task_xpath = "//td[@class=' chart_chase_task']//child::a"

            task_link = driver.find_element_by_xpath(task_xpath)
            WebDriverWait(driver, 300).until(EC.element_to_be_clickable((By.XPATH, task_xpath)))

            try:
                action_click(driver, task_link)

            except:

                sh1.append((str(task_id), "Unable to open task , please check again"))
                workbook.save("Report_HCCCoding.xlsx")
                logger.error("Unable to open task")
                continue

            logger.info("Clicked on Task ")
            time.sleep(2)
            # open the task in new tab
            driver.switch_to.window(driver.window_handles[1])
            WebDriverWait(driver, 300).until(EC.invisibility_of_element((By.XPATH, "//div[@class=\"ajax_preloader\"]")))
            cpt_xpath = "//tr[@class='cpt_header']"
            wait_to_load(cpt_xpath)
            tbody_xpath = "//tbody//child::*"
            wait_to_load(tbody_xpath)
            # store the service date

            encounter_date_xpath = "//div[@id='encounter_wrap']//child::span[@class='td-row2 mls encounter_date']"

            encounter_date = []
            if (check_exists_by_xpath(driver, encounter_date_xpath)):

                enquire_date_data = driver.find_elements_by_xpath(encounter_date_xpath)
                for q in range(0, len(enquire_date_data)):
                    encounter_claim_xpath = "(" + encounter_date_xpath + ")" + "[" + str(q + 1) + "]" + "//child::a"
                    if (check_exists_by_xpath(driver, encounter_claim_xpath)):
                        print("Claim link exists in encounter")
                        encounter_date_value = driver.find_element_by_xpath(encounter_claim_xpath)
                        encounter_date.append(encounter_date_value.get_attribute("innerHTML"))
                    else:
                        encounter_date_value_xpath = "(" + encounter_date_xpath + ")" + "[" + str(q + 1) + "]"
                        encounter_date_value = driver.find_element_by_xpath(encounter_date_value_xpath)
                        encounter_date.append(encounter_date_value.get_attribute("innerHTML"))

            else:
                encounter_date = "Not found"
            print(encounter_date)

            # store the rendering provider

            encounter_provider_xpath = "//div[@id='encounter_wrap']//child::span[@class='td-row2 mls']"
            encounter_provider = []
            if (check_exists_by_xpath(driver, encounter_provider_xpath)):
                print("Encounter Information Rendering PRovider")
                provider_names = driver.find_elements_by_xpath(encounter_provider_xpath)
                if (len(provider_names) == 1):
                    print("1 provider")
                    encounter_provider.append(provider_names[0].get_attribute('prov_name'))
                else:
                    for q in range(0, len(provider_names)):
                        encounter_provider.append(provider_names[q].get_attribute("prov_name"))


            else:
                encounter_provider = "Not found"

            print(encounter_provider)

            logger.info("In coding tool Encounter information : Service Date" + str(
                encounter_date) + " Rendering Provider " + str(encounter_provider))

            # compare lists

            # service date
            service_date.sort()
            encounter_date.sort()
            service_date_matched = ""
            if service_date == encounter_date :
                logger.info("Service date matched")
                service_date_matched = "PASS"
            else:
                if (encounter_date == "Not found"):
                    logger.error("Encounter not displayed in coding sheet ")
                else:
                    logger.error("Chart List shows missing date")
                logger.error("Service date is missing ")
                service_date_matched = "FAIL"

            # rendering provider
            rendering_provider.sort()
            encounter_provider.sort()
            rendering_provider_matched = ""
            # reversed
            if rendering_provider == encounter_provider:
                logger.info("Rendering Provider matched")
                rendering_provider_matched = "PASS"
            else:
                if (encounter_provider == "Not found"):
                    logger.error("Encounter not displayed in coding sheet ")
                else:
                    logger.error("Chart List does not shows rendering provider")
                logger.error("Rendering Provider mismatched ")
                rendering_provider_matched = "FAIL"

            # find   codes and conditions store the ones not displayed
            time.sleep(5)
            condition_found = ""
            not_found = 0
            not_found_list = []
            AWV_count = condition.count("AWV")

            # checking if it is more then 0
            if AWV_count > 0:
                condition.remove("AWV")

            for cond in condition:
                cond_xpath = "//*[text()='" + str(cond) + "']"
                print(cond_xpath)
                if (check_exists_by_xpath(driver, cond_xpath)):
                    logger.info(str(cond) + " Found in coding sheet")
                else:
                    not_found += 1
                    not_found_list.append(cond)
                    logger.error(str(cond) + " Not Found")
            if (not_found == 0):
                condition_found = "PASS"
            else:
                condition_found = "FAIL" + " Missing " + str(cond)

            print(condition_found)

            # check codes
            code_found = 0
            code_not_found_list = []
            for cod in code:
                cod_xpath = "//*[text()='" + str(cod) + "']"
                print(cod_xpath)
                if (check_exists_by_xpath(driver, cod_xpath)):
                    logger.info(str(cod) + " Found in coding sheet")
                else:
                    code_found += 1
                    code_not_found_list.append(cod)
                    logger.error(str(cod) + " Not Found")
            if (code_found == 0):
                code_status = "PASS"
            else:
                code_status = "FAIL" + " Missing " + str(cod)

            print("Code status ", code_status)

            #Check Audit log download, data present
            audit_log_download = ""

            try:
                driver.find_element(By.XPATH, locator.xpath_coding_tool_kebab).click()
                time.sleep(1)
                driver.find_element(By.XPATH, locator.xpath_audit_log_download).click()
                time.sleep(2)

                onlyfiles = [f for f in os.listdir("c:\\VerificationReports\\TaskIngestion\\AuditLogDownload") if
                             isfile(join("c:\\VerificationReports\\TaskIngestion\\AuditLogDownload", f))]
                audit_log_path = "c:\\VerificationReports\\TaskIngestion\\AuditLogDownload\\" + onlyfiles[0]

                def is_csv_file_empty(file_path):
                    with open(file_path, 'r') as csvfile:
                        reader = csv.reader(csvfile)
                        for row in reader:
                            if row:  # Check if the row is not empty
                                return False  # File has data
                    return True  # File is empty

                # Example usage
                file_path = audit_log_path
                if is_csv_file_empty(file_path):
                    audit_log_download = "FAIL"
                else:
                    audit_log_download = "PASS"

                os.remove(audit_log_path)

            except Exception as e:
                traceback.print_exc()
                audit_log_download = "FAIL"

            #Check that no annotations are present
            try:
                annotation_check = ""

                annotation_count = driver.find_element(By.XPATH, locator.xpath_annotation_tab).get_attribute("data-count")

                if annotation_count == "0":
                    annotation_check = "PASS"
                else:
                    annotation_check = "FAIL"
            except Exception as e:
                traceback.print_exc()
                annotation_check = "FAIL"






            # write status in report

            sh1.append((str(task_id), str(service_date_matched), str(month_match), str(rendering_provider_matched), str(condition_found),
                        str(code_status), str(review_status),str(audit_log_download),str(annotation_check), str(driver.current_url)))
            workbook.save("Report_HCCCoding.xlsx")
            # if failed, provide task number,Encounter Mismatch/Code Not found, URL

            # close the tab and move to original tab
            driver.close()
            driver.switch_to.window(driver.window_handles[0])
            logger.info("Window closed and moved to previous tab ")

    rows = sh1.max_row
    cols = sh1.max_column
    try:
        for i in range(1, rows + 1):
            for j in range(1, cols + 1):
                if sh1.cell(i, j).value == 'PASS':
                    sh1.cell(i, j).fill = PatternFill('solid', fgColor='0FC404')
                elif 'FAIL' in sh1.cell(i, j).value:
                    sh1.cell(i, j).fill = PatternFill('solid', fgColor='FC0E03')
    except Exception as e:
        traceback.print_exc()
    workbook.save("Report_HCCCoding.xlsx")
    # apply link status

    # link_dropdown_xpath='//div[text()="Link Status:"]//parent::div//following-sibling::div'
    # link_dropdown=driver.find_element_by_xpath(link_dropdown_xpath)
    # action_click(link_dropdown)
    #
    # selected_value_from_link_xpath='//div[text()="Link Status:"]//parent::div//following-sibling::div//ul//child::span[text()="Linked"]'
    # select_value=driver.find_element_by_xpath(selected_value_from_link_xpath)
    # action_click(select_value)



    # wait_to_load(column_xpath)
    #
    # # find no of pages
    # WebDriverWait(driver, 100).until(EC.invisibility_of_element((By.XPATH, "//div[@class=\"ajax_preloader\"]")))
    #
    # # find no of pages
    # pages_xpath = "//button[@class='mdl-button  mdl-button--raised mdl-button--colored']"
    # pages = driver.find_elements_by_xpath(pages_xpath)
    # print(len(pages))
    # num_pages = pages[len(pages) - 1].text
    #
    #
    # i = 2
    # # for i in range(num_pages+1):
    # #     pages_xpath="//button[@class=\"mdl-button \" and text()="{}"]".format(i)
    #
    # #########task search from filter ############
    #
    # taskid_xpath = "//input[@name='chart_chase_task']"
    # parent_taskid_xpath = "//input[@name='chart_chase_parent_task']"
    # taskid_filter = driver.find_element_by_xpath(taskid_xpath)
    # parent_taskid_filter = driver.find_element_by_xpath(parent_taskid_xpath)
    #
    # # taskid_filter.send_keys("Enter string to be searched")
    # # parent_taskid_filter.send_keys("Enter key to be searched")
    #
    # if(len(pages)>1):
    #     pages_xpath = "//button[@class='mdl-button ']"
    #     pages = driver.find_elements_by_xpath(pages_xpath)
    #     last_page = pages[len(pages) - 1].text
    # else:
    #     last_page=1
    # print("Number of pages ", last_page)
    #
    # current_page_xpath = "//button[@class='mdl-button  mdl-button--raised mdl-button--colored']"
    # current_page = driver.find_element_by_xpath(current_page_xpath).text
    #
    # print(current_page)
    #
    # next_page_value = int(current_page) + 1
    #
    # row = 1
    #
    # for i in range(int(current_page), int(last_page) + 1):
    #     print("Page no", i)
    #     for x in range(1, 21):
    #         err = False
    #         print("\n")
    #         WebDriverWait(driver, 100).until(
    #             EC.invisibility_of_element((By.XPATH, "//div[@class=\"ajax_preloader\"]")))
    #         task_xpath = "//*[@id=\"chart_chase\"]/tbody/tr[%d]/td[2]/div/div/div/a" % x;
    #         # print(task_xpath)
    #         measure_abbrev_xpath = "//*[@id=\"chart_chase\"]/tbody/tr[{}]/td[@class=\" chart_chase_service_date\"]//div[@class=\"col s3 enc_meas tooltipped tooltip_init-processed\"]".format(
    #             x)
    #         measure_abbrev = driver.find_elements_by_xpath(measure_abbrev_xpath)
    #         measure_list = []
    #         if (len(measure_abbrev) > 1):
    #             for m in measure_abbrev:
    #                 measure_list.append(m.text)
    #             measure_abbreviation = measure_list[0]
    #             print(measure_list)
    #         else:
    #             if (len(measure_abbrev) == 0):
    #                 measure_abbreviation = "Blank"
    #             else:
    #                 try:
    #
    #                     measure_abbreviation = measure_abbrev[0].text
    #                 except NoSuchElementException:
    #                     measure_abbreviation = "Blank"
    #
    #         try:
    #             task = WebDriverWait(driver, 10).until(
    #                 EC.invisibility_of_element((By.XPATH, "//div[@class=\"ajax_preloader\"]")))
    #             task = driver.find_element_by_xpath(task_xpath)
    #             action_click(task)
    #         except NoSuchElementException:
    #             if (i == int(last_page)):
    #                 break;
    #             else:
    #                 print("Error: Task not found")
    #         t = task.text
    #
    #         # print(driver.window_handles)
    #         # print(driver.current_window_handle)
    #         driver.switch_to.window(driver.window_handles[1])
    #         coding_table_xpath = "//span[text()=\"Measure\"]"
    #         wait_to_load(coding_table_xpath)
    #         metric_xpath = "//tr[@class=\"metric-tr qc_enabled\" and @data-subtask-id=\"{}\"]".format(t);
    #
    #         # check if old task or not - extract codes - match with chart list
    #
    #         try:
    #             task_id_left_xpath = "(//div[@class=\"task_labels hide lfloat\"])[1]"
    #             task_id = driver.find_element_by_xpath(task_id_left_xpath)
    #         except NoSuchElementException:
    #             task_id_left_xpath = "//div[@class=\"task_labels label-processed\"]"
    #             task_id = driver.find_element_by_xpath(task_id_left_xpath)
    #         task_id_left = task_id.get_attribute("data-task-id")
    #         old_supp_data = False
    #
    #         if check_exists_by_xpath(driver, metric_xpath):
    #             a = driver.find_element_by_xpath(metric_xpath)
    #             mid = a.get_attribute("metric-id")
    #             code_entered_xpath = "//tr[@class=\"metric-tr qc_enabled\" and @data-subtask-id=\"{}\"]//following-sibling::tr[@metric-id=\"{}\"]/child::td[@class=\"codes\"]/div/div/div[1]".format(
    #                 t, mid)
    #             if check_exists_by_xpath(driver, code_entered_xpath) == True:
    #                 code_entered = driver.find_element_by_xpath(code_entered_xpath)
    #                 code = code_entered.get_attribute("data-tooltip")
    #                 metric_abbreviation_xpath = "//tr[@class=\"metric-tr qc_enabled\" and @data-subtask-id={}]//div[@class=\"quality-metric\"]".format(
    #                     t)
    #                 metric_abbreviation_string = driver.find_element_by_xpath(metric_abbreviation_xpath).get_attribute(
    #                     "data-rch")
    #                 bad_chars = ["[", "]"]
    #                 # metric_abbreviation_json = ''.join(i for i in metric_abbreviation_string if not i in bad_chars)
    #                 # metric_abbreviation_json2=metric_abbreviation_json.split(",{")
    #                 # print(metric_abbreviation_json2[0])
    #                 #
    #                 # metric_abbreviation = json.loads(metric_abbreviation_json2[0])["abbreviation"]
    #                 metric_name_xpath = "//tr[@class=\"metric-tr qc_enabled\" and @data-subtask-id={}]//following::div[@class=\"met-name mrb grey-text text-darken-4\"]/span[1]".format(
    #                     t)
    #                 metric_name = driver.find_element_by_xpath(metric_name_xpath).text
    #                 choicevalues = extractchoiceinfo(t, mid)
    #                 codevalues = extractcodeinfo(t, mid,code_entered_xpath)
    #                 dates = extractdateinfo(t, mid)
    #                 ren_provider = extractrenderingproviderinfo(t, mid)
    #                 print(t, " task : entered ")
    #                 print("Choice : ", choicevalues)
    #                 print("Code : ", codevalues)
    #                 print("dates : ", dates)
    #                 print("Rendering Provider : ", ren_provider)
    #                 sh1.append((str(t),str(choicevalues),str(codevalues),str(dates),str(ren_provider),str(driver.current_url)))
    #                 workbook.save("Report_QualityCoding.xlsx")
    #                 codestring = code_to_string(t, measure_abbreviation, dates, choicevalues, codevalues, ren_provider)
    #                 print(codestring)
    #                 # r,rows_list=code_to_rows(t, measure_abbreviation, dates, choicevalues, codevalues, ren_provider)
    #                 # print(rows_list)
    #                 sr = 0
    #                 # print(r)
    #                 # print(rows_list[1][1])
    #                 # while sr<r:
    #                 #     report_write_func(report_sheet,row, rows_list[sr][0],rows_list[sr][1],rows_list[sr][2],rows_list[sr][3],rows_list[sr][4],rows_list[sr][5])
    #                 #     row=row+1
    #                 #     report.save(dest_filename)
    #                 #     sr=sr+1
    #
    #                 driver.close()
    #                 driver.switch_to.window(driver.window_handles[0])
    #             else:
    #                 print(t, "Error code not found")
    #                 get_url = driver.current_url
    #                 print(t, get_url)
    #                 sh1.append((str(t), "Error code not found"))
    #                 row = row + 1
    #                 workbook.save("Report_QualityCoding.xlsx")
    #                 driver.close()
    #                 driver.switch_to.window(driver.window_handles[0])
    #         else:
    #
    #             if t in task_id_left:
    #                 old_supp_data = True
    #                 print("Old task , task id: ", t)
    #                 if old_supp_data:
    #                     code_entered0_xpath = "//tr[@class=\"choice_tr saved_data  custom-disabled\"]"
    #                     if check_exists_by_xpath(driver, code_entered0_xpath):
    #                         choicevalues0 = extractchoiceinfo_old()
    #                         codevalues0 = extractcodeinfo_old()
    #                         dates0 = extractdateinfo_old()
    #                         ren_provider0 = extractrenderingproviderinfo_old()
    #                         # print(t, " task : entered ")
    #                         print("Choice : ", choicevalues0)
    #                         print("Code : ", codevalues0)
    #                         print("dates : ", dates0)
    #                         print("Rendering Provider : ", ren_provider0)
    #                         try:
    #                             codestring = code_to_string(t, measure_abbreviation, dates0, choicevalues0, codevalues0,
    #                                                         ren_provider0)
    #                         except:
    #                             err = True
    #                             codestring = "Blank"
    #
    #                         if err:
    #                             sh1.append((str(t), str(choicevalues0), str(codevalues0), str(dates0),
    #                                         str(ren_provider0), str(driver.current_url)))
    #                             workbook.save("Report_QualityCoding.xlsx")
    #                             print("Error : Blank")
    #                         else:
    #                             print(codestring)
    #                         driver.close()
    #                         driver.switch_to.window(driver.window_handles[0])
    #
    #                     else:
    #                         print("Error : No entry in coding sheet ")
    #                         get_url = driver.current_url
    #                         sh1.append((str(t), str(choicevalues0), str(codevalues0), str(dates0), str(ren_provider0),
    #                                     str(driver.current_url)))
    #                         workbook.save("Report_QualityCoding.xlsx")
    #                         print(t, get_url)
    #                         driver.close()
    #                         driver.switch_to.window(driver.window_handles[0])
    #
    #             else:
    #                 print(t, "Error : Task Not found")
    #                 sh1.append((str(t),"Task Not found"))
    #                 workbook.save("Report_QualityCoding.xlsx")
    #                 get_url = driver.current_url
    #                 print(t, get_url)
    #                 driver.close()
    #                 driver.switch_to.window(driver.window_handles[0])
    #     next_xpath = "//button[@id='chart_chase_next']"
    #     next = driver.find_element_by_xpath(next_xpath)
    #     driver.execute_script("arguments[0].scrollIntoView();", next)
    #     next.click()

def fetch_tasks():

    def save_to_file():
        file_path = source_directory + "\\assets\\Tasks.txt"

        with open(file_path, 'w') as f:
            f.write(text_widget.get(1.0, tk.END).replace(',', '').strip())

        global selected_month
        selected_month = selected_month.get()

        root.destroy()



    # Create tkinter window and configure its settings
    root = tk.Tk()
    root.title("Task Ingestion")
    root.iconbitmap("assets/icon.ico")
    #root.geometry("300x400")

    # Create a label for the dropdown widget
    month_label = tk.Label(root, text="Select Month:")
    month_label.pack(side=tk.TOP, padx=10, pady=10)

    # Create a dropdown widget for month selection
    months = ['January', 'February', 'March', 'April', 'May', 'June', 'July',
              'August', 'September', 'October', 'November', 'December']
    global selected_month
    selected_month = tk.StringVar(root)
    selected_month.set(months[0])  # Set the default value to the first month
    month_dropdown = tk.OptionMenu(root, selected_month, *months)
    month_dropdown.pack(side=tk.TOP, padx=10, pady=10)

    # Create a text widget to accept user input
    text_widget = tk.Text(root, wrap=tk.WORD, width=25)
    text_widget.pack(expand=True, fill=tk.BOTH)

    # Create a button to save the contents of the text widget to the notepad file
    save_button = tk.Button(root, text="Start Script", command=save_to_file)
    save_button.pack(side=tk.BOTTOM, pady=10)

    # Start the tkinter main loop
    root.mainloop()

#create Folder or working directory
dateandtime = date_time()
source_directory = os.getcwd()

fetch_tasks()
taskpath = source_directory + "\\assets\\Tasks.txt"
cdriver_path = source_directory + "\\assets\\chromedriver.exe"
login_file = source_directory + "\\assets\\loginInfo.txt"
master_directory=config.get("Quality","report_directory_input")
os.chdir(master_directory)
path = makedir(dateandtime)
LOG_FORMAT = "%(levelname)s %(asctime)s - %(message)s"
logging.basicConfig(filename=path + "\\" + "CodingTool-Log.log", level=logging.INFO, format=LOG_FORMAT, filemode='w')
logger = logging.getLogger()
#logger.setLevel(logging.INFO)
os.chdir(path)

downloaddefault=config.get("runner","downloaddefault")
makedir(downloaddefault)
driver = driver_setup(cdriver_path)
begin_time = datetime.now()
loc = config.get("runner","login_file")

#login
print("CustId = " + config.get("Quality","customer_id"))
login_to_cozeva(config.get("Quality", "customer_id"), login_file)
logger.info("Login successful")

#Initialize Worksheet

wb=openpyxl.Workbook()

# sm_customer_id = "1300"  # enter customer_id
# session_var = 'app_id=smart_chart&custId=' + str(sm_customer_id) + '&payerId=' + str(
#     sm_customer_id) + '&orgId=' + str(sm_customer_id)
# encoded_string = base64.b64encode(session_var.encode('utf-8'))
# driver.get("https://www.cozeva.com/smart_chart?session=" + encoded_string.decode('utf-8'))

verify_codingsheetHCC(driver, wb, logger, "CozevaSupport", taskpath)




    # convert into string
    # :26613`MeasureName`Servicedate`CodeValue`RenderingProvider`
    # chalk out a string for BP
end = time.time()
driver.quit()
print(f"Total runtime of the program is {end - begin}")
# sys.stdout.close()