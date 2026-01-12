import openpyxl

file_location = "assets/CustomerDB.xlsx"

Excel_file = openpyxl.load_workbook(file_location)

# Get workbook active sheet object
# from the active attribute
Excel_sheet = Excel_file.active

# Cell objects also have a row, column,
# and coordinate attributes that provide
# location information for the cell.

# Note: The first row or
# column integer is 1, not 0.

# Cell object is created by using
# sheet object's cell() method.
'''cell_obj = Excel_sheet.cell(row=1, column=1)'''

# Print value of cell object
# using the value attribute
'''print(cell_obj.value)'''

def getCustomerList():
    customer_list = []
    customer_name = str(Excel_sheet.cell(row=1,column=1).value).strip()
    row_counter = 2
    while customer_name != "None":
        customer_name = str(Excel_sheet.cell(row=row_counter, column=1).value).strip()
        if customer_list.count(customer_name) < 1:
            customer_list.append(customer_name)
            #print(customer_name)
        row_counter = row_counter+1
    return customer_list

def getDefaultUserNames(customer):
    username_dict = {}
    customer_name = (str)(Excel_sheet.cell(row=1, column=1).value).strip()
    row_counter = 2
    while customer_name != "None":
        customer_name = str(Excel_sheet.cell(row=row_counter, column=1).value).strip()
        if customer_name == customer:
            role_name = str(Excel_sheet.cell(row=row_counter, column=3).value).strip()
            user_name = str(Excel_sheet.cell(row=row_counter, column=4).value).strip()
            if user_name != 'None':
                username_dict.update({role_name: user_name})
        row_counter = row_counter + 1
    #print(username_dict)
    return username_dict

def fetchCustomerID(customer):
    ID = 1
    customer_name = str(Excel_sheet.cell(row=1, column=1).value).strip()
    row_counter = 2
    while customer_name != "None":
        customer_name = str(Excel_sheet.cell(row=row_counter, column=1).value).strip()
        if customer_name == customer:
            ID = str(Excel_sheet.cell(row=row_counter, column=2).value).strip()
            break
        row_counter = row_counter + 1
    return ID

def fetchCustomerName(ID):
    name=""
    sheet_ID = str(Excel_sheet.cell(row=1, column=2).value).strip()
    row_counter = 2
    while sheet_ID != "None":
        sheet_ID = str(Excel_sheet.cell(row=row_counter, column=2).value).strip()
        if sheet_ID == ID:
            name = str(Excel_sheet.cell(row=row_counter, column=1).value).strip()
            break
        row_counter+=1
    return name


# clist = getCustomerList()
# nl = []
# for c in clist:
#     nl.append(fetchCustomerID(c))
#
# print(nl)
# print(clist)

#print(fetchCustomerID("EPIC Management, L.P."))





