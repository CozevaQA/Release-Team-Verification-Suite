import time
import traceback

from openpyxl.styles import Font, PatternFill
from selenium.webdriver.common.by import By
from selenium.webdriver.support.wait import WebDriverWait
from tkinter import *
import setups
import logging
import ExcelProcessor as db
import context_functions as cf
import support_functions as sf
import setups as st
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
import variablestorage as locator
from openpyxl import Workbook, load_workbook
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import NoSuchElementException, ElementNotInteractableException, \
    ElementClickInterceptedException, TimeoutException

# Collect reource data from PROD
Client_list = db.getCustomerList()
id_list = []
for customer in Client_list:
    id_list.append(str(db.fetchCustomerID(customer).strip()))
id_list = id_list[1:len(id_list) - 1]
print(id_list)
print(len(id_list))

id_list = ['5000', '4500', '3400', '1200', '3900', '3800', '1300', '2500', '6000', '5400', '1000', '1900', '7000', '4000', '6500', '5800', '4400', '5900']


# def fetch_customer():
#     root = Tk()
#     def on_submit():
#
#         global ClientID
#         ClientID = str(client_id_entry.get())
#         root.destroy()
#         print("test")
#
#
#     Label(root, text="Enter Client ID", font=("Nunito Sans", 10)).grid(row=0, column=0, columnspan=5, sticky='w')
#     client_id_entry = Entry(root)
#     client_id_entry.grid(row=1, column=0, columnspan=5, sticky='w')
#     Button(root, text="Submit", command=on_submit, font=("Nunito Sans", 10)).grid(row=2, column=0, columnspan=5, sticky='w')
#
#
#
#     root.title("Document Library Diff Generator")
#     root.iconbitmap("assets/icon.ico")
#     root.mainloop()

def collect_prod_data():
    prod_data_filenames = []
    prod_data_links = []
    prod_descriptions = []
    cert_data_filenames = []
    cert_data_links = []
    cert_descriptions = []
    driver = setups.driver_setup()
    setups.login_to_cozeva("1500")

    driver.find_element_by_xpath(locator.xpath_app_Tray_Link).click()
    driver.find_element_by_xpath(locator.xpath_app_Time_Capsule).click()
    driver.switch_to.window(driver.window_handles[1])
    setups.login_to_cozeva_cert("1500")

    def switch_to_prod():
        driver.switch_to.window(driver.window_handles[0])

    def switch_to_cert():
        driver.switch_to.window(driver.window_handles[1])

    for id in id_list:
        Client_name = db.fetchCustomerName(id)
        global wb
        wb.create_sheet(str(Client_name))
        ws = wb[str(Client_name)]

        switch_to_prod()
        setups.switch_customer_context(id)
        print("Client: " + str(id))
        driver.find_element_by_xpath(locator.xpath_resources_link).click()
        time.sleep(1)
        driver.find_element_by_id("help_menu_options").find_elements_by_tag_name("li")[1].click()
        sf.ajax_preloader_wait(driver)
        try:
            document_element = driver.find_element_by_xpath("//h5[contains(text(),'DOCUMENT LIBRARY')]//parent::div")
        except Exception as e:
            ws.append(["Document Library not present on PROD"])
            wb.save(locator.parent_dir + "\\Document_lib.xlsx")
            continue
        # document_child = driver.find_element_by_xpath("//h5[contains(text(),'DOCUMENT LIBRARY')]")
        # document_element = document_child.find_element_by_xpath("./..")
        linkrows_tables = document_element.find_elements_by_tag_name('tbody')
        for body in linkrows_tables:
            linkrows = body.find_elements_by_tag_name("tr")
            linkrows = linkrows[1:len(linkrows)]
            for tr_element in linkrows:
                try:
                    attrib_list = tr_element.find_elements_by_tag_name("td")[0].find_elements_by_tag_name("a")
                except IndexError as e:
                    continue
                for attrib in attrib_list:
                    prod_data_filenames.append(attrib.text)
                    prod_data_links.append(attrib.get_attribute("href"))
                try:
                    if "Description" != tr_element.find_elements_by_tag_name("td")[1].text:
                        prod_descriptions.append(tr_element.find_elements_by_tag_name("td")[1].text)
                except IndexError as e:
                    prod_descriptions.append("Description is blank")


        # print(prod_data_filenames)
        # print(prod_data_links)
        # print(prod_descriptions)
        # print(len(prod_data_filenames))
        # print(len(prod_data_links))
        # print(len(prod_descriptions))

        switch_to_cert()
        setups.switch_customer_context_cert(id)
        print("Client(cert): " + str(id))
        driver.find_element_by_xpath(locator.xpath_resources_link).click()
        time.sleep(1)
        driver.find_element_by_id("help_menu_options").find_elements_by_tag_name("li")[1].click()
        sf.ajax_preloader_wait(driver)
        document_element = driver.find_element_by_id("document_library")
        linkrows_tables = document_element.find_elements_by_tag_name('tbody')
        for body in linkrows_tables:
            linkrows = body.find_elements_by_tag_name("tr")

            linkrows = linkrows[1:len(linkrows)]
            print(len(linkrows))
            for tr_element in linkrows:
                try:
                    attrib_list = tr_element.find_elements_by_tag_name("td")[0].find_elements_by_tag_name("a")
                except IndexError as e:
                    continue
                for attrib in attrib_list:
                    cert_data_filenames.append(attrib.text.strip())
                    cert_data_links.append(attrib.get_attribute("href").strip())
                try:
                    if "Description" != tr_element.find_elements_by_tag_name("td")[1].text:
                        cert_descriptions.append(tr_element.find_elements_by_tag_name("td")[1].text.strip())
                except IndexError as e:
                    prod_descriptions.append("Description is blank")


        # print(cert_data_filenames)
        # print(cert_data_links)
        # print(cert_descriptions)
        # print(len(cert_data_filenames))
        # print(len(cert_data_links))
        # print(len(cert_descriptions))
        test_case_id = 1
        # if len(prod_data_filenames) == len(cert_data_filenames) and len(prod_data_links) == len(cert_data_links) and len(prod_data_links) == len(cert_data_filenames):
        #     ws.append(["ID", "Link Name in PROD", "Link Name in CERT", "Link in PROD", "Link is CERT", "Match Status"])
        #     for prod_filename, cert_filename, prod_link, cert_link in zip(prod_data_filenames, cert_data_filenames, prod_data_links, cert_data_links):
        #         if prod_filename == cert_filename and prod_link == cert_link:
        #             ws.append([test_case_id, prod_filename, cert_filename, prod_link, cert_link, "Yes"])
        #         else:
        #             ws.append([test_case_id, prod_filename, cert_filename, prod_link, cert_link, "No"])
        #         test_case_id+=1
        #         wb.save(locator.parent_dir + "\\Document_lib.xlsx")
        #     if len(prod_descriptions) == len(cert_descriptions):
        #         ws.append(["ID", "Description in PROD", "Description in CERT", "Match Status"])
        #         for prod_des, cert_des in zip(prod_descriptions, cert_descriptions):
        #             if prod_des == cert_des:
        #                 ws.append([test_case_id, prod_des, cert_des, "Yes"])
        #             else:
        #                 ws.append([test_case_id, prod_des, cert_des, "No"])
        #             test_case_id+=1
        #             wb.save(locator.parent_dir + "\\Document_lib.xlsx")
        #     else:
        #         ws.append(["ID", "Description in PROD", "Description in CERT", "Match Status",
        #                    "---------------------------------------"])
        #         for prod_des in prod_descriptions:
        #             if prod_des in cert_descriptions:
        #                 cert_des = prod_des
        #                 ws.append([test_case_id, prod_des, cert_des, "Yes"])
        #             elif prod_des not in cert_descriptions:
        #                 ws.append([test_case_id, prod_des, "Not present in Cert", "No"])
        #             wb.save(locator.parent_dir + "\\Document_lib.xlsx")
        #             test_case_id += 1
        if True:
            ws.append(["ID", "Link Name in PROD", "Link Name in CERT", "Match?"])
            for prod_filename in prod_data_filenames:
                if prod_filename in cert_data_filenames:
                    cert_filename = cert_data_filenames[cert_data_filenames.index(prod_filename)]
                    ws.append([test_case_id, prod_filename, cert_filename, "Yes"])
                elif prod_filename not in cert_data_filenames:
                    ws.append([test_case_id, prod_filename, "Not present in Cert", "No"])
                wb.save(locator.parent_dir + "\\Document_lib.xlsx")
                test_case_id+=1

            ws.append(["ID", "Link in PROD", "Link in CERT", "Match?", "---------------------------------------"])
            for prod_link in prod_data_links:
                if prod_link in cert_data_links:
                    cert_link = cert_data_links[cert_data_links.index(prod_link)]
                    ws.append([test_case_id, prod_link, cert_link, "Yes"])
                elif prod_link not in cert_data_links:
                    ws.append([test_case_id, prod_link, "Not present in Cert", "No"])
                wb.save(locator.parent_dir + "\\Document_lib.xlsx")
                test_case_id += 1

            ws.append(["ID", "Description in PROD", "Description in CERT", "Match?", "---------------------------------------"])
            for prod_des in prod_descriptions:
                if prod_des in cert_descriptions:
                    cert_des = cert_descriptions[cert_descriptions.index(prod_des)]
                    ws.append([test_case_id, prod_des, cert_des, "Yes"])
                elif prod_des not in cert_descriptions:
                    ws.append([test_case_id, prod_des, "Not present in Cert", "No"])
                wb.save(locator.parent_dir + "\\Document_lib.xlsx")
                test_case_id += 1


        rows = ws.max_row
        cols = ws.max_column
        for i in range(2, rows + 1):
            for j in range(3, cols + 1):
                if str(ws.cell(i, j).value) == 'Yes':
                    ws.cell(i, j).fill = PatternFill('solid', fgColor='0FC404')
                elif str(ws.cell(i, j).value) == 'No':
                    ws.cell(i, j).fill = PatternFill('solid', fgColor='FC0E03')
        wb.save(locator.parent_dir + "\\Document_lib.xlsx")

        prod_data_filenames.clear()
        prod_data_links.clear()
        prod_descriptions.clear()
        cert_data_filenames.clear()
        cert_data_links.clear()
        cert_descriptions.clear()
        #break


wb = Workbook()
wb.save(locator.parent_dir + "\\Document_lib.xlsx")
collect_prod_data()
