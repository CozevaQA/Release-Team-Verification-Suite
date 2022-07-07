from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter

path = "D:\\Writtwik\\Python_excel"

wb = load_workbook(path+"\\Precreated.xlsx")

ws = wb.active

print(wb.sheetnames)
print(ws)
ws = wb['Sheet2']
print(ws)

wb.create_sheet('Sheet_test')

print(wb.sheetnames)

wb1 = Workbook()
ws1 = wb1.active
ws1.title = "data"
ws1['A1'] = 'huh'

ws1.append(['hello', 'World'])
ws1.append(['hello', 'World'])
ws1.append(['hello', 'World'])
ws1.append(['hello', 'World'])
ws1.append(['hello', 'World'])
ws1.append(['hello', 'World'])
ws1.append(['hello', 'World'])

ws = wb['Grades']
print(ws)

for row in range(1,11):
    for col in range(1,5):
        char = get_column_letter(col)
        ws[char+str(row)] = char+str(row)




wb.save(path+"\\Precreated.xlsx")
wb1.save(path+"\\Brand_new.xlsx")