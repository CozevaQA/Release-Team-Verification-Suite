import pytest
import allure
import os
import random
from statistics import mean
import sys
import math
import time
import traceback
import csv
from os import listdir
from os.path import isfile, join
import csv
import time
from datetime import date
from datetime import datetime
import base64
from tkinter import messagebox, ttk
import guiwindow
import multiprocessing


import pandas as pd
from PIL import Image as img
from PIL import ImageTk
import PIL
from pytest_assume.plugin import assume
from selenium import *
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.alert import Alert
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import *
from selenium.webdriver.support.wait import WebDriverWait
from openpyxl import Workbook
from openpyxl.styles import PatternFill, NamedStyle, Border
from urllib.parse import urlparse, parse_qs
from openpyxl.styles import *
import support_functions as sf
import setups
import variablestorage as locator
import ExcelProcessor as db
from tkinter import *

pracname = "NORTH COUNTY HEALTH PROJECT INC"
provname = "PALOMINO, MARY"
patname = "3Q-SS-GWN"



def run_import_charts_validation_for_client(argument,queue):
    driver = setups.driver_setup()
    setups.login_to_cozeva(str(argument))
    sf.ajax_preloader_wait(driver)
    print("In registries")

    #once in registries, navigate to import charts
    if not sf.check_sidebar_open_status(driver):
        driver.find_element_by_xpath(locator.xpath_side_nav_SlideOut).click()

    driver.find_element(By.XPATH, "//img[@alt='Imported Charts icon']").click()
    sf.ajax_preloader_wait(driver)
    # landing, wait for datatable info
    WebDriverWait(driver, 100).until(EC.presence_of_element_located((By.CLASS_NAME, "dataTables_info")))

    #Begin the loop for this client. using the argument being passed, filter the entries on the input csv to those only, and add them to a list. The format of the csv is like so
    '''
    MarketName	customer_id	CozevaID	CCDID
Monarch	2000	1012BAJ	33225155
Monarch	2000	1012BAJ	36061973
Monarch	2000	1012BAJ	36061974
Monarch	2000	1012BAJ	36061975
Monarch	2000	1012BAJ	36061977
Monarch	2000	1012BAJ	36061978
Monarch	2000	1012BAJ	36061979
Monarch	2000	1012BAJ	37871848

    '''
    input_file_location = os.getcwd() + "\\assets\\CCDs.csv"
    input_df = pd.read_csv(input_file_location)
    filtered_df = input_df[input_df['customer_id'] == argument]
    #now create a list of dictionaries from the filtered dataframe with {'CozevaID': 'CCDID'}
    ccd_list = []
    for index, row in filtered_df.iterrows():
        ccd_entry = {row['CozevaID']:row['CCDID']}
        ccd_list.append(ccd_entry)

    filter_icon_xpath = "//a[contains(@class, 'datatable_filter_dropdown')]"
    ccd_search_xpath = "//input[@title='Search by CCD ID']"
    cz_id_search_xpath = "//input[@title='Search By Cozeva ID']"
    apply_xpath = "//a[@data-once='datatable_filter_apply']"
    report_list = []
    ws = []
    for ccd_pair in ccd_list:
        driver.find_element(By.XPATH, filter_icon_xpath).click()
        time.sleep(0.3)
        czid, ccdid = "", ""
        for czidloop, ccdidloop in ccd_pair.items():
            czid = czidloop
            ccdid = ccdidloop

        ccd_field = driver.find_element(By.XPATH, ccd_search_xpath)
        ccd_field.click()
        ccd_field.clear()
        ccd_field.send_keys(ccdid)

        cz_field = driver.find_element(By.XPATH, cz_id_search_xpath)
        cz_field.click()
        cz_field.clear()
        cz_field.send_keys(czid)

        apply_button = driver.find_element(By.XPATH, apply_xpath)
        driver.execute_script("arguments[0].scrollIntoView();", apply_button)
        time.sleep(0.3)
        apply_button.click()

        time.sleep(0.2)
        sf.ajax_preloader_wait(driver)

        WebDriverWait(driver, 100).until(EC.presence_of_element_located((By.CLASS_NAME, "dataTables_info")))
        empty_datatable_xpath = "//td[@class='dataTables_empty']"

        if len(driver.find_elements(By.XPATH, empty_datatable_xpath)) > 0:
            ws.append([db.fetchCustomerName(str(argument)),czid, ccdid, "Passed", driver.current_url])
        else:
            ws.append([db.fetchCustomerName(str(argument)),czid, ccdid, "Failed", driver.current_url])


        if len(driver.find_elements(By.XPATH, empty_datatable_xpath)) > 0:
            appendable_text = f"{db.fetchCustomerName(str(argument))} | {czid} | {ccdid} | Passed | {driver.current_url}"
        else:
            appendable_text = f"{db.fetchCustomerName(str(argument))} | {czid} | {ccdid} | Failed | {driver.current_url}"

        print(appendable_text)
        report_list.append(appendable_text)

    queue.put((argument, ws))
    #print(f"Completed client {argument}, quitting driver")
    driver.quit()
    return



if __name__ == '__main__':
    multiprocessing.freeze_support()
    processes = []
    queue = multiprocessing.Queue()
    report_workbook = Workbook()
    # client_list = ["APN CT", "AppleCare", "OCUT", "OPTUM", "Optum Care Arizona", "Optum Care Colorado", "Optum Care Network - Idaho", "Optum Care Network New Mexico", "Optum Care Network of Indiana", "Optum Care Network of Ohio", "Optum Care Network of Oregon", "Optum Care New York", "Optum Kansas City", "OptumCare - Nevada"]
    # client_list = [2600, 2200, 4100, 3300, 5100, 7400, 1550, 6600, 5200, 4800 ,5300, 5000, 1650, 4500]
    #had to exclude - Newyork(5000) due to loading times. Had to exclude ohio(4800).
    client_list = [6600, 5200, 5300, 1650, 4500]
    for client in client_list:
        # specific_worksheet = report_workbook.create_sheet(db.fetchCustomerName(str(client)))
        process = multiprocessing.Process(target=run_import_charts_validation_for_client, args=(client,queue,))
        process.start()
        processes.append(process)
        time.sleep(2)


    for process in processes:
        #print("Joining process.....")
        process.join()
        #print("Process joined.")

    while not queue.empty():
        print("Getting data from queue")
        client_id, worksheet_data = queue.get()
        specific_worksheet = report_workbook.create_sheet(db.fetchCustomerName(str(client_id)))
        specific_worksheet.append(["CozevaID", "CCDID", "Status", "URL"])
        print("Creating worksheet for " + db.fetchCustomerName(str(client_id)))
        for row in worksheet_data:
            specific_worksheet.append(row)

    report_workbook.save(locator.parent_dir + "\\CCDREPort.xlsx")



