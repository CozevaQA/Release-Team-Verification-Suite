import base64
import configparser
import datetime
import os
import time
import traceback

from openpyxl.styles import Font, PatternFill
from openpyxl.utils import quote_sheetname
from selenium.webdriver.common.by import By
from selenium.webdriver.support.wait import WebDriverWait
from tkinter import *
import setups
import logging
import ExcelProcessor as db
import context_functions as cf
import support_functions as sf
import setups as st
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
import variablestorage as locator
from openpyxl import Workbook, load_workbook
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import NoSuchElementException, ElementNotInteractableException, \
    ElementClickInterceptedException, TimeoutException
config = configparser.RawConfigParser()
config.read("locator-config.properties")
Health_plans = ["3000", "3600", "3700"]
AGCL_link = "https://www.cozeva.com/registries/users-bridge?session=YXBwX2lkPXJlZ2lzdHJpZXMmY3VzdElkPTMwMDAmb3JnSWQ9MzAwMCZ2Z3BJZD0zMDAwJnZwSWQ9MzAwMA&table_id=attempted_gap_closure_list&dt_filter=eyJhdHRlbXB0ZWRfZ2FwX2Nsb3N1cmVfbGlzdCI6IkptTmhiR04xYkdGMGFXOXVYMlJoZEdVOU1qQXlNeTB4TWkwek1RPT0ifQ=="

def build_agcl_urls():
    url_list = []
    for cusID in Health_plans:
        try:
            sm_customer_id = (str(cusID)).strip()
            #print(sm_customer_id)
            session_var = f'app_id=registries&custId={sm_customer_id}&orgId={sm_customer_id}&vgpId={sm_customer_id}&vpId={sm_customer_id}'
            filter_value_23 = '&calculation_date=2023-12-31'
            filter_value_22 = '&calculation_date=2022-12-31'
            encoded_session_var = base64.b64encode(session_var.encode('utf-8'))
            encoded_filter_value_23 = base64.b64encode(filter_value_23.encode('utf-8'))
            encoded_filter_value_22 = base64.b64encode(filter_value_22.encode('utf-8'))
            filter_var_23 = "{\"attempted_gap_closure_list\":\"" + encoded_filter_value_23.decode('utf-8') + "\"}"
            filter_var_22 = "{\"attempted_gap_closure_list\":\"" + encoded_filter_value_22.decode('utf-8') + "\"}"
            encoded_filter_var_23 = base64.b64encode(filter_var_23.encode('utf-8'))
            encoded_filter_var_22 = base64.b64encode(filter_var_22.encode('utf-8'))

            url_list.append("https://www.cozeva.com/registries/users-bridge?session=" + encoded_session_var.decode('utf-8') + "&table_id=attempted_gap_closure_list&dt_filter=" + encoded_filter_var_23.decode('utf-8'))
            #url_list.append("https://www.cozeva.com/registries/users-bridge?session=" + encoded_session_var.decode('utf-8') + "&table_id=attempted_gap_closure_list&dt_filter=" + encoded_filter_var_22.decode('utf-8'))


        except Exception as e:
            print(e)
            raise

    return url_list

def collect_agcl_data():
    agcl_urls = build_agcl_urls()
    for url in agcl_urls:
        print(url)




# Start of Code ----------------------------------------------------------------------------
ENV = 'CERT'
report_folder = os.path.join(locator.parent_dir,"Bridge_weekly_reports")
isdir = os.path.isdir(report_folder)
if not isdir:
    os.mkdir(report_folder)

AGCL_workbook_path = os.path.join(report_folder, "AGCL_weekly.xlsx")
GC_workbook_path = os.path.join(report_folder, "GC_weekly.xlsx")
OM_workbook_path = os.path.join(report_folder, "OM_weekly.xlsx")

if os.path.exists(AGCL_workbook_path):
    AGCL_workbook = load_workbook(os.path.join(report_folder, AGCL_workbook_path))
else:
    AGCL_workbook = Workbook()
    AGCL_workbook.save(AGCL_workbook_path)
    active_sheet = AGCL_workbook.active
    active_sheet.title = "Summary"
    for cust in Health_plans:
        AGCL_workbook.create_sheet(db.fetchCustomerName(str(cust)))

    summary_sheet = AGCL_workbook["Summary"]
    summary_sheet.append(["Customer", "Last week", "", "", "This week", "", "", "Difference", "", ""])
    summary_sheet.append(
        ["", "Pending", "Closed", "Not-Closed", "Pending", "Closed", "Not-Closed","Pending", "Closed", "Not-Closed"])

    row = 3

    for sheet in AGCL_workbook.sheetnames:
        if sheet != "Summary":  # Skip the Summary sheet itself
            # Create a hyperlink to the first cell of each sheet
            cell = summary_sheet.cell(row, 1)  # Column 1, next available row
            cell.value = sheet  # Text to display
            # Correct hyperlink format for internal links
            #cell.hyperlink = f"#'{sheet}'!A1"
            row += 1  # Move to the next row for the next hyperlink

    AGCL_workbook.save(AGCL_workbook_path)

if os.path.exists(GC_workbook_path):
    GC_workbook = load_workbook(os.path.join(report_folder, GC_workbook_path))
else:
    GC_workbook = Workbook()
    active_sheet = GC_workbook.active
    active_sheet.title = "Summary"

    # Create sheets for each customer
    for cust in db.getCustomerList()[1:-1]:
        GC_workbook.create_sheet(str(cust))

# Add hyperlinks to the Summary sheet
    summary_sheet = GC_workbook["Summary"]
    summary_sheet.append(["Customer", "Last week", "", "This week", "", "Difference", ""])
    summary_sheet.append(["", "Your gaps others can close", "Others gap you can close", "Your gaps others can close", "Others gap you can close", "Your gaps others can close", "Others gap you can close"])

    row = 3  # Starting row for the hyperlinks in the Summary sheet

    for sheet in GC_workbook.sheetnames:
        if sheet != "Summary":  # Skip the Summary sheet itself
            # Create a hyperlink to the first cell of each sheet
            cell = summary_sheet.cell(row, 1)  # Column 1, next available row
            cell.value = sheet  # Text to display
            # Correct hyperlink format for internal links
            #cell.hyperlink = f"#'{sheet}'!A1"
            row += 1  # Move to the next row for the next hyperlink

# Save the workbook
    GC_workbook.save(GC_workbook_path)


if os.path.exists(OM_workbook_path):
    OM_workbook = load_workbook(os.path.join(report_folder, OM_workbook_path))
else:
    OM_workbook = Workbook()
    OM_workbook.save(OM_workbook_path)
    active_sheet = OM_workbook.active
    active_sheet.title = "Summary"
    for cust in db.getCustomerList()[1:-1]:
        OM_workbook.create_sheet(str(cust))

    summary_sheet = OM_workbook["Summary"]
    summary_sheet.append(["Customer", "Last week", "", "This week", "", "Difference", ""])
    summary_sheet.append(["", "Overlapping Members", "Tenancies", "Overlapping Members", "Tenancies", "Overlapping Members", "Tenancies"])

    row = 3

    for sheet in OM_workbook.sheetnames:
        if sheet != "Summary":  # Skip the Summary sheet itself
            # Create a hyperlink to the first cell of each sheet
            cell = summary_sheet.cell(row, 1)  # Column 1, next available row
            cell.value = sheet  # Text to display
            # Correct hyperlink format for internal links
            # cell.hyperlink = f"#'{sheet}'!A1"
            row += 1  # Move to the next row for the next hyperlink

    OM_workbook.save(OM_workbook_path)

#now, add data to the workbooks based on collected data
# AGCL_workbook
# GC_workbook
# OM_workbook

'''
    AGCL_workbook replace data last week this week (More or less the same for GC and OM)
    Data to move
    From E3,E4,E5 to B3,B4,B5
    From F3,F4,F5 to C3,C4,C5
    From G3,G4,G5 to D3,D4,D5
'''

cells_to_move_AGCL = {
    'E': 'B',
    'F': 'C',
    'G': 'D'
}

Bridge_data_workbook_collected = load_workbook('assets/bridge_data.xlsx')
#print(Bridge_data_workbook)

#1. Update AGCL book a. Clear this/next week stuff b. Load values into summary and individual

AGCL_data_worksheet = Bridge_data_workbook_collected['Attempted Gap Closure']
AGCL_data_list = []
for row in AGCL_data_worksheet.iter_rows(values_only=True):
    AGCL_data_list.append(list(row))
AGCL_data_list = AGCL_data_list[1:]
print(AGCL_data_list)

AGCL_workbook_Summary_sheet = AGCL_workbook['Summary']

# Iterate over the specified rows
for row in range(3, 6):  # Rows 3, 4, 5
    for source_col, target_col in cells_to_move_AGCL.items():
        source_cell = f'{source_col}{row}'
        target_cell = f'{target_col}{row}'

        # Move data if not None
        if AGCL_workbook_Summary_sheet[source_cell].value is not None:
            AGCL_workbook_Summary_sheet[target_cell].value = AGCL_workbook_Summary_sheet[source_cell].value
            AGCL_workbook_Summary_sheet[source_cell].value = None

#for AGCL_data in AGCL_data_list:




AGCL_workbook.save(AGCL_workbook_path)



























