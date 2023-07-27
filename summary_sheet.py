from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter


def summarize_report(sum_workbook, folder_path):
    '''
    The flow of this summary processor will be as follows

    We pass the path of the report folder, this probram will load the report.xlsx into a workbook object.

    We will create a worksheet object for the first sheet and rename it to Test summary. We will then look through every other sheet.

    On each sheet, we will look for failed cases. If a row has a failed case, we will append the whole row to the summary sheet object. We will also add the sheet name in bold in every loop.
    '''

    ws_main = sum_workbook['Sheet']
    ws_main.title = 'Summary'
    workbook_name = '\\Report.xlsx'
    sum_workbook.save(folder_path + workbook_name)
    ws_main.append(['Validation Area'])
    sheet_names = sum_workbook.sheetnames


    for worksheet in sheet_names:
        if worksheet == 'Summary':
            continue
        ws_sheets = sum_workbook[worksheet]

        currentSheet = []
        row_counter = 1
        control = str(ws_sheets['A1'].value)
        while control != 'None':
            currentRow = ws_sheets[row_counter]
            temp = []
            for currentCell in currentRow:
                temp.append(str(currentCell.value))
            # temp.append(str(current_row[0].value))
            # temp.append(str(current_row[1].value))
            # temp.append(str(current_row[2].value))
            # temp.append(str(current_row[3].value))
            # temp.append(str(current_row[4].value))
            # temp.append(str(current_row[5].value))
            # temp.append(str(current_row[6].value))
            print(temp)
            currentSheet.append(temp)
            temp = []
            control = str(ws_sheets['A' + str(row_counter + 1)].value)
            row_counter += 1


        print(currentSheet)

        for rows in currentSheet:
            if 'Failed' in rows:
                ws_main.append([worksheet])
                workbook_name = '\\Report_Failed_cases.xlsx'
                break

        for rows in currentSheet:
            if 'Failed' in rows:
                ws_main.append([""]+rows[1:])


        sum_workbook.save(folder_path + "\\Report.xlsx")
    rows = ws_main.max_row
    cols = ws_main.max_column
    for i in range(2, rows + 1):
        for j in range(3, cols + 1):
            if ws_main.cell(i, j).value == 'Passed':
                ws_main.cell(i, j).fill = PatternFill('solid', fgColor='0FC404')
            elif ws_main.cell(i, j).value == 'Failed':
                ws_main.cell(i, j).fill = PatternFill('solid', fgColor='FC0E03')
            elif ws_main.cell(i, j).value == 'Showing 0 to 0':
                ws_main.cell(i, j).fill = PatternFill('solid', fgColor='FCC0BB')


#test Code
#wb = load_workbook("C:\\VerificationReports\\2023-05-03"+"\\Report.xlsx")

#summarize_report(wb, "C:\\VerificationReports\\2023-05-03")






