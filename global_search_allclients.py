# This Script will loop through selected clients and perform global search for each by fetching data from support levels.
import os
import random
from statistics import mean
import sys
import math
import time
import traceback
import csv
from os import listdir
from os.path import isfile, join
import csv
import time
import re
from datetime import date
from datetime import datetime
import base64
from tkinter import messagebox, ttk

import pandas as pd
from PIL import Image as img
from PIL import ImageTk
import PIL
from pytest_assume.plugin import assume
from selenium import *
from selenium import webdriver
from sigfig import round

from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.alert import Alert
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import *
from selenium.webdriver.support.wait import WebDriverWait
from openpyxl import Workbook
from openpyxl.styles import PatternFill, NamedStyle, Border
from urllib.parse import urlparse, parse_qs
from openpyxl.styles import *
import support_functions as sf
import setups
import variablestorage as locator
import ExcelProcessor as db
from tkinter import *


ENV = "PROD"
run_config = []
client_list = ['4050', '4350', '2400', '1750', '2750', '1500', '3700', '3750', '2600', '2200', '3250', '4550', '6800', '1600', '6700', '7100', '4700', '7500', '1450', '2050', '5520', '3100', '1700', '2900', '4300', '3000', '3050', '4150', '3500', '200', '4250', '2350', '4750', '3600', '2250', '2650', '4600', '150', '5700', '2700', '1950', '2000', '2100', '5600', '3850', '4450', '1100', '3300', '5100', '7400', '2800', '1550', '4100', '6600', '5200', '4800', '5300', '4900', '2950', '5000', '1650', '4500', '3950', '3550', '3650', '2450', '3450', '3350', '3400', '1200', '3900', '2550', '3800', '1300', '1850', '6000', '5400', '3150', '1000', '1900', '7000', '4000', '5510', '7350', '4400', '5900', '7200']
#client_list = []
global_test_case_id = 1

report_folder = os.path.join(locator.parent_dir,"Global Search Report")
isdir = os.path.isdir(report_folder)
if not isdir:
    os.mkdir(report_folder)

wb = Workbook()
workbook_title = "Global_Search_Report_"+sf.date_time()+"_"+ENV+".xlsx"
wb.save(report_folder+ "\\"+workbook_title)
def fetch_client_list():
    root = Tk()
    root.title("Multi-client Cozeva Support global Search")
    root.iconbitmap("assets/icon.ico")
    customer_list = db.getCustomerList()
    # print(customer_list)
    Checkbox_variables = []
    select_all_variable = IntVar()
    prod_env_variable = IntVar()
    cert_env_variable = IntVar()
    for i in range(0, len(customer_list)):
        Checkbox_variables.append(IntVar())
    print(len(Checkbox_variables))
    Checkbox_widgets = []

    def on_submitbutton():
        global client_list, ENV
        client_list = []
        for checkbox_index in range(0, len(customer_list)):
            if Checkbox_variables[checkbox_index].get() == 1:
                client_list.append(db.fetchCustomerID(customer_list[checkbox_index]))

        if prod_env_variable.get() == 1:
            ENV = 'PROD'

        if cert_env_variable.get() == 1:
            ENV = 'CERT'

        root.destroy()

        print(client_list)

    def on_select_all():
        if select_all_variable.get() == 1:
            for check_box in Checkbox_widgets[1:len(Checkbox_widgets)-1]:
                check_box.select()
            # for check_box_variable in Checkbox_variables:
            #     check_box_variable.set(1)
        elif select_all_variable.get() == 0:
            for check_box in Checkbox_widgets:
                check_box.deselect()
            # for check_box_variable in Checkbox_variables:
            #     check_box_variable.set(0)


    for i in range(0, len(customer_list)):
        Checkbox_widgets.append(Checkbutton(root, text=customer_list[i], variable=Checkbox_variables[i],
                                            font=("Nunito Sans", 10)))
    submit_button = Button(root, text="Submit", command=on_submitbutton, font=("Nunito Sans", 10))
    select_all_checkbox = Checkbutton(root, text="Select All", command=on_select_all, variable = select_all_variable, font=("Nunito Sans", 10))
    production_checkbox = Checkbutton(root, text="PROD", variable = prod_env_variable, font=("Nunito Sans", 10))
    cert_checkbox = Checkbutton(root, text="CERT", variable = cert_env_variable, font=("Nunito Sans", 10))

    # add all checkboxes to a grid
    # practice_sidemenu_checkbox.grid(row=3, column=0, columnspan=5, sticky="w")
    for i in range(1, len(Checkbox_widgets)):
        if i <= 20:
            Checkbox_widgets[i].grid(row=i, column=0, sticky="w")
        elif 20 < i <= 40:
            Checkbox_widgets[i].grid(row=i - 20, column=1, sticky="w")
        elif 40 < i <= 60:
            Checkbox_widgets[i].grid(row=i - 40, column=2, sticky="w")
        elif 60 < i <= 80:
            Checkbox_widgets[i].grid(row=i - 60, column=3, sticky="w")
        elif 80 < i <= 100:
            Checkbox_widgets[i].grid(row=i - 80, column=4, sticky="w")
    submit_button.grid(row=0, column=4, sticky="e")
    select_all_checkbox.grid(row=0, column=0, sticky="w")
    production_checkbox.grid(row=0, column=2, sticky='e')
    cert_checkbox.grid(row=0, column=3, sticky='w')

    root.mainloop()

fetch_client_list()
driver = setups.driver_setup()

if ENV == 'CERT':
    setups.login_to_cozeva_cert("1500")
elif ENV == 'STAGE':
    setups.login_to_cozeva_stage()
elif ENV == "PROD":
    setups.login_to_cozeva("1500")
else:
    print("ENV INVALID")
    exit(3)

wb.create_sheet("Global_search")
wsheet = wb["Global_search"]
wsheet.append(["Test Case ID", "Context Name", "Scenario", "Status","Time Taken", "Comments", "URL"])

wb.create_sheet("Data_Collection")
wsheet_other = wb["Data_Collection"]
wsheet_other.append(["Test Case ID", "Context Name", "Scenario", "Status","Time Taken", "Comments", "URL"])


def perform_global_search(functiondriver, ws, wsother, client_id):
    global global_test_case_id
    client_name = db.fetchCustomerName(client_id)
    test_case_id = global_test_case_id
    try:
        WebDriverWait(functiondriver, 30).until(
            EC.presence_of_element_located((By.ID, "registry_body")))
        selected_metric_name = 'Couldnt fetch Metric Name'
        context_name = functiondriver.find_element_by_xpath(locator.xpath_context_Name).text
        registry_url = functiondriver.current_url

        metrics = functiondriver.find_element_by_id("registry_body").find_elements(By.CSS_SELECTOR, "li.li-metric")

        percent = '0.00'
        num_den = "(0/0)"
        iter_count = 0
        while percent == '0.00' or percent == '0.00%' or num_den == "(0/0)":
            selectedMetric = metrics[sf.RandomNumberGenerator(len(metrics), 1)[0]]
            functiondriver.execute_script("arguments[0].scrollIntoView({block: 'center'});", selectedMetric)
            try:
                percent = selectedMetric.find_element_by_class_name('percent').text
            except NoSuchElementException as e:
                print("Skipping % for util measure")
                print(e)
                traceback.print_exc()
            num_den = selectedMetric.find_element_by_class_name('num-den').text
            iter_count += 1
            print(iter_count)
            selected_metric_name = selectedMetric.find_element_by_class_name('met-name').text
            print(selected_metric_name)
            if iter_count > 10:
                wsother.append([test_case_id, context_name, "Looking for a non 0/0, non util measure", "Failed",
                           "x", "All/most metrics are 0/0",
                           functiondriver.current_url])
                test_case_id += 1
                raise Exception("All/most metrics are 0/0")
        selected_metric_name = selectedMetric.find_element_by_class_name('met-name').text
        metric_href = selectedMetric.find_element(By.TAG_NAME, "a").get_attribute("href")
        try:
            selectedMetric.click()
        except ElementClickInterceptedException as e:
            print("Exception Occured")
            if ENV == "CERT":
                print(metric_href)
                driver.get(metric_href)
            elif ENV == "PROD":
                driver.get(metric_href)
        except Exception as e:
            traceback.print_exc()



        start_time = time.perf_counter()
        sf.ajax_preloader_wait(functiondriver)
        total_time = time.perf_counter() - start_time
        print("Loaded into Metric- " + selected_metric_name)
        current_context = functiondriver.find_element(By.XPATH, locator.xpath_context_Name).text
        wsother.append([test_case_id, current_context, "Loading into chosen Metric", "Passed",
                   round(total_time, sigfigs=3), selected_metric_name,
                   functiondriver.current_url])
        test_case_id += 1
        # check for number of tabs and thier names
        MSPL_link = functiondriver.current_url
        tabs = functiondriver.find_element(By.CLASS_NAME, "tabs").find_elements(By.CLASS_NAME, "tab")
        strings_for_global_search = []
        for index, tab in enumerate(tabs):
            tab_name = tabs[index].text

            if tab_name == "Practices":
                tabs[index].click()
                start_time = time.perf_counter()
                sf.ajax_preloader_wait(functiondriver)
                total_time = time.perf_counter() - start_time
                access_message = sf.URLAccessCheck(functiondriver.current_url, functiondriver)
                Practice_MSPL_url = functiondriver.current_url

                if not access_message:
                    WebDriverWait(functiondriver, 30).until(EC.presence_of_element_located((By.CLASS_NAME, "tabs")))
                    tabs = functiondriver.find_element(By.CLASS_NAME, "tabs").find_elements(By.CLASS_NAME, "tab")
                    current_context = functiondriver.find_element(By.XPATH, locator.xpath_context_Name).text
                    wsother.append([test_case_id, current_context, "Navigation to practice tab", "Passed",
                               round(total_time, sigfigs=3), "",
                               functiondriver.current_url])
                    test_case_id += 1
                    if len(driver.find_elements(By.CLASS_NAME, 'dataTables_empty')) != 0:
                        print("No Practices")
                        current_context = functiondriver.find_element(By.XPATH, locator.xpath_context_Name).text
                        wsother.append([test_case_id, current_context, "Looking for a non 0 count Practice", "Failed",
                                   'x', "No Practices",
                                   functiondriver.current_url])
                        test_case_id += 1

                    elif len(driver.find_elements(By.CLASS_NAME, 'dataTables_empty')) == 0:
                        entries = functiondriver.find_element(By.XPATH, locator.xpath_data_Table_Info).text
                        practice_table = driver.find_element(By.ID, "metric-support-prac-ls").find_element(By.TAG_NAME,
                                                                                                           "tbody").find_elements(By.TAG_NAME, "tr")
                        print(len(practice_table))
                        print(entries)
                        selected_practice_name = \
                        random.choice(practice_table).find_elements(By.TAG_NAME, "td")[3].find_elements(By.TAG_NAME,
                                                                                                        "a")[1]
                        current_context = functiondriver.find_element(By.XPATH, locator.xpath_context_Name).text
                        print(current_context)

                        global_search_practice_name = selected_practice_name.text
                        strings_for_global_search.append(global_search_practice_name)

            if tab_name == "Providers":
                tabs[index].click()
                start_time = time.perf_counter()
                sf.ajax_preloader_wait(functiondriver)
                total_time = time.perf_counter() - start_time
                access_message = sf.URLAccessCheck(functiondriver.current_url, functiondriver)
                Provider_MSPL_url = functiondriver.current_url

                if not access_message:
                    WebDriverWait(functiondriver, 30).until(EC.presence_of_element_located((By.CLASS_NAME, "tabs")))
                    tabs = functiondriver.find_element(By.CLASS_NAME, "tabs").find_elements(By.CLASS_NAME, "tab")
                    current_context = functiondriver.find_element(By.XPATH, locator.xpath_context_Name).text
                    wsother.append([test_case_id, current_context, "Navigation to provider tab", "Passed",
                               round(total_time, sigfigs=3), "",
                               functiondriver.current_url])
                    test_case_id += 1
                    if len(driver.find_elements(By.CLASS_NAME, 'dataTables_empty')) != 0:
                        print("No Providers")
                        current_context = functiondriver.find_element(By.XPATH, locator.xpath_context_Name).text
                        wsother.append([test_case_id, current_context, "Looking for a non 0 count Provider", "Failed",
                                   'x', "No Practices",
                                   functiondriver.current_url])
                        test_case_id += 1

                    elif len(driver.find_elements(By.CLASS_NAME, 'dataTables_empty')) == 0:
                        entries = functiondriver.find_element(By.XPATH, locator.xpath_data_Table_Info).text
                        provider_table = driver.find_element(By.ID, "metric-support-prov-ls").find_element(By.TAG_NAME,
                                                                                                           "tbody").find_elements(By.TAG_NAME, "tr")
                        print(len(provider_table))
                        print(entries)
                        selected_provider_name = \
                        random.choice(provider_table).find_elements(By.TAG_NAME, "td")[4].find_elements(By.TAG_NAME,
                                                                                                        "a")[1]
                        current_context = functiondriver.find_element(By.XPATH, locator.xpath_context_Name).text
                        print(current_context)

                        global_search_provider_name = selected_provider_name.text
                        strings_for_global_search.append(global_search_provider_name)



                    driver.get(MSPL_link)
                    sf.ajax_preloader_wait(functiondriver)
                    WebDriverWait(functiondriver, 30).until(EC.presence_of_element_located((By.CLASS_NAME, "tabs")))
                    tabs = functiondriver.find_element(By.CLASS_NAME, "tabs").find_elements(By.CLASS_NAME, "tab")

            if tab_name == "Patients":
                tabs[index].click()
                start_time = time.perf_counter()
                sf.ajax_preloader_wait(functiondriver)
                total_time = time.perf_counter() - start_time
                access_message = sf.URLAccessCheck(functiondriver.current_url, functiondriver)
                Patient_MSPL_url = functiondriver.current_url

                if not access_message:
                    WebDriverWait(functiondriver, 30).until(EC.presence_of_element_located((By.CLASS_NAME, "tabs")))
                    tabs = functiondriver.find_element(By.CLASS_NAME, "tabs").find_elements(By.CLASS_NAME, "tab")
                    current_context = functiondriver.find_element(By.XPATH, locator.xpath_context_Name).text
                    wsother.append([test_case_id, current_context, "Navigation to patient tab", "Passed",
                               round(total_time, sigfigs=3), "",
                               functiondriver.current_url])
                    test_case_id += 1
                    if len(driver.find_elements(By.CLASS_NAME, 'dataTables_empty')) != 0:
                        print("No Patients")
                        current_context = functiondriver.find_element(By.XPATH, locator.xpath_context_Name).text
                        wsother.append([test_case_id, current_context, "Looking for a Patient", "Failed",
                                   'x', "No Practices",
                                   functiondriver.current_url])
                        test_case_id += 1

                    elif len(driver.find_elements(By.CLASS_NAME, 'dataTables_empty')) == 0:
                        entries = functiondriver.find_element(By.XPATH, locator.xpath_data_Table_Info).text
                        patient_table = driver.find_element(By.ID, "metric-support-pat-ls").find_element(By.TAG_NAME,
                                                                                                           "tbody").find_elements(By.TAG_NAME, "tr")
                        print(len(patient_table))
                        print(entries)
                        patient_href = random.choice(patient_table).find_elements(By.TAG_NAME, "td")[2].find_elements(By.TAG_NAME,"a")[0].get_attribute("href")
                        parsed_url = urlparse(patient_href)
                        query_params = parse_qs(parsed_url.query)
                        cozeva_id_urlparsed = query_params.get("cozeva_id", [""])[0]
                        cozeva_id = re.search(r'/patient_detail/([^?]+)\?', patient_href)
                        current_context = functiondriver.find_element(By.XPATH, locator.xpath_context_Name).text
                        print(current_context)

                        global_search_patient_name = cozeva_id.group(1).replace("?", "")
                        strings_for_global_search.append(global_search_patient_name)
                        #strings_for_global_search.append(cozeva_id_urlparsed)

        driver.get(registry_url)
        WebDriverWait(driver, 30).until(
            EC.presence_of_element_located((By.XPATH, locator.xpath_filter_measure_list)))
        # time to do the global search
        print(strings_for_global_search)

        def performPracSearch(practice_string):
            try:
                WebDriverWait(driver, 30).until(
                    EC.presence_of_element_located((By.XPATH, locator.xpath_filter_measure_list)))
                window_switched = 0
                driver.find_element_by_id('globalsearch_input').send_keys(practice_string)
                start_time = time.perf_counter()
                WebDriverWait(driver, 45).until(
                    EC.presence_of_element_located((By.CSS_SELECTOR, '.dropdown-content.collection.with-header li')))
                if "No results found for" in driver.find_element(By.CSS_SELECTOR, '.dropdown-content.collection.with-header li').text:
                    raise Exception("No results")
                time_taken = round(time.perf_counter() - start_time)
                driver.find_element_by_id('globalsearch_input').send_keys(Keys.RETURN)
                sf.ajax_preloader_wait(driver)
                time_taken = round(time.perf_counter() - start_time)
                WebDriverWait(driver, 30).until(
                    EC.presence_of_element_located((By.ID, 'search_practices_link')))
                # driver.find_element_by_id('globalsearch_input').send_keys(Keys.RETURN)
                driver.find_element_by_id('search_practices_link').click()
                WebDriverWait(driver, 30).until(
                    EC.presence_of_element_located((By.ID, 'search_practices')))
                driver.find_element_by_id('search_practices').find_elements_by_tag_name('a')[0].click()
                driver.switch_to.window(driver.window_handles[1])
                window_switched = 1
                sf.ajax_preloader_wait(driver)
                WebDriverWait(driver, 30).until(
                    EC.presence_of_element_located((By.XPATH, locator.xpath_filter_measure_list)))
                if len(driver.find_elements_by_xpath(locator.xpath_filter_measure_list)) != 0:
                    ws.append([test_case_id, client_name+' : Practice', 'Context set to: ' + practice_string, 'Passed', time_taken])
                    driver.close()
                    driver.switch_to.window(driver.window_handles[0])
                    driver.get(registry_url)
                else:
                    ws.append([test_case_id, client_name+' : Practice', 'Context set to: ' + practice_string, 'Failed', time_taken,
                               driver.current_url])
                    driver.close()
                    driver.switch_to.window(driver.window_handles[0])
                    driver.get(registry_url)

            except Exception as e:
                print(e)
                traceback.print_exc()
                if window_switched == 1:
                    ws.append([test_case_id, client_name+' : Practice', 'Context set to: ' + practice_string, 'Failed', '',
                               'Unable to click on practice name from global search', driver.current_url])
                    driver.close()
                    driver.switch_to.window(driver.window_handles[0])
                    driver.get(registry_url)
                elif window_switched == 0:
                    ws.append([test_case_id, client_name+' : Practice', 'Context set to: ' + practice_string, 'Failed', '',
                               'Unable to global search', driver.current_url])
                    driver.get(registry_url)

        def performProvSearch(provider_string):
            try:
                window_switched = 0
                driver.find_element_by_id('globalsearch_input').send_keys(provider_string)
                WebDriverWait(driver, 45).until(
                    EC.presence_of_element_located((By.CSS_SELECTOR, '.dropdown-content.collection.with-header li')))
                if "No results found for" in driver.find_element(By.CSS_SELECTOR,
                                                                 '.dropdown-content.collection.with-header li').text:
                    raise Exception("No results")
                driver.find_element_by_id('globalsearch_input').send_keys(Keys.RETURN)
                start_time = time.perf_counter()
                sf.ajax_preloader_wait(driver)
                time_taken = round(time.perf_counter() - start_time)
                # driver.find_element_by_id('globalsearch_input').send_keys(Keys.RETURN)
                driver.find_element_by_id('search_providers_link').click()
                driver.find_element_by_id('search_providers').find_elements_by_tag_name('a')[0].click()
                driver.switch_to.window(driver.window_handles[1])
                window_switched = 1
                sf.ajax_preloader_wait(driver)
                WebDriverWait(driver, 30).until(
                    EC.presence_of_element_located((By.XPATH, locator.xpath_filter_measure_list)))
                if len(driver.find_elements_by_xpath(locator.xpath_filter_measure_list)) != 0:
                    ws.append([test_case_id, client_name+' : Provider', 'Context set to: ' + provider_string, 'Passed', time_taken])
                    driver.close()
                    driver.switch_to.window(driver.window_handles[0])
                    driver.get(registry_url)
                else:
                    ws.append([test_case_id, client_name+' : Provider', 'Context set to: ' + provider_string, 'Failed', time_taken,
                               driver.current_url])
                    driver.close()
                    driver.switch_to.window(driver.window_handles[0])
                    driver.get(registry_url)

            except Exception as e:
                print(e)
                traceback.print_exc()
                if window_switched == 1:
                    ws.append([test_case_id, client_name+' : Provider', 'Context set to: ' + provider_string, 'Failed', '',
                               'Unable to click on practice name from global search', driver.current_url])
                    driver.close()
                    driver.switch_to.window(driver.window_handles[0])
                    driver.get(registry_url)
                elif window_switched == 0:
                    ws.append([test_case_id, client_name+' : Provider', 'Context set to: ' + provider_string, 'Failed', '',
                               'Unable to global search', driver.current_url])
                    driver.get(registry_url)

        def performPatSearch(patient_string):
            try:
                window_switched = 0
                driver.find_element_by_id('globalsearch_input').send_keys(patient_string)
                WebDriverWait(driver, 45).until(
                    EC.presence_of_element_located((By.CSS_SELECTOR, '.dropdown-content.collection.with-header li')))
                if "No results found for" in driver.find_element(By.CSS_SELECTOR,
                                                                 '.dropdown-content.collection.with-header li').text:
                    raise Exception("No results")
                driver.find_element_by_id('globalsearch_input').send_keys(Keys.RETURN)
                start_time = time.perf_counter()
                sf.ajax_preloader_wait(driver)
                time_taken = round(time.perf_counter() - start_time)
                driver.find_element_by_id('search_patients_link').click()

                driver.find_element_by_id('search_patients').find_elements_by_tag_name('li')[
                    1].find_element_by_css_selector("a[href*='patient_detail']").click()
                driver.switch_to.window(driver.window_handles[1])
                window_switched = 1
                sf.ajax_preloader_wait(driver)
                WebDriverWait(driver, 30).until(
                    EC.presence_of_element_located((By.XPATH, locator.xpath_patient_Header_Dropdown_Arrow)))
                if len(driver.find_elements_by_xpath(locator.xpath_patient_Header_Dropdown_Arrow)) != 0:
                    ws.append([test_case_id, client_name+' : Patient', 'Context set to: ' + patient_string, 'Passed', time_taken])
                    driver.close()
                    driver.switch_to.window(driver.window_handles[0])
                    driver.get(registry_url)
                else:
                    ws.append([test_case_id, client_name+' : Patient', 'Context set to: ' + patient_string, 'Failed', time_taken,
                               driver.current_url])
                    driver.close()
                    driver.switch_to.window(driver.window_handles[0])
                    driver.get(registry_url)

            except Exception as e:
                print(e)
                traceback.print_exc()
                if window_switched == 1:
                    ws.append([test_case_id, client_name+' : Patient', 'Context set to: ' + patient_string, 'Failed', '',
                               'Unable to click on practice name from global search', driver.current_url])
                    driver.close()
                    driver.switch_to.window(driver.window_handles[0])
                    driver.get(registry_url)
                elif window_switched == 0:
                    ws.append([test_case_id, client_name+' : Patient', 'Context set to: ' + patient_string, 'Failed', '',
                               'Unable to global search', driver.current_url])
                    driver.get(registry_url)

        performPracSearch(strings_for_global_search[0])
        performProvSearch(strings_for_global_search[1])
        performPatSearch(strings_for_global_search[2])



    except Exception as e:
        traceback.print_exc()
    global_test_case_id = test_case_id


for client in client_list:
    if ENV == 'PROD':
        setups.switch_customer_context(client)
    elif ENV == 'CERT':
        setups.switch_customer_context_cert(client)

    perform_global_search(driver, wsheet, wsheet_other, client)
    rows = wsheet.max_row
    cols = wsheet.max_column
    for i in range(2, rows + 1):
        for j in range(3, cols + 1):
            if wsheet.cell(i, j).value == 'Passed':
                wsheet.cell(i, j).fill = PatternFill('solid', fgColor='0FC404')
            elif wsheet.cell(i, j).value == 'Failed':
                wsheet.cell(i, j).fill = PatternFill('solid', fgColor='FC0E03')
            elif wsheet.cell(i, j).value == 'Showing 0 to 0':
                wsheet.cell(i, j).fill = PatternFill('solid', fgColor='FCC0BB')

    rows = wsheet_other.max_row
    cols = wsheet_other.max_column
    for i in range(2, rows + 1):
        for j in range(3, cols + 1):
            if wsheet_other.cell(i, j).value == 'Passed':
                wsheet_other.cell(i, j).fill = PatternFill('solid', fgColor='0FC404')
            elif wsheet_other.cell(i, j).value == 'Failed':
                wsheet_other.cell(i, j).fill = PatternFill('solid', fgColor='FC0E03')
            elif wsheet_other.cell(i, j).value == 'Showing 0 to 0':
                wsheet_other.cell(i, j).fill = PatternFill('solid', fgColor='FCC0BB')
    wb.save(report_folder + "\\" + workbook_title)





