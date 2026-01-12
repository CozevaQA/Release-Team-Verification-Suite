#!/usr/bin/env python
# coding: utf-8
import pickle

# In[ ]:


# revised scope - Predefined user of each role - Sidebar options, Support level metric specific lists, apptray options access, global search, All user session  -
# Review columns, document preview, screenshot if fail(maybe no) - NPI search - hardcode filter search on lists - GSD@
# Keep separate - NPI - Pickup from provider list.
# future scope - Include as part of daily validSummation of numerator/denominator roll up
import pytest
import allure
import os
import random
from statistics import mean
import sys
import math
import time
import traceback
import csv
from os import listdir
from os.path import isfile, join
import csv
import time
from datetime import date
from datetime import datetime
import base64
from tkinter import messagebox, ttk

import pandas as pd
from PIL import Image as img
from PIL import ImageTk
import PIL
from pytest_assume.plugin import assume
from selenium import *
from selenium import webdriver
from sigfig import round

from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.alert import Alert
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import *
from selenium.webdriver.support.wait import WebDriverWait
from openpyxl import Workbook
from openpyxl.styles import PatternFill, NamedStyle, Border
from urllib.parse import urlparse, parse_qs
from openpyxl.styles import *
import support_functions as sf
import setups
import variablestorage as locator
import ExcelProcessor as db
from tkinter import *
import sarvada_alluser


def check_and_create_All_view_userlist(file_path):
    default_userlist = ['Aagyeman@rhclinic.org', 'aemerson@cnhfclinics.org', 'aalvarez@cmcenters.org',
                        'admin@vistavillagepediatrics.com', 'altamed_kprice']
    if not os.path.exists(file_path):
        with open(file_path, "wb") as Allviewuserlist_file:
            pickle.dump(default_userlist, Allviewuserlist_file)


check_and_create_All_view_userlist('assets/allviewusers.pkl')


def check_context_switch_to_all(functiondriver):
    global global_test_case_id
    test_case_id = global_test_case_id
    # Check current context
    x = 0
    print("Checking Current Context")
    current_context = functiondriver.find_element(By.XPATH, locator.xpath_context_Name).text
    print(current_context)
    # check if current_context is all level
    if current_context == "ALL" or current_context == "All":
        print("its all")
        ws.append([test_case_id, current_context, "Checking All level view", "Passed","x", "User is already on All level", functiondriver.current_url])
        test_case_id+=1
    else:
        try:
            ws.append([test_case_id, current_context, "Switching User to All View", "Passed", "",
                      functiondriver.current_url])
            test_case_id += 1
            functiondriver.find_element(By.ID, 'context_dropdown_arrow').click()
            sf.ajax_preloader_wait(functiondriver)
            # WebDriverWait(functiondriver, 60).until(
            #        EC.presence_of_element_located((By.CLASS_NAME, 'context-menu-container')))
            heirarchy_list = functiondriver.find_element(By.CLASS_NAME, 'slides').find_elements(By.TAG_NAME, 'ul')
            for hlist in heirarchy_list:
                name_list = (hlist.text).split('\n')
                print(len(name_list))
                print(name_list)
                if name_list[0] == "All" or name_list[0] == "ALL":
                    hlist.find_elements(By.TAG_NAME, 'li')[0].click()
                    time.sleep(1)
                    functiondriver.find_element(By.CLASS_NAME, "context-menu-modal-apply").click()
                    start_time = time.perf_counter()
                    sf.ajax_preloader_wait(functiondriver)
                    total_time = time.perf_counter() - start_time
                    print("context switched to all")
                    current_context = functiondriver.find_element(By.XPATH, locator.xpath_context_Name).text
                    ws.append([test_case_id, current_context, "Context Switched to All", "Passed",round(total_time, sigfigs=3), "",
                              functiondriver.current_url])
                    test_case_id += 1
                    break
                elif name_list[0] == "clear":
                    if name_list[1] == "All" or name_list[1] == "ALL":
                        hlist.find_elements(By.TAG_NAME, 'li')[0].click()
                        time.sleep(1)
                        functiondriver.find_element(By.CLASS_NAME, "context-menu-modal-apply").click()
                        start_time = time.perf_counter()
                        sf.ajax_preloader_wait(functiondriver)
                        total_time = time.perf_counter() - start_time
                        print("context switched to all")
                        current_context = functiondriver.find_element(By.XPATH, locator.xpath_context_Name).text
                        ws.append([test_case_id, current_context, "Context Switched to All", "Passed",
                                  round(total_time, sigfigs=3), "",
                                  functiondriver.current_url])
                        test_case_id += 1
                        break
        except Exception as e:
            print(e)
            traceback.print_exc()
            current_context = functiondriver.find_element(By.XPATH, locator.xpath_context_Name).text
            ws.append([test_case_id, current_context, "Switching context to All", "Fail","x", "",
                      functiondriver.current_url])
            test_case_id += 1

            global_test_case_id = test_case_id


        global_test_case_id = test_case_id


    global_test_case_id = test_case_id

        # switch to All context


def navigate_all_menubar(functiondriver):
    global global_test_case_id
    test_case_id = global_test_case_id
    try:
        main_registry_url = driver.current_url
        WebDriverWait(functiondriver, 30).until(
            EC.presence_of_element_located((By.XPATH, locator.xpath_filter_measure_list)))
        WebDriverWait(functiondriver, 10).until(
            EC.presence_of_element_located((By.XPATH, locator.xpath_side_nav_SlideOut)))
        time.sleep(0.5)
        context_name = functiondriver.find_element_by_xpath(locator.xpath_context_Name).text
        print(context_name)
        functiondriver.find_element_by_xpath(locator.xpath_side_nav_SlideOut).click()
        time.sleep(1.5)
        links = functiondriver.find_elements_by_xpath(locator.xpath_menubar_Item_Link)
        names = functiondriver.find_elements_by_xpath(locator.xpath_menubar_Item_Link_Name)
        functiondriver.find_element_by_xpath(locator.xpath_side_nav_SlideOut).click()

    except Exception as e:
        print(e)
        traceback.print_exc()
        current_context = functiondriver.find_element(By.XPATH, locator.xpath_context_Name).text
        ws.append([test_case_id, current_context, "Navigating through Sidebar", "Failed","x", e,
                  functiondriver.current_url])
        test_case_id += 1

        global_test_case_id = test_case_id

    for link in range(len(links)) and range(len(names)):
        time.sleep(1.5)
        functiondriver.find_element_by_xpath(locator.xpath_side_nav_SlideOut).click()
        time.sleep(0.5)
        functiondriver.execute_script("arguments[0].scrollIntoView();", links[link])
        print("Link Index: " + str(link))
        print(names[link].text)
        link_name = names[link].text
        try:
            links[link].click()
            start_time = time.perf_counter()
            sf.ajax_preloader_wait(functiondriver)
            total_time = time.perf_counter() - start_time
            current_url = driver.current_url
            access_message = sf.URLAccessCheck(current_url, functiondriver)
            if not access_message:
                print("Access Check Done")
                if link_name != "Registries":
                    try:
                        start_time = time.perf_counter()
                        sf.sidebar_list_loading(functiondriver)
                        total_time += time.perf_counter() - start_time
                        current_context = functiondriver.find_element(By.XPATH, locator.xpath_context_Name).text
                        ws.append([test_case_id, current_context, "Access check to - "+link_name, "Passed",
                                  round(total_time, sigfigs=3), "",
                                  functiondriver.current_url])
                        test_case_id += 1
                    except Exception as e:
                        traceback.print_exc()
                        current_context = functiondriver.find_element(By.XPATH, locator.xpath_context_Name).text
                        ws.append([test_case_id, current_context, "Access check to - " + link_name, "Failed",
                                  round(total_time, sigfigs=3), "Timeout error",
                                  functiondriver.current_url])
                        test_case_id += 1
                if len(functiondriver.find_elements_by_xpath(locator.xpath_data_Table_Info)) != 0:
                    time.sleep(0.5)
                    datatable_info = functiondriver.find_element_by_xpath(locator.xpath_data_Table_Info).text
                    print(datatable_info)
                    current_context = functiondriver.find_element(By.XPATH, locator.xpath_context_Name).text
                    ws.append([test_case_id, current_context, "List Loading on - " + link_name, "Passed",
                              round(total_time, sigfigs=3), datatable_info,
                              functiondriver.current_url])
                    test_case_id += 1
                else:
                    print("Data table is 0/0")
            else:
                current_context = functiondriver.find_element(By.XPATH, locator.xpath_context_Name).text
                ws.append([test_case_id, current_context, "Access check to - " + link_name, "Failed","x", "Access Denied or Toast message recieved",
                          functiondriver.current_url])
                test_case_id += 1
        except Exception as e:
            print(e)
            traceback.print_exc()
            current_context = functiondriver.find_element(By.XPATH, locator.xpath_context_Name).text
            ws.append([test_case_id, current_context, "Sidebar Navigation", "Failed",
                      "x", "",
                      functiondriver.current_url])
            test_case_id += 1
        finally:
            links = functiondriver.find_elements_by_xpath(locator.xpath_menubar_Item_Link)
            names = functiondriver.find_elements_by_xpath(locator.xpath_menubar_Item_Link_Name)


    functiondriver.get(main_registry_url)
    sf.ajax_preloader_wait(functiondriver)
    WebDriverWait(functiondriver, 30).until(
        EC.presence_of_element_located((By.XPATH, locator.xpath_filter_measure_list)))
    print("Back to All Registry")

    global_test_case_id = test_case_id


def apptray_access_checks(functiondriver):
    global global_test_case_id
    test_case_id = global_test_case_id
    try:
        registry_url = driver.current_url
        window_switched = 0
        functiondriver.find_element_by_xpath(locator.xpath_app_Tray_Link).click()
        time.sleep(1)
        applist = functiondriver.find_element(By.ID, "app_dropdown").find_element(By.TAG_NAME, "div").find_elements(
            By.TAG_NAME, "Div")
        for app in applist:
            appname = app.find_element(By.TAG_NAME, "span")
            current_appname = appname.text
            print(current_appname)
            time.sleep(1)
            appname.click()
            start_time = time.perf_counter()
            sf.ajax_preloader_wait(functiondriver)
            total_time = time.perf_counter() - start_time
            time.sleep(1)
            if len(functiondriver.find_elements(By.CLASS_NAME, "modal-body")) > 0:
                print("No Access to this app")
                functiondriver.find_element(By.XPATH, "//a[@class='modal-dismiss right']").click()
                functiondriver.find_element_by_xpath(locator.xpath_app_Tray_Link).click()
                time.sleep(1)
                applist = functiondriver.find_element(By.ID, "app_dropdown").find_element(By.TAG_NAME,
                                                                                          "div").find_elements(
                    By.TAG_NAME, "Div")
                ws.append([test_case_id, current_appname, "Checking app tray Access", "Passed",
                           round(total_time, sigfigs=3), "No Access to this app",
                           functiondriver.current_url])
                test_case_id += 1
            elif len(functiondriver.window_handles) > 1:
                functiondriver.switch_to.window(driver.window_handles[1])
                window_switched = 1

                start_time = time.perf_counter()
                sf.ajax_preloader_wait(functiondriver)
                total_time = time.perf_counter() - start_time
                current_url = driver.current_url
                access_message = sf.URLAccessCheck(current_url, functiondriver)
                if not access_message:
                    print("Page Launched")
                    ws.append([test_case_id, current_appname, "Checking app tray Access", "Passed",
                               round(total_time, sigfigs=3), "",
                               functiondriver.current_url])
                    test_case_id += 1
                else:
                    print("Access Denied found")
                    ws.append([test_case_id, current_appname, "Checking app tray Access", "Failed",
                               round(total_time, sigfigs=3), "Access Denied found",
                               functiondriver.current_url])
                    test_case_id += 1

                if window_switched == 1:
                    functiondriver.close()
                    functiondriver.switch_to.window(driver.window_handles[0])
                    window_switched = 0
                    sf.ajax_preloader_wait(functiondriver)
                    functiondriver.find_element_by_xpath(locator.xpath_app_Tray_Link).click()
                    time.sleep(1)
                    applist = functiondriver.find_element(By.ID, "app_dropdown").find_element(By.TAG_NAME,
                                                                                              "div").find_elements(
                        By.TAG_NAME, "Div")

        try:
            driver.get(registry_url)
            sf.ajax_preloader_wait(functiondriver)
            WebDriverWait(functiondriver, 30).until(EC.presence_of_element_located((By.ID, "registry_body")))

        except Exception as e:
            print(e)
            traceback.print_exc()
            global_test_case_id = test_case_id





    except Exception as e:
        print(e)
        traceback.print_exc()
        global_test_case_id = test_case_id
        ws.append([test_case_id, current_appname, "Checking app tray Access", "Failed",
                   'x','Issue with navigating through apptray',
                   functiondriver.current_url])
        test_case_id += 1

    global_test_case_id = test_case_id


def check_support_level_mspl(functiondriver):
    global global_test_case_id
    test_case_id = global_test_case_id
    # Support level metric specific lists
    # Select a metric that is not 0/0 or 0%
    # click on that metric
    # Click on each list and check access, and empty list and store global searchable content in a list
    context_name = "Couldn't Fetch"
    registry_url = functiondriver.current_url
    try:
        WebDriverWait(functiondriver, 30).until(
            EC.presence_of_element_located((By.ID, "registry_body")))
        selected_metric_name = 'Couldnt fetch Metric Name'
        context_name = functiondriver.find_element_by_xpath(locator.xpath_context_Name).text
        registry_url = functiondriver.current_url

        metrics = functiondriver.find_element_by_id("registry_body").find_elements(By.CSS_SELECTOR, "li.li-metric")

        percent = '0.00'
        num_den = "(0/0)"
        iter_count = 0
        while percent == '0.00' or percent == '0.00%' or num_den == "(0/0)":
            selectedMetric = metrics[sf.RandomNumberGenerator(len(metrics), 1)[0]]
            functiondriver.execute_script("arguments[0].scrollIntoView({block: 'center'});", selectedMetric)
            try:
                percent = selectedMetric.find_element_by_class_name('percent').text
            except NoSuchElementException as e:
                print("Skipping % for util measure")
                print(e)
                traceback.print_exc()
            num_den = selectedMetric.find_element_by_class_name('num-den').text
            iter_count += 1
            print(iter_count)
            selected_metric_name = selectedMetric.find_element_by_class_name('met-name').text
            print(selected_metric_name)
            if iter_count > 10:
                ws.append([test_case_id, context_name, "Looking for a non 0/0, non util measure", "Failed",
                           "x", "All/most metrics are 0/0",
                           functiondriver.current_url])
                test_case_id += 1
                raise Exception("All/most metrics are 0/0")
        selected_metric_name = selectedMetric.find_element_by_class_name('met-name').text
        selectedMetric.click()
        start_time = time.perf_counter()
        sf.ajax_preloader_wait(functiondriver)
        total_time = time.perf_counter() - start_time
        print("Loaded into Metric- " + selected_metric_name)
        current_context = functiondriver.find_element(By.XPATH, locator.xpath_context_Name).text
        ws.append([test_case_id, current_context, "Loading into chosen Metric", "Passed",
                   round(total_time, sigfigs=3), selected_metric_name,
                   functiondriver.current_url])
        test_case_id += 1
        # check for number of tabs and thier names
        MSPL_link = functiondriver.current_url
        tabs = functiondriver.find_element(By.CLASS_NAME, "tabs").find_elements(By.CLASS_NAME, "tab")
        for index, tab in enumerate(tabs):
            tab_name = tabs[index].text

            if tab_name == "Practices":
                tabs[index].click()
                start_time = time.perf_counter()
                sf.ajax_preloader_wait(functiondriver)
                total_time = time.perf_counter() - start_time
                access_message = sf.URLAccessCheck(functiondriver.current_url, functiondriver)
                Practice_MSPL_url = functiondriver.current_url

                if not access_message:
                    WebDriverWait(functiondriver, 30).until(EC.presence_of_element_located((By.CLASS_NAME, "tabs")))
                    tabs = functiondriver.find_element(By.CLASS_NAME, "tabs").find_elements(By.CLASS_NAME, "tab")
                    current_context = functiondriver.find_element(By.XPATH, locator.xpath_context_Name).text
                    ws.append([test_case_id, current_context, "Navigation to practice tab", "Passed",
                               round(total_time, sigfigs=3), "",
                               functiondriver.current_url])
                    test_case_id += 1
                    if len(driver.find_elements(By.CLASS_NAME, 'dataTables_empty')) != 0:
                        print("No Practices")
                        current_context = functiondriver.find_element(By.XPATH, locator.xpath_context_Name).text
                        ws.append([test_case_id, current_context, "Looking for a non 0 count Practice", "Failed",
                                   'x', "No Practices",
                                   functiondriver.current_url])
                        test_case_id += 1
                    elif len(driver.find_elements(By.CLASS_NAME, 'dataTables_empty')) == 0:
                        entries = functiondriver.find_element(By.XPATH, locator.xpath_data_Table_Info).text
                        practice_table = driver.find_element(By.ID, "metric-support-prac-ls").find_element(By.TAG_NAME,
                                                                                                           "tbody").find_elements(By.TAG_NAME, "tr")
                        print(len(practice_table))
                        print(entries)
                        selected_practice_name = \
                        random.choice(practice_table).find_elements(By.TAG_NAME, "td")[3].find_elements(By.TAG_NAME,
                                                                                                        "a")[1]
                        current_context = functiondriver.find_element(By.XPATH, locator.xpath_context_Name).text
                        ws.append([test_case_id, current_context, "Selecting a random Practice", "Passed",
                                   round(total_time, sigfigs=3), "Selected - "+selected_practice_name.text,
                                   functiondriver.current_url])
                        test_case_id += 1
                        global_search_practice_name = selected_practice_name.text
                        print("Global Search practice name = " + global_search_practice_name)
                        selected_practice_name.click()
                        start_time = time.perf_counter()
                        sf.ajax_preloader_wait(functiondriver)
                        total_time = time.perf_counter() - start_time
                        access_message = sf.URLAccessCheck(functiondriver.current_url, functiondriver)
                        if not access_message:
                            WebDriverWait(functiondriver, 300).until(
                                EC.presence_of_element_located((By.XPATH, locator.xpath_filter_measure_list)))
                            print("Practice Registry launched")
                            current_context = functiondriver.find_element(By.XPATH, locator.xpath_context_Name).text
                            ws.append([test_case_id, current_context, "Navigating to selected random Practice", "Passed",
                                       round(total_time, sigfigs=3), "Selected - " + global_search_practice_name,
                                       functiondriver.current_url])
                            test_case_id += 1
                        else:
                            print("Practice Registry Access Denied")
                            current_context = functiondriver.find_element(By.XPATH, locator.xpath_context_Name).text
                            ws.append([test_case_id, current_context, "Selecting a random Practice", "Failed",
                                       round(total_time, sigfigs=3), "Selected - " + global_search_practice_name + ": Access Denied",
                                       functiondriver.current_url])
                            test_case_id += 1

                        driver.get(MSPL_link)
                        sf.ajax_preloader_wait(functiondriver)
                        WebDriverWait(functiondriver, 30).until(EC.presence_of_element_located((By.CLASS_NAME, "tabs")))
                        tabs = functiondriver.find_element(By.CLASS_NAME, "tabs").find_elements(By.CLASS_NAME, "tab")
                        window_switched = 0
                        functiondriver.find_element_by_id('globalsearch_input').send_keys(global_search_practice_name)
                        start_time = time.perf_counter()
                        WebDriverWait(functiondriver, 45).until(
                            EC.presence_of_element_located((By.CLASS_NAME, 'collection-header')))
                        time_taken = round(time.perf_counter() - start_time)
                        functiondriver.find_element_by_id('globalsearch_input').send_keys(Keys.RETURN)
                        sf.ajax_preloader_wait(functiondriver)
                        time_taken = round(time.perf_counter() - start_time)
                        # driver.find_element_by_id('globalsearch_input').send_keys(Keys.RETURN)
                        functiondriver.find_element_by_id('search_practices_link').click()
                        WebDriverWait(functiondriver, 30).until(
                            EC.presence_of_element_located((By.ID, 'search_practices')))
                        functiondriver.find_element_by_id('search_practices').find_elements_by_tag_name('a')[0].click()
                        functiondriver.switch_to.window(functiondriver.window_handles[1])
                        window_switched = 1
                        sf.ajax_preloader_wait(functiondriver)
                        access_message = sf.URLAccessCheck(functiondriver.current_url, functiondriver)
                        if not access_message:
                            WebDriverWait(functiondriver, 300).until(
                                EC.presence_of_element_located((By.XPATH, locator.xpath_filter_measure_list)))
                            print("Practice Registry launched via Global Search")
                            current_context = functiondriver.find_element(By.XPATH, locator.xpath_context_Name).text
                            ws.append(
                                [test_case_id, current_context, "Global Search of Random practice", "Passed",
                                 round(total_time, sigfigs=3), "Search String - " + global_search_practice_name,
                                 functiondriver.current_url])
                            test_case_id += 1
                        else:
                            print("Practice Registry Access Denied via Global Search")
                            current_context = functiondriver.find_element(By.XPATH, locator.xpath_context_Name).text
                            ws.append(
                                [test_case_id, current_context, "Global Search of Random practice", "Failed",
                                 round(total_time, sigfigs=3), "Search String - " + global_search_practice_name +" - Access Denied recieved" ,
                                 functiondriver.current_url])
                            test_case_id += 1

                        if window_switched == 1:
                            functiondriver.close()
                            functiondriver.switch_to.window(driver.window_handles[0])
                            window_switched = 0

                        driver.get(MSPL_link)
                        sf.ajax_preloader_wait(functiondriver)
                        WebDriverWait(functiondriver, 30).until(EC.presence_of_element_located((By.CLASS_NAME, "tabs")))
                        tabs = functiondriver.find_element(By.CLASS_NAME, "tabs").find_elements(By.CLASS_NAME, "tab")
                else:
                    print("Access Denied found on Practice's Tab")
                    current_context = functiondriver.find_element(By.XPATH, locator.xpath_context_Name).text
                    ws.append([test_case_id, current_context, "Navigation to practice tab", "Failed",
                               'x', "Access Denied",
                               functiondriver.current_url])
                    test_case_id += 1

            if tab_name == "Providers":
                tabs[index].click()
                start_time = time.perf_counter()
                sf.ajax_preloader_wait(functiondriver)
                total_time = time.perf_counter() - start_time
                access_message = sf.URLAccessCheck(functiondriver.current_url, functiondriver)
                if not access_message:
                    WebDriverWait(functiondriver, 30).until(EC.presence_of_element_located((By.CLASS_NAME, "tabs")))
                    tabs = functiondriver.find_element(By.CLASS_NAME, "tabs").find_elements(By.CLASS_NAME, "tab")
                    current_context = functiondriver.find_element(By.XPATH, locator.xpath_context_Name).text
                    ws.append([test_case_id, current_context, "Navigation to provider tab", "Passed",
                               round(total_time, sigfigs=3), "",
                               functiondriver.current_url])
                    test_case_id += 1
                    if len(driver.find_elements(By.CLASS_NAME, 'dataTables_empty')) != 0:
                        print("No Providers")
                        current_context = functiondriver.find_element(By.XPATH, locator.xpath_context_Name).text
                        ws.append([test_case_id, current_context, "Looking for a non 0 count Provider", "Failed",
                                   'x', "No Providers",
                                   functiondriver.current_url])
                        test_case_id += 1
                    elif len(driver.find_elements(By.CLASS_NAME, 'dataTables_empty')) == 0:
                        entries = functiondriver.find_element(By.XPATH, locator.xpath_data_Table_Info).text
                        provider_table = driver.find_element(By.ID, "metric-support-prov-ls").find_element(By.TAG_NAME,
                                                                                                           "tbody").find_elements(By.TAG_NAME, "tr")
                        print(len(provider_table))
                        print(entries)
                        selected_provider_name = \
                            random.choice(provider_table).find_elements(By.TAG_NAME, "td")[4].find_elements(By.TAG_NAME, "a")[1]
                        current_context = functiondriver.find_element(By.XPATH, locator.xpath_context_Name).text
                        ws.append([test_case_id, current_context, "Selecting a random provider", "Passed",
                                   round(total_time, sigfigs=3), "Selected - " + selected_provider_name.text,
                                   functiondriver.current_url])
                        test_case_id += 1
                        global_search_provider_name = selected_provider_name.text
                        print("Global Search Provider name = " + global_search_provider_name)
                        selected_provider_name.click()
                        start_time = time.perf_counter()
                        sf.ajax_preloader_wait(functiondriver)
                        total_time = time.perf_counter() - start_time
                        access_message = sf.URLAccessCheck(functiondriver.current_url, functiondriver)
                        if not access_message:
                            WebDriverWait(functiondriver, 300).until(
                                EC.presence_of_element_located((By.XPATH, locator.xpath_filter_measure_list)))
                            print("Provider Registry launched")
                            current_context = functiondriver.find_element(By.XPATH, locator.xpath_context_Name).text
                            ws.append(
                                [test_case_id, current_context, "Navigating to selected random Provider", "Passed",
                                 round(total_time, sigfigs=3), "Selected - " + global_search_provider_name,
                                 functiondriver.current_url])
                            test_case_id += 1
                        else:
                            print("Provider Registry Access Denied")
                            current_context = functiondriver.find_element(By.XPATH, locator.xpath_context_Name).text
                            ws.append([test_case_id, current_context, "Selecting a random Provider", "Failed",
                                       round(total_time, sigfigs=3),
                                       "Selected - " + global_search_provider_name + ": Access Denied",
                                       functiondriver.current_url])
                            test_case_id += 1

                        driver.get(MSPL_link)
                        sf.ajax_preloader_wait(functiondriver)
                        WebDriverWait(functiondriver, 30).until(EC.presence_of_element_located((By.CLASS_NAME, "tabs")))
                        tabs = functiondriver.find_element(By.CLASS_NAME, "tabs").find_elements(By.CLASS_NAME, "tab")
                        window_switched = 0
                        functiondriver.find_element_by_id('globalsearch_input').send_keys(global_search_provider_name)
                        start_time = time.perf_counter()
                        WebDriverWait(functiondriver, 45).until(
                            EC.presence_of_element_located((By.CLASS_NAME, 'collection-header')))
                        total_time = time.perf_counter() - start_time
                        functiondriver.find_element_by_id('globalsearch_input').send_keys(Keys.RETURN)
                        start_time = time.perf_counter()
                        sf.ajax_preloader_wait(functiondriver)
                        total_time += round(time.perf_counter() - start_time)
                        # driver.find_element_by_id('globalsearch_input').send_keys(Keys.RETURN)
                        functiondriver.find_element_by_id('search_providers_link').click()
                        WebDriverWait(functiondriver, 30).until(
                            EC.presence_of_element_located((By.ID, 'search_providers')))
                        functiondriver.find_element_by_id('search_providers').find_elements_by_tag_name('a')[0].click()
                        functiondriver.switch_to.window(functiondriver.window_handles[1])
                        window_switched = 1
                        start_time = time.perf_counter()
                        sf.ajax_preloader_wait(functiondriver)
                        total_time += round(time.perf_counter() - start_time)
                        access_message = sf.URLAccessCheck(functiondriver.current_url, functiondriver)
                        if not access_message:
                            WebDriverWait(functiondriver, 300).until(
                                EC.presence_of_element_located((By.XPATH, locator.xpath_filter_measure_list)))
                            print("Provider Registry launched via Global Search")
                            current_context = functiondriver.find_element(By.XPATH, locator.xpath_context_Name).text
                            ws.append(
                                [test_case_id, current_context, "Global Search of Random provider", "Passed",
                                 round(total_time, sigfigs=3), "Search String - " + global_search_provider_name,
                                 functiondriver.current_url])
                            test_case_id += 1
                        else:
                            print("Provider Registry Access Denied via Global Search")
                            current_context = functiondriver.find_element(By.XPATH, locator.xpath_context_Name).text
                            ws.append(
                                [test_case_id, current_context, "Global Search of Random provider", "Failed",
                                 round(total_time, sigfigs=3),
                                 "Search String - " + global_search_provider_name + " - Access Denied recieved",
                                 functiondriver.current_url])
                            test_case_id += 1

                        if window_switched == 1:
                            functiondriver.close()
                            functiondriver.switch_to.window(driver.window_handles[0])
                            window_switched = 0

                        driver.get(MSPL_link)
                        sf.ajax_preloader_wait(functiondriver)
                        WebDriverWait(functiondriver, 30).until(EC.presence_of_element_located((By.CLASS_NAME, "tabs")))
                        tabs = functiondriver.find_element(By.CLASS_NAME, "tabs").find_elements(By.CLASS_NAME, "tab")
                else:
                    print("Access Denied found on Provider's Tab")
                    current_context = functiondriver.find_element(By.XPATH, locator.xpath_context_Name).text
                    ws.append([test_case_id, current_context, "Navigation to provider tab", "Failed",
                               'x', "Access Denied",
                               functiondriver.current_url])
                    test_case_id += 1
            if tab_name == "Patients":  # currently out of scope
                tabs[index].click()
                sf.ajax_preloader_wait(functiondriver)
                access_message = sf.URLAccessCheck(functiondriver.current_url, functiondriver)
                if not access_message:
                    WebDriverWait(functiondriver, 30).until(EC.presence_of_element_located((By.CLASS_NAME, "tabs")))
                    tabs = functiondriver.find_element(By.CLASS_NAME, "tabs").find_elements(By.CLASS_NAME, "tab")
                else:
                    print("Access Denied found")
            if tab_name == "Performance Statistics":
                tabs[index].click()
                start_time = time.perf_counter()
                sf.ajax_preloader_wait(functiondriver)
                total_time = round(time.perf_counter() - start_time)

                access_message = sf.URLAccessCheck(functiondriver.current_url, functiondriver)
                if not access_message:
                    WebDriverWait(functiondriver, 30).until(EC.presence_of_element_located((By.CLASS_NAME, "tabs")))
                    tabs = functiondriver.find_element(By.CLASS_NAME, "tabs").find_elements(By.CLASS_NAME, "tab")

                    if len(functiondriver.find_elements(By.ID, "performance_details")) != 0:
                        print("Navigated Successfully to Performance Statistics")
                        current_context = functiondriver.find_element(By.XPATH, locator.xpath_context_Name).text
                        ws.append([test_case_id, current_context, "Navigation to performance statistics tab", "Passed",
                                   round(total_time, sigfigs=3), "",
                                   functiondriver.current_url])
                        test_case_id += 1
                    else:
                        print("Performance Statistics not loading")
                        current_context = functiondriver.find_element(By.XPATH, locator.xpath_context_Name).text
                        ws.append([test_case_id, current_context, "Navigation to performance statistics tab", "Failed",
                                   'x', "Performance Statistics not loading",
                                   functiondriver.current_url])
                        test_case_id += 1
                else:
                    print("Access Denied found")
                    current_context = functiondriver.find_element(By.XPATH, locator.xpath_context_Name).text
                    ws.append([test_case_id, current_context, "Navigation to performance statistics tab", "Failed",
                               'x', "Access Denied",
                               functiondriver.current_url])
                    test_case_id += 1



    except Exception as e:
        print(e)
        traceback.print_exc()
        current_context = functiondriver.find_element(By.XPATH, locator.xpath_context_Name).text
        ws.append([test_case_id, current_context, "Navigation through support level tabs", "Failed",
                   'x', "Access Denied/unknown error occured",
                   functiondriver.current_url])
        test_case_id += 1

    try:
        driver.get(registry_url)
        sf.ajax_preloader_wait(functiondriver)
        WebDriverWait(functiondriver, 30).until(EC.presence_of_element_located((By.ID, "registry_body")))

    except Exception as e:
        print(e)
        traceback.print_exc()



def check_current_session(functiondriver):
    from urllib.parse import urlparse, parse_qs
    import base64
    All_view_url = functiondriver.current_url
    print(All_view_url)
    parsed_url = urlparse(All_view_url)
    query_params = parse_qs(parsed_url.query)
    base64_string = query_params.get("session", [""])[0]
    decoded_bytes = base64.b64decode(base64_string)
    decoded_string = decoded_bytes.decode('utf-8')

    url_parameter_list = decoded_string.split("&")
    decoded_string = ""
    for parameter in url_parameter_list:
        decoded_string = (decoded_string + parameter).strip() + "\n"

    print("Base64 Encoded:", base64_string)
    print("Decoded String:\n", decoded_string.strip())


def check_document_preview(functiondriver):
    global global_test_case_id
    test_case_id = global_test_case_id
    registry_url = functiondriver.current_url
    try:
        WebDriverWait(functiondriver, 30).until(
            EC.presence_of_element_located((By.XPATH, locator.xpath_filter_measure_list)))
        WebDriverWait(functiondriver, 10).until(
            EC.presence_of_element_located((By.XPATH, locator.xpath_side_nav_SlideOut)))
        time.sleep(0.5)
        context_name = functiondriver.find_element_by_xpath(locator.xpath_context_Name).text
        print(context_name)
        functiondriver.find_element_by_xpath(locator.xpath_side_nav_SlideOut).click()
        time.sleep(1.5)
        links = functiondriver.find_elements_by_xpath(locator.xpath_menubar_Item_Link)
        names = functiondriver.find_elements_by_xpath(locator.xpath_menubar_Item_Link_Name)
        sudata_list_flag = 0
        for index, link_name in enumerate(names):
            if link_name.text == "Supplemental Data":
                links[index].click()
                sf.ajax_preloader_wait(functiondriver)
                sf.sidebar_list_loading(functiondriver)
                sudata_list_flag = 1
                break
        if sudata_list_flag == 1:
            print("Su Data list Launched")
            sudata_list_url = functiondriver.current_url
            # Review Columns
            try:
                review_col_count = 0
                review_col_list = ""
                for review_num in range(6):
                    if len(functiondriver.find_elements(By.XPATH,
                                                        "//th[@aria-label='Review " + str(review_num) + "']")) > 0:
                        review_col_list = review_col_list + functiondriver.find_elements(By.XPATH,
                                                        "//th[@aria-label='Review " + str(review_num) + "']").text + " "
                        review_col_count += 1
                print("Total Review columns: " + str(review_col_count))
                ws.append([test_case_id, "SuppData List", "Counting review Columns", "Passed",
                           'x', "Total Review Columns:" + str(review_col_count) + ", and the columns are " + review_col_list,
                           functiondriver.current_url])
                test_case_id+=1
                driver.get(sudata_list_url)
                sf.ajax_preloader_wait(functiondriver)
                WebDriverWait(functiondriver, 300).until(
                    EC.presence_of_element_located((By.XPATH, "//div[@class='ch table_header']")))
            except Exception as e:
                print(e)
                traceback.print_exc()
            # Document Preview
            try:
                attachments = functiondriver.find_elements(By.XPATH, '//i[@data-tooltip="Attachment"]')
                random.choice(attachments).click()
                time.sleep(1)
                if len(functiondriver.find_elements(By.XPATH, "//div[@class='modal dashboard-modal open']")) > 0:
                    document_link = functiondriver.find_element(By.XPATH,
                                                                "//div[@class='modal dashboard-modal open']").find_element(
                        By.XPATH, "//*[@type='application/pdf']").get_attribute('src')
                    print(document_link)
                    ws.append([test_case_id, "SuppData List", "Checking document attachments", "Passed",
                               'x', document_link,
                               functiondriver.current_url])
                    test_case_id += 1
                else:
                    ws.append([test_case_id, "SuppData List", "Checking document attachments", "Failed",
                               'x', "Attachment iframe failed to launch",
                               functiondriver.current_url])
                    test_case_id += 1

                driver.get(sudata_list_url)
                sf.ajax_preloader_wait(functiondriver)
                WebDriverWait(functiondriver, 300).until(
                    EC.presence_of_element_located((By.XPATH, "//div[@class='ch table_header']")))
            except Exception as e:
                print(e)
                traceback.print_exc()
                ws.append([test_case_id, "SuppData List", "Checking document attachments", "Failed",
                           'x', "Attachment iframe failed to launch",
                           functiondriver.current_url])
                test_case_id += 1

            # Hardcoded GSD filter
            try:
                functiondriver.find_element(By.CLASS_NAME, "datatable_filter_dropdown").click()


            except Exception as e:
                print(e)
                traceback.print_exc()




        else:
            print("Su Data list failed")

    except Exception as e:
        print(e)
        traceback.print_exc()
        ws.append([test_case_id, "SuppData List", "Navigation to supplemental data list", "Failed",
                   'x', "Suppdata list failed to open",
                   functiondriver.current_url])
        test_case_id += 1

    try:
        driver.get(registry_url)
        sf.ajax_preloader_wait(functiondriver)
        WebDriverWait(functiondriver, 30).until(EC.presence_of_element_located((By.ID, "registry_body")))

    except Exception as e:
        print(e)
        traceback.print_exc()


def check_REL_chart(functiondriver):
    global global_test_case_id
    test_case_id = global_test_case_id

    registry_url = driver.current_url

    try:
        WebDriverWait(functiondriver, 30).until(
            EC.presence_of_element_located((By.XPATH, locator.xpath_filter_measure_list)))
        patient_count = ""
        summaryList = driver.find_element(By.CLASS_NAME, "registry_header_panel").find_elements(By.TAG_NAME, "div")
        for div, next_div in zip(summaryList, summaryList[1:] + [summaryList[0]]):
            if "Patients" in div.text:
                patient_count = next_div.text.replace(",", "").strip()
        print(patient_count)

    except Exception as e:
        print(e)
        traceback.print_exc()


    try:
        time.sleep(1)
        LOBdropdownelement = functiondriver.find_element_by_xpath("//*[@id='qt-filter-label']")
        LOBdropdownelement.click()

        time.sleep(1)
        WebDriverWait(driver, 30).until(
            EC.presence_of_element_located((By.XPATH, "//*[@id='filter-lob']")))
        LOBname = LOBdropdownelement.find_element_by_xpath("//*[@id='filter-lob']")
        LOBnamelist = LOBname.find_elements_by_tag_name("li")
        print(*LOBnamelist)
        Payername = LOBdropdownelement.find_elements_by_xpath("//*[@id='filter-payer']")
        # LOBdropdownelement.click()
        time.sleep(1)
        for j in range(0, len(LOBnamelist)):
            # LOBdropdownelement.click()
            time.sleep(1)
            print(LOBnamelist[j].text)
            print("--------------------------------")
            try:
                LOBnamelist[j].click()
            except ElementNotInteractableException as e:
                continue
            currentLOBName = LOBnamelist[j].text
            time.sleep(4)
            driver.find_element_by_xpath("//*[@id='reg-filter-apply']").click()
            time.sleep(2)
            loader = WebDriverWait(driver, 300)
            loader.until(
                EC.invisibility_of_element_located((By.XPATH, "//div/div[contains(@class,'ajax_preloader')]")))
            # logger.captureScreenshot(driver, currentLOBName, screenshot_path)
            # Checking Patient count in Lob. Raise error if it is 0
            Lob_type = ["ALL", "Medicare", "Medicare ACO"]
            # Patient count and performance count
            sf.ajax_preloader_wait(driver)
            WebDriverWait(driver, 30).until(
                EC.presence_of_element_located((By.XPATH, locator.xpath_filter_measure_list)))
            try:
                summaryList = driver.find_element_by_class_name("registry_header_panel").find_elements_by_tag_name("div")
                overall_rating = ""
                patient_count = ""
                for div, next_div in zip(summaryList, summaryList[1:] + [summaryList[0]]):
                    if "Overall Rating" in div.text:
                        overall_rating = next_div.text.replace("%", "").strip()
                        if 'Stars' in overall_rating:
                            overall_rating = overall_rating.replace("Stars", "").strip()
                        if 'stars' in overall_rating:
                            overall_rating = overall_rating.replace("stars", "").strip()
                    if "Patients" in div.text:
                        patient_count = next_div.text.replace(",", "").strip()
                print(overall_rating)
                print(patient_count)
                if float(overall_rating) < 1:
                    ws.append(
                        [test_case_id, currentLOBName +" : Overall Rating", overall_rating, "Failed", "Rating is 0", driver.current_url])
                else:
                    ws.append([test_case_id, currentLOBName +" : Overall Rating", overall_rating, "Passed"])

                test_case_id += 1
                if float(patient_count) < 1:
                    ws.append(
                        [test_case_id, currentLOBName +" : Patient Count", patient_count, "Failed", "Patient count is 0", driver.current_url])
                else:
                    ws.append([test_case_id, currentLOBName +" : Patient Count", patient_count, "Passed"])
                test_case_id += 1


            except Exception as e:
                traceback.print_exc()
                ws.append([currentLOBName, "x", "x", "Couldn't fetch patient counts/rating"])

            try:
                driver.find_element(By.XPATH, '//*[@id="rel-chart"]').click()
                time.sleep(1)
                if len(driver.find_elements(By.XPATH, locator.xpath_rel_chart)) > 0:
                    if "No data found" in driver.find_element(By.XPATH, locator.xpath_rel_chart).text:
                        test_case_id += 1
                        # ws.append((test_case_id, context_name, 'REL Chart', 'Failed', 'x', 'No data found'))
                        print("REL chart failed not clickable")
                    else:
                        test_case_id += 1
                        # ws.append((test_case_id, context_name, 'REL Chart', 'Passed', 'x', 'Chart is Clickable'))
                        print("REL chart opened")
                        rel_counts = [int(count.text.replace(",", "")) for count in
                                      functiondriver.find_elements(By.XPATH, "//*[@dy='0.35em']")]
                        print(rel_counts)
                        print(sum(rel_counts))

                    driver.find_element(By.CLASS_NAME, "rel-chart-close").click()
                else:
                    test_case_id += 1
                    # ws.append((test_case_id, context_name, 'REL Chart', 'Failed', 'x', 'Chart not clickable'))
                    print("REL chart failed not clickable")

                driver.get(registry_url)
                sf.ajax_preloader_wait(functiondriver)
                WebDriverWait(functiondriver, 30).until(EC.presence_of_element_located((By.ID, "registry_body")))

            except Exception as e:
                print(e)
                traceback.print_exc()

            try:
                if int(patient_count) == sum(rel_counts):
                    ws.append([test_case_id,  currentLOBName +" : REL Chart", "Checking summation of the REL Chart Counts", "Passed",
                               'x',
                               "Patient Count is " + patient_count + ", and the summation is " + str(sum(rel_counts)),
                               functiondriver.current_url])
                    test_case_id += 1
                else:
                    ws.append([test_case_id,  currentLOBName +" : REL Chart", "Checking summation of the REL Chart Counts", "Failed"
                                                                                                        'x',
                               "Patient Count is " + patient_count + ", and the summation is " + str(sum(rel_counts)),
                               functiondriver.current_url])
                    test_case_id += 1
            except Exception as e:
                ws.append([test_case_id,  currentLOBName +" : REL Chart", "Checking summation of the REL Chart Counts", "Failed",
                           'x', "Unknown Error Occurred",
                           functiondriver.current_url])
                test_case_id += 1

            time.sleep(1)
            LOBdropdownelement = driver.find_element_by_xpath("//*[@id='qt-filter-label']")
            LOBdropdownelement.click()
            time.sleep(3)
            LOBname = LOBdropdownelement.find_element_by_xpath("//*[@id='filter-lob']")
            LOBnamelist = LOBname.find_elements_by_tag_name("li")

    except Exception as e:
        print(e)
        traceback.print_exc()

    try:
        driver.get(registry_url)
        sf.ajax_preloader_wait(functiondriver)
        WebDriverWait(functiondriver, 30).until(EC.presence_of_element_located((By.ID, "registry_body")))

    except Exception as e:
        print(e)
        traceback.print_exc()


def rollup_summation_checks(functiondriver, user):
    global global_test_case_id
    test_case_id = global_test_case_id
    try:
        report_link = sarvada_alluser.extract_context_data(functiondriver, user)

        ws.append([test_case_id, "All context", "Rollup Validation", "HTML Report link",'x', report_link,functiondriver.current_url])
    except Exception as e:
        traceback.print_exc()
        ws.append([test_case_id, "All context", "Rollup Validation", "HTML Report link", 'x', e,
                   functiondriver.current_url])


    test_case_id = 0
    global_test_case_id = test_case_id


def mainui():
    root = Tk()

    root.configure(background="white")

    def on_save_users():
        new_user_list = []
        for new_user in username_entryboxes:
            new_user_list.append(new_user.get().strip())
        print(new_user_list)
        with open('assets/allviewusers.pkl', "wb") as Allviewuserlist_file:
            pickle.dump(new_user_list, Allviewuserlist_file)

    def on_select_all():
        if feature_variables[0].get() == 1:
            for check_box in feature_checkboxes:
                check_box.state(['selected'])
            for iterator in range(1, len(feature_variables)):
                feature_variables[iterator].set(1)
        elif feature_variables[0].get() == 0:
            for check_box in feature_checkboxes:
                check_box.state(['!selected'])
            for iterator in range(1, len(feature_variables)):
                feature_variables[iterator].set(0)

    def on_start_button():
        global run_config
        run_config = []
        runtime_user_list = []
        runtime_feature_list = []
        user_flags = user_selection_input_box.get().strip().split(',')
        for index, user_flag in enumerate(user_flags):
            try:
                user_flags[index] = int(user_flag.strip())-1
            except Exception as e:
                continue

        for index, users in enumerate(username_entryboxes):
            if index in user_flags:
                runtime_user_list.append(users.get().strip())
        for features in feature_variables:
            runtime_feature_list.append(features.get())

        run_config.append(runtime_user_list)
        run_config.append(runtime_feature_list)
        print(run_config)
        root.destroy()

    def create_tooltip(widget, text):
        def on_enter(event):
            global tooltip_window
            x = widget.winfo_rootx() + 20
            y = widget.winfo_rooty() + widget.winfo_height() + 20
            tooltip_window = Toplevel()
            tooltip_window.wm_overrideredirect(True)
            tooltip_window.wm_geometry("+%d+%d" % (x, y))
            label = Label(tooltip_window, text=text, justify='left',
                          background='#5a9c00', relief='solid', borderwidth=1,
                          font=("tahoma", "10", "bold"), fg="#000000")
            label.pack(ipadx=1)

        def on_leave(event):
            global tooltip_window
            tooltip_window.destroy()

        widget.bind("<Enter>", on_enter)
        widget.bind("<Leave>", on_leave)

    style = ttk.Style()
    style.theme_use('alt')
    style.configure('My.TButton', font=('Helvetica', 13, 'bold'), foreground='Black', background='#5a9c32', padding=15,
                    highlightthickness=0, height=1, width=25)
    style.configure('Configs.TButton', font=('Helvetica', 8, 'bold'), foreground='Black', background='#5a9c32',
                    highlightthickness=0)
    style.configure('Next.TButton', font=('Helvetica', 13, 'bold'), foreground='Black', background='#5a9c32',
                    highlightthickness=0)
    style.configure('CheckbuttonStyle.TCheckbutton', font=('Helvetica', 10, 'bold'), foreground='Black',
                    background='white')

    style.map('My.TButton', background=[('active', '#72B132')])
    style.map('Next.TButton', background=[('active', '#72B132')])

    feature_variables = [IntVar(), IntVar(), IntVar(), IntVar(), IntVar(), IntVar(), IntVar(), IntVar()]

    cozeva_logo_image = ImageTk.PhotoImage(img.open("assets/images/cozeva_logo.png").resize((320, 74)))
    help_icon_image = ImageTk.PhotoImage(img.open("assets/images/help_icon.png").resize((15, 15)))

    user1_inputframe = Entry(root, width=30)
    user2_inputframe = Entry(root, width=30)
    user3_inputframe = Entry(root, width=30)
    user4_inputframe = Entry(root, width=30)
    user5_inputframe = Entry(root, width=30)

    user_selection_input_box = Entry(root, width=30)
    user_selection_input_box_tooltip = "Enter Comma seperated values for users to include in this run. eg, '1', '1,2,3', '1,4,3', '1,2,3,4,5'"
    username_entryboxes = []
    username_entryboxes.append(user1_inputframe)
    username_entryboxes.append(user2_inputframe)
    username_entryboxes.append(user3_inputframe)
    username_entryboxes.append(user4_inputframe)
    username_entryboxes.append(user5_inputframe)

    feature_checkbox_select_all = ttk.Checkbutton(root, text="Select All", variable=feature_variables[0],
                                                  style='CheckbuttonStyle.TCheckbutton', command=on_select_all)

    all_level_menubar_checkbox = ttk.Checkbutton(root, text="All Level Menubar", variable=feature_variables[1],
                                                 style='CheckbuttonStyle.TCheckbutton')
    all_level_menubar_checkbox_tooltip = "Checks all level sidebar list loading, time to render lists and missing options if any"

    apptray_access_check_checkbox = ttk.Checkbutton(root, text="Apptray Access Check", variable=feature_variables[2],
                                                    style='CheckbuttonStyle.TCheckbutton')
    apptray_access_check_checkbox_tooltip = "Checks all apptray options, reports access status and compares against expected apps"

    support_mspl_glob_search_checkbox = ttk.Checkbutton(root, text="Support MSPL and Global Search",
                                                        variable=feature_variables[3],
                                                        style='CheckbuttonStyle.TCheckbutton')
    support_mspl_glob_search_checkbox_tooltip = "Checks Metric Specific Practice/Provider/patient lists, respective registries, global search and performance statistics"

    session_info_checkbox = ttk.Checkbutton(root, text="Fetch Session Info", variable=feature_variables[4],
                                            style='CheckbuttonStyle.TCheckbutton')
    session_info_checkbox_tooltip = "Checks and records the current logged in session"

    chart_list_specifics_checkbox = ttk.Checkbutton(root, text="Chart List Specifics", variable=feature_variables[5],
                                                    style='CheckbuttonStyle.TCheckbutton')
    chart_list_specifics_checkbox_tooltip = "Checks available review labels, Document attachement preview and measure filter"

    registry_summary_checkbox = ttk.Checkbutton(root, text="Registry Summary Bar", variable=feature_variables[6],
                                                    style='CheckbuttonStyle.TCheckbutton')
    registry_summary_checkbox_tooltip = "Checks REL chart on the Summary Bar, and patient counts and for valid rating. Also counts total number of 0/0 measures"

    rollup_summation_checkbox = ttk.Checkbutton(root, text="Rollup summation", variable=feature_variables[7],
                                                    style='CheckbuttonStyle.TCheckbutton')
    rollup_summation_checkbox_tooltip = "Validates All level rolled up counts vs other groups, Generates an HTML report"

    feature_checkboxes = []
    feature_checkboxes.append(all_level_menubar_checkbox)
    feature_checkboxes.append(apptray_access_check_checkbox)
    feature_checkboxes.append(support_mspl_glob_search_checkbox)
    feature_checkboxes.append(session_info_checkbox)
    feature_checkboxes.append(chart_list_specifics_checkbox)
    feature_checkboxes.append(rollup_summation_checkbox)

    logo_label = Label(root, image=cozeva_logo_image, background="white")
    header_label = Label(root, text="ALL Level Validation", background="white", font=("Helvetica", 17, 'bold'))
    userlist_main_label = Label(root, text="Users in Sequence", background="white", font=("Helvetica", 13, 'bold'))
    test_config_main_label = Label(root, text="Test Coverage", background="white", font=("Helvetica", 13, 'bold'))
    save_button = ttk.Button(root, text="Save Users", style="Next.TButton", command=on_save_users)
    start_button = ttk.Button(root, text="Start", style="Next.TButton", command=on_start_button)

    logo_label.grid(row=1, column=0, padx=50, columnspan=4)
    header_label.grid(row=2, column=0, columnspan=4)

    userlist_main_label.grid(row=3, column=0, columnspan=2)

    user1_inputframe.grid(row=4, column=0, columnspan=2, padx=20, pady=5)
    user2_inputframe.grid(row=5, column=0, columnspan=2, padx=20, pady=5)
    user3_inputframe.grid(row=6, column=0, columnspan=2, padx=20, pady=5)
    user4_inputframe.grid(row=7, column=0, columnspan=2, padx=20, pady=5)
    user5_inputframe.grid(row=8, column=0, columnspan=2, padx=20, pady=5)

    save_button.grid(row=10, column=0, columnspan=2)

    allviewusers = []
    with open("assets/allviewusers.pkl", 'rb') as allviewusers_file:
        allviewusers = pickle.load(allviewusers_file)

    for index, username in enumerate(allviewusers):
        username_entryboxes[index].insert(0, username)

    test_config_main_label.grid(row=3, column=2, columnspan=2)

    feature_checkbox_select_all.grid(row=4, column=2, columnspan=2, sticky='w', padx=20, pady=2)
    feature_checkbox_select_all.state(['selected'])

    all_level_menubar_checkbox.grid(row=5, column=2, columnspan=2, sticky='w', padx=20, pady=5)
    all_level_menubar_checkbox.state(['selected'])
    feature_variables[1].set(1)
    create_tooltip(all_level_menubar_checkbox, all_level_menubar_checkbox_tooltip)

    apptray_access_check_checkbox.grid(row=6, column=2, columnspan=2, sticky='w', padx=20, pady=5)
    apptray_access_check_checkbox.state(['selected'])
    feature_variables[2].set(1)
    create_tooltip(apptray_access_check_checkbox, apptray_access_check_checkbox_tooltip)

    support_mspl_glob_search_checkbox.grid(row=7, column=2, columnspan=2, sticky='w', padx=20, pady=5)
    support_mspl_glob_search_checkbox.state(['selected'])
    feature_variables[3].set(1)
    create_tooltip(support_mspl_glob_search_checkbox, support_mspl_glob_search_checkbox_tooltip)

    session_info_checkbox.grid(row=8, column=2, columnspan=2, sticky='w', padx=20, pady=5)
    session_info_checkbox.state(['selected'])
    feature_variables[4].set(1)
    create_tooltip(session_info_checkbox, session_info_checkbox_tooltip)

    chart_list_specifics_checkbox.grid(row=9, column=2, columnspan=2, sticky='w', padx=20, pady=5)
    chart_list_specifics_checkbox.state(['selected'])
    feature_variables[5].set(1)
    create_tooltip(chart_list_specifics_checkbox, chart_list_specifics_checkbox_tooltip)

    registry_summary_checkbox.grid(row=10, column=2, columnspan=2, sticky='w', padx=20, pady=5)
    registry_summary_checkbox.state(['selected'])
    feature_variables[6].set(1)
    create_tooltip(registry_summary_checkbox, registry_summary_checkbox_tooltip)

    rollup_summation_checkbox.grid(row=11, column=2, columnspan=2, sticky='w', padx=20, pady=5)
    rollup_summation_checkbox.state(['selected'])
    feature_variables[7].set(1)
    create_tooltip(rollup_summation_checkbox, rollup_summation_checkbox_tooltip)




    user_selection_input_box.grid(row=12, column=2, columnspan=2, sticky='n')
    user_selection_input_box.insert(0, "1,2,3,4,5")
    create_tooltip(user_selection_input_box, user_selection_input_box_tooltip)


    start_button.grid(row=12, column=3, columnspan=2, sticky='w')

    root.title("ALL Level Validation")
    root.iconbitmap("assets/icon.ico")
    root.mainloop()

def user_wise_test_config(current_user):

    feature_list = run_config[1][1:]
    print(feature_list)
    setups.login_to_user(current_user)
    print("Landing page= " + driver.title)
    sf.URLAccessCheck(driver.current_url, driver)
    sf.ajax_preloader_wait(driver)
    sf.skip_intro(driver)
    print("Masq to user successful")
    check_context_switch_to_all(driver)


    if feature_list[0] == 1:
        navigate_all_menubar(driver)
    if feature_list[1] == 1:
        apptray_access_checks(driver)
    if feature_list[2] == 1:
        check_support_level_mspl(driver)
    if feature_list[3] == 1:
        check_current_session(driver)
    if feature_list[4] == 1:
        check_document_preview(driver)
    if feature_list[5] == 1:
        check_REL_chart(driver)
    if feature_list[6] == 1:
        rollup_summation_checks(driver, current_user)




ENV = "PROD"
run_config = []
client_id = '1500'
global_test_case_id = 1

report_folder = os.path.join(locator.parent_dir,"All Level Reports")
isdir = os.path.isdir(report_folder)
if not isdir:
    os.mkdir(report_folder)

wb = Workbook()
workbook_title = "All Level Report_"+sf.date_time()+"_"+ENV+".xlsx"
wb.save(report_folder+ "\\"+workbook_title)


mainui()
driver = setups.driver_setup()

if ENV == 'CERT':
    setups.login_to_cozeva_cert(client_id)
elif ENV == 'STAGE':
    setups.login_to_cozeva_stage()
elif ENV == "PROD":
    setups.login_to_cozeva(client_id)
else:
    print("ENV INVALID")
    exit(3)

user_config = run_config[0]
for current_user in user_config:
    wb.create_sheet(current_user)
    ws = wb[current_user]
    ws.append(["Test Case ID", "Context Name", "Scenario", "Status","Time Taken", "Comments", "URL"])
    user_wise_test_config(current_user)
    rows = ws.max_row
    cols = ws.max_column
    for i in range(2, rows + 1):
        for j in range(3, cols + 1):
            if ws.cell(i, j).value == 'Passed':
                ws.cell(i, j).fill = PatternFill('solid', fgColor='0FC404')
            elif ws.cell(i, j).value == 'Failed':
                ws.cell(i, j).fill = PatternFill('solid', fgColor='FC0E03')
            elif ws.cell(i, j).value == 'Showing 0 to 0':
                ws.cell(i, j).fill = PatternFill('solid', fgColor='FCC0BB')
    wb.save(report_folder + "\\" + workbook_title)
    setups.switch_back()





