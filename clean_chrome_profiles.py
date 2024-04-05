import openpyxl

file_location = "assets/chrome_profile_info.xlsx"
chrome_profiles = openpyxl.load_workbook(file_location)
chrome_profiles_sheet = chrome_profiles.active
# Look for a row with an Available Chromeprofile name, Change it to In use and return the name

for row in range(1, 11):  # Assuming you want to update cells C1 to C10
    cell = chrome_profiles_sheet.cell(row=row, column=3)  # Column C is represented by index 3
    cell.value = 'Available'

# Save the modified Excel file
chrome_profiles.save('assets/chrome_profile_info.xlsx')

file_location = "assets/edge_profile_info.xlsx"
edge_profiles = openpyxl.load_workbook(file_location)
edge_profiles_sheet = edge_profiles.active
# Look for a row with an Available Chromeprofile name, Change it to In use and return the name

for row in range(1, 11):  # Assuming you want to update cells C1 to C10
    cell = edge_profiles_sheet.cell(row=row, column=3)  # Column C is represented by index 3
    cell.value = 'Available'

# Save the modified Excel file
edge_profiles.save('assets/edge_profile_info.xlsx')
