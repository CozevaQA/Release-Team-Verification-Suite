import time
import traceback

from openpyxl.styles import Font, PatternFill
from selenium.webdriver.common.by import By
from selenium.webdriver.support.wait import WebDriverWait

import setups
import logging
import ExcelProcessor as db
import context_functions as cf
import support_functions as sf
import setups as st
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
import variablestorage as locator
from openpyxl import Workbook, load_workbook
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import NoSuchElementException, ElementNotInteractableException, \
    ElementClickInterceptedException, TimeoutException

ENV = 'STAGE'


def performGlobalSearch(role, username, keywords, driver, testID):
    print(username)

    for keyword in keywords:
        try:
            registry_url=driver.current_url
            window_switched = 0
            driver.find_element_by_id('globalsearch_input').send_keys(keyword)
            start_time = time.perf_counter()
            WebDriverWait(driver, 90).until(
                EC.presence_of_element_located((By.CLASS_NAME, 'collection-header')))
            driver.find_element_by_id('globalsearch_input').send_keys(Keys.RETURN)
            # time.sleep(0.25)
            # driver.find_element_by_id('globalsearch_input').send_keys(Keys.RETURN)
            sf.ajax_preloader_wait(driver)
            WebDriverWait(driver, 45).until(EC.presence_of_element_located((By.ID, 'search_all')))
            time_taken = round(time.perf_counter() - start_time - 2, 2)
            search_text = driver.find_element_by_id("search_all").text
            if 'No results' in search_text:
                ws1.append((testID, role, username, keyword, time_taken, "No results"))
            else:
                ws1.append((testID, role, username, keyword, time_taken))

        except (NoSuchElementException, ElementNotInteractableException, ElementClickInterceptedException, TimeoutException) as e:
            traceback.print_exc()
            ws1.append((testID, role, username, keyword, 'x', "Exception occured: "+str(e)))

        testID += 1
        driver.get(registry_url)
        sf.ajax_preloader_wait(driver)
        WebDriverWait(driver, 45).until(EC.presence_of_element_located((By.XPATH, locator.xpath_filter_measure_list)))



if __name__ == '__main__':
    print("Hello World")
    file = str(sf.date_time())+"_GlobalSearch.xlsx"
    driver = setups.driver_setup()
    if ENV == 'CERT':
        setups.login_to_cozeva_cert()
    elif ENV == 'STAGE':
        setups.login_to_cozeva_stage()
    elif ENV == "PROD":
        setups.login_to_cozeva()
    else:
        print("ENV INVALID")
        exit(3)

    sf.ajax_preloader_wait(driver)


    #excel_path = ""

    wb = load_workbook("assets\\GlobalSearch.xlsx")
    ws = wb[ENV]
    sheet_rows = ws.rows
    users = []
    keywords_list = []
    roles = []
    for cellval in sheet_rows:
        users.append(cellval[0].value.strip())
        keywords_list.append(cellval[1].value.strip().split(';'))
        roles.append(cellval[2].value.strip())

    # for user in users:
    #     print(user)
    #
    # for keyword in keywords_list:
    #     print(keyword)

    wb1 = Workbook()
    ws1 = wb1.active
    ws1.title = 'Global Search ' + ENV

    ws1.append(['ID', 'ROLE', 'USERNAME', 'KEYWORD', 'TIME TAKEN', 'COMMENTS'])
    header_font = Font(color='FFFFFF', bold=False, size=12)
    header_cell_color = PatternFill('solid', fgColor='030303')
    ws1['A1'].font = header_font
    ws1['A1'].fill = header_cell_color
    ws1['B1'].font = header_font
    ws1['B1'].fill = header_cell_color
    ws1['C1'].font = header_font
    ws1['C1'].fill = header_cell_color
    ws1['D1'].font = header_font
    ws1['D1'].fill = header_cell_color
    ws1['E1'].font = header_font
    ws1['E1'].fill = header_cell_color
    ws1['F1'].font = header_font
    ws1['F1'].fill = header_cell_color
    ws1.name = "Arial"
    test_case_id = 1

    for user, keywords, role in zip(users, keywords_list, roles):
        setups.login_to_user(user)
        setups.switch_to_registries()
        performGlobalSearch(role, user, keywords, driver, test_case_id)
        wb1.save(locator.parent_dir+file)
        setups.switch_back()

    rows = ws1.max_row
    cols = ws1.max_column
    for i in range(2, rows + 1):
        for j in range(5, cols + 1):
            try:
                dum = int(ws1.cell(i, j).value)
            except Exception as e:
                continue
            if int(ws1.cell(i, j).value) < 8:
                ws1.cell(i, j).fill = PatternFill('solid', fgColor='0FC404')
            elif int(ws1.cell(i, j).value) > 20:
                ws1.cell(i, j).fill = PatternFill('solid', fgColor='FC0E03')
            elif int(ws1.cell(i, j).value) >= 8:
                ws1.cell(i, j).fill = PatternFill('solid', fgColor='FCC0BB')
    wb1.save(locator.parent_dir + file)












