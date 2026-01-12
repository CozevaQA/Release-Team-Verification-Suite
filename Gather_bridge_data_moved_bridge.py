# Excel File addition - Addition and CSV addition
# HCC report check
# Role specific check

import os
import random
from statistics import mean
import sys
import math
import time
import traceback
import csv
from os import listdir
from os.path import isfile, join
import csv
import datetime
import base64
import pandas as pd
from selenium import *
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import *
from selenium.webdriver.support.wait import WebDriverWait
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from urllib.parse import urlparse, parse_qs

import self as self
from openpyxl.formatting import Rule
from openpyxl.styles import *
from openpyxl.styles.differential import DifferentialStyle
from selenium.webdriver.common.keys import Keys


# def WorkBook(Name,WorkBook,SheetName,WorkBookFlag,NewSheetFlag):
#     if NewSheetFlag == 1 and WorkBookFlag == 1:
#         ws = WorkBook.create_sheet(SheetName)
#         ws = WorkBook[Name]
#     else:
#         WorkBook = Workbook()
#         ws = WorkBook.active
#         ws.title = SheetName
#     ws.append(['LOB', 'HCC Measure', 'Patient Non Compliant count(UI)', 'Patient Total count(UI)', 'Gaps', 'Conditions',
#                'Disconfirms', 'Non Compliant Count(Export)', 'Total Count(Export)', 'Status', 'Comments'])
#     ws.auto_filter.ref = "A1:0100"
#     header = NamedStyle(name="header")
#     header.font = Font(bold=True)
#     header.border = Border(bottom=Side(border_style="thin"))
#     header.alignment = Alignment(horizontal="center", vertical="center")
#     header_row = ws[1]
#     for cell in header_row:
#         cell.style = header
#     red_background = PatternFill(fgColor="00FF0000")
#     diff_style = DifferentialStyle(fill=red_background)
#     rule = Rule(type="expression", dxf=diff_style)
#     rule.formula = ["$J1==Fail"]
#     ws.conditional_formatting.add("A1:O100", rule)
#     filename = "C:\\Users\\sbasu\\Documents\\Report\\"+ Name
#     return filename

def ajax_preloader_wait(driver):
    time.sleep(1)
    WebDriverWait(driver, 300).until(EC.invisibility_of_element((By.CLASS_NAME, "ajax_preloader")))
    if len(driver.find_elements(By.CLASS_NAME, "ajax_preloader")) != 0:
        WebDriverWait(driver, 300).until(EC.invisibility_of_element((By.CLASS_NAME, "ajax_preloader")))
    WebDriverWait(driver, 300).until(EC.invisibility_of_element((By.CLASS_NAME, "drupal_message_text")))
    time.sleep(1)

def BridgeLinkNavigator(cust_id, link):
    HP_customers = ['3000', '3700', '3600', '1950']
    link1 = link + "&table_id=overlapping_member_list"
    driver.get(link1)
    ajax_preloader_wait(driver)
    WebDriverWait(driver, 20000).until(EC.presence_of_element_located((By.XPATH, "//*[@id='overlapping_member_list']")))
    customer_name = driver.find_element(By.CLASS_NAME, "specific_most").text
    members = driver.find_element(By.XPATH, "//*[@id='overlapping_member_count']").text
    members = members.replace(",", "")
    print(members)
    tenancies = driver.find_element(By.XPATH, "//*[@id='tenancies_count']").text
    tenancies = tenancies.replace(",", "")
    print(tenancies)
    ws = wb["Overlapping Members"]
    if not members:
        members = 0
    if not tenancies:
        tenancies = 0
    ws.append([customer_name, members, tenancies, week_format])
    link2 = link + "&table_id=gap_comparison_list"
    driver.get(link2)
    ajax_preloader_wait(driver)
    WebDriverWait(driver, 20000).until(EC.presence_of_element_located((By.XPATH, "//*[@id='gap_comparison_list']")))
    OthersClose = driver.find_element(By.XPATH, "//*[@id='attempted']").text
    OthersClose = OthersClose.replace(",", "")
    print(OthersClose)
    YourClose = driver.find_element(By.XPATH, "//*[@id='closed']").text
    YourClose = YourClose.replace(",", "")
    print(YourClose)
    ws = wb["Gap Comparison"]
    if not OthersClose:
        OthersClose = 0
    if not YourClose:
        YourClose = 0
    ws.append([customer_name, OthersClose, YourClose, week_format])
    ajax_preloader_wait(driver)
    if cust_id in HP_customers:
        link3 = link + "&table_id=attempted_gap_closure_list"
        driver.get(link3)
        ajax_preloader_wait(driver)
        WebDriverWait(driver, 20000).until(EC.presence_of_element_located((By.XPATH, "//*[@id='attempted_gap_closure_list']")))
        TotalGaps = driver.find_element(By.XPATH, "//*[@class='total_gap_count']").text
        TotalGaps = TotalGaps.replace(",", "")
        print(TotalGaps)
        current_count = driver.find_element(By.XPATH, "//*[@class='current_decision_count']").text
        current_count = current_count.replace(",", "")
        print(current_count)
        pending = int(TotalGaps) - int(current_count)
        closed = driver.find_element(By.XPATH, "//*[@class='closed_status_value']").text
        closed = closed.replace(",", "")
        print(closed)
        ClosedMostRecent = driver.find_element(By.XPATH, "//*[@class='closed_most_recent_status_value']").text
        ClosedMostRecent = ClosedMostRecent.replace(",", "")
        print(ClosedMostRecent)
        NotClosedMostRecent = driver.find_element(By.XPATH, "//*[@class='not_closed_most_recent_status_value']").text
        NotClosedMostRecent = NotClosedMostRecent.replace(",", "")
        print(NotClosedMostRecent)
        NotClosed = driver.find_element(By.XPATH, "//*[@class='not_closed_status_value']").text
        NotClosed = NotClosed.replace(",", "")
        print(NotClosed)
        ws = wb["Attempted Gap Closure"]
        if not TotalGaps:
            TotalGaps = 0
        if not closed:
            closed = 0
        if not ClosedMostRecent:
            ClosedMostRecent = 0
        if not NotClosedMostRecent:
            NotClosedMostRecent = 0
        if not NotClosed:
            NotClosed = 0
        ws.append([customer_name, TotalGaps, pending, closed, ClosedMostRecent, NotClosedMostRecent, NotClosed, week_format])
    wb.save(path1 + filename)

header = NamedStyle(name="header")
header.font = Font(bold=True)
header.border = Border(bottom=Side(border_style="thin"))
header.alignment = Alignment(horizontal="center", vertical="center")
red_background = PatternFill(patternType='solid', fgColor="00FF0000")
green_background = PatternFill(patternType='solid', fgColor="50C878")
gray_background = PatternFill(patternType='solid', fgColor="5F9EA0")

path1 = "C:\\Users\\wdey\\Documents\\Report\\"
name_date = datetime.datetime.now()
formatted_date = name_date.strftime("%y-%m-%d %H-%M")
sheet_week = name_date.strftime("%m/%d")
week_format = "Week of " + sheet_week
filename = "Bridge Weekly Data Backup_" + str(formatted_date) + ".xlsx"
wb = Workbook()
ws = wb.active
sheet_name = "Overlapping Members"
ws.title = sheet_name
ws.append(["Customer Name", "Overlapping Members", "Tenancies", "Week"])
header_row = ws[1]
for cell in header_row:
    cell.style = header
wb.create_sheet("Gap Comparison")
ws = wb["Gap Comparison"]
ws.append(["Customer Name", "Gaps you can close", "Gaps others can close", "Week"])
header_row = ws[1]
for cell in header_row:
    cell.style = header
wb.create_sheet("Attempted Gap Closure")
ws = wb["Attempted Gap Closure"]
ws.append(["Customer Name", "Total Gaps Count", "Pending Gaps", "Closed Gaps", "Closed Most Recent Gaps", "Not Closed Most Recent Gaps", "Not Closed Gaps", "Week"])
header_row = ws[1]
for cell in header_row:
    cell.style = header
# diff_style1 = DifferentialStyle(fill=red_background)
# rule1 = Rule(type="text", dxf=diff_style1, text="Fail")
# diff_style2 = DifferentialStyle(fill=green_background)
# rule2 = Rule(type="text", dxf=diff_style2, text="Pass")
# diff_style3 = DifferentialStyle(fill=gray_background)
# rule3 = Rule(type="text", dxf=diff_style3, text="Unexecuted")
# foldername = "Masquerade Check_" + str(datetime.datetime.now()).replace(':', '-') + "\\"
# path1 = "C:\\Users\\sbasu\\Documents\\Report\\"
# path2 = os.path.join(path1, foldername)
# os.makedirs(path2)
#url, info
logout_url = "https://www.cozeva.com/user/logout"
login_url = "https://www.cozeva.com/user/login"
base_url = "https://www.cozeva.com/"
# User = os.environ.get('CS2_User')
# Pass = os.environ.get('CS2_Password')
User = "wdey.cs"
Pass = "Celkon@85"
Customer_list = ['3700','3600', '1950', '4150', '4350', '2400', '1750', '4950', '2750', '1500', '3750', '2600', '2200', '3250', '6800', '1600', '6700', '7100', '4700', '7500', '2050', '3100', '1700', '2900', '4300', '3500', '200', '4250', '2350', '2250', '2650', '4600', '150', '5700', '2700', '2000', '2100', '3850', '1100', '3300', '4500', '5100', '7400', '2800', '4450', '1550', '6600', '5200', '4800', '5300', '4900', '2950', '4100', '3950', '5000', '3550', '3650', '2450', '3450', '3350', '1650', '3400', '1200', '3900', '2550', '3800', '1300', '1850', '6000', '5400', '3150', '1000', '1900', '7000', '4000', '4400', '5900', '7200']


#list = ['3000', '3600', '1950', '4350', '2400', '1750', '2750', '1500', '3750', '2600', '2200', '3250', '6800', '1600', '6700', '7100', '4700', '7500', '2050', '3100', '1700', '2900', '4300', '3500', '200', '4250', '2350', '2250', '2650', '4600', '150', '5700', '2700', '2000', '2100', '5600', '3850', '1100', '3300', '4500', '5100', '7400', '2800', '4450', '1550', '6600', '5200', '4800', '5300', '4900', '2950', '4100', '3950', '5000', '3550', '3650', '2450', '3450', '3350', '1650', '3400', '1200', '3900', '2550', '3800', '1300', '1850', '6000', '5400', '3150', '1000', '1900', '7000', '4000', '4400', '5900', '7200']
#Anthem - 3700(Retired from 11/13) #ckc - 3200, family choice- 1800, stanford- 5800 [Discontinued from 12/19] 2500 6500 - discontinued from 01/02
#chrome-setup
# options = webdriver.ChromeOptions()
# prefs = {"download.default_directory" : "C:\\Users\\sbasu\\Documents\\CSV_Files"}
# options.add_argument("user-data-dir=C:\\Users\\sbasu\\AppData\\Local\\Google\\Chrome\\User Data\\SavedData")
# options.add_argument("--disable-notifications")
# options.add_argument("--disk-cache-size=1")
# options.add_argument("--disable-extensions")
# options.add_argument("--disable-gpu")
# options.add_experimental_option("prefs", prefs)
# driver = webdriver.Chrome(executable_path="C:\\Users\\sbasu\\Documents\\Drivers\\chromedriver.exe", options=options)

options = webdriver.ChromeOptions()
prefs = {"download.default_directory" : "C:\\Users\\wdey\\Documents\\CSV_Files"}
options.add_argument("user-data-dir=C:\\Users\\wdey\\AppData\\Local\\Google\\Chrome\\User Data\\SavedData")
options.add_argument("--disable-notifications")
options.add_argument("--disk-cache-size=1")
options.add_argument("--disable-extensions")
options.add_argument("--disable-gpu")
options.add_experimental_option("prefs", prefs)
driver = webdriver.Chrome(executable_path="assets\\chromedriver.exe", options=options)

driver.get(logout_url)
driver.get(login_url)
driver.maximize_window()
print("Initial window= " + driver.title)
driver.find_element(By.ID, "edit-name").send_keys(User)
driver.find_element(By.ID, "edit-pass").send_keys(Pass)
driver.find_element(By.ID, "edit-submit").click()
time.sleep(1)

try:
    WebDriverWait(driver, 120).until(EC.presence_of_element_located((By.ID, "reason_textbox")))
    driver.find_element(By.ID, "reason_textbox").send_keys("https://redmine2.cozeva.com/issues/24376")
    driver.find_element(By.ID, "edit-submit").click()
except NoSuchElementException:
    traceback.print_exc()
    driver.quit()
print("Landing page= " + driver.title)
ajax_preloader_wait(driver)
if driver.title != "Cozeva Bridge | Cozeva":
    driver.find_element(By.XPATH, "//*[@data-target='app_dropdown']").click()
    driver.find_element(By.XPATH, "//*[@class='no-hover app_cozeva_bridge']").click()
    driver.switch_to.window(driver.window_handles[1])
ajax_preloader_wait(driver)

Bridge_URL = driver.current_url
parsed_url = urlparse(Bridge_URL)
query_parameters = parse_qs(parsed_url.query)

session_id = query_parameters.get("session", [None])[0]
print(session_id)
decoded_val = base64.b64decode(session_id).decode('utf-8')
print(decoded_val)


for cust_id in Customer_list:
    temp = decoded_val.replace('1500',str(cust_id))
    print(temp)
    session_id_temp = base64.b64encode(temp.encode('utf-8'))
    temp1 = str(session_id_temp.decode('utf-8'))
    print(temp1)
    temp_link = base_url + "cozeva_bridge/users-bridge?session=" + temp1
    try:
        BridgeLinkNavigator(cust_id, temp_link)
    except TimeoutException as e:
        traceback.print_exc(e)
        continue


driver.quit()