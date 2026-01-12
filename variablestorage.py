import pickle

#xpaths
import openpyxl
from datetime import date, datetime, time

xpath_filter_measure_list = "//a[@data-target='qt-reg-nav-filters']"
xpath_switch_back = '//*[@id="nav"]/div/ul/li[2]/a'
xpath_switch_back2 = '//*[@id="quick_switch_links"]/div/ul/li/a'
xpath_side_nav_SlideOut="//a[@data-target='sidenav_slide_out']"
xpath_context_Name="//a[@id='context_trigger']/div/span[@class='specific_most']"
#xpath_menubar_Item_Link="//a[contains(@class,'nav_link')]"
xpath_menubar_Item_Link="//li[contains(@class, 'sidebar-menu-item')]/a[contains(@class, 'nav_link')]"
#xpath_menubar_Item_Link_Name="//a[contains(@class,'nav_link')]/span"
xpath_menubar_Item_Link_Name="//li[contains(@class,'sidebar-menu-item')]/a[contains(@class, 'nav_link')]/span"
xpath_data_Table_Info="//div[@class='dataTables_info']"
xpath_registry_Link="//a[@id='registry-link']"
xpath_toast_Message="//div[@class='drupal_message_text']"
xpath_providers_Tab="//a[@id='providers-list']"
xpath_link_To_Provider_Practice_Dashboard="//td/a[contains(@href,'registries')]"
xpath_measure_In_Registry="//div/span[@class='met-name']"
xpath_practice_Tab="(//li[@class='tab'])[1]"
xpath_providers_Patients_Tab_Link="//a[@id='all_patients_tab']"
xpath_cozeva_Id="(//div/span[@data-tooltip='Cozeva Id (Click to Copy)'])[1]"
xpath_provider_Practice_Link_In_Global_Search="(//div[contains(@class,'col s12 l6')]/a[contains(@href,'/registries?')])[2]"
xpath_practice_Link_In_Global_Search="//ul[@data-type='Practices']/li/div/div[contains(@class,'col s12 l6')]/a[contains(@href,'/registries?')]"
xpath_provider_Link_In_Global_Search="//ul[@data-type='Providers']/li/div/div[contains(@class,'col s12 l6')]/a[contains(@href,'/registries?')]"
xpath_app_Tray_Link="//a[@data-target='app_dropdown']"
xpath_app_Time_Capsule="//a[@title='Time Capsule']"
xpath_latest_Card_Title="((//div[@class='card-title'])[1])/div[1]"
xpath_app_Analytics="//a[@title='Analytics']"
xpath_total_Workbooks="(//tr[@class='sm_list_row']/td[@class='sm_tab_link'][1])"
xpath_total_Drilldowns="(//"
xpath_app_Secure_Messaging="//a[@title='Secure Messaging']"
xpath_inbox_Message="//tbody/tr[@role='row']"
xpath_patient_Header_Dropdown_Arrow="//a[@data-target='patient_header_dropdown_compact']/div[@id='patient_header_compact']"
xpath_patient_History="//li[@id='history']"
xpath_patient_History_Item_Link="//li[@id='history']/ul[@class='patient_submenu']/li/a"
xpath_patient_Info_Link="//ul[@id='patient_header_dropdown_compact']/li/a[contains(text(),'Patient Information')]"
xpath_patient_Info_Coverage_Link="//ul[@id='patient_info_contents']/li/span[contains(text(),'Coverage')]"
xpath_patient_Info_Care_Team_Link="//ul[@id='patient_info_contents']/li/span[contains(text(),'Care Team')]"
xpath_data_Table_Row="//tbody/tr"
xpath_empty_Data_Table_Row="//td[@class='dataTables_empty']"
xpath_registry_Filter="//a[@data-target='qt-reg-nav-filters']"
xpath_unhide_Submeasures_Checkbox="//ul[@id='qt-reg-nav-filters']/li/label[@class='col s12']"
xpath_registry_Submeasures_Filter_Apply_Button="//button[@id='qt-apply-search']"
xpath_registry_Measures="//div[@class='qt-metric']/a"
xpath_patient_Toggle="(//ul[@id='qt-mt-support-ls']/li[@class='tab']/a)[contains(text(),'Patients')]"
xpath_patients_Toggle_Patient_Links="//td[@class=' msg_pat_name']/div/a[contains(@href,'/patient_detail/')]"
xpath_provider_Toggle_Provider_Links="//table[@id='metric-support-prov-ls']/tbody/tr/td/a[contains(@href,'/registries?')]"
xpath_resources_link="//a[@data-target='help_menu_dropdown']"
xpath_resources_page_prac_sup="//span[@data-badge-caption='Practice Support Page']"
xpath_had_er_visit="//li[@id='had_er_visit_tab']"
xpath_skip_button=".//div[@id='agreement-buttons']//button[@value='Skip']"
xpath_submit_button = "//*[@data-badge-caption='Submit' and contains(@style,'float: left !important; padding: 0 10px !important;')]"
#xpath_careops = "//*[@class='col right_border child firstelem']//span[1]"
#xpath_careops2 = "//*[@class='col right_border child firstelem tooltipped']//span[1]"

xpath_careops = "//div[contains(@class, 'col') and contains(@class, 'right_border') and contains(@class, 'child') and (contains(@class, 'firstelem') or contains(@class, 'secondelem'))]"
xpath_coding_tool_kebab = "//a[@class='action_list_dropdown not_disable']"
xpath_audit_log_download = "//a[@onclick='return download_audit_log();']"
xpath_annotation_tab = "//a[@class='chart_notes_tab tab_on_demand']"
xpath_medication_chart_icon = "//div[@class='med_adherence_chart medical_adherence_contain relative_elem']"
xpath_rel_chart = "//div[@class='modal-content rel-chart-modal-modal-content']"

#css_selectors

#ids

def fetch_free_profile():
    # with open('assets/driver_choice.pkl', "rb") as file:
    #     driver_choice = pickle.load(file)
    #     file.close()
    # print("driver choice in variable storage: "+driver_choice)
    with open("assets/driver_choice.txt", 'r+') as driver_choice_file:
        driver_choice = driver_choice_file.read().strip()
    driver_choice_file.close()
    print("In Variable Storage with :" + driver_choice)
    if driver_choice == "CHROME":
        file_location = "assets/chrome_profile_info.xlsx"
        chrome_profiles = openpyxl.load_workbook(file_location)
        chrome_profiles_sheet = chrome_profiles.active
        chrome_profile_available = False
        chrome_profile = None
        #Look for a row with an Available Chromeprofile name, Change it to In use and return the name
        for profile_index in range(1, 11):
            days_since = (date.today() - chrome_profiles_sheet.cell(row=profile_index,column=4).value.date()).days
            #print(chrome_profiles_sheet.cell(row=profile_index, column=2).value, days_since)
            if str(chrome_profiles_sheet.cell(row=profile_index,column=3).value).strip() == "Available" and days_since <= 29:
                chrome_profiles_sheet.cell(row=profile_index, column=3).value = 'In Use'
                chrome_profile_available = True
                chrome_profile = str(chrome_profiles_sheet.cell(row=profile_index, column=2).value).strip()
                chrome_profiles.save("assets/chrome_profile_info.xlsx")
                break
            else:
                chrome_profile_available = False
                chrome_profile = None
        return chrome_profile
    elif driver_choice == "EDGE":
        file_location = "assets/edge_profile_info.xlsx"
        edge_profiles = openpyxl.load_workbook(file_location)
        edge_profiles_sheet = edge_profiles.active
        edge_profile_available = False
        edge_profile = None
        # Look for a row with an Available Chromeprofile name, Change it to In use and return the name
        for profile_index in range(1, 11):
            days_since = (date.today() - edge_profiles_sheet.cell(row=profile_index, column=4).value.date()).days
            # print(chrome_profiles_sheet.cell(row=profile_index, column=2).value, days_since)
            if str(edge_profiles_sheet.cell(row=profile_index,
                                            column=3).value).strip() == "Available" and days_since <= 29:
                edge_profiles_sheet.cell(row=profile_index, column=3).value = 'In Use'
                edge_profile_available = True
                edge_profile = str(edge_profiles_sheet.cell(row=profile_index, column=2).value).strip()
                edge_profiles.save("assets/edge_profile_info.xlsx")
                break
            else:
                edge_profile_available = False
                edge_profile = None


        return edge_profile


def fetch_free_edge_profile():
    file_location = "assets/edge_profile_info.xlsx"
    edge_profiles = openpyxl.load_workbook(file_location)
    edge_profiles_sheet = edge_profiles.active
    edge_profile_available = False
    edge_profile = None
    # Look for a row with an Available Chromeprofile name, Change it to In use and return the name
    for profile_index in range(1, 11):
        days_since = (date.today() - edge_profiles_sheet.cell(row=profile_index, column=4).value.date()).days
        # print(chrome_profiles_sheet.cell(row=profile_index, column=2).value, days_since)
        if str(edge_profiles_sheet.cell(row=profile_index,
                                        column=3).value).strip() == "Available" and days_since <= 29:
            edge_profiles_sheet.cell(row=profile_index, column=3).value = 'In Use'
            edge_profile_available = True
            edge_profile = str(edge_profiles_sheet.cell(row=profile_index, column=2).value).strip()
            edge_profiles.save("assets/edge_profile_info.xlsx")
            break
        else:
            edge_profile_available = False
            edge_profile = None

    return edge_profile


def fetch_free_chrome_profile():
    file_location = "assets/chrome_profile_info.xlsx"
    chrome_profiles = openpyxl.load_workbook(file_location)
    chrome_profiles_sheet = chrome_profiles.active
    chrome_profile_available = False
    chrome_profile = None
    # Look for a row with an Available Chromeprofile name, Change it to In use and return the name
    for profile_index in range(1, 11):
        days_since = (date.today() - chrome_profiles_sheet.cell(row=profile_index, column=4).value.date()).days
        # print(chrome_profiles_sheet.cell(row=profile_index, column=2).value, days_since)
        if str(chrome_profiles_sheet.cell(row=profile_index,
                                          column=3).value).strip() == "Available" and days_since <= 29:
            chrome_profiles_sheet.cell(row=profile_index, column=3).value = 'In Use'
            chrome_profile_available = True
            chrome_profile = str(chrome_profiles_sheet.cell(row=profile_index, column=2).value).strip()
            chrome_profiles.save("assets/chrome_profile_info.xlsx")
            break
        else:
            chrome_profile_available = False
            chrome_profile = None
    return chrome_profile


free_chrome_profile = fetch_free_chrome_profile()
if free_chrome_profile is None:
    #print("All ChromeProfiles in use.")
    free_chrome_profile = str(free_chrome_profile)
    # exit(4)

free_edge_profile = fetch_free_edge_profile()
if free_edge_profile is None:
    #print("All EdgeProfiles in use.")
    free_edge_profile = str(free_edge_profile)
    # exit(4)


file = open(r"assets\loginInfo.txt", "r+")
pc_username = file.readlines()[2].strip()
chrome_profile_path = "user-data-dir=C:\\Users\\"+pc_username+"\\AppData\\Local\\Google\\Chrome\\User Data\\"+free_chrome_profile
edge_profile_path = "user-data-dir=C:\\Users\\"+pc_username+"\\AppData\\Local\\Microsoft\\Edge\\User Data\\"+free_edge_profile
file.seek(0)
file.close()
#print(chrome_profile_path)
#print(edge_profile_path)
#chrome_driver_path = "C:\\cdriver\\chromedriver.exe"
edge_driver_path = "assets\\msedgedriver.exe"
chrome_driver_path = "assets\\chromedriver.exe"
login_link = "https://www.cozeva.com/user/login"
logout_link = "https://www.cozeva.com/user/logout"
login_link_cert = "https://cert.cozeva.com/user/login"
logout_link_cert = "https://cert.cozeva.com/user/logout"
login_link_stage = "https://stage.cozeva.com/user/login"
logout_link_stage = "https://stage.cozeva.com/user/logout"
login_link_amp = "https://amp.cozeva.com/user/logout"
logout_link_amp = "https://amp.cozeva.com/user/logout"
parent_dir = "C:\\VerificationReports\\"
download_dir = "C:\\VerificationReports\\DownloadDirectory\\"

#accordian validation stuff

select_measurement_year_flag_support = "False"
select_measurement_year_flag_provider = "False"
MeasurementYear_Support = "2024"
MeasurementYear_Provider = "2024"
MeasurementYearQuarter_Support = "2024 Q4"
MeasurementYearQuarter_Provider = "2024 Q4"