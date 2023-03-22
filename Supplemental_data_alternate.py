import configparser
import datetime
import os
import time
import traceback

from openpyxl.styles import Font, PatternFill
from selenium.webdriver.common.by import By
from selenium.webdriver.support.wait import WebDriverWait
from tkinter import *
import setups
import logging
import ExcelProcessor as db
import context_functions as cf
import support_functions as sf
import setups as st
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
import variablestorage as locator
from openpyxl import Workbook, load_workbook
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import NoSuchElementException, ElementNotInteractableException, \
    ElementClickInterceptedException, TimeoutException
config = configparser.RawConfigParser()
config.read("locator-config.properties")

ENV = 'CERT'
client_id = str(1000)
MY = "2023"
wait_time = 60

''' 
    Checkbox value vs measure mapping
    0 = BCS
    1 = COL
    2 = CCS
    3 = HbA1c Poor control
    4 = BP Control
    5 = Placeholder
    6 = Placeholder
    7 = Placeholder 
    NA = Not Accordion
    A = Accordion 
'''


measure_map = [["BCS", "1", "NA"],
               ["COL", "3", "739"],
               ["CCS", "2", "NA"],
               ["HBD1 (in progress)"],
               ["HBD2 (in progress)"],
               ["CBP (in progress)"],
               ["Placeholder"],
               ["Placeholder"]]

final_measure_list = []



def fetch_measures():
    root = Tk()

    Checkbox_variables = [None] * 8
    for i in range(0, len(Checkbox_variables)):
        Checkbox_variables[i] = IntVar()


    def on_start():
        global client_id, MY, wait_time
        client_id = db.fetchCustomerID(selected_cust.get())
        MY = my_entry.get()
        wait_time = int(wait_time_entry.get())
        for i in range(0, len(Checkbox_variables)):
            if Checkbox_variables[i].get() == 1:
                final_measure_list.append(measure_map[i])
        print(final_measure_list)
        root.destroy()


    customer_label = Label(root, text="Select customer", font=("Nunito Sans", 10))
    selected_cust = StringVar()
    selected_cust.set("Customer")
    customer_list = db.getCustomerList()  # vs.customer_list
    customer_drop = OptionMenu(root, selected_cust, *customer_list)


    customer_label.grid(row=0, column=0)
    customer_drop.grid(row=1, column=0)
    Label(root, text="Measurement Year",  font=("Nunito Sans", 10)).grid(row=0, column=3)
    my_entry = Entry(root)
    my_entry.grid(row=1, column=3)
    my_entry.insert(0, "2023")

    Button(root, text="Start Test", command=on_start, font=("Nunito Sans", 10)).grid(row=6, column=3)
    wait_time_entry = Entry(root)
    wait_time_entry.grid(row=6, column=1)
    wait_time_entry.insert(0, "60")
    Label(root, text="Wait time (s)", font=("Nunito Sans", 10)).grid(row=6, column=0)
    for i, measure in enumerate(measure_map):
        if i<=3:
            Checkbutton(root, text=measure[0], variable=Checkbox_variables[i], font=("Nunito Sans", 10)).grid(row=i+2, column=0, sticky='w')
        if i > 3:
            Checkbutton(root, text=measure[0], variable=Checkbox_variables[i], font=("Nunito Sans", 10)).grid(row=i - 2,
                                                                                                              column=3,
                                                                                                              sticky='w')






    root.title("Supplemental Data Validation")
    root.iconbitmap("assets/icon.ico")
    # root.geometry("400x400")
    root.mainloop()

def add_supplemental_data(measure_tiny_text, value):
    try:
        task_id = "Couldn't Fetch Task ID"
        sf.ajax_preloader_wait(driver)
        task_id = driver.find_element_by_xpath("//div[@class='task_def left pts prm']").text

        attachment_element = driver.find_element_by_class_name("file_upload_attachment")
        attachment_path = os.getcwd()
        attachment_path = os.path.join(attachment_path, 'Assets\Doc_pdf.pdf')
        print(attachment_path)
        attachment_element.send_keys(attachment_path)

        time.sleep(2)

        if sf.URLAccessCheck(driver.current_url, driver):
            ws.append(
                [test_case_id, LOB_Name, measure[0] + "_" + measure[1],
                 provider_name, patient_id, "Attaching a File",
                 "Failed", task_id, "Error Toast message was displayed", driver.current_url])
            raise Exception("Attachment Failed")
        else:
            # Add Suppdata
            ws.append(
                [test_case_id, LOB_Name, measure[0] + "_" + measure[1],
                 provider_name, patient_id, "Attaching a File",
                 "Passed", task_id, "", driver.current_url])
            supdata = driver.find_element_by_class_name("saved_data")
            DoS = datetime.datetime.now()
            DoS = DoS.strftime("%m/%d/%Y")
            supdata.find_element_by_xpath(
                "//input[@class='materialize_datefilter datepicker no-future-dt form-control form-text required']"). \
                send_keys(DoS)
            time.sleep(2)

            code_box = supdata.find_element_by_class_name("codebox-wrapper")
            time.sleep(1)
            code_box.click()
            codes = driver.find_element_by_xpath(
                "//ul[@class='dropdown-content mat-ac-dropdown quality_ac_dropdown']").find_elements_by_tag_name('li')
            time.sleep(1)
            codes[0].click()
            time.sleep(1)

            if value is not None:
                driver.find_element_by_xpath("//input[@class='code-box-hidden-field value_box']").send_keys(str(value))
                time.sleep(1)

            task_url = driver.current_url
            sf.action_click(driver.find_element_by_xpath(locator.xpath_submit_button),
                            driver)
            if sf.URLAccessCheck(driver.current_url, driver):
                ws.append(
                    [test_case_id, LOB_Name, measure[0] + "_" + measure[1],
                     provider_name, patient_id, "Submit Supp Data",
                     "Failed", task_id, "Error Toast message was displayed",
                     driver.current_url])
                raise Exception("Suppdata Submit Failed")
            else:

                ws.append(
                    [test_case_id, LOB_Name, measure[0] + "_" + measure[1],
                     provider_name, patient_id, "Submit Supp Data",
                     "Passed", task_id, "", driver.current_url])
                sf.ajax_preloader_wait(driver)

            quality_table = driver.find_element_by_id("table_1")
            driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", quality_table)

            all_measures = quality_table.find_elements_by_xpath(
                ".//tr[contains(@class, 'row-group')]")
            selected_measure_xpath = "//div[contains(text(),'"+measure_tiny_text+"')]//..//..//..//..//..//.."

            global wait_time
            start_time = time.perf_counter()
            end_time = int(start_time) + wait_time
            while True:
                try:
                    measure_element = driver.find_element_by_xpath(selected_measure_xpath)
                except Exception as e:
                    traceback.print_exc()
                    ws.append(
                        [test_case_id, LOB_Name, measure[0] + "_" + measure[1],
                         provider_name, patient_id, "Red dot Status",
                         "Failed", task_id,
                         "Unknown Error occured! Attempting delete",
                         driver.current_url])
                    break
                red_dot_xpath = ".//div[@class='non_compliant red_dot']"

                if end_time - time.perf_counter() < 1:
                    ws.append(
                        [test_case_id, LOB_Name, measure[0] + "_" + measure[1],
                         provider_name, patient_id, "Red dot Status",
                         "Failed", task_id, "Taking longer than wait time! skipping and attempting delete", driver.current_url])
                    break

                if len(measure_element.find_elements_by_xpath(red_dot_xpath)) == 0:
                    print("Dot gone after " + str(int(time.perf_counter() - start_time))+" seconds")
                    ws.append(
                        [test_case_id, LOB_Name, measure[0] + "_" + measure[1],
                         provider_name, patient_id, "Red dot Status",
                         "Passed", task_id, "", driver.current_url])
                    break
                else:
                    print("Looking for red dot, " + str(int(time.perf_counter() - start_time)) + " seconds elapsed")
                    driver.refresh()
                    sf.ajax_preloader_wait(driver)
                    WebDriverWait(driver, 120).until(
                        EC.presence_of_element_located(
                            (By.XPATH, locator.xpath_cozeva_Id)))

                    continue

            # deleting the task
            driver.get(task_url)
            sf.ajax_preloader_wait(driver)
            time.sleep(2)
            delete_button_xpath = config.get("MAP", "delete_xpath")
            try:
                sf.action_click(driver.find_element_by_xpath(delete_button_xpath), driver)
                time.sleep(1)
                # give reason
                reason_modal_xpath = config.get("MAP", "reason_input_modal")
                reason_modal = driver.find_element_by_xpath(reason_modal_xpath)
                reason_modal.send_keys("Cozeva QA")
                time.sleep(1)
                sf.action_click(
                    driver.find_element_by_xpath(config.get("MAP", "confirm_modal_xpath")),
                    driver)
                time.sleep(1)
                if sf.URLAccessCheck(driver.current_url, driver):
                    ws.append(
                        [test_case_id, LOB_Name, measure[0] + "_" + measure[1],
                         provider_name, patient_id, "Delete Supp Data",
                         "Failed", task_id, "Error Toast message was displayed, Manual Intervention needed",
                         driver.current_url])
                    raise Exception("Supp Data Delete Failed")
                else:

                    ws.append(
                        [test_case_id, LOB_Name, measure[0] + "_" + measure[1],
                         provider_name, patient_id, "Delete Supp Data",
                         "Passed", task_id, "", driver.current_url])
                    sf.ajax_preloader_wait(driver)
                sf.ajax_preloader_wait(driver)


            except NoSuchElementException as e:
                traceback.print_exc()
                ws.append(
                    [test_case_id, LOB_Name, measure[0] + "_" + measure[1],
                     provider_name, patient_id, "Delete Supp Data",
                     "Failed", task_id, "Manual Intervention needed", driver.current_url])

            except Exception as e:
                traceback.print_exc()
                ws.append(
                    [test_case_id, LOB_Name, measure[0] + "_" + measure[1],
                     provider_name, patient_id, "Delete Supp Data",
                     "Failed", task_id, "Manual Intervention needed", driver.current_url])
    except Exception as e:
        traceback.print_exc()
        print(e)
        ws.append(
            [test_case_id, LOB_Name, measure[0] + "_" + measure[1],
             provider_name, patient_id, "Attaching a file",
             "Failed", task_id, "Error occured in Coding tool, Manual Intervention might be needed", driver.current_url])



fetch_measures()

print(client_id, MY)

print("Hello World")

#make reporting folder

report_folder = os.path.join(locator.parent_dir,"Supplemental Data Reports")
isdir = os.path.isdir(report_folder)
if not isdir:
    os.mkdir(report_folder)

wb = Workbook()
ws = wb.active
workbook_title = str(db.fetchCustomerName(client_id))+"_Supplemental Data_"+sf.date_time()+"_"+ENV+".xlsx"
ws.title = str(client_id)+"_"+ENV
ws.append(["ID", "LoB", "Metric Name/ID", "Provider","Patient ID", "Scenario", "Status", "Task ID", "Comments", "URL"])
wb.save(report_folder+ "\\"+workbook_title)
test_case_id = 1
driver = setups.driver_setup()
if ENV == 'CERT':
    setups.login_to_cozeva_cert(client_id)
elif ENV == 'STAGE':
    setups.login_to_cozeva_stage()
elif ENV == "PROD":
    setups.login_to_cozeva(client_id)
else:
    print("ENV INVALID")
    exit(3)

found_flags = [None]*len(final_measure_list)
for iterator in range(0, len(found_flags)):
    found_flags[iterator] = 0
sf.ajax_preloader_wait(driver)
WebDriverWait(driver, 120).until(EC.presence_of_element_located((By.XPATH, "//*[@id='qt-filter-label']")))
print("Selected page= " + driver.title)
Registry_URL = driver.current_url
sf.ajax_preloader_wait(driver)
driver.find_element(By.XPATH, "//*[@id='qt-filter-label']").click()
Quarter_list = driver.find_element(By.XPATH, "//*[@id='filter-quarter']").find_elements(By.TAG_NAME, "li")
for quarter in Quarter_list:
    if quarter.text == MY:
        quarter.click()
        break
LOB_list = driver.find_element(By.XPATH, "//*[@id='filter-lob']").find_elements(By.TAG_NAME, 'li')
for i in range(0, len(LOB_list)):
    LOB_Name = LOB_list[i].text
    print("LOB Name: " + LOB_Name)
    try:
        LOB_list[i].click()
    except ElementNotInteractableException as e:
        continue
    #LOB_list[i].click()
    driver.find_element(By.ID, "reg-filter-apply").click()
    LOB_Specific_URL = driver.current_url
    sf.ajax_preloader_wait(driver)
    if driver.find_element(By.XPATH, "//*[@id='conti_enroll']").is_selected():
        driver.find_element(By.XPATH, "//*[@class='cont_disc_toggle']").click()
    print("LOB URL: " + LOB_Specific_URL)
    for measure_index, measure in enumerate(final_measure_list):
        driver.find_element_by_xpath(locator.xpath_filter_measure_list).click()
        time.sleep(2)
        driver.find_element_by_id("qt-search-met").clear()
        driver.find_element_by_id("qt-search-met").send_keys(measure[0])
        driver.find_element_by_xpath("//*[@id='qt-reg-nav-filters']/li[1]/label").click()
        time.sleep(2)
        driver.execute_script("arguments[0].scrollIntoView();", driver.find_element_by_id("qt-apply-search"))
        driver.find_element_by_id("qt-apply-search").click()
        time.sleep(1)
        registry_element = driver.find_element_by_id("registry_body").find_element_by_tag_name("ul")

        if len(registry_element.find_elements_by_id(measure[1])) > 0:
            if measure[2] == "NA":
                driver.find_element_by_id(measure[1]).click()
            else:
                anchor_tags = driver.find_element_by_id(measure[1]).find_elements_by_tag_name("a")
                for anchor in anchor_tags:
                    if str(measure[2]) in str(anchor.get_attribute("href")):
                        anchor.click()

            print('validating')
            sf.ajax_preloader_wait(driver)
            # click on first man icon
            first_provider_element = \
            driver.find_element(By.XPATH, "//*[@id='metric-support-prov-ls']").find_element(By.TAG_NAME,
                                                                                            "tbody").find_elements(
                By.TAG_NAME, 'tr')[0]
            if "No data available" in first_provider_element.text:
                driver.get(LOB_Specific_URL)
                sf.ajax_preloader_wait(driver)
                WebDriverWait(driver, 60).until(
                    EC.presence_of_element_located((By.XPATH, locator.xpath_filter_measure_list)))
                # driver.find_element_by_xpath(locator.xpath_filter_measure_list).click()
                # time.sleep(2)
                # driver.find_element_by_id("qt-search-met").clear()
                # driver.find_element_by_id("qt-search-met").send_keys(measure[0])
                # driver.execute_script("arguments[0].scrollIntoView();",
                #                       driver.find_element_by_id("qt-apply-search"))
                # driver.find_element_by_id("qt-apply-search").click()
                # time.sleep(1)
                # registry_element = driver.find_element_by_id("registry_body").find_element_by_tag_name("ul")
                ws.append([test_case_id, LOB_Name, measure[0]+"_"+measure[1], "x", "x", "Navigate to MSPL of a provider", "Failed", "x", "No providers in Support level", driver.current_url])
                continue
            else:
                provider_elements = \
                    driver.find_element(By.XPATH, "//*[@id='metric-support-prov-ls']").find_element(By.TAG_NAME,
                                                                                                    "tbody").find_elements(
                        By.TAG_NAME, 'tr')
                selected_provider_element = driver.find_element_by_xpath("//body")
                selected_provider_element_name = "PCP"
                while "PCP" in selected_provider_element_name or "pcp" in selected_provider_element_name or "Pcp" in selected_provider_element_name:
                    if len(provider_elements) == 1:
                        selected_provider_element = provider_elements[0]
                        break
                    selected_provider_element = provider_elements[
                        sf.RandomNumberGenerator(len(provider_elements), 1)[0]]
                    selected_provider_element_name = selected_provider_element.find_elements_by_tag_name('a')[
                        2].text.strip()
                driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", selected_provider_element)
                selected_provider_element.find_elements(By.TAG_NAME, 'a')[1].click()
                sf.ajax_preloader_wait(driver)
                provider_name = driver.find_element(By.XPATH, locator.xpath_context_Name).text
                window_switched = 0
                table = driver.find_element_by_id(
                    "quality_registry_list").find_element_by_tag_name(
                    "tbody").find_elements_by_tag_name('tr')
                if "No Data Available" in table[0].text:
                    ws.append(
                        [test_case_id, LOB_Name, measure[0] + "_" + measure[1], provider_name,"x", "Navigate to a patient's Dashboard",
                         "Failed", "x", "No Patients in MSPL", driver.current_url])
                    driver.get(LOB_Specific_URL)
                    sf.ajax_preloader_wait(driver)
                    WebDriverWait(driver, 60).until(
                        EC.presence_of_element_located((By.XPATH, locator.xpath_filter_measure_list)))
                    continue
                first_patient_element = table[0]
                if len(first_patient_element.find_elements_by_xpath("//td[@class=' pencil_icon pencil_icon_pt']")) > 0:
                    print(first_patient_element.find_element_by_xpath("//td[@class=' pencil_icon pencil_icon_pt']").get_attribute("innerHTML"))
                    if "Add Supplemental Data" in first_patient_element.find_element_by_xpath("//td[@class=' pencil_icon pencil_icon_pt']").get_attribute("innerHTML"):
                        first_patient_element.find_element_by_class_name("pat_name").click()
                        time.sleep(1)
                        driver.switch_to.window(driver.window_handles[1])
                        sf.ajax_preloader_wait(driver)
                        WebDriverWait(driver, 120).until(
                            EC.presence_of_element_located((By.CLASS_NAME, "patient_header_wrapper")))
                        window_switched = 1
                        print("Patient Dashboard URL: " + driver.current_url)
                        patient_id = driver.find_element_by_xpath(locator.xpath_cozeva_Id).text
                        print(patient_id)

                        try:
                            driver.find_element_by_xpath("//i[@class='material-icons hcc_toggle tooltipped']").click()
                            time.sleep(1)
                        except Exception as e:
                            #traceback.print_exc()
                            print(e)

                        quality_table = driver.find_element_by_id("table_1")
                        non_complaint_measures = quality_table.find_elements_by_xpath(".//tr[@class='row-group compliant_true']")
                        driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", quality_table)

                        # now, metric wise coding tools.
                        # BCS supplemental Data
                        for dashboard_measure in non_complaint_measures:
                            found_flags[measure_index] = 0
                            if measure[0] in dashboard_measure.find_element_by_xpath(".//div[@class='tiny-text text-grey valign-wrapper']").get_attribute("outerHTML"):
                                found_flags[measure_index] = 1
                                driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", dashboard_measure)
                                sf.action_click(dashboard_measure.find_element_by_xpath(".//a[@class='addSuppData-trigger pts']"), quality_table)
                                time.sleep(0.5)
                                pencil_options = dashboard_measure.find_element_by_xpath(".//ul[@class='dropdown-content patient-menu-list']").find_elements_by_tag_name("a")
                                for pencil_option in pencil_options:
                                    if "Add Supplemental Data" in pencil_option.text:
                                        pencil_option.click()
                                        # After entering the coding tool
                                        #task_id = "Unable to fetch"
                                        if measure[0] == "BCS":
                                            add_supplemental_data("BCS · Preventive Health Screening", None)
                                        elif measure[0] == "CCS":
                                            add_supplemental_data("CCS · Preventive Health Screening", None)
                                        elif measure[0] == "COL":
                                            add_supplemental_data("COL_50_75 · Preventive Health Screening", 4)
                                        break
                            if found_flags[measure_index] == 1:
                                break
                        if window_switched == 1:
                            driver.close()
                            driver.switch_to.window(driver.window_handles[0])
                            window_switched = 0
                            driver.get(LOB_Specific_URL)
                            sf.ajax_preloader_wait(driver)
                            WebDriverWait(driver, 60).until(
                                EC.presence_of_element_located((By.XPATH, locator.xpath_filter_measure_list)))



                    else:
                        ws.append(
                            [test_case_id, LOB_Name, measure[0] + "_" + measure[1], provider_name, "x",
                             "Navigate to a patient's Dashboard",
                             "Failed", "x", "No \"Add Supplemental Data\" option", driver.current_url])
                        driver.get(LOB_Specific_URL)
                        sf.ajax_preloader_wait(driver)
                        WebDriverWait(driver, 60).until(
                            EC.presence_of_element_located((By.XPATH, locator.xpath_filter_measure_list)))
                        continue

                else:
                    ws.append(
                        [test_case_id, LOB_Name, measure[0] + "_" + measure[1], provider_name, "x",
                         "Navigate to a patient's Dashboard",
                         "Failed", "x", "No Pencil icon", driver.current_url])
                    driver.get(LOB_Specific_URL)
                    sf.ajax_preloader_wait(driver)
                    WebDriverWait(driver, 60).until(
                        EC.presence_of_element_located((By.XPATH, locator.xpath_filter_measure_list)))
                    continue

        wb.save(report_folder + "\\" + workbook_title)

    wb.save(report_folder + "\\" + workbook_title)
    sf.ajax_preloader_wait(driver)
    driver.find_element(By.XPATH, "//*[@id='qt-filter-label']").click()
    LOB_list = driver.find_element(By.XPATH, "//*[@id='filter-lob']").find_elements(By.TAG_NAME, 'li')

driver.quit()
rows = ws.max_row
cols = ws.max_column
for i in range(2, rows + 1):
    for j in range(3, cols + 1):
        if ws.cell(i, j).value == 'Passed':
            ws.cell(i, j).fill = PatternFill('solid', fgColor='0FC404')
        elif ws.cell(i, j).value == 'Failed':
            ws.cell(i, j).fill = PatternFill('solid', fgColor='FC0E03')
        elif ws.cell(i, j).value == 'Showing 0 to 0':
            ws.cell(i, j).fill = PatternFill('solid', fgColor='FCC0BB')

wb.save(report_folder + "\\" + workbook_title)


























