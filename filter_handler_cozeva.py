import base64
import os
import random
import time
import traceback
from tkinter import ttk, messagebox

from openpyxl.styles import Font, PatternFill
from selenium.webdriver.common.by import By
from selenium.webdriver.support.wait import WebDriverWait
from tkinter import *
import setups
import pickle
import logging
import ExcelProcessor as db
import context_functions as cf
import support_functions as sf
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
import variablestorage as locator
from openpyxl import Workbook, load_workbook
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import NoSuchElementException, ElementNotInteractableException, \
    ElementClickInterceptedException, TimeoutException

# ENV = 'PROD'
# URL = ""
# client_list = ["1300"]
# Selected_checklist = []
# provider_count = 1
# measurement_year = "2022"
# PatientDashboardFlag = 0



def fetch_client_name():
    root = Tk()


    def on_start():
        client_id = db.fetchCustomerID(selected_cust.get())
        client_list.clear()
        client_list.append(client_id)
        # MY = my_entry.get()
        root.destroy()


    customer_label = Label(root, text="Select customer", font=("Nunito Sans", 10))
    selected_cust = StringVar()
    selected_cust.set("Customer")
    customer_list = db.getCustomerList()  # vs.customer_list
    customer_drop = OptionMenu(root, selected_cust, *customer_list)


    customer_label.grid(row=0, column=0)
    customer_drop.grid(row=1, column=0)
    # Label(root, text="Measurement Year",  font=("Nunito Sans", 10)).grid(row=0, column=3)
    # my_entry = Entry(root)
    # my_entry.grid(row=1, column=3)

    Button(root, text="Start Test", command=on_start, font=("Nunito Sans", 10)).grid(row=2, column=0, columnspan=5)





    root.title("Special column Validation")
    root.iconbitmap("assets/icon.ico")
    # root.geometry("400x400")
    root.mainloop()

def supplemental_data_list(driver, client_id, wb, report_folder, workbook_title):
    wb.create_sheet('Supplemental Data List'+ client_id)
    ws = wb['Supplemental Data List'+ client_id]
    ws.append(["ID", "Filter", "Filtered Value", "Data Present", "Status", "Comments"])
    url_session = "app_id=registries&custId={}&payerId={}&orgId={}&vgpId={}&vpId={}".format(client_id, client_id, client_id, client_id, client_id)
    encoded_string = base64.b64encode(url_session.encode('utf-8'))
    supplemental_data_list_url = "https://www.cozeva.com/smart_chart?session={}&list_type=1".format(encoded_string.decode('utf-8'))
    print(supplemental_data_list_url)
    driver.get(supplemental_data_list_url)
    sf.ajax_preloader_wait(driver)
    # the below list is filter name, internal filter name
    supplemental_data_list_filter_map = [["Task #", "chart_chase_task"],
                                         ["Patient", "chart_chase_patient"],
                                         ["Service Date", "chart_chase_service_date_from", "chart_chase_service_date_to"],
                                         ["Rendering / Reviewing Provider", "chart_chase_provider"],
                                         ["Measure", "chart_chase_measure"],
                                         ["Review 1", "chart_chase_review_6"],
                                         ["Status", "chart_chase_status"]]
    #Gather filter values
    table_data = driver.find_element(By.ID, "chart_chase").find_element(By.TAG_NAME, "tbody").find_elements(By.TAG_NAME, "tr")
    tasks, rendering_providers, service_dates, measures, patients, created_dates, review_status, task_status = [], [], [], [], [], [], [], []
    tasks_unfiltered, rendering_providers_unfiltered, service_dates_unfiltered, measures_unfiltered, patients_unfiltered, review_status_unfiltered, task_status_unfiltered = [], [], [], [], [], ["13"], ["4"]
    # service_dates[0] = []
    review_filter_mapping = [["13", "Completed"]]
    status_filter_mapping = [["4", "Passed"]]
    for list_row in table_data:
        tasks_unfiltered.append(
            str(list_row.find_element(By.CLASS_NAME, "chart_chase_task").find_element(By.TAG_NAME, "a").text).strip())
        patients_unfiltered.append(
            str(list_row.find_element(By.CLASS_NAME, "chart_chase_patient").find_element(By.TAG_NAME,"a").text).strip())
        service_dates_unfiltered.append(str(list_row.find_element(By.CLASS_NAME,"enc_service_date ").text).strip())
        rendering_providers_unfiltered.append(str(list_row.find_element(By.CLASS_NAME, "enc_prov").text).strip())
        measures_unfiltered.append(str(list_row.find_element(By.CLASS_NAME, "enc_meas").text).strip())


    tasks.append(tasks_unfiltered)
    patients.append(patients_unfiltered)
    service_dates.append(service_dates_unfiltered)
    rendering_providers.append(rendering_providers_unfiltered)
    measures.append(measures_unfiltered)
    review_status.append(review_status_unfiltered)
    task_status.append(task_status_unfiltered)


    #index 0 = Unfiltered list
    print(tasks[0])
    print(patients[0])
    print(service_dates[0])
    print(rendering_providers[0])
    print(measures[0])
    print(review_status[0])
    print(task_status[0])

    raw_filter_values = [[random.choice(tasks[0])],
                         [random.choice(patients[0])],
                         [random.choice(service_dates[0]), random.choice(service_dates[0])],
                         [random.choice(rendering_providers[0])],
                         [random.choice(measures[0])],
                         [random.choice(review_status[0])],
                         [random.choice(task_status[0])]]
    sup_filter_values = ["&{}={}".format(supplemental_data_list_filter_map[0][1], raw_filter_values[0][0]),
                         "&{}={}".format(supplemental_data_list_filter_map[1][1], raw_filter_values[1][0]),
                         "&{}={}&{}={}".format(supplemental_data_list_filter_map[2][1], raw_filter_values[2][0], supplemental_data_list_filter_map[2][2], raw_filter_values[2][1]),
                         "&{}={}".format(supplemental_data_list_filter_map[3][1], raw_filter_values[3][0]),
                         "&{}={}".format(supplemental_data_list_filter_map[4][1], raw_filter_values[4][0]),
                         "&{}={}".format(supplemental_data_list_filter_map[5][1], raw_filter_values[5][0]),
                         "&{}={}".format(supplemental_data_list_filter_map[6][1], raw_filter_values[6][0])]
    created_dates = driver.find_elements(By.XPATH, "//div[@class='pls']//span[@class='dt_tag_value']")
    created_from_filter_string = "&{}={}&{}={}".format("chart_chase_uploaded_from", created_dates[0].text[-10:], "chart_chase_uploaded_to", created_dates[1].text[-10:])


    #check filters
    for index, filter_string in enumerate(sup_filter_values):
        filtered_list = []
        print("Reload")
        driver.get(supplemental_data_list_url)
        sf.ajax_preloader_wait(driver)
        filter_url = base64.b64encode(str('{"chart_chase":"' + base64.b64encode((filter_string+created_from_filter_string).encode('utf-8')).decode('utf-8') + '"}').encode('utf-8')).decode('utf-8')
        supplemental_data_list_url_filter_applied = supplemental_data_list_url + "&dt_filter=" + filter_url
        print(supplemental_data_list_url_filter_applied)
        print("List = " + supplemental_data_list_filter_map[index][0])
        print(filter_string)
        driver.get(supplemental_data_list_url_filter_applied)
        sf.ajax_preloader_wait(driver)
        filtered_table_data = driver.find_element(By.ID, "chart_chase").find_element(By.TAG_NAME, "tbody").find_elements(
            By.TAG_NAME, "tr")
        print(filtered_table_data)
        if len(filtered_table_data) < 1:
            ws.append([str(index), str(supplemental_data_list_filter_map[index][0]), str(', '.join(raw_filter_values[index])), "No", "Failed", "Page is Blank", str(driver.current_url)])
        elif len(filtered_table_data) >= 1 and "No matching records found" not in filtered_table_data[0].text:
            for list_row in filtered_table_data:
                if index == 0:
                    filtered_list.append(str(list_row.find_element(By.CLASS_NAME, "chart_chase_task").find_element(By.TAG_NAME, "a").text).strip())
                elif index == 1:
                    filtered_list.append(str(list_row.find_element(By.CLASS_NAME, "chart_chase_patient").find_element(By.TAG_NAME, "a").text).strip())
                elif index == 2:
                    filtered_list.append(str(list_row.find_element(By.CLASS_NAME, "enc_service_date ").text).strip())
                elif index == 3:
                    filtered_list.append(str(list_row.find_element(By.CLASS_NAME, "enc_prov").text).strip())
                elif index == 4:
                    filtered_list.append(str(list_row.find_element(By.CLASS_NAME, "enc_meas").text).strip())
                elif index == 5:
                    filtered_list.append(str(list_row.find_element(By.XPATH, "//td[@class=' chart_chase_review_6']//div//div[1]").text).strip())
                elif index == 6:
                    filtered_list.append(str(list_row.find_element(By.CLASS_NAME, "enc_code_status").text).strip())
            if index == 5:
                for filter_value_index, filter_value in enumerate(filtered_list):
                    for mapping in review_filter_mapping:
                        if mapping[1] in filter_value:
                            filtered_list[filter_value_index] = mapping[0]
                            continue
            if index == 6:
                for filter_value_index, filter_value in enumerate(filtered_list):
                    for mapping in status_filter_mapping:
                        if mapping[1] in filter_value:
                            filtered_list[filter_value_index] = mapping[0]
                            continue

            if raw_filter_values[index][0] in filtered_list:
                ws.append([str(index), str(supplemental_data_list_filter_map[index][0]),str(', '.join(raw_filter_values[index])), "Yes", "Passed", "Filtered Data is present on the list", str(driver.current_url)])
            else:
                ws.append([str(index), str(supplemental_data_list_filter_map[index][0]),str(', '.join(raw_filter_values[index])), "yes", "Failed", "Filtered data not present on the list", str(driver.current_url)])
        filtered_list.clear()
        wb.save(report_folder + "\\" + workbook_title)

    rows = ws.max_row
    cols = ws.max_column
    for i in range(2, rows + 1):
        for j in range(3, cols + 1):
            if ws.cell(i, j).value == 'Passed':
                ws.cell(i, j).fill = PatternFill('solid', fgColor='0FC404')
            elif ws.cell(i, j).value == 'Failed':
                ws.cell(i, j).fill = PatternFill('solid', fgColor='FC0E03')
            elif ws.cell(i, j).value == 'Data table is empty':
                ws.cell(i, j).fill = PatternFill('solid', fgColor='FCC0BB')

    wb.save(report_folder + "\\" + workbook_title)

def hcc_chart_list(driver, client_id, wb, report_folder, workbook_title):
    wb.create_sheet('HCC Chart List'+ client_id)
    ws = wb['HCC Chart List'+ client_id]
    ws.append(["ID", "Filter", "Filtered Value", "Data Present", "Status", "Comments"])
    url_session = "app_id=registries&custId={}&payerId={}&orgId={}&vgpId={}&vpId={}".format(client_id, client_id, client_id, client_id, client_id)
    encoded_string = base64.b64encode(url_session.encode('utf-8'))
    supplemental_data_list_url = "https://www.cozeva.com/smart_chart?session={}&list_type=2".format(encoded_string.decode('utf-8'))
    print(supplemental_data_list_url)
    driver.get(supplemental_data_list_url)
    sf.ajax_preloader_wait(driver)
    # the below list is filter name, internal filter name
    supplemental_data_list_filter_map = [["Task #", "chart_chase_task"],
                                         ["Patient", "chart_chase_patient"],
                                         ["Service Date", "chart_chase_service_date_from", "chart_chase_service_date_to"],
                                         ["Rendering / Reviewing Provider", "chart_chase_provider"],
                                         ["Measure", "chart_chase_measure"],
                                         ["Review 1", "chart_chase_review_6"],
                                         ["Status", "chart_chase_status"]]
    #Gather filter values
    table_data = driver.find_element(By.ID, "chart_chase").find_element(By.TAG_NAME, "tbody").find_elements(By.TAG_NAME, "tr")
    tasks, rendering_providers, service_dates, measures, patients, created_dates, review_status, task_status = [], [], [], [], [], [], [], []
    tasks_unfiltered, rendering_providers_unfiltered, service_dates_unfiltered, measures_unfiltered, patients_unfiltered, review_status_unfiltered, task_status_unfiltered = [], [], [], [], [], ["13"], ["4"]
    # service_dates[0] = []
    review_filter_mapping = [["13", "Completed"]]
    status_filter_mapping = [["4", "Passed"]]
    for list_row in table_data:
        try:
            tasks_unfiltered.append(
                str(list_row.find_element(By.CLASS_NAME, "chart_chase_task").find_element(By.TAG_NAME, "a").text).strip())
            #print(tasks_unfiltered)
            patients_unfiltered.append(
                str(list_row.find_element(By.CLASS_NAME, "chart_chase_patient").find_element(By.TAG_NAME,"a").text).strip())
            #print(patients_unfiltered)
            service_dates_unfiltered.append(str(list_row.find_element(By.CLASS_NAME,"enc_service_date ").text).strip())
            rendering_providers_unfiltered.append(str(list_row.find_element(By.CLASS_NAME, "enc_prov").text).strip())
            measures_unfiltered.append(str(list_row.find_element(By.CLASS_NAME, "enc_meas").text).strip())
        except Exception as e:
            #traceback.print_exc()
            print(e)

    tasks.append(tasks_unfiltered)
    patients.append(patients_unfiltered)
    service_dates.append(service_dates_unfiltered)
    rendering_providers.append(rendering_providers_unfiltered)
    measures.append(measures_unfiltered)
    review_status.append(review_status_unfiltered)
    task_status.append(task_status_unfiltered)


    #index 0 = Unfiltered list
    print(tasks[0])
    print(patients[0])
    print(service_dates[0])
    print(rendering_providers[0])
    print(measures[0])
    print(review_status[0])
    print(task_status[0])

    raw_filter_values = [[random.choice(tasks[0])],
                         [random.choice(patients[0])],
                         [service_dates[0][int((len(service_dates[0])/2)+((len(service_dates[0])/2)/2))], service_dates[0][int((len(service_dates[0])/2)-((len(service_dates[0])/2)/2))]],
                         [random.choice(rendering_providers[0])],
                         [random.choice(measures[0])],
                         [random.choice(review_status[0])],
                         [random.choice(task_status[0])]]
    sup_filter_values = ["&{}={}".format(supplemental_data_list_filter_map[0][1], raw_filter_values[0][0]),
                         "&{}={}".format(supplemental_data_list_filter_map[1][1], raw_filter_values[1][0]),
                         "&{}={}&{}={}".format(supplemental_data_list_filter_map[2][1], raw_filter_values[2][0], supplemental_data_list_filter_map[2][2], raw_filter_values[2][1]),
                         "&{}={}".format(supplemental_data_list_filter_map[3][1], raw_filter_values[3][0]),
                         "&{}={}".format(supplemental_data_list_filter_map[4][1], raw_filter_values[4][0]),
                         "&{}={}".format(supplemental_data_list_filter_map[5][1], raw_filter_values[5][0]),
                         "&{}={}".format(supplemental_data_list_filter_map[6][1], raw_filter_values[6][0])]
    created_dates = driver.find_elements(By.XPATH, "//div[@class='pls']//span[@class='dt_tag_value']")
    created_from_filter_string = "&{}={}&{}={}".format("chart_chase_uploaded_from", created_dates[0].text[-10:], "chart_chase_uploaded_to", created_dates[1].text[-10:])


    #check filters
    for index, filter_string in enumerate(sup_filter_values):
        filtered_list = []
        print("Reload")
        driver.get(supplemental_data_list_url)
        sf.ajax_preloader_wait(driver)
        filter_url = base64.b64encode(str('{"chart_chase":"' + base64.b64encode((filter_string+created_from_filter_string).encode('utf-8')).decode('utf-8') + '"}').encode('utf-8')).decode('utf-8')
        supplemental_data_list_url_filter_applied = supplemental_data_list_url + "&dt_filter=" + filter_url
        print(supplemental_data_list_url_filter_applied)
        print("List = " + supplemental_data_list_filter_map[index][0])
        print(filter_string)
        driver.get(supplemental_data_list_url_filter_applied)
        sf.ajax_preloader_wait(driver)
        filtered_table_data = driver.find_element(By.ID, "chart_chase").find_element(By.TAG_NAME, "tbody").find_elements(
            By.TAG_NAME, "tr")
        print(filtered_table_data)
        if len(filtered_table_data) < 1:
            ws.append([str(index), str(supplemental_data_list_filter_map[index][0]), str(', '.join(raw_filter_values[index])), "No", "Failed", "Page is Blank", str(driver.current_url)])
        elif len(filtered_table_data) >= 1 and "No matching records found" not in filtered_table_data[0].text:
            for list_row in filtered_table_data:
                try:
                    if index == 0:
                        filtered_list.append(str(list_row.find_element(By.CLASS_NAME, "chart_chase_task").find_element(By.TAG_NAME, "a").text).strip())
                    elif index == 1:
                        filtered_list.append(str(list_row.find_element(By.CLASS_NAME, "chart_chase_patient").find_element(By.TAG_NAME, "a").text).strip())
                    elif index == 2:
                        filtered_list.append(str(list_row.find_element(By.CLASS_NAME, "enc_service_date ").text).strip())
                    elif index == 3:
                        filtered_list.append(str(list_row.find_element(By.CLASS_NAME, "enc_prov").text).strip())
                    elif index == 4:
                        filtered_list.append(str(list_row.find_element(By.CLASS_NAME, "enc_meas").text).strip())
                    elif index == 5:
                        filtered_list.append(str(list_row.find_element(By.XPATH, "//td[@class=' chart_chase_review_6']//div//div[1]").text).strip())
                    elif index == 6:
                        filtered_list.append(str(list_row.find_element(By.CLASS_NAME, "enc_code_status").text).strip())
                except Exception as e:
                    traceback.print_exc()

            print("Filtered List 1 = "+str(filtered_list))
            if index == 5:
                for filter_value_index, filter_value in enumerate(filtered_list):
                    for mapping in review_filter_mapping:
                        if mapping[1] in filter_value:
                            filtered_list[filter_value_index] = mapping[0]
                            continue
            if index == 6:
                for filter_value_index, filter_value in enumerate(filtered_list):
                    for mapping in status_filter_mapping:
                        if mapping[1] in filter_value:
                            filtered_list[filter_value_index] = mapping[0]
                            continue
            print("Filtered List 2 = " + str(filtered_list))

            if raw_filter_values[index][0] in filtered_list:
                ws.append([str(index), str(supplemental_data_list_filter_map[index][0]),str(', '.join(raw_filter_values[index])), "Yes", "Passed", "Filtered Data is present on the list", str(driver.current_url)])
            else:
                ws.append([str(index), str(supplemental_data_list_filter_map[index][0]),str(', '.join(raw_filter_values[index])), "yes", "Failed", "Filtered data not present on the list", str(driver.current_url)])
        filtered_list.clear()
        wb.save(report_folder + "\\" + workbook_title)

    rows = ws.max_row
    cols = ws.max_column
    for i in range(2, rows + 1):
        for j in range(3, cols + 1):
            if ws.cell(i, j).value == 'Passed':
                ws.cell(i, j).fill = PatternFill('solid', fgColor='0FC404')
            elif ws.cell(i, j).value == 'Failed':
                ws.cell(i, j).fill = PatternFill('solid', fgColor='FC0E03')
            elif ws.cell(i, j).value == 'Data table is empty':
                ws.cell(i, j).fill = PatternFill('solid', fgColor='FCC0BB')

    wb.save(report_folder + "\\" + workbook_title)


def awv_chart_list(driver, client_id, wb, report_folder, workbook_title):
    wb.create_sheet('AWV Chart List '+client_id)
    ws = wb['AWV Chart List '+client_id]
    ws.append(["ID", "Filter", "Filtered Value", "Data Present", "Status", "Comments"])
    url_session = "app_id=registries&custId={}&payerId={}&orgId={}&vgpId={}&vpId={}".format(client_id, client_id, client_id, client_id, client_id)
    encoded_string = base64.b64encode(url_session.encode('utf-8'))
    supplemental_data_list_url = "https://www.cozeva.com/smart_chart?session={}&list_type=3".format(encoded_string.decode('utf-8'))
    print(supplemental_data_list_url)
    driver.get(supplemental_data_list_url)
    sf.ajax_preloader_wait(driver)
    # the below list is filter name, internal filter name
    supplemental_data_list_filter_map = [["Task #", "chart_chase_task"],
                                         ["Patient", "chart_chase_patient"],
                                         ["Service Date", "chart_chase_service_date_from", "chart_chase_service_date_to"],
                                         ["Rendering / Reviewing Provider", "chart_chase_provider"],
                                         ["Measure", "chart_chase_measure"],
                                         ["Review 1", "chart_chase_review_6"],
                                         ["Status", "chart_chase_status"]]
    #Gather filter values
    table_data = driver.find_element(By.ID, "chart_chase").find_element(By.TAG_NAME, "tbody").find_elements(By.TAG_NAME, "tr")
    tasks, rendering_providers, service_dates, measures, patients, created_dates, review_status, task_status = [], [], [], [], [], [], [], []
    tasks_unfiltered, rendering_providers_unfiltered, service_dates_unfiltered, measures_unfiltered, patients_unfiltered, review_status_unfiltered, task_status_unfiltered = [], [], [], [], [], ["13"], ["4"]
    # service_dates[0] = []
    review_filter_mapping = [["13", "Completed"]]
    status_filter_mapping = [["4", "Passed"]]
    for list_row in table_data:
        try:
            tasks_unfiltered.append(
                str(list_row.find_element(By.CLASS_NAME, "chart_chase_task").find_element(By.TAG_NAME, "a").text).strip())
            #print(tasks_unfiltered)
            patients_unfiltered.append(
                str(list_row.find_element(By.CLASS_NAME, "chart_chase_patient").find_element(By.TAG_NAME,"a").text).strip())
            #print(patients_unfiltered)
            service_dates_unfiltered.append(str(list_row.find_element(By.CLASS_NAME,"enc_service_date ").text).strip())
            rendering_providers_unfiltered.append(str(list_row.find_element(By.CLASS_NAME, "enc_prov").text).strip())
            measures_unfiltered.append(str(list_row.find_element(By.CLASS_NAME, "enc_meas").text).strip())
        except Exception as e:
            #traceback.print_exc()
            print(e)

    tasks.append(tasks_unfiltered)
    patients.append(patients_unfiltered)
    service_dates.append(service_dates_unfiltered)
    rendering_providers.append(rendering_providers_unfiltered)
    measures.append(measures_unfiltered)
    review_status.append(review_status_unfiltered)
    task_status.append(task_status_unfiltered)


    #index 0 = Unfiltered list
    print(tasks[0])
    print(patients[0])
    print(service_dates[0])
    print(rendering_providers[0])
    print(measures[0])
    print(review_status[0])
    print(task_status[0])

    raw_filter_values = [[random.choice(tasks[0])],
                         [random.choice(patients[0])],
                         [service_dates[0][int((len(service_dates[0])/2)+((len(service_dates[0])/2)/2))], service_dates[0][int((len(service_dates[0])/2)-((len(service_dates[0])/2)/2))]],
                         [random.choice(rendering_providers[0])],
                         [random.choice(measures[0])],
                         [random.choice(review_status[0])],
                         [random.choice(task_status[0])]]
    sup_filter_values = ["&{}={}".format(supplemental_data_list_filter_map[0][1], raw_filter_values[0][0]),
                         "&{}={}".format(supplemental_data_list_filter_map[1][1], raw_filter_values[1][0]),
                         "&{}={}&{}={}".format(supplemental_data_list_filter_map[2][1], raw_filter_values[2][0], supplemental_data_list_filter_map[2][2], raw_filter_values[2][1]),
                         "&{}={}".format(supplemental_data_list_filter_map[3][1], raw_filter_values[3][0]),
                         "&{}={}".format(supplemental_data_list_filter_map[4][1], raw_filter_values[4][0]),
                         "&{}={}".format(supplemental_data_list_filter_map[5][1], raw_filter_values[5][0]),
                         "&{}={}".format(supplemental_data_list_filter_map[6][1], raw_filter_values[6][0])]
    created_dates = driver.find_elements(By.XPATH, "//div[@class='pls']//span[@class='dt_tag_value']")
    created_from_filter_string = "&{}={}&{}={}".format("chart_chase_uploaded_from", created_dates[0].text[-10:], "chart_chase_uploaded_to", created_dates[1].text[-10:])


    #check filters
    for index, filter_string in enumerate(sup_filter_values):
        filtered_list = []
        print("Reload")
        driver.get(supplemental_data_list_url)
        sf.ajax_preloader_wait(driver)
        filter_url = base64.b64encode(str('{"chart_chase":"' + base64.b64encode((filter_string+created_from_filter_string).encode('utf-8')).decode('utf-8') + '"}').encode('utf-8')).decode('utf-8')
        supplemental_data_list_url_filter_applied = supplemental_data_list_url + "&dt_filter=" + filter_url
        print(supplemental_data_list_url_filter_applied)
        print("List = " + supplemental_data_list_filter_map[index][0])
        print(filter_string)
        driver.get(supplemental_data_list_url_filter_applied)
        sf.ajax_preloader_wait(driver)
        filtered_table_data = driver.find_element(By.ID, "chart_chase").find_element(By.TAG_NAME, "tbody").find_elements(
            By.TAG_NAME, "tr")
        print(filtered_table_data)
        if len(filtered_table_data) < 1:
            ws.append([str(index), str(supplemental_data_list_filter_map[index][0]), str(', '.join(raw_filter_values[index])), "No", "Failed", "Page is Blank", str(driver.current_url)])
        elif len(filtered_table_data) >= 1 and "No matching records found" not in filtered_table_data[0].text:
            for list_row in filtered_table_data:
                try:
                    if index == 0:
                        filtered_list.append(str(list_row.find_element(By.CLASS_NAME, "chart_chase_task").find_element(By.TAG_NAME, "a").text).strip())
                    elif index == 1:
                        filtered_list.append(str(list_row.find_element(By.CLASS_NAME, "chart_chase_patient").find_element(By.TAG_NAME, "a").text).strip())
                    elif index == 2:
                        filtered_list.append(str(list_row.find_element(By.CLASS_NAME, "enc_service_date ").text).strip())
                    elif index == 3:
                        filtered_list.append(str(list_row.find_element(By.CLASS_NAME, "enc_prov").text).strip())
                    elif index == 4:
                        filtered_list.append(str(list_row.find_element(By.CLASS_NAME, "enc_meas").text).strip())
                    elif index == 5:
                        filtered_list.append(str(list_row.find_element(By.XPATH, "//td[@class=' chart_chase_review_6']//div//div[1]").text).strip())
                    elif index == 6:
                        filtered_list.append(str(list_row.find_element(By.CLASS_NAME, "enc_code_status").text).strip())
                except Exception as e:
                    traceback.print_exc()

            print("Filtered List 1 = "+str(filtered_list))
            if index == 5:
                for filter_value_index, filter_value in enumerate(filtered_list):
                    for mapping in review_filter_mapping:
                        if mapping[1] in filter_value:
                            filtered_list[filter_value_index] = mapping[0]
                            continue
            if index == 6:
                for filter_value_index, filter_value in enumerate(filtered_list):
                    for mapping in status_filter_mapping:
                        if mapping[1] in filter_value:
                            filtered_list[filter_value_index] = mapping[0]
                            continue
            print("Filtered List 2 = " + str(filtered_list))

            if raw_filter_values[index][0] in filtered_list:
                ws.append([str(index), str(supplemental_data_list_filter_map[index][0]),str(', '.join(raw_filter_values[index])), "Yes", "Passed", "Filtered Data is present on the list", str(driver.current_url)])
            else:
                ws.append([str(index), str(supplemental_data_list_filter_map[index][0]),str(', '.join(raw_filter_values[index])), "yes", "Failed", "Filtered data not present on the list", str(driver.current_url)])
        filtered_list.clear()
        wb.save(report_folder + "\\" + workbook_title)

    rows = ws.max_row
    cols = ws.max_column
    for i in range(2, rows + 1):
        for j in range(3, cols + 1):
            if ws.cell(i, j).value == 'Passed':
                ws.cell(i, j).fill = PatternFill('solid', fgColor='0FC404')
            elif ws.cell(i, j).value == 'Failed':
                ws.cell(i, j).fill = PatternFill('solid', fgColor='FC0E03')
            elif ws.cell(i, j).value == 'Data table is empty':
                ws.cell(i, j).fill = PatternFill('solid', fgColor='FCC0BB')

    wb.save(report_folder + "\\" + workbook_title)

def pending_list(client_id, wb, report_folder, workbook_title):
    # make a copy the hcc chart list but replace the list type with 4, and the filter names with the pending list filter names
    wb.create_sheet('Pending List')
    ws = wb['Pending List']
    ws.append(["ID", "Filter", "Filtered Value", "Data Present", "Status", "Comments"]) # add the filter names

# fetch_client_name()

def create_report_folder_and_file(client_list):
    report_folder = os.path.join(locator.parent_dir, "Filter Validation")
    isdir = os.path.isdir(report_folder)
    if not isdir:
        os.mkdir(report_folder)
    client_string = ""
    for ID in client_list:
        client_string = client_string + "_" + ID
    workbook_title = "Filter_report" + client_string + "_" + sf.date_time() + ".xlsx"
    wb = Workbook()
    wb.save(report_folder + "\\" + workbook_title)
    return wb, report_folder, workbook_title








#ws.title = client_list[0] + ENV



# for ID in client_list:
#     driver = setups.driver_setup()
#     if ENV == 'CERT':
#         setups.login_to_cozeva_cert(ID)
#     elif ENV == 'STAGE':
#         setups.login_to_cozeva_stage()
#     elif ENV == "PROD":
#         setups.login_to_cozeva(ID)
#     else:
#         print("ENV INVALID")
#         driver.quit()
#         exit(3)
#
#     try:
#         supplemental_data_list()
#         try:
#             hcc_chart_list()
#         except IndexError as e:
#             pass
#         try:
#             awv_chart_list()
#         except IndexError as e:
#             pass
#     except Exception as e:
#         traceback.print_exc()
#         print(e)






