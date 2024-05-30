
# add other versioning info as print statements. This will help in debugging and version control.
# This is the main file for the RTVS application. This file will be the main file to be run for the application to start.
from datetime import date, datetime, time, timedelta
import multiprocessing
import time
import os
import traceback
from tkinter import *
from threading import Timer
import pickle

from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from openpyxl import Workbook
from msedge.selenium_tools import Edge, EdgeOptions

import ExcelProcessor as db
import openpyxl
from PIL import ImageTk, Image
from tkinter import ttk, messagebox
import webbrowser
import os


# Yes, this is a convoluted way to maintain login info but im lazy and this is necessary
def check_and_create_login_file(file_path):
    if not os.path.exists(file_path):
        try:
            with open(file_path, 'w') as file:
                # Optionally, you can write some initial content to the file
                file.write(' ')
        except IOError:
            print("Login file exists")


def check_and_create_profile_info_file(file_path):
    if not os.path.exists(file_path):
        profile_info_workbook = Workbook()
        # Create a default sheet
        profile_info_sheet = profile_info_workbook.active
        profile_info_sheet.title = "Profile Info"
        profile_info_sheet_content = [["1", "UserTestProfile1", "Available", datetime(2020, 1, 1)],
                                      ["2", "UserTestProfile2", "Available", datetime(2020, 1, 1)],
                                      ["3", "UserTestProfile3", "Available", datetime(2020, 1, 1)],
                                      ["4", "UserTestProfile4", "Available", datetime(2020, 1, 1)],
                                      ["5", "UserTestProfile5", "Available", datetime(2020, 1, 1)],
                                      ["6", "UserTestProfile6", "Available", datetime(2020, 1, 1)],
                                      ["7", "UserTestProfile7", "Available", datetime(2020, 1, 1)],
                                      ["8", "UserTestProfile8", "Available", datetime(2020, 1, 1)],
                                      ["9", "UserTestProfile9", "Available", datetime(2020, 1, 1)],
                                      ["10", "UserTestProfile10", "Available", datetime(2020, 1, 1)]]
        for profile_row_data in profile_info_sheet_content:
            profile_info_sheet.append(profile_row_data)
        # Save the workbook to the specified path
        profile_info_workbook.save(file_path)


def check_and_create_edge_profile_info_file(file_path):
    if not os.path.exists(file_path):
        profile_info_workbook = Workbook()
        # Create a default sheet
        profile_info_sheet = profile_info_workbook.active
        profile_info_sheet.title = "Profile Info"
        profile_info_sheet_content = [["1", "UserTestProfile1", "Available", datetime(2020, 1, 1)],
                                      ["2", "UserTestProfile2", "Available", datetime(2020, 1, 1)],
                                      ["3", "UserTestProfile3", "Available", datetime(2020, 1, 1)],
                                      ["4", "UserTestProfile4", "Available", datetime(2020, 1, 1)],
                                      ["5", "UserTestProfile5", "Available", datetime(2020, 1, 1)],
                                      ["6", "UserTestProfile6", "Available", datetime(2020, 1, 1)],
                                      ["7", "UserTestProfile7", "Available", datetime(2020, 1, 1)],
                                      ["8", "UserTestProfile8", "Available", datetime(2020, 1, 1)],
                                      ["9", "UserTestProfile9", "Available", datetime(2020, 1, 1)],
                                      ["10", "UserTestProfile10", "Available", datetime(2020, 1, 1)]]
        for profile_row_data in profile_info_sheet_content:
            profile_info_sheet.append(profile_row_data)
        # Save the workbook to the specified path
        profile_info_workbook.save(file_path)

def check_and_create_driver_choice_pkl(file_path):
    if not os.path.exists(file_path):
        # with open("assets/driver_choice.pkl", "wb") as driver_choice_file:
        #     pickle.dump("CHROME", driver_choice_file)
        with open("assets/driver_choice.txt", 'a') as driver_choice_file:
            driver_choice_file.seek(0)
            driver_choice_file.truncate()
            driver_choice_file.write("CHROME")
        print("Changed to Chrome 1")
        driver_choice_file.close()

def check_and_create_load_contact_log_pkl(file_path):
    if not os.path.exists(file_path):
        with open(file_path, "wb") as contact_log_file:
            pickle.dump(["1500"], contact_log_file)





# Usage
check_and_create_login_file('assets\loginInfo.txt')
check_and_create_profile_info_file('assets/chrome_profile_info.xlsx')
check_and_create_edge_profile_info_file('assets/edge_profile_info.xlsx')
check_and_create_driver_choice_pkl('assets/driver_choice.txt')
check_and_create_load_contact_log_pkl('assets/contact_log.pkl')
# This snippet will check if the logininfo file is empty. If it is, it will run first time setup
# then it will import the locator library
file = open(r"assets\loginInfo.txt", "r+")
file_content = str(file.read())
file.seek(0)
file.close()
if len(file_content) < 4:
    import FirstTimeSetup
    import variablestorage as locator
else:
    import variablestorage as locator

client_list = []


def rtvsmaster():
    # Store the directory the codebase is in for future calls
    # with open("assets/driver_choice.pkl", "wb") as driver_choice_file:
    #     pickle.dump("CHROME", driver_choice_file)
    with open("assets/driver_choice.txt", 'a') as driver_choice_file:
        driver_choice_file.seek(0)
        driver_choice_file.truncate()
        driver_choice_file.write("CHROME")
    print("Changed to Chrome 2")
    driver_choice_file.close()
    code_directory = os.getcwd()
    try:
        root = Tk()
        root.configure(background='white')
        style = ttk.Style()
        style.theme_use('alt')
        style.configure('My.TButton', font=('Helvetica', 13, 'bold'), foreground='Black', background='#5a9c32',
                        padding=15, highlightthickness=0, height=1, width=25)
        style.configure('My_split.TButton', font=('Helvetica', 12, 'bold'), foreground='Black', background='#5a9c32',
                        padding=15, highlightthickness=0, height=1, width=15)
        style.configure('My_split_filt.TButton', font=('Helvetica', 10, 'bold'), foreground='Black', background='#5a9c32',
                        padding=15, highlightthickness=0, height=1, width=20)
        style.configure('My_edit.TButton', font=('Helvetica', 8, 'bold'), foreground='Black', background='#5a9c32',
                        padding=15, highlightthickness=0, height=1, width=1)
        style.configure('Configs.TButton', font=('Helvetica', 10, 'bold'), foreground='Black', background='#5a9c32',
                        highlightthickness=0)
        style.configure('ChromeProfiles.TButton', font=('Helvetica', 8, 'bold'), foreground='Black',
                        background='#b33d25',
                        highlightthickness=0)

        # style.configure('My.TButton', font=('American typewriter', 14), background='#232323', foreground='white')
        style.map('My.TButton', background=[('active', '#72B132')])
        style.map('My_edit.TButton', background=[('active', '#72B132')])
        style.map('My_split.TButton', background=[('active', '#72B132')])
        style.map('My_split_filt.TButton', background=[('active', '#72B132')])

        def create_tooltip(widget, text):
            def on_enter(event):
                global tooltip_window
                x = widget.winfo_rootx() + 20
                y = widget.winfo_rooty() + widget.winfo_height() + 20
                tooltip_window = Toplevel()
                tooltip_window.wm_overrideredirect(True)
                tooltip_window.wm_geometry("+%d+%d" % (x, y))
                label = Label(tooltip_window, text=text, justify='left',
                                 background='yellow', relief='solid', borderwidth=1,
                                 font=("tahoma", "8", "normal"))
                label.pack(ipadx=1)

            def on_leave(event):
                global tooltip_window
                tooltip_window.destroy()

            widget.bind("<Enter>", on_enter)
            widget.bind("<Leave>", on_leave)
        def image_sizer(image_path):
            image_small = Image.open(image_path).resize((25, 25))

            return image_small

        # making image widgets
        first_time_setup_image = ImageTk.PhotoImage(image_sizer("assets/images/first_time_setup.png"))
        # verification_suite_image = ImageTk.PhotoImage(image_sizer("assets/images/verification_suite.png"))
        verification_suite_image = ImageTk.PhotoImage(Image.open("assets/images/verification_suite.png").resize((20, 22)))
        hcc_validation_image = ImageTk.PhotoImage(image_sizer("assets/images/hcc_validation.png"))
        global_search_image = ImageTk.PhotoImage(image_sizer("assets/images/global_search.png"))
        filter_validation_image = ImageTk.PhotoImage(image_sizer("assets/images/filter_validation_icon.png"))
        task_ingestion_image = ImageTk.PhotoImage(image_sizer("assets/images/task_ingestion.png"))
        analytics_image = ImageTk.PhotoImage(image_sizer("assets/images/analytics.png"))
        slow_log_image = ImageTk.PhotoImage(image_sizer("assets/images/slow_log_trends.png"))
        pdf_printer_image = ImageTk.PhotoImage(image_sizer("assets/images/pdf_printer.png"))
        multi_role_image = ImageTk.PhotoImage(image_sizer("assets/images/Multi_role_access.png"))
        special_column_image = ImageTk.PhotoImage(image_sizer("assets/images/special_columns.png"))
        hospital_activity_image = ImageTk.PhotoImage(image_sizer("assets/images/hospital_activity.png"))
        supp_data_image = ImageTk.PhotoImage(image_sizer("assets/images/supp_data.png"))
        contact_log_image = ImageTk.PhotoImage(image_sizer("assets/images/contact_log.png"))
        cozeva_logo_image = ImageTk.PhotoImage(Image.open("assets/images/cozeva_logo.png").resize((320, 71)))
        help_icon_image = ImageTk.PhotoImage(Image.open("assets/images/help_icon.png").resize((20, 20)))
        update_image = ImageTk.PhotoImage(Image.open("assets/images/update_image_2.png").resize((20, 20)))
        green_dot_image = ImageTk.PhotoImage(Image.open("assets/images/GreenDot.png").resize((10, 10)))
        red_dot_image = ImageTk.PhotoImage(Image.open("assets/images/RedDot.png").resize((10, 10)))
        orange_dot_image = ImageTk.PhotoImage(Image.open("assets/images/OrangeDot.png").resize((10, 10)))
        chrome_logo_image = ImageTk.PhotoImage(Image.open("assets/images/chrome_logo.png").resize((15, 15)))
        edge_logo_image = ImageTk.PhotoImage(Image.open("assets/images/edge_logo.png").resize((15, 15)))
        export_logo_image = ImageTk.PhotoImage(image_sizer("assets/images/export_logo.png"))


        # Widgets+

        logo_label = Label(root, image=cozeva_logo_image, background="white")
        logo_label.grid(row=0, column=2)

        root.columnconfigure(1, weight=1)
        root.rowconfigure(0, weight=1)
        logo_label.grid(sticky="n")
        please_select_label = Label(root, text="Release Team Verification Suite", background="white",
                                    font=("Times New Roman", 15))
        please_select_label.grid(row=1, column=2)
        root.rowconfigure(1, weight=1)
        please_select_label.grid(sticky='n')

        # TRYING SOMETHING ELSE, HOPING THIS WORKS ITS 3 AM

        def on_first_time_setup():
            root.destroy()
            import FirstTimeSetup

        def on_verification_suite():
            flush_unused_driver()
            root.destroy()
            import main

        def on_hcc_validation():
            flush_unused_driver()
            root.destroy()
            import HCC_Validation_multi

        def on_global_search():
            root.destroy()
            import global_search

        def on_global_search_edit():
            print(os.getcwd())
            os.startfile(os.path.join(os.getcwd(),'assets/GlobalSearch.xlsx'))

        def on_filter_validaton():
            flush_unused_driver()
            root.destroy()
            import filter_handler_cozeva

        def on_task_ingestion():
            flush_unused_driver()
            root.destroy()
            import ProspectInjestHCC

        def on_analytics():
            flush_unused_driver()
            root.destroy()
            import runner

        def on_analytics_edit():
            print(os.getcwd())
            os.startfile(os.path.join(os.getcwd(), 'assets/SchemaForAllWorksheets.xlsx'))

        def on_slow_trends():
            root.destroy()
            import slowLogPlotter

        def on_role_access():
            root.destroy()

        def on_special_columns():
            flush_unused_driver()
            root.destroy()
            import special_columns

        def on_hospital_activity():
            flush_unused_driver()
            root.destroy()
            import Hospital_Activity

        def on_supp_data():
            root.destroy()
            # import Supplemental_data_alternate
            # import secret_menu
            import xml_parser

        def on_contact_log():
            root.destroy()
            import contact_log_validator

        # Settings will open a small UI to add in client IDS. Maybe the checkbox UI?
        root_contact_log = Toplevel(root)
        root_contact_log.title("Contact Log Client Selector")
        root_contact_log.iconbitmap("assets/icon.ico")
        root_contact_log.withdraw()

        def on_submitbutton_contact_log():
            global client_list_contact_log
            for i_clog2 in range(1, len(Checkbox_variables_contact_log)):
                # print("In here?")
                if Checkbox_variables_contact_log[i_clog2].get() == 1:
                    #print(db.fetchCustomerID(customer_list_contact_log[i_clog2]))
                    client_list_contact_log.append(db.fetchCustomerID(customer_list_contact_log[i_clog2]))
            #print("asdfdf")
            # print(client_list_contact_log)
            with open("assets/contact_log.pkl", "wb") as contact_log_file:
                pickle.dump(client_list_contact_log, contact_log_file)

            new_config_popup = '''Configuration changed to:\n'''
            for contact_log_id_popup in client_list_contact_log:
                new_config_popup += db.fetchCustomerName(contact_log_id_popup) + "\n"
            new_config_popup = new_config_popup[:-1]

            messagebox.showinfo("Contact Log Config", new_config_popup)

            root_contact_log.destroy()
            root.destroy()
            import contact_log_validator

        def on_chart_list_export():
            print("Chart List validation clicked")
            root.destroy()
            import ChartListExportVerification

        def on_chart_list_report_preview():
            print("Chart List preview clicked")
            #root.destroy()
            #import view_chartlist_report
            exec(open("view_chartlist_report.py").read(), globals())

        customer_list_contact_log = db.getCustomerList()
        # print(customer_list_contact_log)
        # Checkbox_variables_contact_log = [None] * len(customer_list_contact_log)
        Checkbox_variables_contact_log = [IntVar() for _ in range(len(customer_list_contact_log))]
        # for i_clog in range(0, len(Checkbox_variables_contact_log)):
        #    Checkbox_variables_contact_log[i_clog] = IntVar()
        # print(len(Checkbox_variables_contact_log))
        Checkbox_widgets_contact_log = []

        # GenerateCheckboxesForall
        for i_clog3 in range(0, len(customer_list_contact_log)):
            Checkbox_widgets_contact_log.append(
                Checkbutton(root_contact_log, text=customer_list_contact_log[i_clog3],
                            variable=Checkbox_variables_contact_log[i_clog3],
                            font=("Nunito Sans", 10)))
        submit_button_contact_log = Button(root_contact_log, text="Submit", command=on_submitbutton_contact_log,
                                           font=("Nunito Sans", 10))
        # print(Checkbox_widgets_contact_log)
        # add all checkboxes to a grid
        # practice_sidemenu_checkbox.grid(row=3, column=0, columnspan=5, sticky="w")
        for i_clog4 in range(1, len(Checkbox_widgets_contact_log)):
            if i_clog4 <= 20:
                Checkbox_widgets_contact_log[i_clog4].grid(row=i_clog4, column=0, sticky="w")
            elif 20 < i_clog4 <= 40:
                Checkbox_widgets_contact_log[i_clog4].grid(row=i_clog4 - 20, column=1, sticky="w")
            elif 40 < i_clog4 <= 60:
                Checkbox_widgets_contact_log[i_clog4].grid(row=i_clog4 - 40, column=2, sticky="w")
            elif 60 < i_clog4 <= 80:
                Checkbox_widgets_contact_log[i_clog4].grid(row=i_clog4 - 60, column=3, sticky="w")
        submit_button_contact_log.grid(row=0, column=3, sticky="e")

        def on_submitbutton():
            flush_unused_driver()
            global multi
            multi = 1
            for checkbox_index in range(0, len(customer_list)):
                if Checkbox_variables[checkbox_index].get() == 1:
                    client_list.append(db.fetchCustomerID(customer_list[checkbox_index]))

            override_choice = offshore_override_var.get()
            print(client_list)
            with open("assets/offshore_override.pkl", "wb") as override_file:
                pickle.dump(override_choice, override_file)
            if override_choice == 1:
                print("Validating Offshore clients with CS2 only")
            else:
                print("Validating Offshore clients as usual")

            my_choice = overwatch_my_var.get()
            with open("assets/overwatch_my.pkl", "wb") as my_file:
                pickle.dump(my_choice, my_file)

            analytics_choice = overwatch_analytics_choice_var.get()
            with open("assets/overwatch_analytics_choice.pkl", "wb") as analytics_file:
                pickle.dump(analytics_choice, analytics_file)



            root_overwatch.destroy()
            root.destroy()
            # import secret_menu

        def on_pdf_print():
            root.destroy()
            import printcareops

        def on_help():
            # root.destroy()
            pdf_path = os.getcwd()
            pdf_file_path2 = "assets/RTVS Documentation.pdf"
            pdf_path = os.path.join(pdf_path, pdf_file_path2)
            webbrowser.open_new(pdf_path)

        def on_update():
            import git
            import subprocess

            def has_new_commits():
                local_repo = os.getcwd()
                print(local_repo)
                subprocess.run(["git", "fetch"], check=True, cwd=local_repo, shell=True)
                result = subprocess.run(["git", "status", "-uno"], check=True, cwd=local_repo, shell=False,
                                        capture_output=True, text=True)
                output = result.stdout.strip()
                # print(status.stdout)
                if "Your branch is up to date" in output:
                    return False

                elif "Your branch is behind" in output:
                    return True
                return False

            if has_new_commits():
                update_button.configure(text="Updates Available")
            else:
                update_button.configure(text="No Updates")

        button_widgets = []
        button_widgets.append(
            ttk.Button(root, text="First time Setup", command=on_first_time_setup, image=first_time_setup_image,
                       compound="left", style='My.TButton'))
        button_widgets.append(
            ttk.Button(root, text="Verification Suite", command=on_verification_suite, image=verification_suite_image,
                       compound="left", style='My.TButton'))
        button_widgets.append(ttk.Button(root, text="HCC Validation Multi-Client", command=on_hcc_validation,
                                         image=hcc_validation_image, compound="left", style='My.TButton'))
        # button_widgets.append(
        #     ttk.Button(root, text="Filter Validation", command=on_filter_validaton, image=filter_validation_image,
        #                compound="left",
        #                style='My.TButton'))

        chart_list_export_buttons = [
            ttk.Button(root, text="Filters and Exports", command=on_chart_list_export, image=export_logo_image,
                       compound="left", style='My_split_filt.TButton'),
            ttk.Button(root, text="", command=on_chart_list_report_preview, image=first_time_setup_image,
                       compound="left", style='My_edit.TButton')]

        button_widgets.append(chart_list_export_buttons)

        button_widgets.append(ttk.Button(root, text="Task Ingestion(AWV)", command=on_task_ingestion,
                                         image=task_ingestion_image, compound="left", style='My.TButton'))
        # button_widgets.append(
        #     ttk.Button(root, text="Analytics(Default)", command=on_analytics, image=analytics_image,
        #                compound="left", style='My.TButton'))

        analytics_buttons = [
            ttk.Button(root, text="Analytics Full", command=on_analytics, image=analytics_image,
                       compound="left", style='My_split.TButton'),
            ttk.Button(root, text="", command=on_analytics_edit, image=first_time_setup_image,
                       compound="left", style='My_edit.TButton')]

        button_widgets.append(analytics_buttons)

        button_widgets.append(
            ttk.Button(root, text="Slow Log Trends", command=on_slow_trends, image=slow_log_image, compound="left",
                       style='My.TButton'))
        button_widgets.append(
            ttk.Button(root, text="PDF Print Validation", command=on_pdf_print, image=pdf_printer_image,
                       compound="left", style='My.TButton'))

        button_widgets.append(
            ttk.Button(root, text="Special Columns", command=on_special_columns, image=special_column_image,
                       compound="left", style='My.TButton'))

        button_widgets.append(ttk.Button(root, text="Hospital Activity (All Clients)", command=on_hospital_activity,
                                         image=hospital_activity_image, compound="left", style='My.TButton'))
        button_widgets.append(
            ttk.Button(root, text="XML Parser", command=on_supp_data, image=global_search_image,
                       compound="left", style='My.TButton'))
        button_widgets.append(
            ttk.Button(root, text="Overwatch", command=lambda: root_overwatch.deiconify(), image=global_search_image,
                       compound="left", style='My.TButton'))
        global_search_buttons = [
            ttk.Button(root, text="Global Search", command=on_global_search, image=global_search_image,
                       compound="left", style='My_split.TButton'),
            ttk.Button(root, text="", command=on_global_search_edit, image=first_time_setup_image,
                       compound="left", style='My_edit.TButton')]

        button_widgets.append(global_search_buttons)

        button_widgets.append(
            ttk.Button(root, text="TMP", image=help_icon_image,
                       compound="left", style='My.TButton'))

        contact_log_buttons = [
            ttk.Button(root, text="Contact Log", command=on_contact_log, image=contact_log_image,
                       compound="left", style='My_split.TButton'),
            ttk.Button(root, text="", command=lambda: root_contact_log.deiconify(), image=first_time_setup_image,
                       compound="left", style='My_edit.TButton')]

        button_widgets.append(contact_log_buttons)



        button_widgets.append(
            ttk.Button(root, text="TMP", image=help_icon_image,
                       compound="left", style='My.TButton'))

        button_widgets.append(
            ttk.Button(root, text="TMP", image=help_icon_image,
                       compound="left", style='My.TButton'))

        button_widgets.append(
            ttk.Button(root, text="TMP", image=help_icon_image,
                       compound="left", style='My.TButton'))

        button_widgets.append(
            ttk.Button(root, text="TMP", image=help_icon_image,
                       compound="left", style='My.TButton'))

        button_widgets.append(
            ttk.Button(root, text="TMP", image=help_icon_image,
                       compound="left", style='My.TButton'))

        button_widgets.append(
            ttk.Button(root, text="TMP", image=help_icon_image,
                       compound="left", style='My.TButton'))

        button_widgets.append(
            ttk.Button(root, text="TMP", image=help_icon_image,
                       compound="left", style='My.TButton'))



        # button_widgets.append(
        #     ttk.Button(root, text="XML Parser", command=on_supp_data, image=global_search_image,
        #                compound="left", style='My.TButton'))

        help_button = ttk.Button(root, text="Help", command=on_help, image=help_icon_image,
                                 compound="left", style='Configs.TButton')
        update_button = ttk.Button(root, text="Check for Updates", command=on_update, image=update_image,
                                   compound="left", style='Configs.TButton')
        # widget counter to add the buttons in gridwise
        widget_counter = 0
        loopbreak = 0
        tooltip_counter = 0
        with open("assets/contact_log.pkl", 'rb') as contact_log_file:
            client_list_contact_log = pickle.load(contact_log_file)
        contact_log_tooltip = '''Saved Config: \n'''
        for contact_log_id in client_list_contact_log:
            contact_log_tooltip += contact_log_id+" - "+ db.fetchCustomerName(contact_log_id) + "\n"
        contact_log_tooltip = contact_log_tooltip[:-1]
        view_report_tooltip = "View Previous Report"
        filters_exports_tooltip = '''Predefined List: \n'''
        customer_ids_filter_export = ["3000", "1300", "200", "4600", "6800", "6700", "1000", "3300", "1850"]
        customer_ids_filter_export.sort()
        for customer_id_filter_export in customer_ids_filter_export:
            filters_exports_tooltip+= customer_id_filter_export+" - "+db.fetchCustomerName(customer_id_filter_export) + "\n"
        filters_exports_tooltip = filters_exports_tooltip[:-1]

        for i in range(2, 7):
            for j in range(0, 6, 2):
                try:
                    if isinstance(button_widgets[widget_counter], list):
                        tooltip_counter+=1
                        button_widgets[widget_counter][0].grid(row=i, column=j, padx=5, pady=5)
                        button_widgets[widget_counter][1].grid(row=i, column=j+1, padx=5, pady=5)
                        if tooltip_counter == 1:
                            create_tooltip(button_widgets[widget_counter][0], filters_exports_tooltip)
                            create_tooltip(button_widgets[widget_counter][1], view_report_tooltip)
                        if tooltip_counter == 4:
                            create_tooltip(button_widgets[widget_counter][0], contact_log_tooltip)
                    else:
                        button_widgets[widget_counter].grid(row=i, column=j, columnspan=2, padx=5, pady=5)
                except IndexError as e:
                    loopbreak = 1
                    break
                widget_counter += 1
            if loopbreak == 1:
                break

        help_button.grid(row=0, column=0, sticky='nw', padx=5, pady=5)
        update_button.grid(row=0, column=4, columnspan=2, sticky='Ne', padx=5, pady=5)

        def chrome_profile_thread_info(state):
            chrome_profile_frame = Frame(root, background="white")
            if state == "SET":
                # Chromeprofile multi threading

                Label(chrome_profile_frame, text="Chrome Profile Status", background="white",
                      font=("Times New Roman", 15)).grid(row=0, column=0, columnspan=2)

                GUI_workbook = openpyxl.load_workbook('assets/chrome_profile_info.xlsx')
                GUI_sheet = GUI_workbook.active

                chrome_profile_info = []

                for row in GUI_sheet.iter_rows():
                    row_data = []
                    for cell in row:
                        row_data.append(cell.value)
                    chrome_profile_info.append(row_data)
                row_counter = 1
                profile_name_button_list = []
                for profile_row in chrome_profile_info:
                    expired_profile_for_command = profile_row[1]
                    profile_name_button_list.append(ttk.Button(chrome_profile_frame, text=expired_profile_for_command,
                                                               command=lambda
                                                                   profile=expired_profile_for_command: on_expired_profile(
                                                                   profile),
                                                               style='ChromeProfiles.TButton'))
                for profile_row in chrome_profile_info:

                    otp_date = profile_row[3]
                    date_diff = (date.today() - otp_date.date()).days

                    profile_name_label = Label(chrome_profile_frame, text=profile_row[1], background="white",
                                               font=("Times New Roman", 10))
                    if date_diff > 29:
                        profile_name_button_list[row_counter - 1].grid(row=row_counter, column=0)
                    else:
                        profile_name_label.grid(row=row_counter, column=0)
                    profile_status_label = Label(chrome_profile_frame, text=profile_row[2], background="white",
                                                 image=green_dot_image, compound="left", font=("Times New Roman", 10))
                    profile_status_label.grid(row=row_counter, column=1, sticky="w", padx=10)
                    row_counter += 1

                    if profile_row[2] == "In Use":
                        profile_status_label.configure(image=red_dot_image, fg="Red")
                    if profile_row[2] == "Available":
                        profile_status_label.configure(fg="#408022")
                    if profile_row[1] == locator.free_chrome_profile:
                        profile_status_label.configure(image=orange_dot_image, fg="#fc9003", text="Current")
                    if date_diff > 29:
                        profile_status_label.configure(image=red_dot_image, fg="Red", text="Expired!")
                    # print(profile_row[1], date_diff)

                def on_expired_profile(expired_profile):
                    # here, take input of the profile and launch chrome with those parameters
                    print(len(profile_name_button_list))
                    print(profile_name_button_list[2])
                    print(expired_profile)
                    login_info_file = open(r"assets\loginInfo.txt", "r+")
                    expired_chrome_profile_path = "user-data-dir=C:\\Users\\" + login_info_file.readlines()[
                        2].strip() + "\\AppData\\Local\\Google\\Chrome\\User Data\\" + expired_profile
                    login_info_file.seek(0)
                    login_info_file.close()

                    options = webdriver.ChromeOptions()
                    options.add_argument("--disable-notifications")
                    options.add_argument("--start-maximized")
                    options.add_argument(expired_chrome_profile_path)  # Path to your chrome profile
                    options.add_argument('--disable-gpu')
                    # options.add_argument("--window-size=1920,1080")
                    # options.add_argument("--start-maximized")
                    options.add_argument("--no-sandbox")
                    options.add_argument("--dns-prefetch-disable")
                    driver = webdriver.Chrome(executable_path=locator.chrome_driver_path, options=options)
                    print("Chrome Driver for OTP setup with: " + expired_chrome_profile_path)
                    driver.get(locator.logout_link)
                    driver.get(locator.login_link)
                    driver.maximize_window()
                    login_info_file = open(r"assets\loginInfo.txt", "r+")
                    details = login_info_file.readlines()
                    js_clear_and_type = "arguments[0].value = arguments[1];"

                    # Clear and input username
                    username_field = driver.find_element_by_id("edit-name")
                    driver.execute_script(js_clear_and_type, username_field, details[0].strip())

                    # Clear and input password
                    password_field = driver.find_element_by_id("edit-pass")
                    driver.execute_script(js_clear_and_type, password_field, details[1].strip())
                    login_info_file.seek(0)
                    login_info_file.close()
                    driver.find_element(By.ID, "edit-submit").click()
                    time.sleep(4)

                    # If OTP box appears, wait for OTP, then kill the chrome session. No changes needed.

                    otpurl = driver.current_url
                    sub_str = "/twostepAuthSettings"
                    if otpurl.find(sub_str) != -1:
                        # print("Need to enter OTP for login. Please paste the OTP here")
                        # wait_time = 60
                        # start_time = time.perf_counter()
                        # otp = ""
                        # while (time.perf_counter() - start_time) < wait_time:
                        #     otp = input()
                        #     if len(otp) > 0:
                        #         print("OTP Recieved")
                        # if len(otp) == 0:
                        #     print("You did not enter an OTP!!")
                        #     exit(999)

                        timeout = 60
                        t = Timer(timeout, print, ['You did not enter the OTP'])
                        t.start()
                        prompt = "You have %d seconds to enter the OTP here\n" % timeout
                        otp = input(prompt)
                        t.cancel()

                        driver.find_element(By.ID, "edit-twostep-code").send_keys(otp)
                        time.sleep(1)

                        driver.find_element(By.ID, "edit-twostep").click()

                        WebDriverWait(driver, 120).until(EC.presence_of_element_located((By.ID, "reason_textbox")))
                        driver.find_element(By.ID, "reason_textbox").send_keys(details[4].strip())
                        time.sleep(0.5)

                        chrome_profiles_file_location = "assets/chrome_profile_info.xlsx"
                        chrome_profile_workbook = openpyxl.load_workbook(chrome_profiles_file_location)
                        chrome_profiles_current_sheet = chrome_profile_workbook.active
                        for profile_index in range(1, 11):
                            if str(chrome_profiles_current_sheet.cell(row=profile_index,
                                                                      column=2).value).strip() == expired_profile:
                                chrome_profiles_current_sheet.cell(row=profile_index, column=4).value = date.today()
                                break

                        chrome_profile_workbook.save("assets/chrome_profile_info.xlsx")
                    else:
                        WebDriverWait(driver, 120).until(EC.presence_of_element_located((By.ID, "reason_textbox")))
                        driver.find_element(By.ID, "reason_textbox").send_keys(details[4].strip())
                        time.sleep(0.5)
                        chrome_profiles_file_location = "assets/chrome_profile_info.xlsx"
                        chrome_profile_workbook = openpyxl.load_workbook(chrome_profiles_file_location)
                        chrome_profiles_current_sheet = chrome_profile_workbook.active

                        for profile_index in range(1, 11):
                            if str(chrome_profiles_current_sheet.cell(row=profile_index,
                                                                      column=2).value).strip() == expired_profile:
                                chrome_profiles_current_sheet.cell(row=profile_index,
                                                                   column=4).value = date.today() - timedelta(days=29)
                                break

                        chrome_profile_workbook.save("assets/chrome_profile_info.xlsx")

                    driver.quit()

                def on_reset_chrome_profile():
                    import clean_chrome_profiles

                    reset_chrome_profile_button.configure(text="DONE!! Please Relaunch")
                    root.update_idletasks()
                    time.sleep(1)

                def on_kill_chromedriver():
                    import killchromedriver

                    kill_chromedriver_button.configure(text="DONE!! Please Relaunch")
                    time.sleep(1)

                reset_chrome_profile_button = ttk.Button(chrome_profile_frame, text="Reset Chromeprofile Availability",
                                                         command=on_reset_chrome_profile, style='Configs.TButton')
                kill_chromedriver_button = ttk.Button(chrome_profile_frame, text="Kill Chromedriver Tasks",
                                                      command=on_kill_chromedriver, style='Configs.TButton')

                reset_chrome_profile_button.grid(row=11, column=0, columnspan=2)
                kill_chromedriver_button.grid(row=12, column=0, columnspan=2)
                chrome_profile_frame.grid(row=2, rowspan=6, column=6, sticky="NE")
            elif state == "REMOVE":
                #chrome_profile_frame.grid_forget()
                #chrome_profile_frame.grid_remove()
                chrome_profile_frame.destroy()
                print("Forget Chrome")

        def edge_profile_thread_info(state):
            edge_profile_frame = Frame(root, background="white")
            if state == "SET":
                # edgeprofile multi threading

                Label(edge_profile_frame, text="Edge Profile Status", background="white",
                      font=("Times New Roman", 15)).grid(row=0, column=0, columnspan=2)

                GUI_workbook = openpyxl.load_workbook('assets/edge_profile_info.xlsx')
                GUI_sheet = GUI_workbook.active

                edge_profile_info = []

                for row in GUI_sheet.iter_rows():
                    row_data = []
                    for cell in row:
                        row_data.append(cell.value)
                    edge_profile_info.append(row_data)
                row_counter = 1
                profile_name_button_list = []
                for profile_row in edge_profile_info:
                    expired_profile_for_command = profile_row[1]
                    profile_name_button_list.append(ttk.Button(edge_profile_frame, text=expired_profile_for_command,
                                                               command=lambda
                                                                   profile=expired_profile_for_command: on_expired_profile(
                                                                   profile),
                                                               style='ChromeProfiles.TButton'))
                for profile_row in edge_profile_info:

                    otp_date = profile_row[3]
                    date_diff = (date.today() - otp_date.date()).days

                    profile_name_label = Label(edge_profile_frame, text=profile_row[1], background="white",
                                               font=("Times New Roman", 10))
                    if date_diff > 29:
                        profile_name_button_list[row_counter - 1].grid(row=row_counter, column=0)
                    else:
                        profile_name_label.grid(row=row_counter, column=0)
                    profile_status_label = Label(edge_profile_frame, text=profile_row[2], background="white",
                                                 image=green_dot_image, compound="left", font=("Times New Roman", 10))
                    profile_status_label.grid(row=row_counter, column=1, sticky="w", padx=10)
                    row_counter += 1

                    if profile_row[2] == "In Use":
                        profile_status_label.configure(image=red_dot_image, fg="Red")
                    if profile_row[2] == "Available":
                        profile_status_label.configure(fg="#408022")
                    if profile_row[1] == locator.free_chrome_profile:
                        profile_status_label.configure(image=orange_dot_image, fg="#fc9003", text="Current")
                    if date_diff > 29:
                        profile_status_label.configure(image=red_dot_image, fg="Red", text="Expired!")
                    # print(profile_row[1], date_diff)

                def on_expired_profile(expired_profile):
                    # here, take input of the profile and launch chrome with those parameters
                    print(len(profile_name_button_list))
                    print(profile_name_button_list[2])
                    print(expired_profile)
                    login_info_file = open(r"assets\loginInfo.txt", "r+")
                    expired_edge_profile_path = "user-data-dir=C:\\Users\\" + login_info_file.readlines()[
                        2].strip() + "\\AppData\\Local\\Microsoft\\Edge\\User Data\\" + expired_profile
                    login_info_file.seek(0)
                    login_info_file.close()
                    print(expired_edge_profile_path)

                    options = EdgeOptions()
                    options.use_chromium = True  # Ensure we're using the Chromium-based version of Edge
                    options.add_argument("--disable-notifications")
                    options.add_argument("--start-maximized")
                    options.add_argument("--"+expired_edge_profile_path)  # Path to your edge profile
                    options.add_argument('--disable-gpu')
                    # options.add_argument("--window-size=1920,1080")
                    # options.add_argument("--start-maximized")
                    options.add_argument("--no-sandbox")
                    options.add_argument("--disable-extensions")  # Disabling extensions in Edge
                    options.add_argument("--dns-prefetch-disable")
                    preferences = {
                        "download.default_directory": locator.download_dir}
                    options.add_experimental_option("prefs", preferences)
                    driver = Edge(executable_path=locator.edge_driver_path, options=options)
                    print("Edge Driver for OTP setup with: "+expired_edge_profile_path)
                    driver.get(locator.logout_link)
                    driver.get(locator.login_link)
                    driver.maximize_window()
                    login_info_file = open(r"assets\loginInfo.txt", "r+")
                    details = login_info_file.readlines()
                    js_clear_and_type = "arguments[0].value = arguments[1];"

                    # Clear and input username
                    username_field = driver.find_element_by_id("edit-name")
                    driver.execute_script(js_clear_and_type, username_field, details[0].strip())

                    # Clear and input password
                    password_field = driver.find_element_by_id("edit-pass")
                    driver.execute_script(js_clear_and_type, password_field, details[1].strip())
                    login_info_file.seek(0)
                    login_info_file.close()
                    driver.find_element(By.ID, "edit-submit").click()
                    time.sleep(4)

                    # If OTP box appears, wait for OTP, then kill the chrome session. No changes needed.

                    otpurl = driver.current_url
                    sub_str = "/twostepAuthSettings"
                    if otpurl.find(sub_str) != -1:
                        # print("Need to enter OTP for login. Please paste the OTP here")
                        # wait_time = 60
                        # start_time = time.perf_counter()
                        # otp = ""
                        # while (time.perf_counter() - start_time) < wait_time:
                        #     otp = input()
                        #     if len(otp) > 0:
                        #         print("OTP Recieved")
                        # if len(otp) == 0:
                        #     print("You did not enter an OTP!!")
                        #     exit(999)

                        timeout = 60
                        t = Timer(timeout, print, ['You did not enter the OTP'])
                        t.start()
                        prompt = "You have %d seconds to enter the OTP here\n" % timeout
                        otp = input(prompt)
                        t.cancel()

                        driver.find_element(By.ID, "edit-twostep-code").send_keys(otp)
                        time.sleep(1)

                        driver.find_element(By.ID, "edit-twostep").click()

                        WebDriverWait(driver, 120).until(EC.presence_of_element_located((By.ID, "reason_textbox")))
                        driver.find_element(By.ID, "reason_textbox").send_keys(details[4].strip())
                        time.sleep(0.5)

                        edge_profiles_file_location = "assets/edge_profile_info.xlsx"
                        edge_profile_workbook = openpyxl.load_workbook(edge_profiles_file_location)
                        edge_profiles_current_sheet = edge_profile_workbook.active
                        for profile_index in range(1, 11):
                            if str(edge_profiles_current_sheet.cell(row=profile_index,
                                                                    column=2).value).strip() == expired_profile:
                                edge_profiles_current_sheet.cell(row=profile_index, column=4).value = date.today()
                                break

                        edge_profile_workbook.save("assets/edge_profile_info.xlsx")
                    else:
                        WebDriverWait(driver, 120).until(EC.presence_of_element_located((By.ID, "reason_textbox")))
                        driver.find_element(By.ID, "reason_textbox").send_keys(details[4].strip())
                        time.sleep(0.5)
                        edge_profiles_file_location = "assets/edge_profile_info.xlsx"
                        edge_profile_workbook = openpyxl.load_workbook(edge_profiles_file_location)
                        edge_profiles_current_sheet = edge_profile_workbook.active

                        for profile_index in range(1, 11):
                            if str(edge_profiles_current_sheet.cell(row=profile_index,
                                                                    column=2).value).strip() == expired_profile:
                                edge_profiles_current_sheet.cell(row=profile_index,
                                                                 column=4).value = date.today() - timedelta(days=29)
                                break

                        edge_profile_workbook.save("assets/edge_profile_info.xlsx")

                    driver.quit()

                def on_reset_edge_profile():
                    import clean_chrome_profiles

                    reset_edge_profile_button.configure(text="DONE!! Please Relaunch")
                    root.update_idletasks()
                    time.sleep(1)

                def on_kill_edgedriver():
                    import killchromedriver

                    kill_edgedriver_button.configure(text="DONE!! Please Relaunch")
                    time.sleep(1)

                reset_edge_profile_button = ttk.Button(edge_profile_frame, text="Reset Edgeprofile Availability",
                                                       command=on_reset_edge_profile, style='Configs.TButton')
                kill_edgedriver_button = ttk.Button(edge_profile_frame, text="Kill Edgedriver Tasks",
                                                    command=on_kill_edgedriver, style='Configs.TButton')

                reset_edge_profile_button.grid(row=11, column=0, columnspan=2)
                kill_edgedriver_button.grid(row=12, column=0, columnspan=2)
                edge_profile_frame.grid(row=2, rowspan=4, column=6, sticky="NE")
            elif state == "REMOVE":
                edge_profile_frame.destroy()
                print("Forget Edge")

        def on_driver_choice(choice):
            # Get the current value of the driver choice from the radio button
            driver_choice = choice

            # Remove old frame
            edge_profile_thread_info("REMOVE")
            chrome_profile_thread_info("REMOVE")

            # Apply new frame based on the driver choice
            if driver_choice == "EDGE":
                edge_profile_thread_info("SET")
            elif driver_choice == "CHROME":
                chrome_profile_thread_info("SET")

            # Update pkl file with the current driver choice
            # with open("assets/driver_choice.pkl", "wb") as driver_choice_file:
            #     pickle.dump(driver_choice, driver_choice_file)
            #     print("Set to :"+driver_choice)
            #     driver_choice_file.close()

            with open("assets/driver_choice.txt", 'a') as driver_choice_file:
                driver_choice_file.seek(0)
                driver_choice_file.truncate()
                driver_choice_file.write(driver_choice)
            driver_choice_file.close()
            print("Set to :" + driver_choice)




        # Define a style for Checkbuttons
        style.configure('My.TCheckbutton', font=('Helvetica', 10, 'bold'), foreground='Black',
                        background='#5a9c32', focuscolor='none', padding=5, highlightthickness=0, height=1, width=20)

        # Adjust padding and indicator size
        style.configure('My.TCheckbutton')
        style.map('My.TCheckbutton',
                  foreground=[('selected', 'Black'), ('active', 'black')],
                  background=[('active', '#497f29'), ('selected', '#497f29')],
                  )

        style.configure('DriverChoice.TButton', font=('Helvetica', 10, 'bold'), foreground='Black',
                        background='#5a9c32',
                        highlightthickness=0)

        # Create a StringVar to hold the value of the selected driver
        driver_var = StringVar(value="CHROME")  # Default value
        driver_choice_frame = Frame(root, background='white')

        # Create checkbuttons for driver selection
        edge_check = ttk.Button(driver_choice_frame, text="Microsoft Edge",
                                command=lambda: on_driver_choice("EDGE"), image=edge_logo_image,
                                compound="left", style='DriverChoice.TButton')
        chrome_check = ttk.Button(driver_choice_frame, text="Google Chrome",
                                  command=lambda: on_driver_choice("CHROME"), image=chrome_logo_image,
                                  compound="left", style='DriverChoice.TButton')

        # Place the checkbuttons in the window
        Label(driver_choice_frame, text="Browser Selection", background="white",
              font=("Times New Roman", 15)).grid(row=0, column=0, padx=1, pady=2)
        edge_check.grid(row=1, column=0, padx=2, pady=1)
        chrome_check.grid(row=2, column=0, padx=2, pady=1)
        driver_choice_frame.grid(row=0, column=6)

        on_driver_choice("CHROME")

        root_overwatch = Toplevel(root)
        root_overwatch.title("Cozeva Overwatch")
        root_overwatch.iconbitmap("assets/icon.ico")
        root_overwatch.withdraw()
        customer_list = db.getCustomerList()
        # print(customer_list)
        Checkbox_variables = []
        for i in range(0, len(customer_list)):
            Checkbox_variables.append(IntVar())
        print(len(Checkbox_variables))
        Checkbox_widgets = []

        for i in range(0, len(customer_list)):
            Checkbox_widgets.append(Checkbutton(root_overwatch, text=customer_list[i], variable=Checkbox_variables[i],
                                                font=("Nunito Sans", 10)))
        submit_button = Button(root_overwatch, text="Submit", command=on_submitbutton, font=("Nunito Sans", 10))

        # add all checkboxes to a grid
        # practice_sidemenu_checkbox.grid(row=3, column=0, columnspan=5, sticky="w")
        for i in range(1, len(Checkbox_widgets)):
            if i <= 20:
                Checkbox_widgets[i].grid(row=i, column=0, sticky="w")
            elif 20 < i <= 40:
                Checkbox_widgets[i].grid(row=i - 20, column=1, sticky="w")
            elif 40 < i <= 60:
                Checkbox_widgets[i].grid(row=i - 40, column=2, sticky="w")
            elif 60 < i <= 80:
                Checkbox_widgets[i].grid(row=i - 60, column=3, sticky="w")
        submit_button.grid(row=0, column=3, sticky="e")
        offshore_override_var = IntVar()
        offshore_override_checkbutton = Checkbutton(root_overwatch, text="Validate Offshore clients through CS2", variable=offshore_override_var, font=("Nunito Sans", 11))
        offshore_override_checkbutton.grid(row=0, column=0, sticky='w')
        offshore_override_checkbutton.select()
        overwatch_my_var = StringVar()
        overwatch_my_var.set("Default")
        overwatch_selected_my_list = ["2020", "2021", "2022", "2023", "2024", "2025", "2026", "2027", "2028"]
        overwatch_selected_my_list = overwatch_selected_my_list[::-1]
        overwatch_my_drop = ttk.Combobox(root_overwatch, textvariable=overwatch_my_var, values=overwatch_selected_my_list, state='readonly',
                               style='TCombobox', width=10, height=35)
        Label(root_overwatch, text=" || Select MY ||", font=("Nunito Sans", 12)).grid(row=0, column=1, sticky='w')
        overwatch_my_drop.grid(row=0, column=1, sticky='e')

        #make a checkbox that checks if analytics is yes or no
        overwatch_analytics_choice_var = IntVar()
        overwatch_analytics_choice_checkbutton = Checkbutton(root_overwatch, text="Analytics Choice", variable=overwatch_analytics_choice_var, font=("Nunito Sans", 11))
        overwatch_analytics_choice_checkbutton.grid(row=0, column=2, sticky='w')
        overwatch_analytics_choice_checkbutton.select()


        root.title("Release Team Master Suite")
        root.iconbitmap("assets/icon.ico")
        # root.geometry("400x400+300+100")
        root.mainloop()

    except Exception as e:
        print(e)
        traceback.print_exc()
    finally:
        os.chdir(code_directory)
        chrome_profiles = openpyxl.load_workbook("assets/chrome_profile_info.xlsx")
        chrome_profiles_sheet = chrome_profiles.active
        chrome_profile_available = False
        for profile_index in range(1, 11):
            if str(chrome_profiles_sheet.cell(row=profile_index,
                                              column=2).value).strip() == locator.free_chrome_profile:
                chrome_profiles_sheet.cell(row=profile_index, column=3).value = 'Available'
                break

        chrome_profiles.save("assets/chrome_profile_info.xlsx")
        edge_profiles = openpyxl.load_workbook("assets/edge_profile_info.xlsx")
        edge_profiles_sheet = edge_profiles.active
        edge_profile_available = False
        # Look for a row with an Available edgeprofile name, Change it to In use and return the name
        for profile_index in range(1, 11):
            if str(edge_profiles_sheet.cell(row=profile_index,
                                            column=2).value).strip() == locator.free_edge_profile:
                edge_profiles_sheet.cell(row=profile_index, column=3).value = 'Available'
                break

        edge_profiles.save("assets/edge_profile_info.xlsx")


        # with open("assets/driver_choice.txt", 'a') as driver_choice_file:
        #     driver_choice_file.seek(0)
        #     driver_choice_file.truncate()
        #     driver_choice_file.write("CHROME")
        # print("Changed to Chrome 3")
        # driver_choice_file.close()

        # with open("assets/driver_choice.pkl", "wb") as driver_choice_file:
        #     pickle.dump("CHROME", driver_choice_file)
def flush_unused_driver():
    with open("assets/driver_choice.txt", 'r+') as driver_flush_file:
        driver_flush_choice = driver_flush_file.read().strip()
    if driver_flush_choice == "EDGE":
        chrome_profiles = openpyxl.load_workbook("assets/chrome_profile_info.xlsx")
        chrome_profiles_sheet = chrome_profiles.active
        chrome_profile_available = False
        for profile_index in range(1, 11):
            if str(chrome_profiles_sheet.cell(row=profile_index,
                                              column=2).value).strip() == locator.free_chrome_profile:
                chrome_profiles_sheet.cell(row=profile_index, column=3).value = 'Available'
                break

        chrome_profiles.save("assets/chrome_profile_info.xlsx")
    elif driver_flush_choice == "CHROME":
        edge_profiles = openpyxl.load_workbook("assets/edge_profile_info.xlsx")
        edge_profiles_sheet = edge_profiles.active
        edge_profile_available = False
        # Look for a row with an Available edgeprofile name, Change it to In use and return the name
        for profile_index in range(1, 11):
            if str(edge_profiles_sheet.cell(row=profile_index,
                                            column=2).value).strip() == locator.free_edge_profile:
                edge_profiles_sheet.cell(row=profile_index, column=3).value = 'Available'
                break

        edge_profiles.save("assets/edge_profile_info.xlsx")


def run_script(argument):
    def add_number_to_file(file_path, number):
        try:
            with open(file_path, 'a') as file:
                file.write(str(number) + '\n')
        except IOError as e:
            print(f"Error: {e}")

    file_path = 'assets\\overwatch_cache.txt'
    number = argument
    add_number_to_file(file_path, number)
    flush_unused_driver()

    import multimain
    time.sleep(2)


if __name__ == '__main__':
    multiprocessing.freeze_support()
    multi = 0
    client_list_contact_log = []
    #print("Release Team Verification Suite(RTVS) Version: 1.4.1, Latest Update: Edgedriver support, Overwatch onshore override")
    #print("Release Team Verification Suite(RTVS) Version: 1.4.4, Latest Update: Contact Log, Global search, Custom Settings, Selectable MY, Help Icon dropdown")
    print("Release Team Verification Suite(RTVS) Version: 1.4.5, Latest Update: Exports and filters(Chart Lists), Global search, Custom Settings, Help Icon dropdown")
    print("Requires : Chrome Version 124, Edge Version 124, Python 3.9+, Git(for live updates), Windows 10, 11")
    print("Developed by: Writtwik Dey for the Cozeva Release Team")
    print("Current Client Count: " + str(len(db.getCustomerList()) - 2))

    print("-------------------------Logs-------------------------")

    rtvsmaster()
    num_processes = len(client_list)

    if multi == 1:

        processes = []
        for arg in client_list:
            process = multiprocessing.Process(target=run_script, args=(arg,))
            process.start()
            processes.append(process)
            time.sleep(2)

        for process in processes:
            process.join()

        import clean_chrome_profiles
    else:
        print(multi)
