import base64
import configparser
import os
import random
import re
import timeit
import traceback
from os import listdir
from os.path import isfile, join
from random import randint
from termcolor import colored

from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.formatting import Rule
from openpyxl.styles import PatternFill, Font, NamedStyle, Border, Alignment, Side
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.styles import PatternFill
from selenium.webdriver import ActionChains
from selenium.webdriver.common.by import By
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import NoSuchElementException, ElementNotInteractableException, ElementClickInterceptedException, TimeoutException
from sigfig import round

import support_functions as sf
import variablestorage as locator
import time
import datetime as dt
from datetime import date, datetime, timedelta
config = configparser.RawConfigParser()
config.read("locator-config.properties")

global global_search_prov, global_search_prac, global_search_pat

#in development
def submit_support_ticket(driver):
    driver.find_element(By.XPATH, '//*[@data-target="help_menu_dropdown"]').click()
    driver.find_element(By.XPATH, '//*[@id="help_menu_options"]/li[4]/a').click();
    time.sleep(4)

    today = date.today()
    driver.find_element(By.XPATH, '//*[@id="edit-subject"]').send_keys("Test Automation" + str(today))
    driver.find_element(By.XPATH, '//*[@id="edit-problem-cozeva-id"]').send_keys("Test Automation" + str(today))
    driver.find_element(By.XPATH, '//*[@id="recreate_issue_note"]').send_keys("Test Automation" + str(today))
    driver.find_element(By.XPATH, '//*[@id="select_div_lower"]/div[9]/div/div/input').click()

    driver.find_element(By.XPATH, "//*[text()='Normal (Minimal impact or maintenance of business request)']").click()
    driver.find_element(By.XPATH, '//*[@id="edit-file-upload"]').send_keys(os.getcwd() + "/Doc_pdf.pdf")

    ele = driver.find_element(By.TAG_NAME, 'body')
    ele.send_keys(Keys.END)
    time.sleep(3)
    driver.find_element(By.XPATH, '//*[@id="phi_warning_cust_data"]/button').click()

def init_global_search():
    global global_search_pat
    global global_search_prov
    global global_search_prac
    global_search_prov = None
    global_search_pat = None
    global_search_prac = None




def support_menubar(driver, workbook, ws, logger, run_from, report_folder, context):
    if ws is None:
        workbook.create_sheet('Support Menubar')
        ws = workbook['Support Menubar']

    ws.append(['ID', 'Context', 'Scenario', 'Status', 'Time Taken', 'Comments'])
    logger.info("MenubarNavigation function started.")
    header_font = Font(color='FFFFFF', bold=False, size=12)
    header_cell_color = PatternFill('solid', fgColor='030303')
    ws['A1'].font = header_font
    ws['A1'].fill = header_cell_color
    ws['B1'].font = header_font
    ws['B1'].fill = header_cell_color
    ws['C1'].font = header_font
    ws['C1'].fill = header_cell_color
    ws['D1'].font = header_font
    ws['D1'].fill = header_cell_color
    ws['E1'].font = header_font
    ws['E1'].fill = header_cell_color
    ws.name = "Arial"
    test_case_id = 1
    main_registry_url = driver.current_url
    if context == "":
        context = driver.find_element(By.XPATH, "//span[@class='specific_most']").text

    try:
        logger.info("Menubar navigation block started.")
        time.sleep(1)
        sf.ajax_preloader_wait(driver)
        current_url = driver.current_url
        access_message = sf.CheckAccessDenied(current_url)
        error_message = sf.CheckErrorMessage(driver)

        if access_message == 0 and error_message == 0:
            WebDriverWait(driver, 30).until(
                EC.presence_of_element_located((By.XPATH, locator.xpath_filter_measure_list)))
            WebDriverWait(driver, 10).until(
                EC.presence_of_element_located((By.XPATH, locator.xpath_side_nav_SlideOut)))


            time.sleep(0.5)
            context_name = driver.find_element_by_xpath(locator.xpath_context_Name).text
            print(context_name)
            if context_name is None or context_name == "":
                print("EXECUTION FAILED. PLEASE RERUN")
                driver.quit()
                exit(99999)
            test_case_id += 1
            ws.append((test_case_id, '', 'Set context to: ' + context_name, 'Passed'))
            logger.info("if#1 block ended.")

        elif access_message == 1:
            logger.info("elif#1 block started.")
            test_case_id += 1
            ws.append([test_case_id, '', 'Access Check', 'Failed'])
            logger.info("elif#1 block ended.")

        elif error_message == 1:
            logger.info("elif#2 block started.")
            test_case_id += 1
            ws.append((test_case_id, '', 'Default context without error message', 'Failed'))
            logger.info("elif#2 block ended.")

        WebDriverWait(driver, 30).until(
            EC.presence_of_element_located((By.XPATH, locator.xpath_filter_measure_list)))
        WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.XPATH, locator.xpath_side_nav_SlideOut)))
        time.sleep(0.5)
        context_name = driver.find_element_by_xpath(locator.xpath_context_Name).text
        print(context_name)
        driver.find_element_by_xpath(locator.xpath_side_nav_SlideOut).click()
        try:
            time.sleep(1.5)
            links = driver.find_elements_by_xpath(locator.xpath_menubar_Item_Link)
            names = driver.find_elements_by_xpath(locator.xpath_menubar_Item_Link_Name)
            driver.find_element_by_xpath(locator.xpath_side_nav_SlideOut).click()

        except Exception as e:
            test_case_id += 1
            ws.append((test_case_id, context_name, 'Sidemenubar navigation', 'Failed'))
            print(e)
            traceback.print_exc()

        for link in range(len(links)) and range(len(names)):
            time.sleep(1.5)
            driver.find_element_by_xpath(locator.xpath_side_nav_SlideOut).click()
            time.sleep(0.5)
            driver.execute_script("arguments[0].scrollIntoView();", links[link])
            print("Link Index: " + str(link))
            print(names[link].text)
            link_name = names[link].text
            try:
                links[link].click()
                start_time = time.perf_counter()
                sf.ajax_preloader_wait(driver)
                if link_name == "AWV Chart List" or "Imported Charts":
                    sf.ajax_preloader_wait2(driver)
                total_time = time.perf_counter() - start_time
                current_url = driver.current_url
                access_message = sf.CheckAccessDenied(current_url)
                if access_message == 1:
                    print("Access Denied found!")
                    logger.error(context_name + "-->" + link_name + ": " + "Access Denied found!")
                    test_case_id += 1
                    ws.append((test_case_id, context_name, 'Access Check for ' + link_name, 'Failed'))
                else:
                    print("Access Check done!")
                    error_message = sf.CheckErrorMessage(driver)
                    if error_message == 1:
                        print("Error toast message is displayed")
                        # logger.critical("ERROR TOAST MESSAGE IS DISPLAYED!")
                        test_case_id += 1
                        ws.append((test_case_id, context_name, link_name + ' without error message', 'Failed', driver.current_url))
                        logger.error(context_name + "-->" + link_name + ": " + "Error message found!")
                    else:

                        if len(driver.find_elements_by_xpath(locator.xpath_data_Table_Info)) != 0:
                            time.sleep(0.5)
                            datatable_info = driver.find_element_by_xpath(locator.xpath_data_Table_Info).text
                            print(datatable_info)
                            test_case_id += 1
                            time.sleep(2)
                            sf.captureScreenshot(driver, link_name + " " + context, report_folder)
                            ws.append((test_case_id, context_name, 'Navigation to ' + link_name, 'Passed',
                                       str(round(total_time, sigfigs=3)),
                                       datatable_info))
                            logger.info(context_name + "-->" + link_name + ": " + "Navigation done.")

                        else:
                            print("No datatable!")
                            test_case_id += 1
                            sf.captureScreenshot(driver, link_name + " No Data " + context, report_folder)
                            ws.append((test_case_id, context_name, 'Navigation to ' + link_name, 'Passed',
                                       str(round(total_time, sigfigs=3))))
                        if link_name == "Bridge":
                            dashboard_links = driver.find_element(By.XPATH, "//ul[@id='user_bridge_ls']").find_elements(By.TAG_NAME, "li")
                            for i in range(1, len(dashboard_links), 1):
                                dashboard_links[i].click()
                                sf.ajax_preloader_wait(driver)
                                WebDriverWait(driver, 120).until(EC.presence_of_element_located((By.XPATH, "//*[@id='user_bridge_ls']")))
                                dashboard_links = driver.find_element(By.XPATH, "//*[@id='user_bridge_ls']").find_elements(By.TAG_NAME, "li")
                                time.sleep(2)
                                sf.captureScreenshot(driver, dashboard_links[i].text, report_folder)
                        if link_name == "Providers":
                            WebDriverWait(driver, 60).until(EC.presence_of_element_located((By.XPATH, "//*[@class='dataTables_info']")))
                            path = "//*[@id='qt-mt-support-ls']"
                            if sf.check_exists_by_xpath(driver, path):
                                driver.find_element(By.XPATH, "//*[@id='qt-mt-support-ls']//li[1]").click()
                                sf.ajax_preloader_wait(driver)
                                sf.captureScreenshot(driver, "Practice tab " + context, report_folder)

                        if link_name == "Patients":
                            if len(driver.find_elements_by_xpath(locator.xpath_had_er_visit)) != 0:
                                test_case_id += 1
                                ws.append((test_case_id, context_name, 'Presence of Had ER Visit Tab', 'Passed'))
                                driver.find_element(By.XPATH, "//*[@id='patient_panel']//li[2]").click()
                                sf.ajax_preloader_wait(driver)
                                sf.captureScreenshot(driver, "Had ER Visit tab " + context, report_folder)
            except Exception as e:
                print(e)
                traceback.print_exc()
                test_case_id += 1
                ws.append((test_case_id, context_name, 'Navigation to ' + link_name, 'Failed', driver.current_url))
                sf.captureScreenshot(driver, link_name + " failed" + " " + context, report_folder)

            finally:
                links = driver.find_elements_by_xpath(locator.xpath_menubar_Item_Link)
                names = driver.find_elements_by_xpath(locator.xpath_menubar_Item_Link_Name)

        driver.find_element_by_xpath(locator.xpath_side_nav_SlideOut).click()
        WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.XPATH, locator.xpath_registry_Link)))
        time.sleep(2)
        driver.find_element_by_xpath(locator.xpath_registry_Link).click()
        sf.ajax_preloader_wait(driver)

    except Exception as e:
        print(e)
        traceback.print_exc()
        test_case_id += 1
        ws.append((test_case_id, "", 'Menubar Navigation', 'Failed', driver.current_url))
        sf.captureScreenshot(driver, "Menubar failed" + " " + context, report_folder)
    driver.get(main_registry_url)
    sf.ajax_preloader_wait(driver)
    rows = ws.max_row
    cols = ws.max_column
    for i in range(2, rows + 1):
        for j in range(3, cols + 1):
            if ws.cell(i, j).value == 'Passed':
                ws.cell(i, j).fill = PatternFill('solid', fgColor='0FC404')
            elif ws.cell(i, j).value == 'Failed':
                ws.cell(i, j).fill = PatternFill('solid', fgColor='FC0E03')
            elif ws.cell(i, j).value == 'Showing 0 to 0':
                ws.cell(i, j).fill = PatternFill('solid', fgColor='FCC0BB')


def practice_menubar(driver, workbook, logger, run_from, report_folder):
    workbook.create_sheet('Practice Menubar')
    ws = workbook['Practice Menubar']
    main_registry_url = driver.current_url
    if run_from == "Cozeva Support" or run_from == "Limited Cozeva Support" or run_from == "Customer Support" or run_from == "Regional Support":
        # Switching to random Practice name from default set context, main page
        try:
            WebDriverWait(driver, 30).until(
                EC.presence_of_element_located((By.XPATH, locator.xpath_side_nav_SlideOut)))
            driver.find_element_by_xpath(locator.xpath_side_nav_SlideOut).click()
            driver.find_element_by_id("providers-list").click()
            sf.ajax_preloader_wait(driver)
            WebDriverWait(driver, 60).until(
                EC.presence_of_element_located((By.CLASS_NAME, 'tabs')))
            driver.find_element_by_class_name("tabs").find_elements_by_tag_name('li')[0].click()
            sf.ajax_preloader_wait(driver)
            WebDriverWait(driver, 30).until(
                EC.presence_of_element_located((By.ID, "metric-support-prac-ls")))
            list_of_practice_elements = driver.find_element_by_id("metric-support-prac-ls").find_element_by_tag_name("tbody").find_elements_by_tag_name('tr')
            #Practice patient counts
            try:
                list_index = len(list_of_practice_elements)
                zero_prac = [0, 0, 0]
                zero_prac_flag = [0, 0, 0]
                for prac_entry in range(0, list_index):
                    patient_count = list_of_practice_elements[prac_entry].find_elements_by_tag_name("td")[5].text
                    performance = list_of_practice_elements[prac_entry].find_elements_by_tag_name("td")[1].text
                    gap = list_of_practice_elements[prac_entry].find_elements_by_tag_name("td")[2].text
                    if int(patient_count.replace(",", "")) == 0:
                        zero_prac[2] += 1
                    if float(performance.replace("%", "")) == 0:
                        zero_prac[0] += 1
                    if int(gap.replace(",", "")) == 0:
                        zero_prac[1] += 1

                if zero_prac[2] > list_index / 2:
                    print("More than half the praciders have 0 patient counts")
                    prac_url = driver.current_url
                    zero_prac_flag[2] = 1
                if zero_prac[0] > list_index / 2:
                    print("More than half the praciders have 0 patient counts")
                    prac_url = driver.current_url
                    zero_prac_flag[0] = 1
                if zero_prac[1] > list_index / 2:
                    print("More than half the praciders have 0 patient counts")
                    prac_url = driver.current_url
                    zero_prac_flag[1] = 1
            except Exception as e:
                print(e)
                traceback.print_exc()
                zero_prac_flag[2], zero_prac_flag[1], zero_prac_flag[0] = 2, 2, 2


            if len(list_of_practice_elements) > 1:
                selected_practice = list_of_practice_elements[
                    sf.RandomNumberGenerator(len(list_of_practice_elements), 1)[0]].find_element_by_tag_name('a')
            elif len(list_of_practice_elements) == 1:
                selected_practice = list_of_practice_elements[0].find_element_by_tag_name('a')

            global global_search_prac
            global_search_prac = selected_practice.text
            selected_practice.click()
        except Exception as e:
            ws.append(['1', "Attempting to navigate to a random practice", 'Navigation to practice context', 'Failed',
                       "Unable to navigate to a practice. Either the Practice list is unreachable or navigation access is denied", driver.current_url])
            print(e)
            traceback.print_exc()
            return
    elif run_from == "Office Admin Provider Delegate" or run_from == "Provider":
        ws.append(["1", run_from + " Role does not have access to practice Submenus"])
        return
    support_menubar(driver, workbook, ws, logger, run_from, report_folder, "")

    if run_from == "Cozeva Support" or run_from == "Limited Cozeva Support" or run_from == "Customer Support" or run_from == "Regional Support":
        if zero_prac_flag[2] == 1:
            ws.append(["1", "Practice List", "Checking for 0 patient counts", "Failed", "x","More than half the practices on the front page have 0 patients", prac_url])
        elif zero_prac_flag[2] == 0:
            ws.append(["1", "Practice List", "Checking for 0 patient counts", "Passed", "x"])
        elif zero_prac_flag[2] == 2:
            ws.append(["1", "Practice List", "Checking for 0 patient counts", "Failed", "x", "Unable to find practice list"])
        if zero_prac_flag[0] == 1:
            ws.append(["1", "Provider List", "Checking for 0 performance counts", "Failed", "x","More than half the practices on the front page have 0% performance", prac_url])
        elif zero_prac_flag[0] == 0:
            ws.append(["1", "Provider List", "Checking for 0 performance counts", "Passed", "x"])
        elif zero_prac_flag[0] == 2:
            ws.append(["1", "Provider List", "Checking for 0 performance counts", "Failed", "x", "Unable to find practices list"])
        if zero_prac_flag[1] == 1:
            ws.append(["1", "Provider List", "Checking for 0 gap counts", "Failed", "x","More than half the practices on the front page have 0 gap", prac_url])
        elif zero_prac_flag[1] == 0:
            ws.append(["1", "Provider List", "Checking for 0 gap counts", "Passed", "x"])
        elif zero_prac_flag[1] == 2:
            ws.append(["1", "Provider List", "Checking for 0 gap counts", "Failed", "x", "Unable to find practices list"])

    driver.get(main_registry_url)
    sf.ajax_preloader_wait(driver)
    rows = ws.max_row
    cols = ws.max_column
    for i in range(2, rows + 1):
        for j in range(3, cols + 1):
            if ws.cell(i, j).value == 'Passed':
                ws.cell(i, j).fill = PatternFill('solid', fgColor='0FC404')
            elif ws.cell(i, j).value == 'Failed':
                ws.cell(i, j).fill = PatternFill('solid', fgColor='FC0E03')
            elif ws.cell(i, j).value == 'Data table is empty':
                ws.cell(i, j).fill = PatternFill('solid', fgColor='FCC0BB')


def provider_menubar(driver, workbook, logger, run_from, report_folder):
    workbook.create_sheet('Provider Menubar')
    ws = workbook['Provider Menubar']
    main_registry_url = driver.current_url
    if run_from == "Cozeva Support" or run_from == "Limited Cozeva Support" or run_from == "Customer Support" or run_from == "Regional Support" or run_from == "Office Admin Practice Delegate":
        # Switching to random Provider name from default set context, main page
        try:
            WebDriverWait(driver, 30).until(
                EC.presence_of_element_located((By.XPATH, locator.xpath_side_nav_SlideOut)))
            driver.find_element_by_xpath(locator.xpath_side_nav_SlideOut).click()
            driver.find_element_by_id("providers-list").click()
            sf.ajax_preloader_wait(driver)
            WebDriverWait(driver, 60).until(
                EC.presence_of_element_located((By.ID, 'metric-support-prov-ls')))
            list_of_provider_elements = driver.find_element_by_id("metric-support-prov-ls").find_element_by_tag_name("tbody").\
                find_elements_by_tag_name('tr')

            try:
                list_index = len(list_of_provider_elements)
                zero_prov = [0, 0, 0]
                zero_prov_flag = [0, 0, 0]
                zero_prov_flag[0],zero_prov_flag[1],zero_prov_flag[2] = 0, 0, 0
                for prov_entry in range(0, list_index):
                    patient_count = list_of_provider_elements[prov_entry].find_elements_by_tag_name("td")[9].text
                    performance = list_of_provider_elements[prov_entry].find_elements_by_tag_name("td")[1].text
                    gap = list_of_provider_elements[prov_entry].find_elements_by_tag_name("td")[2].text
                    if int(patient_count.replace(",","")) == 0:
                        zero_prov[2] += 1
                    if float(performance.replace("%","")) == 0:
                        zero_prov[0] += 1
                    if int(gap.replace(",","")) == 0:
                        zero_prov[1] += 1

                if zero_prov[2] > list_index/2:
                    print("More than half the providers have 0 patient counts")
                    prov_url = driver.current_url
                    zero_prov_flag[2] = 1
                if zero_prov[0] > list_index/2:
                    print("More than half the providers have 0 patient counts")
                    prov_url = driver.current_url
                    zero_prov_flag[0] = 1
                if zero_prov[1] > list_index/2:
                    print("More than half the providers have 0 patient counts")
                    prov_url = driver.current_url
                    zero_prov_flag[1] = 1
            except Exception as e:
                traceback.print_exc()
                print(e)
                zero_prov_flag[2],zero_prov_flag[1],zero_prov_flag[0] = 2,2,2

            if len(list_of_provider_elements) > 1:
                selected_provider = list_of_provider_elements[
                    sf.RandomNumberGenerator(len(list_of_provider_elements), 1)[0]].find_elements_by_tag_name('a')[1]
            elif len(list_of_provider_elements) == 1:
                selected_provider = list_of_provider_elements[0].find_elements_by_tag_name('a')[1]

            global global_search_prov
            global_search_prov = selected_provider.text
            selected_provider.click()
        except Exception as e:
            ws.append(['1', "Attempting to navigate to a random provider", 'Navigation to provider context', 'Failed',
                       "Unable to navigate to a provider. Either the Provider list is unreachable or navigation access is denied", driver.current_url])
            driver.get(main_registry_url)
            sf.ajax_preloader_wait(driver)
            WebDriverWait(driver, 30).until(
                EC.presence_of_element_located((By.XPATH, locator.xpath_side_nav_SlideOut)))
            print(e)
            traceback.print_exc()
            return

    support_menubar(driver, workbook, ws, logger, run_from, report_folder, "")

    # if run_from == "Cozeva Support" or run_from == "Limited Cozeva Support" or run_from == "Customer Support" or run_from == "Regional Support":
    #     driver.find_element_by_xpath(locator.xpath_side_nav_SlideOut).click()
    #     driver.find_element_by_id("home").click()
    #     sf.ajax_preloader_wait(driver)

    if run_from == "Cozeva Support" or run_from == "Limited Cozeva Support" or run_from == "Customer Support" or run_from == "Regional Support" or run_from == "Office Admin Practice Delegate":
        if zero_prov_flag[2] == 1:
            ws.append(["1", "Provider List", "Checking for 0 patient counts", "Failed", "x","More than half the provider on the front page have 0 patients", prov_url])
        elif zero_prov_flag[2] == 0:
            ws.append(["1", "Provider List", "Checking for 0 patient counts", "Passed", "x"])
        elif zero_prov_flag[2] == 2:
            ws.append(["1", "Provider List", "Checking for 0 patient counts", "Failed", "x", "Unable to find provider list"])
        if zero_prov_flag[0] == 1:
            ws.append(["1", "Provider List", "Checking for 0 performance counts", "Failed", "x","More than half the provider on the front page have 0% performance", prov_url])
        elif zero_prov_flag[0] == 0:
            ws.append(["1", "Provider List", "Checking for 0 performance counts", "Passed", "x"])
        elif zero_prov_flag[0] == 2:
            ws.append(["1", "Provider List", "Checking for 0 performance counts", "Failed", "x", "Unable to find provider list"])
        if zero_prov_flag[1] == 1:
            ws.append(["1", "Provider List", "Checking for 0 gap counts", "Failed", "x","More than half the provider on the front page have 0 gap", prov_url])
        elif zero_prov_flag[1] == 0:
            ws.append(["1", "Provider List", "Checking for 0 gap counts", "Passed", "x"])
        elif zero_prov_flag[1] == 2:
            ws.append(["1", "Provider List", "Checking for 0 gap counts", "Failed", "x", "Unable to find provider list"])



    driver.get(main_registry_url)
    sf.ajax_preloader_wait(driver)
    WebDriverWait(driver, 30).until(
        EC.presence_of_element_located((By.XPATH, locator.xpath_side_nav_SlideOut)))
    rows = ws.max_row
    cols = ws.max_column
    for i in range(2, rows + 1):
        for j in range(3, cols + 1):
            if ws.cell(i, j).value == 'Passed':
                ws.cell(i, j).fill = PatternFill('solid', fgColor='0FC404')
            elif ws.cell(i, j).value == 'Failed':
                ws.cell(i, j).fill = PatternFill('solid', fgColor='FC0E03')
            elif ws.cell(i, j).value == 'Data table is empty':
                ws.cell(i, j).fill = PatternFill('solid', fgColor='FCC0BB')


def patient_dashboard(driver, workbook, logger, run_from, report_folder):
    workbook.create_sheet('Patient Dashboard')
    ws = workbook['Patient Dashboard']
    ws.append(['ID', 'Context', 'Scenario', 'Status', 'Time Taken', 'Comments'])
    header_font = Font(color='FFFFFF', bold=False, size=12)
    header_cell_color = PatternFill('solid', fgColor='030303')
    ws['A1'].font = header_font
    ws['A1'].fill = header_cell_color
    ws['B1'].font = header_font
    ws['B1'].fill = header_cell_color
    ws['C1'].font = header_font
    ws['C1'].fill = header_cell_color
    ws['D1'].font = header_font
    ws['D1'].fill = header_cell_color
    ws.name = "Arial"
    test_case_id = 1

    def hoverCheck(driver, ws, run_from, Pcp_hover, test_case_id):
        x = 1

    # From Starting point Registry, navigate to a random patient of a random metric
    main_registry_url = driver.current_url
    window_switched = 0
    try:
        WebDriverWait(driver, 30).until(
            EC.presence_of_element_located((By.ID, "registry_body")))
        metrics = driver.find_element_by_id("registry_body").find_elements_by_tag_name('li')
        print("Provider Registry metrics loaded into a variable")
        percent = '0.00'
        loop_count = 0
        while percent == '0.00' or percent == '0.00 %' and loop_count < 20:
            if len(metrics) > 1:
                selectedMetric = metrics[sf.RandomNumberGenerator(len(metrics), 1)[0]]
                percent = selectedMetric.find_element_by_class_name('percent').text
            else:
                selectedMetric = metrics[0]
                percent = selectedMetric.find_element_by_class_name('percent').text
            if loop_count > 18:
                break
            else:
                loop_count += 1
        print("Found a Suitable Metric to click on")
        print("Attempting to click on " + selectedMetric.text)
        driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", selectedMetric)
        #selectedMetric.click()
        sf.action_click(selectedMetric, driver)
        print("Click Performed")
        sf.ajax_preloader_wait(driver)
        WebDriverWait(driver, 30).until(
            EC.presence_of_element_located((By.CLASS_NAME, 'tabs')))
        if run_from == "Cozeva Support" or run_from == "Customer Support" or run_from == "Regional Support" or run_from == "Limited Cozeva Support":
            WebDriverWait(driver, 30).until(
                EC.presence_of_element_located((By.CLASS_NAME, 'tabs')))
            tab_list = driver.find_element_by_class_name('tabs').find_elements(By.TAG_NAME, "li")
            for tab in tab_list:
                if "Patients" in tab.text:
                    tab.click()
                    break
            sf.ajax_preloader_wait(driver)
            if len(driver.find_elements_by_class_name('dt_tag_value')) > 0:
                driver.find_element_by_class_name('dt_tag_close').click()
                sf.ajax_preloader_wait(driver)
            WebDriverWait(driver, 30).until(
                EC.presence_of_element_located((By.ID, "metric-support-pat-ls")))
            sf.captureScreenshot(driver, "MSPL " + driver.find_element(By.XPATH, "//span[@class='specific_most']").text, report_folder)
            patients = driver.find_element_by_id("metric-support-pat-ls").find_elements_by_tag_name('tr')
            patients[sf.RandomNumberGenerator(len(patients), 1)[0]].find_element_by_class_name('pat_name').click()
        elif run_from == "Office Admin Practice Delegate":
            WebDriverWait(driver, 30).until(
                EC.presence_of_element_located((By.CLASS_NAME, 'tabs')))
            driver.find_element_by_class_name('tabs').find_elements_by_class_name('tab')[1].click()
            sf.ajax_preloader_wait(driver)
            if len(driver.find_elements_by_class_name('dt_tag_value')) > 0:
                driver.find_element_by_class_name('dt_tag_close').click()
                sf.ajax_preloader_wait(driver)
            WebDriverWait(driver, 30).until(
                EC.presence_of_element_located((By.ID, "metric-support-pat-ls")))
            sf.captureScreenshot(driver, "MSPL " + driver.find_element(By.XPATH, "//span[@class='specific_most']").text, report_folder)
            patients = driver.find_element_by_id("metric-support-pat-ls").find_elements_by_tag_name('tr')
            patients[sf.RandomNumberGenerator(len(patients), 1)[0]].find_element_by_class_name('pat_name').click()
        elif run_from == "Office Admin Provider Delegate" or run_from == "Provider":
            WebDriverWait(driver, 30).until(
                EC.presence_of_element_located((By.CLASS_NAME, 'tabs')))
            if len(driver.find_elements_by_class_name('dt_tag_value')) > 0:
                driver.find_element_by_class_name('dt_tag_close').click()
                sf.ajax_preloader_wait(driver)
            WebDriverWait(driver, 30).until(
                EC.presence_of_element_located((By.ID, "quality_registry_list")))
            sf.captureScreenshot(driver, "MSPL " + driver.find_element(By.XPATH, "//span[@class='specific_most']").text, report_folder)
            patients = driver.find_element_by_id("quality_registry_list").find_elements_by_tag_name('tr')
            patients[sf.RandomNumberGenerator(len(patients), 1)[0]].find_element_by_class_name('pat_name').click()

        driver.switch_to.window(driver.window_handles[1])
        window_switched = 1
        sf.ajax_preloader_wait(driver)
        WebDriverWait(driver, 30).until(
            EC.presence_of_element_located((By.XPATH, locator.xpath_cozeva_Id)))
        patient_id = driver.find_element_by_xpath(locator.xpath_cozeva_Id).text
        global global_search_pat
        global_search_pat = patient_id
        current_url = driver.current_url
        access_message = sf.CheckAccessDenied(current_url)

        if access_message == 1:
            print("Access Denied found!")
            # logger.critical("Access Denied found!")
            test_case_id += 1
            ws.append((test_case_id, patient_id, 'Navigation to dashboard page',
                       'Failed', 'x', 'Access Denied', driver.current_url))
            sf.captureScreenshot(driver, "Patient tab access denied " + run_from, report_folder)

        else:
            print("Access Check done!")
            # logger.info("Access Check done!")
            error_message = sf.CheckErrorMessage(driver)

            if error_message == 1:
                print("Error toast message is displayed")
                # logger.critical("ERROR TOAST MESSAGE IS DISPLAYED!")
                test_case_id += 1
                ws.append \
                    ((test_case_id, patient_id, 'Navigation to dashboard page ',
                      'Failed', 'x', 'Error toast message is displayed', driver.current_url))
                sf.captureScreenshot(driver, "Patient tab error toast " + run_from, report_folder)

            else:
                measure_count_dashboard = len \
                    (driver.find_elements_by_xpath("//tbody[@class='measurement-body careops-new']/tr"))
                test_case_id += 1
                ws.append((test_case_id, patient_id, 'Navigation to dashboard page',
                           'Passed', 'x', 'Measures count in dashboard: ' + str(measure_count_dashboard)))
                sf.captureScreenshot(driver, "Patient tab navigated to " + run_from, report_folder)
                logger.info(patient_id + ": Navigated to patient dashboard.")
                """ **** PCP INFO BLOCK **** """
                try:
                    Pcp_Name = driver.find_element_by_id("pcp_name").text
                    Pcp_Webelement = driver.find_element_by_id("pcp_name")
                    Pcp_hover = Pcp_Webelement.get_attribute("data-tooltip")
                    # Pcp_Name = "-"
                    # Pcp_hover = "N/A, N/A, N/A"

                    if Pcp_Name == '-':
                        test_case_id += 1
                        ws.append((test_case_id, patient_id, 'PCP Name',
                                   'Failed', 'x', "PCP Name is Blank", driver.current_url))
                    elif Pcp_Name == "N/A":
                        test_case_id += 1
                        ws.append((test_case_id, patient_id, 'PCP Name',
                                   'Failed', 'x', "PCP Name is NA", driver.current_url))
                    else:
                        test_case_id += 1
                        ws.append((test_case_id, patient_id, 'PCP Name',
                                   'Passed', 'x', Pcp_Name))

                    if Pcp_hover == "N/A, N/A, No Practice":
                        test_case_id += 1
                        ws.append((test_case_id, patient_id, 'PCP Attribution on hover',
                                   'Failed', 'x', "PCP does not have Region/Panel Attribution", driver.current_url))
                    elif Pcp_hover == "N/A, N/A, N/A":
                        test_case_id += 1
                        ws.append((test_case_id, patient_id, 'PCP Attribution on hover',
                                   'Failed', 'x', "PCP does not have any attribution", driver.current_url))
                    else:
                        test_case_id += 1
                        ws.append((test_case_id, patient_id, 'PCP Attribution on hover',
                                   'Passed', 'x', Pcp_hover))
                        # function for hovercheck
                        hoverCheck(driver, ws, run_from, Pcp_hover, test_case_id)
                except Exception as e:
                    print(e)
                    traceback.print_exc()
                    test_case_id += 1
                    ws.append((test_case_id, patient_id, 'PCP hover',
                               'Failed', 'x', "PCP Name is not present/Not interactable", driver.current_url))

                # Aspy Edit ------------------------------------------------------------------------------------
                # '''
                """ **** PATIENT MENUBAR NAVIGATION **** """
                WebDriverWait(driver, 30).until(EC.presence_of_element_located(
                    (By.XPATH, locator.xpath_patient_Header_Dropdown_Arrow)))
                driver.find_element_by_xpath(locator.xpath_patient_Header_Dropdown_Arrow).click()
                WebDriverWait(driver, 30).until(
                    EC.presence_of_element_located((By.XPATH, locator.xpath_patient_History)))
                history_link = driver.find_element_by_xpath(locator.xpath_patient_History)
                ActionChains(driver).move_to_element(history_link).perform()
                patient_history_items = driver.find_elements_by_xpath(
                    locator.xpath_patient_History_Item_Link)
                driver.find_element_by_xpath(locator.xpath_patient_Header_Dropdown_Arrow).click()
                for item_counter in range(len(patient_history_items)):
                    WebDriverWait(driver, 30).until(
                        EC.presence_of_element_located(
                            (By.XPATH, locator.xpath_patient_Header_Dropdown_Arrow)))
                    driver.find_element_by_xpath(locator.xpath_patient_Header_Dropdown_Arrow).click()
                    WebDriverWait(driver, 30).until(
                        EC.presence_of_element_located((By.XPATH, locator.xpath_patient_History)))
                    history_link = driver.find_element_by_xpath(locator.xpath_patient_History)
                    time.sleep(0.5)
                    ActionChains(driver).move_to_element(history_link).perform()
                    item_name = patient_history_items[item_counter].text
                    print(item_name)
                    time.sleep(0.5)
                    patient_history_items[item_counter].click()
                    start_time = time.perf_counter()
                    sf.ajax_preloader_wait(driver)
                    total_time = time.perf_counter() - start_time
                    current_url = driver.current_url
                    access_message = sf.CheckAccessDenied(current_url)
                    sf.captureScreenshot(driver, "Patient tab " + item_name + " " + run_from, report_folder)
                    if access_message == 1:
                        print("Access Denied found!")
                        # logger.critical("Access Denied found!")
                        test_case_id += 1
                        ws.append((test_case_id,
                                   patient_id, 'Navigation to  ' + item_name,
                                   'Failed', 'x', 'Access Denied', driver.current_url))

                    else:
                        print("Access Check done!")
                        # logger.info("Access Check done!")
                        error_message = sf.CheckErrorMessage(driver)

                        if error_message == 1:
                            print("Error toast message is displayed")
                            # logger.critical("ERROR TOAST MESSAGE IS DISPLAYED!")
                            test_case_id += 1
                            ws.append((test_case_id,
                                       patient_id, 'Navigation to  ' + item_name,
                                       'Failed', 'x', 'Error toast message is displayed', driver.current_url))

                        else:

                            if len(driver.find_elements_by_xpath(locator.xpath_data_Table_Row)) != 0:
                                if len(driver.find_elements_by_xpath(
                                        locator.xpath_empty_Data_Table_Row)) != 0:
                                    print("Data table is empty")
                                    test_case_id += 1
                                    ws.append((test_case_id, patient_id, "Navigation to" + item_name, 'Passed',
                                               round(total_time, sigfigs=3),
                                               'Data table is empty'))
                                    logger.info("Navigated to:  " + item_name)
                                else:
                                    table_row_count = len(
                                        driver.find_elements_by_xpath(locator.xpath_data_Table_Row))
                                    print(table_row_count)
                                    test_case_id += 1
                                    ws.append((test_case_id, patient_id, "Navigation to" + item_name, 'Passed',
                                               round(total_time, sigfigs=3),
                                               'Data table row count in the first page: ' + str
                                               (table_row_count)))
                                    logger.info("Navigated to: " + item_name)

                            patient_history_items = driver.find_elements_by_xpath(
                                locator.xpath_patient_History_Item_Link)

                """ **** PATIENT INFO **** """
                driver.find_element_by_xpath(locator.xpath_patient_Header_Dropdown_Arrow).click()
                WebDriverWait(driver, 30).until(
                    EC.presence_of_element_located((By.XPATH, locator.xpath_patient_Info_Link)))
                driver.find_element_by_xpath(locator.xpath_patient_Info_Link).click()
                start_time = time.perf_counter()
                sf.ajax_preloader_wait(driver)
                total_time = time.perf_counter() - start_time
                current_url = driver.current_url
                access_message = sf.CheckAccessDenied(current_url)
                sf.captureScreenshot(driver, "Patient information tab " + run_from, report_folder)
                if access_message == 1:
                    print("Access Denied found!")
                    # logger.critical("Access Denied found!")
                    test_case_id += 1
                    ws.append(
                        (test_case_id, patient_id, 'Navigation to Patient Info page',
                         'Failed', total_time, 'Access Denied', driver.current_url))

                else:
                    print("Access Check done!")
                    # logger.info("Access Check done!")
                    error_message = sf.CheckErrorMessage(driver)

                    if error_message == 1:
                        print("Error toast message is displayed")
                        # logger.critical("ERROR TOAST MESSAGE IS DISPLAYED!")
                        test_case_id += 1
                        ws.append((test_case_id,
                                   patient_id, 'Navigation to Patient Info page',
                                   'Failed', total_time, 'Error toast message is displayed', driver.current_url))

                    else:
                        """ **** COVERAGE **** """
                        WebDriverWait(driver, 30).until(
                            EC.presence_of_element_located
                            ((By.XPATH, locator.xpath_patient_Info_Coverage_Link)))
                        driver.find_element_by_xpath(locator.xpath_patient_Info_Coverage_Link).click()
                        start_time = time.perf_counter()
                        sf.ajax_preloader_wait(driver)
                        total_time = time.perf_counter() - start_time
                        time.sleep(1)
                        sf.captureScreenshot(driver, "Patient coverage " + run_from, report_folder)
                        coverage_number = len \
                            (driver.find_elements_by_xpath("//table[@id='patient_info_payment_table']"))
                        if coverage_number != 0:
                            test_case_id += 1
                            ws.append((test_case_id, patient_id, "Patient Info-->Coverage", 'Passed', total_time,
                                       'Number of Coverage card(s): ' + str(coverage_number)))
                        elif coverage_number == 0:
                            test_case_id += 1
                            ws.append((test_case_id, patient_id + ": Patient Info-->Coverage", 'Failed', total_time,
                                       'Number of Coverage card(s): ' + str(coverage_number), driver.current_url))

                        """ **** PATIENT INFO **** """
                        WebDriverWait(driver, 30).until(
                            EC.presence_of_element_located(
                                (By.XPATH, locator.xpath_patient_Info_Care_Team_Link)))
                        driver.find_element_by_xpath(locator.xpath_patient_Info_Care_Team_Link).click()
                        time.sleep(1)
                        sf.captureScreenshot(driver, "Patient care team " + run_from, report_folder)
                        careteam_provider_number = len(driver.find_elements_by_xpath("//div[@class='mlm']/div"))
                        if careteam_provider_number != 0:
                            test_case_id += 1
                            ws.append((test_case_id, total_time, "Patient Info-->Care Team", 'Passed', 'x'
                                                                                                       'Number of Providers present in Care Team: ' + str
                                       (careteam_provider_number)))
                        elif coverage_number == 0:
                            test_case_id += 1
                            ws.append((test_case_id, patient_id + ": Patient Info-->Care Team", 'Failed',
                                       'Number of Providers present in Care Team: ' + str
                                       (careteam_provider_number), driver.current_url))
                        logger.info("Navigated to Patient Info.")

    except Exception as e:
        print(e)
        traceback.print_exc()
        traceback.print_exc()

        ws.append(['1', 'Unable to navigate to a patient from random metric'])
    finally:
        if window_switched == 1:
            driver.close()
            driver.switch_to.window(driver.window_handles[0])

    driver.get(main_registry_url)
    sf.ajax_preloader_wait(driver)
    rows = ws.max_row
    cols = ws.max_column
    for i in range(2, rows + 1):
        for j in range(3, cols + 1):
            if ws.cell(i, j).value == 'Passed':
                ws.cell(i, j).fill = PatternFill('solid', fgColor='0FC404')
            elif ws.cell(i, j).value == 'Failed':
                ws.cell(i, j).fill = PatternFill('solid', fgColor='FC0E03')
            elif ws.cell(i, j).value == 'Data table is empty':
                ws.cell(i, j).fill = PatternFill('solid', fgColor='FCC0BB')


def provider_registry(driver, workbook, logger, run_from, report_folder):
    workbook.create_sheet('Provider Registry')
    ws = workbook['Provider Registry']

    ws.append(['ID', 'Context', 'Scenario', 'Status', 'Time Taken', 'Comments'])
    header_font = Font(color='FFFFFF', bold=False, size=12)
    header_cell_color = PatternFill('solid', fgColor='030303')
    ws['A1'].font = header_font
    ws['A1'].fill = header_cell_color
    ws['B1'].font = header_font
    ws['B1'].fill = header_cell_color
    ws['C1'].font = header_font
    ws['C1'].fill = header_cell_color
    ws['D1'].font = header_font
    ws['D1'].fill = header_cell_color
    ws['E1'].font = header_font
    ws['E1'].fill = header_cell_color
    ws['F1'].font = header_font
    ws['F1'].fill = header_cell_color
    ws.name = "Arial"
    test_case_id = 1

    # checking for default context and then navigating to a random provider registry
    main_registry_url = driver.current_url
    if run_from == "Cozeva Support" or run_from == "Limited Cozeva Support" or run_from == "Customer Support" or run_from == "Regional Support" or run_from == "Office Admin Practice Delegate":
        try:
            print("1")
            driver.find_element_by_xpath(locator.xpath_side_nav_SlideOut).click()
            time.sleep(0.5)
            driver.find_element_by_id("providers-list").click()
            sf.ajax_preloader_wait(driver)
            WebDriverWait(driver, 30).until(
                EC.presence_of_element_located((By.ID, "metric-support-prov-ls")))
            sf.captureScreenshot(driver, "Provider List " + run_from + driver.find_element(By.XPATH, "//span[@class='specific_most']").text, report_folder)
            list_of_provider_elements = driver.find_element_by_id("metric-support-prov-ls").find_elements_by_tag_name(
                'tr')
            global global_search_prov
            if len(list_of_provider_elements) > 1:
                global_search_prov = list_of_provider_elements[
                    sf.RandomNumberGenerator(len(list_of_provider_elements), 1)[0]].find_elements_by_tag_name('a')[
                    1].text
            elif len(list_of_provider_elements) == 1:
                global_search_prov = list_of_provider_elements[0].find_elements_by_tag_name('a')[1].text
            if len(list_of_provider_elements) > 1:
                list_of_provider_elements[
                    sf.RandomNumberGenerator(len(list_of_provider_elements), 1)[0]].find_elements_by_tag_name('a')[
                    1].click()
            elif len(list_of_provider_elements) == 1:
                list_of_provider_elements[0].find_elements_by_tag_name('a')[1].click()

        except Exception as e:
            ws.append([test_case_id, "Attempting to navigate to a random provider", 'Navigation to provider context',
                       'Failed', 'x',
                       "Unable to navigate to a provider. Either the Provider list is unreachable or navigation access is denied", driver.current_url])
            test_case_id += 1
            driver.get(main_registry_url)
            sf.ajax_preloader_wait(driver)
            sf.captureScreenshot(driver, "Provider navigation failed "+ driver.find_element(By.XPATH, "//span[@class='specific_most']").text, report_folder)
            WebDriverWait(driver, 30).until(
                EC.presence_of_element_located((By.XPATH, locator.xpath_filter_measure_list)))
            print(e)
            traceback.print_exc()
            return

    # Store registry url for back navigation
    #Newly added
    sf.ajax_preloader_wait(driver)
    driver.find_element(By.XPATH, "//*[@id='qt-filter-label']").click()
    time.sleep(1)
    flag = 0
    Comment = ""
    target_url = ""
    try:
        LOB_list = driver.find_element(By.XPATH, "//*[@id='filter-lob']").find_elements(By.TAG_NAME, "li")
        temp = 0
        count = 0
        for i in range(0, len(LOB_list)):
            if count <= len(LOB_list):
                LOB_list[i].click()
                name = LOB_list[i].text
                driver.find_element(By.XPATH, "//*[@id='reg-filter-apply']").click()
                sf.ajax_preloader_wait(driver)
                if driver.find_element(By.XPATH, "//*[@id='conti_enroll']").is_selected():
                    driver.find_element(By.XPATH, "//*[@class='cont_disc_toggle']").click()
                time.sleep(0.5)
                current_context = driver.find_element(By.XPATH, "//span[@class='specific_most']").text
                sf.captureScreenshot(driver, name + " " + current_context, report_folder)
                patient_count = driver.find_element(By.XPATH, "//div[contains(text(), 'Patients')]/../div[2]").text
                patient_count = patient_count.replace(",", "")
                print(patient_count)
                if int(patient_count) > int(temp):
                    temp = patient_count
                    target_url = driver.current_url
                home_url = driver.current_url
                # driver.find_element(By.XPATH, "//a[@data-target='qt-reg-nav-filters']").click()
                # time.sleep(0.5)
                # option = driver.find_elements(By.XPATH, "//input[@class='select-dropdown dropdown-trigger']")[0]
                # option.click()
                # lists = driver.find_elements(By.XPATH, "//ul[@class='dropdown-content select-dropdown']")[0]
                # options = lists.find_elements(By.TAG_NAME, 'li')
                # for option in options:
                #     if option.text == "Denominator":
                #         option.click()
                # apply_btn = driver.find_element(By.XPATH, "//button[@id='qt-apply-search']")
                # driver.execute_script("arguments[0].scrollIntoView();", apply_btn)
                # apply_btn.click()
                # time.sleep(0.5)
                try:
                    metric = driver.find_element(By.XPATH, "//*[@id='registry_body']").find_elements(By.TAG_NAME, "li")[0]
                    metric.click()
                    sf.ajax_preloader_wait(driver)
                    WebDriverWait(driver, 30).until(EC.presence_of_element_located((By.XPATH, "//*[@class='tabs']")))
                    if len(driver.find_elements_by_class_name('dt_tag_value')) > 0:
                        driver.find_element_by_class_name('dt_tag_close').click()
                        sf.ajax_preloader_wait(driver)
                    time.sleep(3)
                    WebDriverWait(driver, 30).until(EC.presence_of_element_located((By.ID, "quality_registry_list")))
                    sf.captureScreenshot(driver, current_context + " MSPL " + name, report_folder)
                    ws.append([test_case_id, current_context, 'Navigation to MSPL' + driver.find_element(By.XPATH, "//*[@class='ch metric_specific_patient_list_title']").text,
                               'Passed', 'x', name, driver.current_url])
                except Exception as e:
                    ws.append([test_case_id, current_context, 'Navigation to MSPL failed', 'Failed', 'x', name, driver.current_url])
                    sf.captureScreenshot(driver, "MSPL " + name, report_folder)
                driver.get(home_url)
                sf.ajax_preloader_wait(driver)
                driver.find_element(By.XPATH, "//*[@id='qt-filter-label']").click()
                time.sleep(1)
                LOB_list = driver.find_element(By.XPATH, "//*[@id='filter-lob']").find_elements(By.TAG_NAME, "li")
                count += 1
            else:
                driver.get(target_url)
                sf.ajax_preloader_wait(driver)
    except Exception as e:
        print("No LOBs for the provider")
        print(e)
        flag = 1
    WebDriverWait(driver, 30).until(EC.presence_of_element_located((By.XPATH, "//*[@id='quality_registry']")))
    patient_count = driver.find_element(By.XPATH, "//div[contains(text(), 'Patients')]/../div[2]").text
    patient_count = patient_count.replace(",", "")
    if flag == 1:
        Comment = "Provider does not have any patients in current registry and has only one LOB"
    else:
        Comment = "Provider does not have any patients in current registry"
    if patient_count == '0':
        ws.append([test_case_id, "Attempting to navigate to a random provider", 'Navigation to provider context',
                   'Failed', 'x',
                   Comment,
                   driver.current_url])
    registry_url = driver.current_url
    # Navigation test 1 : Navigation to patient context through providers patients tab
    try:
        sf.ajax_preloader_wait(driver)
        current_context = driver.find_element_by_class_name("current_context").text
        print('1.5')
        driver.find_element_by_xpath(locator.xpath_side_nav_SlideOut).click()
        print('2')
        driver.find_element_by_id("all_patients_tab").click()
        start_time = time.perf_counter()
        sf.ajax_preloader_wait(driver)
        total_time = time.perf_counter() - start_time
        current_url = driver.current_url
        access_message = sf.CheckAccessDenied(current_url)
        sf.captureScreenshot(driver, "Provider All patients " + driver.find_element(By.XPATH, "//span[@class='specific_most']").text, report_folder)

        if access_message == 1:
            print("Access Denied found!")
            # logger.critical("Access Denied found!")
            test_case_id += 1
            ws.append((test_case_id, current_context, 'Navigation to all patients tab',
                       'Failed', 'x', 'Access Denied', driver.current_url))

        else:
            print("Access Check done!")
            # logger.info("Access Check done!")
            error_message = sf.CheckErrorMessage(driver)

            if error_message == 1:
                print("Error toast message is displayed")
                # logger.critical("ERROR TOAST MESSAGE IS DISPLAYED!")
                test_case_id += 1
                ws.append \
                    ((test_case_id, current_context, 'Navigation to all patients tab ',
                      'Failed', 'x', 'Error toast message is displayed', driver.current_url))

            else:
                test_case_id += 1
                ws.append((test_case_id, current_context, 'Navigation to all patients tab',
                           'Passed', total_time))
                logger.info(current_context + ": Navigated to all patients tab.")
                # Now navigating to a patient dashboard at random
                WebDriverWait(driver, 30).until(
                    EC.presence_of_element_located((By.ID, "all_patients")))
                patient_elements = driver.find_element_by_id("all_patients").find_element_by_tag_name(
                    'tbody').find_elements_by_tag_name('tr')
                if len(patient_elements) > 1:
                    patient_elements[sf.RandomNumberGenerator(len(patient_elements), 1)[0]].find_element_by_class_name(
                        'pat_name').click()
                else:
                    patient_elements[0].find_element_by_class_name('pat_name').click()
                driver.switch_to.window(driver.window_handles[1])
                start_time = time.perf_counter()
                sf.ajax_preloader_wait(driver)
                total_time = time.perf_counter() - start_time
                WebDriverWait(driver, 30).until(
                    EC.presence_of_element_located((By.XPATH, locator.xpath_cozeva_Id)))
                sf.captureScreenshot(driver, "Patient Dashboard " + run_from, report_folder)
                patient_id = driver.find_element_by_xpath(locator.xpath_cozeva_Id).text
                current_url = driver.current_url
                access_message = sf.CheckAccessDenied(current_url)

                if access_message == 1:
                    print("Access Denied found!")
                    # logger.critical("Access Denied found!")
                    test_case_id += 1
                    ws.append((test_case_id, patient_id, 'Navigation to dashboard page',
                               'Failed', 'x', 'Access Denied', driver.current_url))


                else:
                    print("Access Check done!")
                    # logger.info("Access Check done!")
                    error_message = sf.CheckErrorMessage(driver)

                    if error_message == 1:
                        print("Error toast message is displayed")
                        # logger.critical("ERROR TOAST MESSAGE IS DISPLAYED!")
                        test_case_id += 1
                        ws.append \
                            ((test_case_id, patient_id, 'Navigation to dashboard page ',
                              'Failed', 'x', 'Error toast message is displayed', driver.current_url))

                    else:
                        measure_count_dashboard = len \
                            (driver.find_elements_by_xpath("//tbody[@class='measurement-body careops-new']/tr"))
                        test_case_id += 1
                        ws.append((test_case_id, patient_id, 'Navigation to dashboard page',
                                   'Passed', total_time,
                                   'Measures count in dashboard: ' + str(measure_count_dashboard)))
                        logger.info(patient_id + ": Navigated to patient dashboard.")
                        if sf.check_exists_by_class(driver, 'primary_val'):
                            test_case_id += 1
                            ws.append([test_case_id, patient_id, 'CareOps count present', 'Passed', 'x',
                                       'Count: ' + driver.find_element_by_class_name("primary_val").text])
                        else:
                            test_case_id += 1
                            ws.append([test_case_id, patient_id, 'CareOps count present', 'Failed', 'x',
                                       'Careops count not present', driver.current_url])
                driver.close()
                driver.switch_to.window(driver.window_handles[0])
        driver.get(registry_url)
        sf.ajax_preloader_wait(driver)
        WebDriverWait(driver, 30).until(
            EC.presence_of_element_located((By.XPATH, locator.xpath_filter_measure_list)))

    except Exception as e:
        ws.append([test_case_id, "Provider registry navigation",
                   "Navigation to patient context through providers patients tab", 'Failed', 'x',
                   'Unable to navigate to patients list/Patient dashboard', driver.current_url])
        test_case_id += 1
        print(e)
        traceback.print_exc()
        driver.get(registry_url)
        sf.ajax_preloader_wait(driver)
        WebDriverWait(driver, 30).until(
            EC.presence_of_element_located((By.XPATH, locator.xpath_filter_measure_list)))

    # Navigation test 2: Navigation to patient context through providers MSPL
    try:
        current_context = driver.find_element_by_class_name("current_context").text
        WebDriverWait(driver, 30).until(
            EC.presence_of_element_located((By.ID, "registry_body")))
        metrics = driver.find_element_by_id("registry_body").find_elements_by_tag_name('li')
        percent = '0.00'
        loop_counter = 0
        if driver.find_element(By.XPATH, "//*[@id='conti_enroll']").is_selected():
            driver.find_element(By.XPATH, "//*[@class='cont_disc_toggle']").click()
        selected_metric_name = ""
        driver.find_element(By.XPATH, "//a[@data-target='qt-reg-nav-filters']").click()
        time.sleep(0.5)
        option = driver.find_elements(By.XPATH, "//input[@class='select-dropdown dropdown-trigger']")[0]
        option.click()
        lists = driver.find_elements(By.XPATH, "//ul[@class='dropdown-content select-dropdown']")[0]
        options = lists.find_elements(By.TAG_NAME, 'li')
        for option in options:
            if option.text == "Numerator":
                option.click()
        apply_btn = driver.find_element(By.XPATH, "//button[@id='qt-apply-search']")
        driver.execute_script("arguments[0].scrollIntoView();", apply_btn)
        apply_btn.click()
        time.sleep(0.5)
        selectedMetric = driver.find_element(By.XPATH, "//*[@id='registry_body']/ul").find_elements(By.TAG_NAME, "li")
        selected_metric_name = selectedMetric[0].find_element(By.XPATH, "//*[@class='met-name']").text
        selectedMetric[0].click()
        start_time = time.perf_counter()
        sf.ajax_preloader_wait(driver)
        total_time = time.perf_counter() - start_time
        test_case_id += 1
        sf.captureScreenshot(driver, "Provider MSPL " + driver.find_element(By.XPATH, "//span[@class='specific_most']").text, report_folder)
        ws.append([test_case_id, current_context, "Navigation to MSPL of " + selected_metric_name, 'Passed', total_time])
        window_switched = 0
        try:
            patient_id = "Could not Fetch"
            WebDriverWait(driver, 30).until(
                EC.presence_of_element_located((By.ID, "quality_registry_list")))
            if len(driver.find_elements_by_class_name('dt_tag_value')) > 0:
                driver.find_element_by_class_name('dt_tag_close').click()
                sf.ajax_preloader_wait(driver)
            time.sleep(3)
            sf.ajax_preloader_wait(driver)
            WebDriverWait(driver, 30).until(
                EC.presence_of_element_located((By.ID, "quality_registry_list")))
            patients = driver.find_element_by_id("quality_registry_list").find_element_by_tag_name(
                'tbody').find_elements_by_tag_name('tr')
            if len(patients) > 1:
                patients[sf.RandomNumberGenerator(len(patients), 1)[0]].find_element_by_class_name('pat_name').click()
            elif len(patients) == 1:
                if patients[0].text == "No data available":
                    raise Exception("No patients available in this MSPL")
                patients[0].find_element_by_class_name('pat_name').click()
            else:
                ws.append(['Mspl has no patients: ' + global_search_prov])
            #Add clause to check for no patients
            driver.switch_to.window(driver.window_handles[1])
            window_switched = 1
            start_time = time.perf_counter()
            sf.ajax_preloader_wait(driver)
            sf.captureScreenshot(driver, "Patient Dashboard from MSPL " + run_from, report_folder)
            WebDriverWait(driver, 30).until(
                EC.presence_of_element_located((By.XPATH, locator.xpath_cozeva_Id)))
            total_time = time.perf_counter() - start_time

            patient_id = driver.find_element_by_xpath(locator.xpath_cozeva_Id).text

            measure_count_dashboard = len \
                (driver.find_elements_by_xpath("//tbody[@class='measurement-body careops-new']/tr"))
            test_case_id += 1
            ws.append((test_case_id, patient_id, 'Navigation to dashboard page',
                       'Passed', total_time, 'Measures count in dashboard: ' + str(measure_count_dashboard)))
            logger.info(patient_id + ": Navigated to patient dashboard.")
            if sf.check_exists_by_class(driver, 'primary_val'):
                test_case_id += 1
                ws.append([test_case_id, patient_id, 'CareOps count present', 'Passed', 'x',
                           'Count: ' + driver.find_element_by_class_name("primary_val").text])
            else:
                test_case_id += 1
                ws.append(
                    [test_case_id, patient_id, 'CareOps count present', 'Failed', 'x', 'Careops count not present', driver.current_url])
            driver.close()
            driver.switch_to.window(driver.window_handles[0])
            driver.get(registry_url)
            sf.ajax_preloader_wait(driver)
            WebDriverWait(driver, 30).until(
                EC.presence_of_element_located((By.XPATH, locator.xpath_filter_measure_list)))

        except Exception as e:
            test_case_id += 1
            print(e)
            traceback.print_exc()
            ws.append(
                [test_case_id, patient_id, 'clicking on random patient from patient list of Provider\'s MSPL', 'Failed',
                 '', 'Unable to click on a random patient from the MSPL', driver.current_url])
            if window_switched == 1:
                driver.close()
                driver.switch_to.window(driver.window_handles[0])
            driver.get(main_registry_url)
            sf.ajax_preloader_wait(driver)
            WebDriverWait(driver, 30).until(
                EC.presence_of_element_located((By.XPATH, locator.xpath_filter_measure_list)))


    except Exception as e:
        ws.append([test_case_id, "Provider registry navigation", "Navigation to patient context through providers MSPL",
                   'Failed', 'x', 'Unable to navigate to patients list', driver.current_url])
        test_case_id += 1
        print(e)
        traceback.print_exc()
        driver.get(main_registry_url)
        sf.ajax_preloader_wait(driver)
        WebDriverWait(driver, 30).until(
            EC.presence_of_element_located((By.XPATH, locator.xpath_filter_measure_list)))

    driver.get(main_registry_url)
    sf.ajax_preloader_wait(driver)
    WebDriverWait(driver, 30).until(
        EC.presence_of_element_located((By.XPATH, locator.xpath_filter_measure_list)))
    rows = ws.max_row
    cols = ws.max_column
    for i in range(2, rows + 1):
        for j in range(3, cols + 1):
            if ws.cell(i, j).value == 'Passed':
                ws.cell(i, j).fill = PatternFill('solid', fgColor='0FC404')
            elif ws.cell(i, j).value == 'Failed':
                ws.cell(i, j).fill = PatternFill('solid', fgColor='FC0E03')
            elif ws.cell(i, j).value == 'Data table is empty':
                ws.cell(i, j).fill = PatternFill('solid', fgColor='FCC0BB')


def practice_registry(driver, workbook, logger, run_from, report_folder):
    workbook.create_sheet('Practice Registry')
    ws = workbook['Practice Registry']

    ws.append(['ID', 'Context', 'Scenario', 'Status', 'Time Taken', 'Comments'])
    header_font = Font(color='FFFFFF', bold=False, size=12)
    header_cell_color = PatternFill('solid', fgColor='030303')
    ws['A1'].font = header_font
    ws['A1'].fill = header_cell_color
    ws['B1'].font = header_font
    ws['B1'].fill = header_cell_color
    ws['C1'].font = header_font
    ws['C1'].fill = header_cell_color
    ws['D1'].font = header_font
    ws['D1'].fill = header_cell_color
    ws['E1'].font = header_font
    ws['E1'].fill = header_cell_color
    ws['F1'].font = header_font
    ws['F1'].fill = header_cell_color
    ws.name = "Arial"
    test_case_id = 1
    window_switched = 0
    print('Start Practice Registry Block')
    main_registry_url = driver.current_url
    # Checking run_froms and navigating to a practice registry at random
    if run_from == "Cozeva Support" or run_from == "Limited Cozeva Support" or run_from == "Customer Support" or run_from == "Regional Support":
        # Switching to random Practice name from default set context, main page
        try:
            driver.find_element_by_xpath(locator.xpath_side_nav_SlideOut).click()
            time.sleep(1)
            driver.find_element_by_id("providers-list").click()
            sf.ajax_preloader_wait(driver)
            sf.captureScreenshot(driver, "Provider list " + driver.find_element(By.XPATH, "//span[@class='specific_most']").text, report_folder)
            WebDriverWait(driver, 30).until(
                EC.presence_of_element_located((By.CLASS_NAME, 'tabs')))
            driver.find_element_by_class_name("tabs").find_elements_by_tag_name('li')[0].click()
            sf.ajax_preloader_wait(driver)
            WebDriverWait(driver, 30).until(
                EC.presence_of_element_located((By.ID, "metric-support-prac-ls")))
            list_of_practice_elements = driver.find_element_by_id("metric-support-prac-ls").find_element_by_tag_name('tbody').find_elements_by_tag_name('tr')
            global global_search_prac
            if len(list_of_practice_elements) > 1:
                global_search_prac = list_of_practice_elements[
                    sf.RandomNumberGenerator(len(list_of_practice_elements), 1)[0]].find_element_by_tag_name('a').text
            elif len(list_of_practice_elements) == 1:
                global_search_prac = list_of_practice_elements[0].find_element_by_tag_name('a').text


            if len(list_of_practice_elements) > 1:
                list_of_practice_elements[
                    sf.RandomNumberGenerator(len(list_of_practice_elements), 1)[0]].find_element_by_tag_name('a').click()
            elif len(list_of_practice_elements) == 1:
                list_of_practice_elements[0].find_element_by_tag_name('a').click()

            sf.ajax_preloader_wait(driver)
            sf.captureScreenshot(driver, "Practice Registry " + driver.find_element(By.XPATH, "//span[@class='specific_most']").text, report_folder)
        except Exception as e:
            ws.append(['1', "Attempting to navigate to a random practice", 'Navigation to practice context', 'Failed',
                       "Unable to navigate to a practice. Either the Practice list is unreachable or navigation access is denied", driver.current_url])
            driver.get(main_registry_url)

            print(e)
            traceback.print_exc()
            sf.ajax_preloader_wait(driver)
            WebDriverWait(driver, 30).until(
                EC.presence_of_element_located((By.XPATH, locator.xpath_filter_measure_list)))
            return
    elif run_from == "Office Admin Provider Delegate" or run_from == "Provider":
        ws.append(["1", run_from + " Role does not have access to practice Submenus"])
        return
    context_name = "Couldn't Fetch"
    #newly added
    sf.ajax_preloader_wait(driver)
    driver.find_element(By.XPATH, "//*[@id='qt-filter-label']").click()
    time.sleep(1)
    flag = 0
    Comment = ""
    target_url = ""
    try:
        LOB_list = driver.find_element(By.XPATH, "//*[@id='filter-lob']").find_elements(By.TAG_NAME, "li")
        temp = 0
        count = 0
        for i in range(0, len(LOB_list)):
            if count <= len(LOB_list):
                LOB_list[i].click()
                name = LOB_list[i].text
                driver.find_element(By.XPATH, "//*[@id='reg-filter-apply']").click()
                sf.ajax_preloader_wait(driver)
                if driver.find_element(By.XPATH, "//*[@id='conti_enroll']").is_selected():
                    driver.find_element(By.XPATH, "//*[@class='cont_disc_toggle']").click()
                current_context = driver.find_element(By.XPATH, "//span[@class='specific_most']").text
                sf.captureScreenshot(driver, name + " " + current_context, report_folder)
                patient_count = driver.find_element(By.XPATH, "//div[contains(text(), 'Patients')]/../div[2]").text
                patient_count = patient_count.replace(",", "")
                print(patient_count)
                if int(patient_count) > int(temp):
                    temp = patient_count
                    print(temp)
                    target_url = driver.current_url
                    print(target_url)
                home_url = driver.current_url
                # driver.find_element(By.XPATH, "//a[@data-target='qt-reg-nav-filters']").click()
                # time.sleep(0.5)
                # option = driver.find_elements(By.XPATH, "//input[@class='select-dropdown dropdown-trigger']")[0]
                # option.click()
                # lists = driver.find_elements(By.XPATH, "//ul[@class='dropdown-content select-dropdown']")[0]
                # options = lists.find_elements(By.TAG_NAME, 'li')
                # for option in options:
                #     if option.text == "Denominator":
                #         option.click()
                # apply_btn = driver.find_element(By.XPATH, "//button[@id='qt-apply-search']")
                # driver.execute_script("arguments[0].scrollIntoView();", apply_btn)
                # apply_btn.click()
                # time.sleep(0.5)
                try:
                    metric = driver.find_element(By.XPATH, "//*[@id='registry_body']").find_elements(By.TAG_NAME, "li")[0]
                    metric.click()
                    sf.ajax_preloader_wait(driver)
                    WebDriverWait(driver, 30).until(EC.presence_of_element_located((By.CLASS_NAME, 'tabs')))
                    driver.find_element_by_class_name('tabs').find_elements_by_class_name('tab')[1].click()
                    WebDriverWait(driver, 30).until(EC.presence_of_element_located((By.ID, "metric-support-pat-ls")))
                    sf.captureScreenshot(driver, current_context + " Patient tab " + name, report_folder)
                    ws.append([test_case_id, current_context, 'Navigation to MSPL' + driver.find_element(By.XPATH,
                                                                                                         "//*[@class='ch metric_specific_patient_list_title']").text,
                               'Passed', 'x', name, driver.current_url])
                except Exception as e:
                    ws.append([test_case_id, current_context, 'Navigation to Patient tab failed', 'Failed', 'x', name,
                               driver.current_url])
                    sf.captureScreenshot(driver, "Patient tab " + name, report_folder)
                driver.get(home_url)
                sf.ajax_preloader_wait(driver)
                driver.find_element(By.XPATH, "//*[@id='qt-filter-label']").click()
                time.sleep(1)
                LOB_list = driver.find_element(By.XPATH, "//*[@id='filter-lob']").find_elements(By.TAG_NAME, "li")
                count += 1
            else:
                driver.get(target_url)
                sf.ajax_preloader_wait(driver)
    except Exception as e:
        print("No LOBs for the practice")
        print(e)
        flag = 1
    WebDriverWait(driver, 30).until(EC.presence_of_element_located((By.XPATH, "//*[@id='quality_registry']")))
    patient_count = driver.find_element(By.XPATH, "//div[contains(text(), 'Patients')]/../div[2]").text
    patient_count = patient_count.replace(",", "")
    if flag == 1:
        Comment = "Practice does not have any patients in current registry and has only one LOB"
    else:
        Comment = "Practice does not have any patients in current registry"
    if patient_count == '0':
        ws.append([test_case_id, "Attempting to navigate to a random Practice", 'Navigation to practice context',
                   'Failed', 'x',
                   Comment,
                   driver.current_url])
    registry_url = driver.current_url
    # Nav check two : Navigation to patient context through patient toggle of practice Metric Specific List

    try:
        sf.ajax_preloader_wait(driver)
        # selecting a random non zero metric from the registry
        print(driver.current_url)
        context_name = driver.find_element_by_class_name("specific_most").text
        WebDriverWait(driver, 30).until(
            EC.presence_of_element_located((By.ID, "registry_body")))
        metrics = driver.find_element_by_id("registry_body").find_elements_by_tag_name('li')
        percent = '0.00'
        loop_count = 0
        if driver.find_element(By.XPATH, "//*[@id='conti_enroll']").is_selected():
            driver.find_element(By.XPATH, "//*[@class='cont_disc_toggle']").click()
        selected_metric_name = ""
        driver.find_element(By.XPATH, "//a[@data-target='qt-reg-nav-filters']").click()
        time.sleep(0.5)
        option = driver.find_elements(By.XPATH, "//input[@class='select-dropdown dropdown-trigger']")[0]
        option.click()
        lists = driver.find_elements(By.XPATH, "//ul[@class='dropdown-content select-dropdown']")[0]
        options = lists.find_elements(By.TAG_NAME, 'li')
        for option in options:
            if option.text == "Numerator":
                option.click()
        apply_btn = driver.find_element(By.XPATH, "//button[@id='qt-apply-search']")
        driver.execute_script("arguments[0].scrollIntoView();", apply_btn)
        apply_btn.click()
        time.sleep(0.5)
        selectedMetric = driver.find_element(By.XPATH, "//*[@id='registry_body']/ul").find_elements(By.TAG_NAME, "li")
        selected_metric_name = selectedMetric[0].find_element(By.XPATH, "//*[@class='met-name']").text
        selectedMetric[0].click()
        sf.ajax_preloader_wait(driver)
        try:
            WebDriverWait(driver, 30).until(
                EC.presence_of_element_located((By.CLASS_NAME, 'tabs')))
            driver.find_element_by_class_name('tabs').find_elements_by_class_name('tab')[2].click()
            start_time = time.perf_counter()
            sf.ajax_preloader_wait(driver)
            WebDriverWait(driver, 30).until(
                EC.presence_of_element_located((By.CLASS_NAME, 'tabs')))
            sf.captureScreenshot(driver, "Performance stat tab " + driver.find_element(By.XPATH, "//span[@class='specific_most']").text, report_folder)
            time_taken = time.perf_counter() - start_time
            if driver.find_elements_by_id('performance_details') != 0:
                ws.append([test_case_id, context_name,
                           'Navigation to Performance Stats from Practice Metric : ' + selected_metric_name, 'Passed',
                           time_taken])
                test_case_id += 1
            else:
                ws.append([test_case_id, context_name,
                           'Navigation to Performance Stats from Practice Metric : ' + selected_metric_name, 'Failed', driver.current_url])
                test_case_id += 1

        except Exception as e:
            print(e)
            traceback.print_exc()
            ws.append([test_case_id, context_name, 'Navigation to Performance Stats from Practice MSPL', 'Failed', '',
                       'Couldnt click on the performance tab of metric :' + selected_metric_name, driver.current_url])
            test_case_id += 1

        try:
            driver.find_element_by_class_name('tabs').find_elements_by_class_name('tab')[1].click()
            sf.ajax_preloader_wait(driver)
            print("Preloader gone")
            time.sleep(0.5)
            if len(driver.find_elements_by_class_name("ajax_preloader")) != 0:
                print("Preloader Reappeared")
                sf.ajax_preloader_wait(driver)
            WebDriverWait(driver, 30).until(
                EC.presence_of_element_located((By.ID, "metric-support-pat-ls")))
            sf.captureScreenshot(driver, "Patients list tab " + driver.find_element(By.XPATH, "//span[@class='specific_most']").text, report_folder)
            patients = driver.find_element_by_id("metric-support-pat-ls").find_element_by_tag_name(
                'tbody').find_elements_by_tag_name('tr')
            selected_patient = ""
            if len(patients) > 1:
                selected_patient = patients[sf.RandomNumberGenerator(len(patients), 1)[0]].find_element_by_class_name(
                    'pat_name')
            elif len(patients) == 1:
                if patients[0].text == "No Data Available":
                    raise Exception("No patients in this MSPL")
                selected_patient = patients[0].find_element_by_class_name('pat_name')
            selected_patient_name = selected_patient.text
            selected_patient.click()
            driver.switch_to.window(driver.window_handles[1])
            window_switched = 1
            start_time = time.perf_counter()
            sf.ajax_preloader_wait(driver)
            print("Preloader gone")
            time.sleep(0.5)
            if len(driver.find_elements_by_class_name("ajax_preloader")) != 0:
                print("Preloader Reappeared")
                sf.ajax_preloader_wait(driver)
            time_taken = round((time.perf_counter() - start_time), 3)
            WebDriverWait(driver, 20).until(
                EC.presence_of_element_located((By.CLASS_NAME, "primary_val")))
            sf.captureScreenshot(driver, "Patient dashboard " + run_from, report_folder)
            if len(driver.find_elements_by_class_name("primary_val")) != 0:
                ws.append([test_case_id, selected_patient_name,
                           "Navigation to patient context through patient toggle of practice Metric Specific List",
                           'Passed', time_taken])
                test_case_id += 1
                driver.close()
                driver.switch_to.window(driver.window_handles[0])
                driver.get(registry_url)
                sf.ajax_preloader_wait(driver)
            else:
                ws.append([test_case_id, selected_patient_name,
                           "Navigation to patient context through patient toggle of practice Metric Specific List",
                           'Failed', time_taken, driver.current_url])
                test_case_id += 1
                if window_switched == 1:
                    driver.close()
                    driver.switch_to.window(driver.window_handles[0])
                driver.get(registry_url)
                sf.ajax_preloader_wait(driver)

        except Exception as e:
            print(e)
            traceback.print_exc()
            ws.append(
                [test_case_id, context_name,
                 "Navigation to patient context through patient toggle of practice Metric Specific List", 'Failed', '',
                 'No patients in patient tab', driver.current_url])
            test_case_id += 1
            if window_switched == 1:
                driver.close()
                driver.switch_to.window(driver.window_handles[0])
            driver.get(registry_url)
            sf.ajax_preloader_wait(driver)


    except Exception as e:
        print(e)
        traceback.print_exc()
        print(driver.current_url)
        ws.append(
            [test_case_id, context_name,
             "Navigation to patient context through patient toggle of practice Metric Specific List", 'Failed', '',
             'Couldn\'t navigate into a random metric from the provivdr registry', driver.current_url])
        test_case_id += 1

        driver.get(registry_url)
        sf.ajax_preloader_wait(driver)

    # nav check 3 : Navigation to provider registry through providers tab in of a practice
    registry_url = driver.current_url
    try:
        driver.find_element_by_xpath(locator.xpath_side_nav_SlideOut).click()
        time.sleep(1)
        driver.execute_script("arguments[0].scrollIntoView();", driver.find_element_by_id("providers-list"))
        driver.find_element_by_id("providers-list").click()
        sf.ajax_preloader_wait(driver)
        time.sleep(1)
        WebDriverWait(driver, 30).until(
            EC.presence_of_element_located((By.ID, "metric-support-prov-ls")))
        sf.captureScreenshot(driver, "Provider list " + driver.find_element(By.XPATH, "//span[@class='specific_most']").text, report_folder)
        list_of_provider_elements = driver.find_element_by_id("metric-support-prov-ls").find_elements_by_tag_name('tr')
        if len(list_of_provider_elements) > 1:
            selected_provider = list_of_provider_elements[
                sf.RandomNumberGenerator(len(list_of_provider_elements), 1)[0]].find_elements_by_tag_name('a')[1]
            selected_provider_name = selected_provider.text
            selected_provider.click()
        else:
            selected_provider = list_of_provider_elements[0].find_elements_by_tag_name('a')[1]
            selected_provider_name = selected_provider.text
            selected_provider.click()
        time_start = time.perf_counter()
        sf.ajax_preloader_wait(driver)
        WebDriverWait(driver, 30).until(
            EC.presence_of_element_located((By.XPATH, locator.xpath_filter_measure_list)))
        sf.captureScreenshot(driver, "Provider registry " + driver.find_element(By.XPATH, "//span[@class='specific_most']").text, report_folder)
        time_taken = time.perf_counter() - time_start
        if driver.find_elements_by_xpath(locator.xpath_filter_measure_list) != 0:
            ws.append([test_case_id, selected_provider_name,
                       'Navigation to provider registry through providers tab in of a practice', 'Passed',
                       round(time_taken, 3)])
            test_case_id += 1
            driver.get(registry_url)
            sf.ajax_preloader_wait(driver)
        else:
            ws.append([test_case_id, selected_provider_name,
                       'Navigation to provider registry through providers tab in of a practice', 'Failed',
                       round(time_taken, 3), 'Unable to locate filter element on provider\'s registry', driver.current_url])
            test_case_id += 1
            driver.get(registry_url)
            sf.ajax_preloader_wait(driver)

    except Exception as e:
        print(e)
        traceback.print_exc()
        ws.append([test_case_id, context_name, "Navigation to provider registry through providers tab in of a practice",
                   'Failed', "", 'Unable to click on providers\' tab and navigate to their registry', driver.current_url])
        test_case_id += 1
        driver.get(registry_url)
        sf.ajax_preloader_wait(driver)

    # nav check 4 : Navigation to Performance Stats from Practice Metric specific list

    if run_from == "Cozeva Support" or run_from == "Limited Cozeva Support" or run_from == "Customer Support" or run_from == "Regional Support":
        driver.find_element_by_xpath(locator.xpath_side_nav_SlideOut).click()
        time.sleep(1)
        driver.execute_script("arguments[0].scrollIntoView();", driver.find_element_by_id("home"))
        driver.find_element_by_id("home").click()
        sf.ajax_preloader_wait(driver)
    driver.get(main_registry_url)
    sf.ajax_preloader_wait(driver)
    WebDriverWait(driver, 30).until(
        EC.presence_of_element_located((By.XPATH, locator.xpath_filter_measure_list)))
    rows = ws.max_row
    cols = ws.max_column
    for i in range(2, rows + 1):
        for j in range(3, cols + 1):
            if ws.cell(i, j).value == 'Passed':
                ws.cell(i, j).fill = PatternFill('solid', fgColor='0FC404')
            elif ws.cell(i, j).value == 'Failed':
                ws.cell(i, j).fill = PatternFill('solid', fgColor='FC0E03')
            elif ws.cell(i, j).value == 'Data table is empty':
                ws.cell(i, j).fill = PatternFill('solid', fgColor='FCC0BB')


def support_level(driver, workbook, logger, run_from, report_folder):
    workbook.create_sheet('Support Level')
    ws = workbook['Support Level']

    ws.append(['ID', 'Context', 'Scenario', 'Status', 'Time Taken', 'Comments'])
    header_font = Font(color='FFFFFF', bold=False, size=12)
    header_cell_color = PatternFill('solid', fgColor='030303')
    ws['A1'].font = header_font
    ws['A1'].fill = header_cell_color
    ws['B1'].font = header_font
    ws['B1'].fill = header_cell_color
    ws['C1'].font = header_font
    ws['C1'].fill = header_cell_color
    ws['D1'].font = header_font
    ws['D1'].fill = header_cell_color
    ws['E1'].font = header_font
    ws['E1'].fill = header_cell_color
    ws['F1'].font = header_font
    ws['F1'].fill = header_cell_color
    ws.name = "Arial"
    test_case_id = 1

    registry_url = driver.current_url
    # Selecting tabs from Support MSPL
    context_name = "Couldn't Fetch"
    try:
        WebDriverWait(driver, 30).until(
            EC.presence_of_element_located((By.ID, "registry_body")))
        sf.captureScreenshot(driver, "Registry page " + driver.find_element(By.XPATH, "//span[@class='specific_most']").text, report_folder)
        selected_metric_name = 'Couldnt fetch Metric Name'
        context_name = driver.find_element_by_xpath(locator.xpath_context_Name).text

        metrics = driver.find_element_by_id("registry_body").find_elements_by_tag_name('li')
        percent = '0.00'
        num_den = "(0/0)"
        iter_count = 0
        while percent == '0.00' or percent == '0.00%' or num_den == "(0/0)":
            selectedMetric = metrics[sf.RandomNumberGenerator(len(metrics), 1)[0]]
            percent = selectedMetric.find_element_by_class_name('percent').text
            num_den = selectedMetric.find_element_by_class_name('num-den').text
            iter_count+=1
            if iter_count > 10:
                raise Exception("All/most metrics are 0/0")
        selected_metric_name = selectedMetric.find_element_by_class_name('met-name').text
        selectedMetric.click()
        sf.ajax_preloader_wait(driver)
        WebDriverWait(driver, 30).until(EC.presence_of_element_located((By.CLASS_NAME, 'tabs')))
        sf.captureScreenshot(driver, "Provider list from " + driver.find_element(By.XPATH, "//span[@class='specific_most']").text, report_folder)
        metric_url = driver.current_url
        # nav 1 : Practice Tab
        try:
            selectedPracticeName = 'Couldn\'t Fetch'
            driver.find_element_by_class_name('tabs').find_elements_by_class_name('tab')[0].click()
            sf.ajax_preloader_wait(driver)
            WebDriverWait(driver, 30).until(
                EC.presence_of_element_located((By.ID, "metric-support-prac-ls")))
            sf.captureScreenshot(driver, "Practice list " + driver.find_element(By.XPATH, "//span[@class='specific_most']").text, report_folder)
            practices = driver.find_element_by_id("metric-support-prac-ls").find_element_by_tag_name(
                'tbody').find_elements_by_tag_name('tr')

            try:
                list_index = len(practices)
                zero_prac = 0
                for prac_entry in range(0, list_index):
                    patient_count = practices[prac_entry].find_elements_by_tag_name("td")[5].text
                    if int(patient_count.replace(",","")) == 0:
                        zero_prac += 1

                if zero_prac > list_index/2:
                    print("More than half the practices have 0 patient counts")
                    ws.append([test_case_id,context_name, "Checking for 0 patient counts in practice tab", "Failed", 'x', "Practices with 0 patients: "+str(zero_prac), driver.current_url])
                else:
                    ws.append([test_case_id,context_name, "Checking for 0 patient counts in practice tab", "Passed", 'x', "Practices with 0 patients: "+str(zero_prac)])

            except:
                traceback.print_exc()
                ws.append([test_case_id, context_name, "Checking for 0 patient counts in practice tab", "Failed", 'x', "Couldn't fetch patient counts", driver.current_url])
                test_case_id+=1
            if len(practices) > 1:
                selectedPractice = \
                    practices[sf.RandomNumberGenerator(len(practices), 1)[0]].find_elements_by_tag_name('a')[1]
                selectedPracticeName = selectedPractice.text
                global global_search_prac
                global_search_prac = selectedPracticeName
                selectedPractice.click()
            else:
                selectedPractice = practices[0].find_elements_by_tag_name('a')[1]
                selectedPracticeName = selectedPractice.text
                # global global_search_prac
                global_search_prac = selectedPracticeName
                selectedPractice.click()
            start_time = time.perf_counter()
            sf.ajax_preloader_wait(driver)
            WebDriverWait(driver, 30).until(
                EC.presence_of_element_located((By.XPATH, locator.xpath_filter_measure_list)))
            sf.captureScreenshot(driver, "Practice registry " + driver.find_element(By.XPATH, "//span[@class='specific_most']").text, report_folder)
            time_taken = round((time.perf_counter() - start_time), 3)
            if len(driver.find_elements_by_xpath(locator.xpath_filter_measure_list)) != 0:
                ws.append([test_case_id, selectedPracticeName,
                           'Nagivation to practice Registry from the practice tab of support MSPL: ' + selected_metric_name,
                           'Passed', time_taken])
                test_case_id += 1
                driver.get(metric_url)
            else:
                ws.append([test_case_id, selectedPracticeName,
                           'Nagivation to practice Registry from the practice tab of support MSPL: ' + selected_metric_name,
                           'Failed', time_taken, 'Couldnt load into registry of a practice', driver.current_url])
                test_case_id += 1
                driver.get(metric_url)

        except Exception as e:
            ws.append([test_case_id, context_name,
                       'Navigation to a practice registry from the pratice tab of support MSPL :' + selected_metric_name,
                       'Failed', '',
                       'Couldnt click on practice tab or a random practice name: ' + selectedPracticeName, driver.current_url])
            test_case_id += 1
            print(e)
            traceback.print_exc()
            driver.get(metric_url)

        # Nav to provider registry
        try:
            sf.ajax_preloader_wait(driver)
            WebDriverWait(driver, 30).until(EC.presence_of_element_located((By.CLASS_NAME, 'tabs')))
            selectedProviderName = 'Couldn\'t Fetch'
            driver.find_element_by_class_name('tabs').find_elements_by_class_name('tab')[1].click()
            sf.ajax_preloader_wait(driver)
            WebDriverWait(driver, 30).until(
                EC.presence_of_element_located((By.ID, "metric-support-prov-ls")))
            providers = driver.find_element_by_id("metric-support-prov-ls").find_element_by_tag_name(
                'tbody').find_elements_by_tag_name('tr')
            try:
                list_index = len(providers)
                zero_prov = 0
                for prov_entry in range(0, list_index):
                    patient_count = providers[prov_entry].find_elements_by_tag_name("td")[9].text
                    if int(patient_count.replace(",","")) == 0:
                        zero_prov += 1

                if zero_prov > list_index/2:
                    print("More than half the practices have 0 patient counts")
                    ws.append([test_case_id,context_name, "Checking for 0 patient counts in provider tab", "Failed", 'x', "Providers with 0 patients: "+str(zero_prov), driver.current_url])
                else:
                    ws.append([test_case_id,context_name, "Checking for 0 patient counts in provider tab", "Passed", 'x', "Providers with 0 patients: "+str(zero_prov)])

            except:
                traceback.print_exc()
                ws.append([test_case_id, context_name, "Checking for 0 patient counts in provider tab", "Failed", 'x', "Couldn't fetch patient counts", driver.current_url])
                test_case_id+=1
            if len(providers) > 1:
                selectedProvider = \
                    providers[sf.RandomNumberGenerator(len(providers), 1)[0]].find_elements_by_tag_name('a')[2]
                selectedProviderName = selectedProvider.text
                global global_search_prov
                global_search_prov = selectedProviderName
                selectedProvider.click()
            else:
                selectedProvider = providers[0].find_elements_by_tag_name('a')[2]
                selectedProviderNameName = selectedProvider.text
                # global global_search_prov
                global_search_prov = selectedProviderName
                selectedProvider.click()
            start_time = time.perf_counter()
            sf.ajax_preloader_wait(driver)
            WebDriverWait(driver, 30).until(
                EC.presence_of_element_located((By.XPATH, locator.xpath_filter_measure_list)))
            sf.captureScreenshot(driver, "Provider Registry " + driver.find_element(By.XPATH, "//span[@class='specific_most']").text, report_folder)
            time_taken = round((time.perf_counter() - start_time), 3)
            if len(driver.find_elements_by_xpath(locator.xpath_filter_measure_list)) != 0:
                ws.append([test_case_id, selectedProviderName,
                           'Nagivation to provider Registry from the provider tab of support MSPL: ' + selected_metric_name,
                           'Passed', time_taken])
                test_case_id += 1
                driver.get(metric_url)
            else:
                ws.append([test_case_id, selectedProviderName,
                           'Nagivation to provider Registry from the provider tab of support MSPL: ' + selected_metric_name,
                           'Failed', time_taken, 'Couldnt load into registry of a provider', driver.current_url])
                test_case_id += 1
                driver.get(metric_url)

        except Exception as e:
            print(e)
            traceback.print_exc()
            ws.append([test_case_id, context_name,
                       'Navigation to a provider registry from the provider tab of support MSPL :' + selected_metric_name,
                       'Failed', '',
                       'Couldnt click on provider tab or a random provider name: ' + selectedProviderName, driver.current_url])
            test_case_id += 1
            driver.get(metric_url)

        # nav 3 : Patient context
        try:
            sf.ajax_preloader_wait(driver)
            WebDriverWait(driver, 30).until(EC.presence_of_element_located((By.CLASS_NAME, 'tabs')))
            patient_id = 'Couldn\'t Fetch'
            driver.find_element_by_class_name('tabs').find_elements_by_class_name('tab')[2].click()
            sf.ajax_preloader_wait(driver)
            WebDriverWait(driver, 30).until(
                EC.presence_of_element_located((By.ID, "metric-support-pat-ls")))
            sf.captureScreenshot(driver, "Patient list " + driver.find_element(By.XPATH, "//span[@class='specific_most']").text, report_folder)
            patients = driver.find_element_by_id("metric-support-pat-ls").find_element_by_tag_name(
                'tbody').find_elements_by_tag_name('tr')
            if len(patients) > 1:
                selectedPatient = \
                    patients[sf.RandomNumberGenerator(len(patients), 1)[0]].find_elements_by_class_name('pat_name')[0]
                selectedPatient.click()
            else:
                selectedPatient = patients[0].find_elements_by_class_name('pat_name')[0]
                selectedPatient.click()
            driver.switch_to.window(driver.window_handles[1])
            start_time = time.perf_counter()
            sf.ajax_preloader_wait(driver)
            WebDriverWait(driver, 30).until(
                EC.presence_of_element_located((By.XPATH, locator.xpath_cozeva_Id)))
            sf.captureScreenshot(driver, "Patient dashboard " + run_from, report_folder)
            time_taken = round((time.perf_counter() - start_time), 3)
            patient_id = driver.find_element_by_xpath(locator.xpath_cozeva_Id).text
            global global_search_pat
            global_search_pat = patient_id
            if sf.check_exists_by_class(driver, 'primary_val'):
                measure_count_dashboard = len(
                    driver.find_elements_by_xpath("//tbody[@class='measurement-body careops-new']/tr"))
                test_case_id += 1
                ws.append(
                    (test_case_id, patient_id, 'Navigation to patient context from the patients tab of support MSPL',
                     'Passed', time_taken, 'Measures count in dashboard: ' + str(measure_count_dashboard)))
                logger.info(patient_id + ": Navigated to patient dashboard.")
                test_case_id += 1
                ws.append(
                    [test_case_id, patient_id, 'Navigation to patient context from the patients tab of support MSPL',
                     'Passed', 'x', 'Count: ' + driver.find_element_by_class_name("primary_val").text])
                driver.close()
                driver.switch_to.window(driver.window_handles[0])
                driver.get(metric_url)
            else:
                test_case_id += 1
                ws.append(
                    [test_case_id, patient_id, 'Navigation to patient context from the patients tab of support MSPL',
                     'Failed', 'x', 'Careops count not present', driver.current_url])
                driver.close()
                driver.switch_to.window(driver.window_handles[0])
                driver.get(metric_url)


        except Exception as e:
            print(e)
            traceback.print_exc()
            ws.append([test_case_id, context_name,
                       'Navigation to patient context from the patients tab of support MSPL :' + selected_metric_name,
                       'Failed', '', 'Couldnt click on patient tab or a random patient : ' + patient_id, driver.current_url])
            test_case_id += 1
            driver.get(metric_url)

        # nav 4 : Performance Statistics
        try:
            sf.ajax_preloader_wait(driver)
            WebDriverWait(driver, 30).until(
                EC.presence_of_element_located((By.CLASS_NAME, 'tab')))
            tab_list = driver.find_element_by_class_name('tabs').find_elements(By.TAG_NAME, "li")
            for tab in tab_list:
                if "Performance Statistics" in tab.text:
                    tab.click()
                    break
            start_time = time.perf_counter()
            sf.ajax_preloader_wait(driver)
            WebDriverWait(driver, 30).until(
                EC.presence_of_element_located((By.CLASS_NAME, 'tabs')))
            sf.captureScreenshot(driver, "Performance tab " + driver.find_element(By.XPATH, "//span[@class='specific_most']").text, report_folder)
            time_taken = time.perf_counter() - start_time
            if driver.find_elements_by_id('performance_details') != 0:
                ws.append([test_case_id, context_name,
                           'Navigation to Performance Stats from Support Metric : ' + selected_metric_name, 'Passed',
                           time_taken])
                test_case_id += 1
            else:
                ws.append([test_case_id, context_name,
                           'Navigation to Performance Stats from Support Metric : ' + selected_metric_name, 'Failed', driver.current_url])
                test_case_id += 1

        except Exception as e:
            print(e)
            traceback.print_exc()
            ws.append([test_case_id, context_name, 'Navigation to Performance Stats from Practice MSPL', 'Failed', '',
                       'Couldnt click on the performance tab of metric :' + selected_metric_name, driver.current_url])
            test_case_id += 1

    except Exception as e:
        print(e)
        traceback.print_exc()
        ws.append([test_case_id, context_name, 'Navigation to Support MSPL', 'Failed', '',
                   'Unable to click on a random metric: ' + selected_metric_name, driver.current_url])
        test_case_id += 1
        driver.get(registry_url)

    driver.get(registry_url)

    rows = ws.max_row
    cols = ws.max_column
    for i in range(2, rows + 1):
        for j in range(3, cols + 1):
            if ws.cell(i, j).value == 'Passed':
                ws.cell(i, j).fill = PatternFill('solid', fgColor='0FC404')
            elif ws.cell(i, j).value == 'Failed':
                ws.cell(i, j).fill = PatternFill('solid', fgColor='FC0E03')
            elif ws.cell(i, j).value == 'Data table is empty':
                ws.cell(i, j).fill = PatternFill('solid', fgColor='FCC0BB')


def global_search(driver, workbook, logger, run_from, report_folder):
    workbook.create_sheet('Global Search')
    ws = workbook['Global Search']
    ws.append(['ID', 'Context', 'Scenario', 'Status', 'Time Taken', 'Comments'])
    header_font = Font(color='FFFFFF', bold=False, size=12)
    header_cell_color = PatternFill('solid', fgColor='030303')
    ws['A1'].font = header_font
    ws['A1'].fill = header_cell_color
    ws['B1'].font = header_font
    ws['B1'].fill = header_cell_color
    ws['C1'].font = header_font
    ws['C1'].fill = header_cell_color
    ws['D1'].font = header_font
    ws['D1'].fill = header_cell_color
    ws['E1'].font = header_font
    ws['E1'].fill = header_cell_color
    ws.name = "Arial"
    test_case_id = 1
    main_registry_url = driver.current_url

    global global_search_prov, global_search_prac, global_search_pat

    def fetchPracName():
        return "test"

    def fetchProvName():
        return "test"

    def fetchPatId():
        return "test"

    def performPracSearch():
        try:
            window_switched = 0
            driver.find_element_by_id('globalsearch_input').send_keys(global_search_prac)
            start_time = time.perf_counter()
            WebDriverWait(driver, 45).until(
                EC.presence_of_element_located((By.CLASS_NAME, 'collection-header')))
            time_taken = round(time.perf_counter() - start_time)
            driver.find_element_by_id('globalsearch_input').send_keys(Keys.RETURN)
            sf.ajax_preloader_wait(driver)
            time_taken = round(time.perf_counter() - start_time)
            # driver.find_element_by_id('globalsearch_input').send_keys(Keys.RETURN)
            driver.find_element_by_id('search_practices_link').click()
            WebDriverWait(driver, 30).until(
                EC.presence_of_element_located((By.ID, 'search_practices')))
            driver.find_element_by_id('search_practices').find_elements_by_tag_name('a')[0].click()
            driver.switch_to.window(driver.window_handles[1])
            window_switched = 1
            sf.ajax_preloader_wait(driver)
            WebDriverWait(driver, 30).until(
                EC.presence_of_element_located((By.XPATH, locator.xpath_filter_measure_list)))
            if len(driver.find_elements_by_xpath(locator.xpath_filter_measure_list)) != 0:
                sf.captureScreenshot(driver, "Practice global search- Passed: " + run_from, report_folder)
                ws.append([test_case_id, 'Practice', 'Context set to: ' + global_search_prac, 'Passed', time_taken])
                driver.close()
                driver.switch_to.window(driver.window_handles[0])
                driver.get(main_registry_url)
            else:
                sf.captureScreenshot(driver, "Practice global search- Failed: " + run_from, report_folder)
                ws.append([test_case_id, 'Practice', 'Context set to: ' + global_search_prac, 'Failed', time_taken, driver.current_url])
                driver.close()
                driver.switch_to.window(driver.window_handles[0])
                driver.get(main_registry_url)

        except Exception as e:
            print(e)
            traceback.print_exc()
            if window_switched == 1:
                sf.captureScreenshot(driver, "Practice global search- Click failed: " + run_from, report_folder)
                ws.append([test_case_id, 'Practice', 'Context set to: ' + global_search_prac, 'Failed', '',
                           'Unable to click on practice name from global search', driver.current_url])
                driver.close()
                driver.switch_to.window(driver.window_handles[0])
                driver.get(main_registry_url)
            elif window_switched == 0:
                sf.captureScreenshot(driver, "Practice global search- Failed: " + run_from, report_folder)
                ws.append([test_case_id, 'Practice', 'Context set to: ' + global_search_prac, 'Failed', '',
                           'Unable to global search', driver.current_url])
                driver.get(main_registry_url)

    def performProvSearch():
        try:
            window_switched = 0
            driver.find_element_by_id('globalsearch_input').send_keys(global_search_prov)
            WebDriverWait(driver, 45).until(
                EC.presence_of_element_located((By.CLASS_NAME, 'collection-header')))
            driver.find_element_by_id('globalsearch_input').send_keys(Keys.RETURN)
            start_time = time.perf_counter()
            sf.ajax_preloader_wait(driver)
            time_taken = round(time.perf_counter() - start_time)
            # driver.find_element_by_id('globalsearch_input').send_keys(Keys.RETURN)
            driver.find_element_by_id('search_providers_link').click()
            driver.find_element_by_id('search_providers').find_elements_by_tag_name('a')[0].click()
            driver.switch_to.window(driver.window_handles[1])
            window_switched = 1
            sf.ajax_preloader_wait(driver)
            WebDriverWait(driver, 30).until(
                EC.presence_of_element_located((By.XPATH, locator.xpath_filter_measure_list)))
            if len(driver.find_elements_by_xpath(locator.xpath_filter_measure_list)) != 0:
                sf.captureScreenshot(driver, "Provider global search- Passed: " + run_from, report_folder)
                ws.append([test_case_id, 'Provider', 'Context set to: ' + global_search_prov, 'Passed', time_taken])
                driver.close()
                driver.switch_to.window(driver.window_handles[0])
                driver.get(main_registry_url)
            else:
                sf.captureScreenshot(driver, "Provider global search- Failed: " + run_from, report_folder)
                ws.append([test_case_id, 'Practice', 'Context set to: ' + global_search_prov, 'Failed', time_taken, driver.current_url])
                driver.close()
                driver.switch_to.window(driver.window_handles[0])
                driver.get(main_registry_url)

        except Exception as e:
            print(e)
            traceback.print_exc()
            if window_switched == 1:
                sf.captureScreenshot(driver, "Provider global search- Click failed:" + run_from, report_folder)
                ws.append([test_case_id, 'Provider', 'Context set to: ' + global_search_prov, 'Failed', '',
                           'Unable to click on practice name from global search', driver.current_url])
                driver.close()
                driver.switch_to.window(driver.window_handles[0])
                driver.get(main_registry_url)
            elif window_switched == 0:
                sf.captureScreenshot(driver, "Provider global search- Failed: " + run_from, report_folder)
                ws.append([test_case_id, 'Provider', 'Context set to: ' + global_search_prov, 'Failed', '',
                           'Unable to global search', driver.current_url])
                driver.get(main_registry_url)

    def performPatSearch():
        try:
            window_switched = 0
            driver.find_element_by_id('globalsearch_input').send_keys(global_search_pat)
            WebDriverWait(driver, 45).until(
                EC.presence_of_element_located((By.CLASS_NAME, 'collection-header')))
            driver.find_element_by_id('globalsearch_input').send_keys(Keys.RETURN)
            start_time = time.perf_counter()
            sf.ajax_preloader_wait(driver)
            time_taken = round(time.perf_counter() - start_time)
            driver.find_element_by_id('search_patients_link').click()
            driver.find_element_by_id('search_patients').find_elements_by_tag_name('li')[
                1].find_element_by_css_selector("a[href*='patient_detail']").click()
            driver.switch_to.window(driver.window_handles[1])
            window_switched = 1
            sf.ajax_preloader_wait(driver)
            WebDriverWait(driver, 30).until(
                EC.presence_of_element_located((By.XPATH, locator.xpath_patient_Header_Dropdown_Arrow)))
            if len(driver.find_elements_by_xpath(locator.xpath_patient_Header_Dropdown_Arrow)) != 0:
                sf.captureScreenshot(driver, "Patient global search- Passed: " + run_from, report_folder)
                ws.append([test_case_id, 'Patient', 'Context set to: ' + global_search_pat, 'Passed', time_taken])
                driver.close()
                driver.switch_to.window(driver.window_handles[0])
                driver.get(main_registry_url)
            else:
                sf.captureScreenshot(driver, "Patient global search- Failed: " + run_from, report_folder)
                ws.append([test_case_id, 'Patient', 'Context set to: ' + global_search_pat, 'Failed', time_taken, driver.current_url])
                driver.close()
                driver.switch_to.window(driver.window_handles[0])
                driver.get(main_registry_url)

        except Exception as e:
            print(e)
            traceback.print_exc()
            if window_switched == 1:
                sf.captureScreenshot(driver, "Patient global search- Click failed: " + run_from, report_folder)
                ws.append([test_case_id, 'Patient', 'Context set to: ' + global_search_pat, 'Failed', '',
                           'Unable to click on practice name from global search', driver.current_url])
                driver.close()
                driver.switch_to.window(driver.window_handles[0])
                driver.get(main_registry_url)
            elif window_switched == 0:
                sf.captureScreenshot(driver, "Patient global search- Failed: " + run_from, report_folder)
                ws.append([test_case_id, 'Patient', 'Context set to: ' + global_search_pat, 'Failed', '',
                           'Unable to global search', driver.current_url])
                driver.get(main_registry_url)

    if run_from == "Cozeva Support" or run_from == "Limited Cozeva Support" or run_from == "Customer Support" or run_from == "Regional Support":
        # Perform global search for Practice, Provider and Patient
        # Fetch Practice, provider and patient ID
        if global_search_prac is None:
            global_search_prac = fetchPracName()
        if global_search_prov is None:
            global_search_prov = fetchProvName()
        if global_search_pat is None:
            global_search_pat = fetchPatId()

        performPracSearch()
        performProvSearch()
        performPatSearch()



    elif run_from == "Office Admin Practice Delegate":
        # perform global search for provider and patient
        # fetch Provider and Patient ID
        if global_search_prov is None:
            global_search_prov = fetchProvName()
        if global_search_pat is None:
            global_search_pat = fetchPatId()

        performProvSearch()
        performPatSearch()
    elif run_from == "Office Admin Provider Delegate" or run_from == "Provider":
        # perform global search for patients
        # Fetch PatientID
        if global_search_pat is None:
            global_search_pat = fetchPatId()

        performPatSearch()

    rows = ws.max_row
    cols = ws.max_column
    for i in range(2, rows + 1):
        for j in range(3, cols + 1):
            if ws.cell(i, j).value == 'Passed':
                ws.cell(i, j).fill = PatternFill('solid', fgColor='0FC404')
            elif ws.cell(i, j).value == 'Failed':
                ws.cell(i, j).fill = PatternFill('solid', fgColor='FC0E03')
            elif ws.cell(i, j).value == 'Data table is empty':
                ws.cell(i, j).fill = PatternFill('solid', fgColor='FCC0BB')


def provider_mspl(driver, workbook, logger, run_from, report_folder):
    workbook.create_sheet('Provider\'s MSPL')
    ws = workbook['Provider\'s MSPL']
    selected_metric_name = ""

    ws.append(['ID', 'Context', 'Scenario', 'Status', 'Time Taken', 'Comments'])
    header_font = Font(color='FFFFFF', bold=False, size=12)
    header_cell_color = PatternFill('solid', fgColor='030303')
    ws['A1'].font = header_font
    ws['A1'].fill = header_cell_color
    ws['B1'].font = header_font
    ws['B1'].fill = header_cell_color
    ws['C1'].font = header_font
    ws['C1'].fill = header_cell_color
    ws['D1'].font = header_font
    ws['D1'].fill = header_cell_color
    ws.name = "Arial"
    test_case_id = 1
    main_registry_url = driver.current_url
    selected_provider = "Couldn't Fetch"
    # check for default page and navigate to a Provider's Registry
    if run_from == "Cozeva Support" or run_from == "Limited Cozeva Support" or run_from == "Customer Support" or run_from == "Regional Support" or run_from == "Office Admin Practice Delegate":
        # Switching to random Practice name from default set context, main page
        try:
            driver.find_element_by_xpath(locator.xpath_side_nav_SlideOut).click()
            time.sleep(0.5)
            driver.find_element_by_id("providers-list").click()
            sf.ajax_preloader_wait(driver)
            WebDriverWait(driver, 30).until(
                EC.presence_of_element_located((By.ID, "metric-support-prov-ls")))
            sf.captureScreenshot(driver, "Provider list " + driver.find_element(By.XPATH, "//span[@class='specific_most']").text, report_folder)
            list_of_provider_elements = driver.find_element_by_id("metric-support-prov-ls").find_element_by_tag_name('tbody').find_elements_by_tag_name(
                'tr')
            if len(list_of_provider_elements) > 1:
                selected_provider = list_of_provider_elements[
                    sf.RandomNumberGenerator(len(list_of_provider_elements), 1)[0]].find_elements_by_tag_name('a')[1]
            elif len(list_of_provider_elements) == 1:
                if list_of_provider_elements.text == "No data available":
                    raise Exception("No Providers!!")
                selected_provider = list_of_provider_elements[0].find_elements_by_tag_name('a')[1]



            global global_search_prov
            global_search_prov = selected_provider.text
            selected_provider.click()
            sf.ajax_preloader_wait(driver)
            WebDriverWait(driver, 30).until(
                EC.presence_of_element_located((By.XPATH, locator.xpath_filter_measure_list)))
            sf.captureScreenshot(driver, "Provider registry " + driver.find_element(By.XPATH, "//span[@class='specific_most']").text, report_folder)
        except Exception as e:
            ws.append(['1', "Attempting to navigate to a random provider", 'Navigation to provider context', 'Failed', "Either the Provider list is empty", driver.current_url])
            print(e)
            traceback.print_exc()
            return
    provider_registry_url = driver.current_url
    if len(driver.find_elements_by_xpath(locator.xpath_filter_measure_list)) > 0:
        logger.info("Nagivated to Provider Registry")
        # Now, Select a random metric that is not 0/0
        try:
            window_switched = 0
            metrics = driver.find_element_by_id("registry_body").find_elements_by_tag_name('li')
            percent = '0.00'
            loop_count = 0
            if driver.find_element(By.XPATH, "//*[@id='conti_enroll']").is_selected():
                driver.find_element(By.XPATH, "//*[@class='cont_disc_toggle']").click()
            selected_metric_name = ""
            driver.find_element(By.XPATH, "//a[@data-target='qt-reg-nav-filters']").click()
            time.sleep(0.5)
            option = driver.find_elements(By.XPATH, "//input[@class='select-dropdown dropdown-trigger']")[0]
            option.click()
            lists = driver.find_elements(By.XPATH, "//ul[@class='dropdown-content select-dropdown']")[0]
            options = lists.find_elements(By.TAG_NAME, 'li')
            for option in options:
                if option.text == "Numerator":
                    option.click()
            apply_btn = driver.find_element(By.XPATH, "//button[@id='qt-apply-search']")
            driver.execute_script("arguments[0].scrollIntoView();", apply_btn)
            apply_btn.click()
            time.sleep(0.5)
            selectedMetric = driver.find_element(By.XPATH, "//*[@id='registry_body']/ul").find_elements(By.TAG_NAME, "li")
            selected_metric_name = selectedMetric[0].find_element(By.XPATH, "//*[@class='met-name']").text
            selectedMetric[0].click()
            sf.ajax_preloader_wait(driver)
            WebDriverWait(driver, 30).until(
                EC.presence_of_element_located((By.CLASS_NAME, 'tabs')))
            mspl_link = driver.current_url
            sf.captureScreenshot(driver, "MSPL from provider registry " + driver.find_element(By.XPATH, "//span[@class='specific_most']").text, report_folder)
            # now, at the MSPL, Process starts.
            # Nav 1 : Performance Statistics
            try:
                driver.find_element_by_class_name('tabs').find_elements_by_class_name('tab')[1].click()
                start_time = time.perf_counter()
                sf.ajax_preloader_wait(driver)
                WebDriverWait(driver, 30).until(
                    EC.presence_of_element_located((By.CLASS_NAME, 'tabs')))
                time_taken = round((time.perf_counter() - start_time), 3)
                if driver.find_elements_by_id('performance_details') != 0:
                    ws.append([test_case_id, global_search_prov,
                               'Navigation to Performance Stats from MSPL : ' + selected_metric_name,
                               'Passed', time_taken])
                    test_case_id += 1
                else:
                    ws.append([test_case_id, global_search_prov,
                               'Navigation to Performance Stats from MSPL : ' + selected_metric_name,
                               'Failed', '', 'Performance Ribbon Missing', driver.current_url])
                    test_case_id += 1
            except Exception as e:
                print(e)
                traceback.print_exc()
                ws.append([test_case_id, global_search_prov,
                           'Navigation to Performance Stats from MSPL : ' + selected_metric_name,
                           'Failed', '', 'Unable to click on performance Statistics tab', driver.current_url])
                test_case_id += 1

            # nav 2: Navigation to Network Comparision
            try:
                driver.find_element_by_class_name('tabs').find_elements_by_class_name('tab')[2].click()
                start_time = time.perf_counter()
                sf.ajax_preloader_wait(driver)
                WebDriverWait(driver, 30).until(
                    EC.presence_of_element_located((By.CLASS_NAME, 'tabs')))
                sf.captureScreenshot(driver, "Performance stat " + driver.find_element(By.XPATH, "//span[@class='specific_most']").text, report_folder)
                time_taken = round((time.perf_counter() - start_time), 3)
                if driver.find_elements_by_id('network-table_info') != 0:
                    data_info = driver.find_element_by_id('network-table_info').text
                    ws.append([test_case_id, global_search_prov,
                               'Navigation to Network Comparision from MSPL : ' + selected_metric_name,
                               'Passed', time_taken, data_info])
                    test_case_id += 1
                else:
                    ws.append([test_case_id, global_search_prov,
                               'Navigation to Network Comparision from MSPL : ' + selected_metric_name,
                               'Failed', '', '', driver.current_url])
                    test_case_id += 1
            except Exception as e:
                print(e)
                traceback.print_exc()
                ws.append([test_case_id, global_search_prov,
                           'Navigation to Network comparision from MSPL : ' + selected_metric_name,
                           'Failed', '', 'Unable to click on Network Comparision tab', driver.current_url])
                test_case_id += 1

            # Mspl to basic careops count checking
            driver.get(mspl_link)
            sf.ajax_preloader_wait(driver)
            WebDriverWait(driver, 30).until(
                EC.presence_of_element_located((By.CLASS_NAME, 'tabs')))
            window_switched = 0
            try:
                if len(driver.find_elements_by_class_name('dt_tag_value')) > 0:
                    driver.find_element_by_class_name('dt_tag_close').click()
                    sf.ajax_preloader_wait(driver)
                time.sleep(3)
                WebDriverWait(driver, 30).until(
                    EC.presence_of_element_located((By.ID, "quality_registry_list")))
                table = driver.find_element_by_id(
                    "quality_registry_list").find_element_by_tag_name(
                    "tbody").find_elements_by_tag_name('tr')
                if len(table) == 0:
                    ws.append(
                        [test_case_id, global_search_prov, 'Careops comparision between MSPL and Dashbaord', 'Failed',
                         '', 'MSPL is Empty for: ' + selected_metric_name, driver.current_url])
                    return

                chosen_patient = randint(0, len(table) - 1)
                print(chosen_patient)
                table = driver.find_element_by_id(
                    "quality_registry_list").find_element_by_tag_name(
                    "tbody").find_elements_by_tag_name('tr')
                # Picks up the Patient's name from the MSPL
                MSPLname = table[chosen_patient].find_element_by_tag_name("a").text
                # print(MSPLname)
                # Locates the caregaps, quality gaps and hcc gaps
                MSPL_Caregap_count = table[chosen_patient].find_element_by_class_name("care_ops").text
                MSPL_caregap_hover = table[chosen_patient].find_element_by_class_name(
                    "care_ops").find_element_by_class_name(
                    "tooltipped").get_attribute(
                    "data-tooltip")
                MSPL_Quality_Gaps = table[chosen_patient].find_element_by_class_name(
                    "measure_abbr_group_pt").get_attribute(
                    "innerHTML")
                table = driver.find_element_by_id(
                    "quality_registry_list").find_element_by_tag_name(
                    "tbody").find_elements_by_tag_name('tr')
                MSPL_HCC_Gaps = table[chosen_patient].find_element_by_class_name(
                    "condition_abbr_group").get_attribute(
                    "innerHTML")
                # print("MSPL Caregaps = "+str(MSPL_Caregap_count))
                # print("MSPL Caregaps OnHover = "+MSPL_caregap_hover)
                # print("MSPL hover Quality Gaps = "+MSPL_Quality_Gaps)
                # print("MSPL hover HCC Gaps = "+MSPL_HCC_Gaps)
                MSPL_Quality_Gaps = MSPL_Quality_Gaps.split(',')
                MSPL_HCC_Gaps = MSPL_HCC_Gaps.split(',')
                MSPL_caregap_hover_count = str(len(MSPL_caregap_hover.split(',')))
                # print("MSPL Caregaps OnHover Count = " + str(MSPL_caregap_hover_count))

                # TEST CASE 1--------------------------------------------------------------------
                # print("TEST CASE 1 - MSPL CARE GAP VS # OF GAPS ON HOVER :", end=" ")
                # MSPL_Caregap_count = 21
                if MSPL_Caregap_count == MSPL_caregap_hover_count:
                    test_case_id += 1
                    ws.append((test_case_id,
                               str(MSPLname), "MSPL CARE GAP VS # OF GAPS ON HOVER",
                               'Passed', '',
                               "Caregap count on MSPL: " + MSPL_Caregap_count + " and, Number of caregaps present on hover: " + MSPL_caregap_hover_count + " . The Caregaps are: " + MSPL_caregap_hover))
                    # print("FAILED")))
                    # print("PASSED")
                    # print("TEST CASE 1 COMMENTS -", end=" ")
                    # print("Caregaps present : " + MSPL_caregap_hover)
                else:
                    test_case_id += 1
                    ws.append((test_case_id,
                               str(MSPLname), "MSPL CARE GAP VS # OF GAPS ON HOVER",
                               'Failed', '',
                               "Count mismatch between hover and caregaps, Caregap count on MSPL: " + MSPL_Caregap_count + " and, Number of caregaps present on hover: " + MSPL_caregap_hover_count + " . The Caregaps are: " + MSPL_caregap_hover, driver.current_url))
                    # print("FAILED")
                    # print("TEST CASE 1 COMMENTS -", end=" ")
                    # print("Count mismatch between hover and caregaps")
                # TEST CASE 1--------------------------------------------------------------------

                for x in range(0, len(MSPL_Quality_Gaps)):
                    MSPL_Quality_Gaps[x] = MSPL_Quality_Gaps[x].strip()
                for x in range(0, len(MSPL_HCC_Gaps)):
                    MSPL_HCC_Gaps[x] = MSPL_HCC_Gaps[x].strip()
                MSPL_Quality_Gaps.sort()
                MSPL_HCC_Gaps.sort()
                if len(MSPL_HCC_Gaps) == 1:
                    if MSPL_HCC_Gaps[0] == "":
                        MSPL_HCC_Gaps.clear()
                if len(MSPL_Quality_Gaps) == 1:
                    if MSPL_Quality_Gaps[0] == "":
                        MSPL_Quality_Gaps.clear()

                # for x in MSPL_Quality_Gaps:
                #     print("Quality gap = "+x)
                # for x in MSPL_HCC_Gaps:
                #     print("HCC gap = "+x)

                # print("After List conversion lengths, Quality has "+str(len(MSPL_Quality_Gaps))+" Measures and hcc has
                # "+str(len(MSPL_HCC_Gaps))+" Measures") patientDashboard
                table = driver.find_element_by_id(
                    "quality_registry_list").find_element_by_tag_name(
                    "tbody").find_elements_by_tag_name('tr')
                table[chosen_patient].find_element_by_class_name("pat_name").click()

                # -------------------------Window Switch---------------------------
                # print("current window is "+driver.title)
                # parent_window = driver.current_window_handle
                window_switched = 0
                driver.switch_to.window(driver.window_handles[1])
                sf.ajax_preloader_wait(driver)
                WebDriverWait
                window_switched = 1
                # print("current window is "+driver.title)
                # -------------------------Window Switch---------------------------
                WebDriverWait(driver, 30).until(
                    EC.presence_of_element_located((By.CLASS_NAME, "primary_val")))
                sf.captureScreenshot(driver, "Patient Dashboard " + run_from, report_folder)
                Dashboard_caregap = driver.find_element_by_class_name("primary_val").text
                # print("Dashboard caregaps = "+Dashboard_caregap)
                driver.find_element_by_class_name("select-dropdown").click()
                dropdown_contents = driver.find_element_by_class_name("filter-panel").find_elements_by_tag_name(
                    "li")
                for dropdowntext in dropdown_contents:
                    if dropdowntext.text == "Non-Compliant":
                        dropdowntext.click()
                        break

                # TEST CASE 2--------------------------------------------------------------
                # print("TEST CASE 2 - MSPL CAREGAP COUNT VS DASHBOARD HEADER CAREOPS COUNT :", end=" ")
                if Dashboard_caregap == MSPL_caregap_hover_count:
                    test_case_id += 1
                    ws.append((test_case_id,
                               str(MSPLname), "MSPL CAREGAP COUNT VS PATIENT DASHBOARD HEADER CAREOPS COUNT",
                               'Passed', '', "Careop Count : " + Dashboard_caregap))
                    # print("PASSED")
                else:
                    test_case_id += 1
                    ws.append((test_case_id,
                               str(MSPLname), "MSPL CAREGAP COUNT VS PATIENT DASHBOARD HEADER CAREOPS COUNT",
                               'Failed', '', "Careop Count : " + Dashboard_caregap, driver.current_url))
                    # print("FAILED")
                # print("TEST CASE 2 COMMENTS -", end=" ")
                # print("Careop Count : " + Dashboard_caregap)
                # TEST CASE 2-----------------------------------------------------------------
                Dashboard_quality_List = []
                Dashboard_HCC_List = []

                # Listize HCC Gaps------------------------------------------------------
                if sf.check_exists_by_id(driver, "table_4"):
                    hcctable = driver.find_element_by_id("table_4").find_elements_by_class_name(
                        "compliant_true")
                    # print("Dashboard Hcc Count = "+str(len(hcctable)))
                    for hcc_measures in hcctable:
                        hcc_abr = ""
                        for hcc_abr_text in hcc_measures.find_element_by_class_name(
                                "hcc_details").find_element_by_tag_name(
                            "span").text:
                            if hcc_abr_text != 'H' and hcc_abr_text != 'C' and hcc_abr_text != 'C':
                                hcc_abr = hcc_abr + hcc_abr_text
                        Dashboard_HCC_List.append("#" + hcc_abr.strip())
                    Dashboard_HCC_List.sort()
                    # for x in Dashboard_HCC_List:
                    #     print("Dashboard HCC = "+x)

                # Listize Quality Gaps----------------------------------------------------
                if sf.check_exists_by_id(driver, "table_1"):
                    qualitytable = driver.find_element_by_id("table_1").find_elements_by_class_name(
                        "compliant_true")
                    # print("Dashboard Quality Gap count = "+str(len(qualitytable)))
                    for quality_measures in qualitytable:
                        quality_abr = ""
                        for Quality_abr_text in quality_measures.find_element_by_class_name("tiny-text").text:
                            if Quality_abr_text == '·':
                                break
                            else:
                                quality_abr = quality_abr + Quality_abr_text
                        Dashboard_quality_List.append(quality_abr.strip())
                    Dashboard_quality_List.sort()
                    # for x in Dashboard_quality_List:
                    #     print("Dashboard Quality = "+x)

                # TEST CASE 3--------------------------------------------------------------
                # print("TEST CASE 3 - NUMBER OF MEASURES ON DASHBOARD VS MSPL CARE GAP COUNT :", end=" ")
                if MSPL_Caregap_count == str(len(Dashboard_quality_List) + len(Dashboard_HCC_List)):
                    test_case_id += 1
                    ws.append((test_case_id,
                               str(MSPLname), "NUMBER OF MEASURES ON PATIENT DASHBOARD VS MSPL CARE GAP COUNT",
                               'Passed', '', "Measures on dashboard : " + str(
                        len(Dashboard_quality_List) + len(Dashboard_HCC_List))))
                    # print("PASSED")
                    # print("TEST CASE 3 COMMENTS -", end=" ")
                    # print("Measures on dashboard : " + str(len(Dashboard_quality_List) + len(Dashboard_HCC_List)))
                else:
                    test_case_id += 1
                    ws.append((test_case_id,
                               str(MSPLname), "NUMBER OF MEASURES ON PATIENT DASHBOARD VS MSPL CARE GAP COUNT",
                               'Failed', '',
                               "Measures on dashboard : " + str(
                                   len(Dashboard_quality_List) + len(Dashboard_HCC_List)) +
                               ". MSPL caregap count :" + MSPL_Caregap_count + ". For a difference of " +
                               str(abs(int(MSPL_Caregap_count) - (
                                       len(Dashboard_quality_List) + len(Dashboard_HCC_List)))), driver.current_url))
                    # print("FAILED")
                    # print("TEST CASE 3 COMMENTS -", end=" ")
                    # print("Measures on dashboard : " + str(
                    #     len(Dashboard_quality_List) + len(Dashboard_HCC_List)) + ". MSPL caregap count :" + str(
                    #     MSPL_Caregap_count) + ". For a difference of " + str(
                    #     abs(MSPL_Caregap_count - (len(Dashboard_quality_List) + len(Dashboard_HCC_List)))))
                # TEST CASE 3------------------------------------------------------------------
                # TEST CASE 4------------------------------------------------------------------
                # print("TEST CASE 4 - COMPARISION BETWEEN QUALITY GAPS ON MSPL HOVER AND DASHBOARD :", end=" ")
                if len(MSPL_Quality_Gaps) == len(Dashboard_quality_List):
                    flag = 0
                    for x in range(0, len(MSPL_Quality_Gaps)):
                        if MSPL_Quality_Gaps[x] != Dashboard_quality_List[x]:
                            Different_Measure = list(set(MSPL_Quality_Gaps) ^ set(Dashboard_quality_List))
                            flag = 1
                            break
                    if flag == 1:
                        test_case_id += 1
                        ws.append((test_case_id,
                                   str(MSPLname),
                                   "COMPARISION BETWEEN QUALITY GAPS ON MSPL HOVER AND PATIENT DASHBOARD",
                                   'Failed', '',
                                   str(len(MSPL_Quality_Gaps)) + " of " + str(len(Dashboard_quality_List)) +
                                   " Measures. Different Measures are " + ', '.join(
                                       map(str, Different_Measure)), driver.current_url))
                        # print("FAILED")
                        # print("TEST CASE 4 COMMENTS -", end=" ")
                        # print(str(len(MSPL_Quality_Gaps)) + " of " + str(
                        #     len(Dashboard_quality_List)) + " Measures. Different Measures are " + ', '.join(
                        #     map(str, Different_Measure)))
                    elif flag == 0:
                        test_case_id += 1
                        ws.append((test_case_id,
                                   str(MSPLname),
                                   "COMPARISION BETWEEN QUALITY GAPS ON MSPL HOVER AND PATIENT DASHBOARD",
                                   'Passed', '',
                                   str(len(MSPL_Quality_Gaps)) + " of " + str(len(Dashboard_quality_List)) +
                                   " Measures which are " + ', '.join(map(str, Dashboard_quality_List))))
                        # print("PASSED")
                        Different_Measure = list(set(MSPL_Quality_Gaps) ^ set(Dashboard_quality_List))
                        # print("TEST CASE 4 COMMENTS -", end=" ")
                        # print(str(len(MSPL_Quality_Gaps)) + " of " + str(
                        #     len(Dashboard_quality_List)) + " Measures which are " + ', '.join(
                        #     map(str, Dashboard_quality_List)))
                else:
                    Different_Measure = list(set(MSPL_Quality_Gaps) ^ set(Dashboard_quality_List))
                    test_case_id += 1
                    ws.append((test_case_id,
                               str(MSPLname), "COMPARISION BETWEEN QUALITY GAPS ON MSPL HOVER AND PATIENT DASHBOARD",
                               'Failed', '',
                               str(len(MSPL_Quality_Gaps)) + " of " + str(len(Dashboard_quality_List)) +
                               " Measures. Different Measures are " + ', '.join(map(str, Different_Measure)), driver.current_url))
                    # print("FAILED")
                    # print("TEST CASE 4 COMMENTS -", end=" ")
                    # print(str(len(MSPL_Quality_Gaps)) + " of " + str(len(Dashboard_quality_List))
                    #       + " Measures. Different Measures are " + ', '.join(map(str, Different_Measure)))
                # TEST CASE 4------------------------------------------------------------------
                # Dashboard_HCC_List.append("#45")
                # TEST CASE 5------------------------------------------------------------------
                # print("TEST CASE 5 - COMPARISION BETWEEN HCC GAPS ON MSPL HOVER AND DASHBOARD :", end=" ")
                if len(MSPL_HCC_Gaps) == len(Dashboard_HCC_List):
                    flag = 0
                    for x in range(0, len(MSPL_HCC_Gaps)):
                        if MSPL_HCC_Gaps[x] != Dashboard_HCC_List[x]:
                            Different_Measure = list(set(MSPL_HCC_Gaps) ^ set(Dashboard_HCC_List))
                            flag = 1
                            break
                    if flag == 1:
                        test_case_id += 1
                        ws.append((test_case_id,
                                   str(MSPLname), "COMPARISION BETWEEN HCC GAPS ON MSPL HOVER AND PATIENT DASHBOARD",
                                   'Failed', '', str(len(MSPL_HCC_Gaps)) + " of " + str(len(Dashboard_HCC_List)) +
                                   " Measures. Different Measures are " + ', '.join(
                            map(str, Different_Measure)), driver.current_url))
                        # print("FAILED")
                        # print("TEST CASE 5 COMMENTS -", end=" ")
                        # print(str(len(MSPL_HCC_Gaps)) + " of " + str(
                        #     len(Dashboard_HCC_List)) + " Measures. Different Measures are " + ', '.join(
                        #     map(str, Different_Measure)))
                    elif flag == 0 & len(MSPL_HCC_Gaps) == 0:
                        test_case_id += 1
                        ws.append((test_case_id,
                                   str(MSPLname), "COMPARISION BETWEEN HCC GAPS ON MSPL HOVER AND PATIENT DASHBOARD",
                                   'Passed', '', str(len(MSPL_HCC_Gaps)) + " of " +
                                   str(len(Dashboard_HCC_List)) + " Measures"))
                        # print("PASSED")
                        # print("TEST CASE 5 COMMENTS -", end=" ")
                        # print(str(len(MSPL_HCC_Gaps)) + " of " + str(
                        #     len(Dashboard_HCC_List)) + " Measures")
                    elif flag == 0:
                        test_case_id += 1
                        ws.append((test_case_id,
                                   str(MSPLname), "COMPARISION BETWEEN HCC GAPS ON MSPL HOVER AND PATIENT DASHBOARD",
                                   'Passed', '', str(len(MSPL_HCC_Gaps)) + " of " + str(len(Dashboard_HCC_List)) +
                                   " Measures which are " + ', '.join(map(str, Dashboard_HCC_List))))
                        # print("PASSED")
                        # print("TEST CASE 5 COMMENTS -", end=" ")
                        # print(str(len(MSPL_HCC_Gaps)) + " of " + str(
                        #     len(Dashboard_HCC_List)) + " Measures which are " + ', '.join(
                        #     map(str, Dashboard_HCC_List)))
                else:
                    Different_Measure = list(set(MSPL_HCC_Gaps) ^ set(Dashboard_HCC_List))
                    test_case_id += 1
                    ws.append((test_case_id,
                               str(MSPLname), "COMPARISION BETWEEN HCC GAPS ON MSPL HOVER AND PATIENT DASHBOARD",
                               'Failed', '', str(len(MSPL_HCC_Gaps)) + " of " + str(len(Dashboard_HCC_List)) +
                               " Measures. Different Measures are " + ', '.join(map(str, Different_Measure)), driver.current_url))
                    # print("FAILED")
                    Different_Measure = list(set(MSPL_HCC_Gaps) ^ set(Dashboard_HCC_List))
                    # print("TEST CASE 5 COMMENTS -", end=" ")
                    # print(str(len(MSPL_HCC_Gaps)) + " of " + str(
                    #     len(Dashboard_HCC_List)) + " Measures. Different Measures are " + ', '.join(
                    #     map(str, Different_Measure)))
                # TEST CASE 5------------------------------------------------------------------

                # -------------------------Window Switch---------------------------
                # print("current window is "+driver.title)
                driver.close()
                driver.switch_to.window(driver.window_handles[0])
                window_switched=0
                sf.ajax_preloader_wait(driver)
                # print("current window is "+driver.title)
                # -------------------------Window Switch---------------------------
                # ASPYEDIT END-----------------------------------------------------------------------------------------------



            except Exception as e:
                print(e)
                traceback.print_exc()
                ws.append([test_case_id, global_search_prov, 'Careops comparision from mspl: ' + selected_metric_name,
                           'Failed', '', driver.current_url])
                if window_switched == 1:
                    driver.switch_to.window(driver.window_handles[0])
                    sf.ajax_preloader_wait(driver)
                    window_switched = 0

        except Exception as e:
            print(e)
            traceback.print_exc()
            logger.critical("Unable to click on a metric from the provider dashboard")
            ws.append(
                [test_case_id, str(selected_provider), "Attempting to click on metric in dashboard: " + str(selected_metric_name),
                 'Failed', '', 'Unable to click on metric', driver.current_url])
            test_case_id += 1
            if window_switched == 1:
                driver.switch_to.window(driver.window_handles[0])
                sf.ajax_preloader_wait(driver)
                window_switched = 0

    else:
        logger.info("Provider Registry Navigation Failed")

    driver.get(main_registry_url)
    sf.ajax_preloader_wait(driver)
    WebDriverWait(driver, 30).until(
        EC.presence_of_element_located((By.XPATH, locator.xpath_filter_measure_list)))


    rows = ws.max_row
    cols = ws.max_column
    for i in range(2, rows + 1):
        for j in range(3, cols + 1):
            if ws.cell(i, j).value == 'Passed':
                ws.cell(i, j).fill = PatternFill('solid', fgColor='0FC404')
            elif ws.cell(i, j).value == 'Failed':
                ws.cell(i, j).fill = PatternFill('solid', fgColor='FC0E03')
            elif ws.cell(i, j).value == 'Data table is empty':
                ws.cell(i, j).fill = PatternFill('solid', fgColor='FCC0BB')


def time_capsule(driver, workbook, logger, run_from, report_folder):
    workbook.create_sheet('Time Capsule')
    ws = workbook['Time Capsule']

    ws.append(['ID', 'Context', 'Scenario', 'Status', 'Comments'])
    header_font = Font(color='FFFFFF', bold=False, size=12)
    header_cell_color = PatternFill('solid', fgColor='030303')
    ws['A1'].font = header_font
    ws['A1'].fill = header_cell_color
    ws['B1'].font = header_font
    ws['B1'].fill = header_cell_color
    ws['C1'].font = header_font
    ws['C1'].fill = header_cell_color
    ws['D1'].font = header_font
    ws['D1'].fill = header_cell_color
    ws['E1'].font = header_font
    ws['E1'].fill = header_cell_color

    ws.name = "Arial"
    test_case_id = 1
    try:
        last_url = driver.current_url
        window_switched = 0
        driver.find_element_by_xpath(locator.xpath_app_Tray_Link).click()
        driver.find_element_by_xpath(locator.xpath_app_Time_Capsule).click()
        driver.switch_to.window(driver.window_handles[1])
        window_switched = 1
        sf.ajax_preloader_wait(driver)
        try:
            sf.ajax_preloader_wait(driver)
            current_url = driver.current_url
            sf.captureScreenshot(driver, "Time Capsule " + driver.find_element(By.XPATH, "//span[@class='specific_most']").text, report_folder)
            access_message = sf.CheckAccessDenied(current_url)

            if access_message == 1:
                print("Access Denied found!")
                # logger.critical("Access Denied found!")
                test_case_id+= 1
                ws.append(
                    (test_case_id, 'Time Capsule', 'Access check for Time Capsule', 'Failed', 'Access Denied', driver.current_url))

            else:
                print("Access Check done!")
                # logger.info("Access Check done!")
                error_message = sf.CheckErrorMessage(driver)

                if error_message == 1:
                    print("Error toast message is displayed")
                    # logger.critical("ERROR TOAST MESSAGE IS DISPLAYED!")
                    test_case_id += 1
                    ws.append((test_case_id, 'Time Capsule', 'Navigation to Time Capsule without error message',
                                'Failed', 'Error toast message is displayed', driver.current_url))

                else:
                    ws.append((test_case_id, 'Time Capsule', 'Time Capsule page loading',
                                    'Passed'))
                    if len(driver.find_elements_by_xpath(locator.xpath_latest_Card_Title)) != 0:
                        latest_computation_dete = driver.find_element_by_xpath(
                            locator.xpath_latest_Card_Title).text
                        test_case_id += 1
                        ws.append((test_case_id, 'Time Capsule', 'Computation Card',
                                    'Passed', 'Latest Computation date: ' + latest_computation_dete))
                    else:
                        test_case_id += 1
                        ws.append((test_case_id, 'Time Capsule', 'Computation Card', 'Failed',
                                    'Computation card details is not found!', driver.current_url))

                    # time capsule cards comparision, Making variables
                    measure_performance = []
                    claim_count = []
                    claim_provider_count = []
                    claim_member_count = []
                    pmc_count = []
                    lab_count = []
                    lab_members_count = []

                    cards = driver.find_element_by_class_name("card_section_wrapper").find_elements_by_class_name("card_wrapper")
                    cards = cards[0:4]
                    for card in cards:
                        measure_performance.append(str(card.find_element_by_class_name("card_body_value").text))
                        card.click()
                        time.sleep(2)
                        sf.ajax_preloader_wait(driver)
                        WebDriverWait(driver, 60).until(
                            EC.presence_of_element_located((By.CLASS_NAME, 'card_details_body')))
                        card_details = driver.find_element_by_id('card_details_modal')
                        claim_count.append(str(
                            card_details.find_element_by_class_name('card_claims').find_elements_by_tag_name('span')[
                                1].text))
                        claim_provider_count.append(str(
                            card_details.find_element_by_class_name('card_providers').find_elements_by_tag_name('span')[
                                1].text))
                        claim_member_count.append(str(
                            card_details.find_element_by_class_name('card_members').find_elements_by_tag_name('span')[
                                1].text))
                        pmc_count.append(str(
                            card_details.find_element_by_class_name('card_pmc').find_elements_by_tag_name('span')[
                                1].text))
                        lab_count.append(str(
                            card_details.find_element_by_class_name('card_labs').find_elements_by_tag_name('span')[
                                1].text))
                        lab_members_count.append(str(
                            card_details.find_element_by_class_name('card_lab_members').find_elements_by_tag_name('span')[
                                1].text))

                        measure_performance_copy = [*set(measure_performance)]
                        claim_count_copy = [*set(claim_count)]
                        claim_member_count_copy = [*set(claim_member_count)]
                        claim_provider_count_copy = [*set(claim_provider_count)]
                        pmc_count_copy = [*set(pmc_count)]
                        lab_count_copy = [*set(lab_count)]
                        lab_members_count_copy = [*set(lab_members_count)]

                        #add appends
                    print(measure_performance)
                    print(claim_count)
                    print(pmc_count)
                    if len(measure_performance_copy) == len(measure_performance):
                        ws.append((test_case_id, 'Time Capsule', 'Measure Performance Duplicates', 'Passed',
                                   'All cards have different values', driver.current_url))
                    else:
                        ws.append((test_case_id, 'Time Capsule', 'Measure Performance Duplicates', 'Failed',
                                   'Unique values: \"'+'\",\"'.join([str(elem) for elem in measure_performance_copy])+'\"', driver.current_url))
                    test_case_id+=1
                    if len(claim_count_copy) == len(claim_count):
                        ws.append((test_case_id, 'Time Capsule', 'Claim count Duplicates', 'Passed',
                                   'All cards have different values', driver.current_url))
                    else:
                        ws.append((test_case_id, 'Time Capsule', 'Claim count Duplicates', 'Failed',
                                   'Unique values: \"'+'\",\"'.join([str(elem) for elem in claim_count_copy])+'\"', driver.current_url))
                    test_case_id+=1
                    if len(claim_provider_count_copy) == len(claim_provider_count):
                        ws.append((test_case_id, 'Time Capsule', 'Claim Provider Count Duplicates', 'Passed',
                                   'All cards have different values', driver.current_url))
                    else:
                        ws.append((test_case_id, 'Time Capsule', 'Claim Provider Count Duplicates', 'Failed',
                                   'Unique values: \"'+'\",\"'.join([str(elem) for elem in claim_provider_count_copy])+'\"', driver.current_url))
                    test_case_id+=1
                    if len(claim_member_count_copy) == len(claim_member_count):
                        ws.append((test_case_id, 'Time Capsule', 'Claim Member Count Duplicates', 'Passed',
                                   'All cards have different values', driver.current_url))
                    else:
                        ws.append((test_case_id, 'Time Capsule', 'Claim Member Count Duplicates', 'Failed',
                                   'Unique values: \"'+'\",\"'.join([str(elem) for elem in claim_member_count_copy])+'\"', driver.current_url))
                    test_case_id+=1
                    if len(pmc_count_copy) == len(pmc_count):
                        ws.append((test_case_id, 'Time Capsule', 'PMC Count Duplicates', 'Passed',
                                   'All cards have different values', driver.current_url))
                    else:
                        ws.append((test_case_id, 'Time Capsule', 'PMC Count Duplicates', 'Failed',
                                   'Unique values: \"'+'\",\"'.join([str(elem) for elem in pmc_count_copy])+'\"', driver.current_url))
                    test_case_id+=1
                    if len(lab_count_copy) == len(lab_count):
                        ws.append((test_case_id, 'Time Capsule', 'Lab Count Duplicates', 'Passed',
                                   'All cards have different values', driver.current_url))
                    else:
                        ws.append((test_case_id, 'Time Capsule', 'Lab Count Duplicates', 'Failed',
                                   'Unique values: \"'+'\",\"'.join([str(elem) for elem in lab_count_copy])+'\"', driver.current_url))
                    test_case_id+=1
                    if len(lab_members_count_copy) == len(lab_members_count):
                        ws.append((test_case_id, 'Time Capsule', 'Lab Members Count Duplicates', 'Passed',
                                   'All cards have different values', driver.current_url))
                    else:
                        ws.append((test_case_id, 'Time Capsule', 'Lab Members Count Duplicates', 'Failed',
                                   'Unique values: \"'+'\",\"'.join([str(elem) for elem in lab_members_count_copy])+'\"', driver.current_url))
                    test_case_id+=1





                    #populate these counts to the 4th capsule card and then make copies of the list with _copy
                    #remove duplicates and compare the list lengths. If length is different, fail the test case


        except Exception as e:
            print(e)
            test_case_id += 1
            ws.append(
                (test_case_id, 'Time Capsule', 'Navigation to Time Capsule', 'Failed', 'Exception occurred!', driver.current_url))
            traceback.print_exc()
        finally:
            driver.close()
            time.sleep(1)
            if window_switched == 1:
                driver.switch_to.window(driver.window_handles[0])

    except Exception as e:
        print(e)
        test_case_id += 1
        ws.append(
            (test_case_id, 'Time Capsule', 'Navigation to Time Capsule', 'Failed', 'Exception occurred!', driver.current_url))
        driver.get(last_url)
        sf.ajax_preloader_wait(driver)
        WebDriverWait(driver, 30).until(
            EC.presence_of_element_located((By.XPATH, locator.xpath_app_Tray_Link)))

    rows = ws.max_row
    cols = ws.max_column
    for i in range(2, rows + 1):
        for j in range(3, cols + 1):
            if ws.cell(i, j).value == 'Passed':
                ws.cell(i, j).fill = PatternFill('solid', fgColor='0FC404')
            elif ws.cell(i, j).value == 'Failed':
                ws.cell(i, j).fill = PatternFill('solid', fgColor='FC0E03')
            elif ws.cell(i, j).value == 'Data table is empty':
                ws.cell(i, j).fill = PatternFill('solid', fgColor='FCC0BB')


def secure_messaging(driver, workbook, logger, run_from, report_folder):
    workbook.create_sheet('Secure Messaging')
    ws = workbook['Secure Messaging']

    ws.append(['ID', 'Context', 'Scenario', 'Status', 'Comments'])
    header_font = Font(color='FFFFFF', bold=False, size=12)
    header_cell_color = PatternFill('solid', fgColor='030303')
    ws['A1'].font = header_font
    ws['A1'].fill = header_cell_color
    ws['B1'].font = header_font
    ws['B1'].fill = header_cell_color
    ws['C1'].font = header_font
    ws['C1'].fill = header_cell_color
    ws['D1'].font = header_font
    ws['D1'].fill = header_cell_color
    ws['E1'].font = header_font
    ws['E1'].fill = header_cell_color
    test_case_ID = 0

    ws.name = "Arial"

    window_switched = 0
    try:
        last_url = driver.current_url
        WebDriverWait(driver, 30).until(
            EC.presence_of_element_located((By.XPATH, locator.xpath_app_Tray_Link)))
        driver.find_element_by_xpath(locator.xpath_app_Tray_Link).click()
        WebDriverWait(driver, 30).until(
            EC.presence_of_element_located((By.XPATH, locator.xpath_app_Secure_Messaging)))
        driver.find_element_by_xpath(locator.xpath_app_Secure_Messaging).click()
        driver.switch_to.window(driver.window_handles[1])
        window_switched = 1
        try:
            sf.ajax_preloader_wait(driver)
            current_url = driver.current_url
            sf.captureScreenshot(driver, "Secure Messaging " + driver.find_element(By.XPATH, "//span[@class='specific_most']").text, report_folder)
            access_message = sf.CheckAccessDenied(current_url)

            if access_message == 1:
                print("Access Denied found!")
                # logger.critical("Access Denied found!")
                test_case_ID += 1
                ws.append(
                    (test_case_ID, 'Secure Messaging', 'Access check for Secure Messaging', 'Failed',
                     'Access Denied', driver.current_url))

            else:
                print("Access Check done!")
                # logger.info("Access Check done!")
                error_message = sf.CheckErrorMessage(driver)

                if error_message == 1:
                    print("Error toast message is displayed")
                    # logger.critical("ERROR TOAST MESSAGE IS DISPLAYED!")
                    test_case_ID += 1
                    ws.append(
                        (test_case_ID, 'Secure Messaging',
                         'Navigation to Secure Messaging without error message',
                         'Failed', 'Error toast message is displayed', driver.current_url))

                else:
                    total_inbox_messages = len(driver.find_elements_by_xpath(locator.xpath_inbox_Message))
                    test_case_ID += 1
                    ws.append((test_case_ID, 'Secure Messaging', 'Navigation to Secure Messaging', 'Passed',
                                '[Inbox]Number of messages in the first page: ' + str(total_inbox_messages)))
        except Exception as e:
            print(e)
            test_case_ID += 1
            ws.append((test_case_ID, 'Secure Messaging', 'Navigation to Secure Messaging', 'Failed',
                        'Exception occurred!', driver.current_url))
        finally:
            driver.close()
            time.sleep(1)
            driver.switch_to.window(driver.window_handles[0])

    except Exception as e:
        print(e)
        test_case_ID += 1
        ws.append((test_case_ID, 'Secure Messaging', 'Navigation to Secure Messaging', 'Failed',
                    'Exception occurred!', driver.current_url))
        driver.get(last_url)
        sf.ajax_preloader_wait(driver)
        WebDriverWait(driver, 30).until(
            EC.presence_of_element_located((By.XPATH, locator.xpath_app_Tray_Link)))


    rows = ws.max_row
    cols = ws.max_column
    for i in range(2, rows + 1):
        for j in range(3, cols + 1):
            if ws.cell(i, j).value == 'Passed':
                ws.cell(i, j).fill = PatternFill('solid', fgColor='0FC404')
            elif ws.cell(i, j).value == 'Failed':
                ws.cell(i, j).fill = PatternFill('solid', fgColor='FC0E03')
            elif ws.cell(i, j).value == 'Data table is empty':
                ws.cell(i, j).fill = PatternFill('solid', fgColor='FCC0BB')


def analytics(driver, workbook, logger, run_from, report_folder):
    workbook.create_sheet('Analytics')
    ws = workbook['Analytics']

    ws.append(['ID', 'Context', 'Scenario', 'Status', 'Time Taken', 'Comments'])
    header_font = Font(color='FFFFFF', bold=False, size=12)
    header_cell_color = PatternFill('solid', fgColor='030303')
    ws['A1'].font = header_font
    ws['A1'].fill = header_cell_color
    ws['B1'].font = header_font
    ws['B1'].fill = header_cell_color
    ws['C1'].font = header_font
    ws['C1'].fill = header_cell_color
    ws['D1'].font = header_font
    ws['D1'].fill = header_cell_color
    ws['E1'].font = header_font
    ws['E1'].fill = header_cell_color
    ws['F1'].font = header_font
    ws['F1'].fill = header_cell_color
    ws.name = "Arial"
    test_case_id = 1
    last_url = driver.current_url

    try:
        WebDriverWait(driver, 30).until(EC.presence_of_element_located((By.XPATH, locator.xpath_app_Tray_Link)))
        driver.find_element_by_xpath(locator.xpath_app_Tray_Link).click()
        WebDriverWait(driver, 30).until(
            EC.presence_of_element_located((By.XPATH, locator.xpath_app_Analytics)))
        driver.find_element_by_xpath(locator.xpath_app_Analytics).click()
        driver.switch_to.window(driver.window_handles[1])
        try:
            sf.ajax_preloader_wait(driver)
            current_url = driver.current_url
            sf.captureScreenshot(driver, "Analytics main page " + driver.find_element(By.XPATH, "//span[@class='specific_most']").text, report_folder)
            access_message = sf.CheckAccessDenied(current_url)

            if access_message == 1:
                print("Access Denied found!")
                # logger.critical("Access Denied found!")
                test_case_id += 1
                ws.append(
                    (test_case_id, 'Analytics', 'Access check for Analytics', 'Failed', 'x', 'Access Denied', driver.current_url))

            else:
                print("Access Check done!")
                # logger.info("Access Check done!")
                error_message = sf.CheckErrorMessage(driver)

                if error_message == 1:
                    print("Error toast message is displayed")
                    # logger.critical("ERROR TOAST MESSAGE IS DISPLAYED!")
                    test_case_id += 1
                    ws.append((test_case_id, 'Analytics', 'Navigation to Analytics without error message',
                               'Failed', 'x', 'Error toast message is displayed', driver.current_url))

                else:

                    total_workbooks = len(driver.find_elements_by_xpath(locator.xpath_total_Workbooks))
                    all_workbooks = driver.find_elements_by_xpath(locator.xpath_total_Workbooks)
                    test_case_id += 1
                    ws.append((test_case_id, 'Analytics', 'Navigation to Analytics', 'Passed', 'x',
                               'Number of Workbook links: ' + str(total_workbooks)))
                    workbook_link = 0
                    while workbook_link < len(all_workbooks):
                        WebDriverWait(driver, 60).until(
                            EC.presence_of_element_located((By.XPATH, "//tr[@worksheet_title='Quality Overview']")))
                        workbook_name = (all_workbooks[workbook_link]).text
                        all_workbooks[workbook_link].click()
                        start_time = time.perf_counter()
                        print(workbook_name)
                        try:
                            WebDriverWait(driver, 100).until(EC.invisibility_of_element_located(
                                (By.XPATH, "// div[@class ='sm_download_cssload_loader']")))
                            WebDriverWait(driver, 30).until(
                                EC.presence_of_element_located((By.XPATH, "//a[@id='sm_back']")))
                            time_taken = time.perf_counter() - start_time
                            sf.captureScreenshot(driver, workbook_name,report_folder)
                            if len(driver.find_elements_by_xpath("//div[@class='nodata']")) == 0:
                                print(workbook_name + "Passed")
                                test_case_id += 1
                                ws.append((test_case_id, 'Analytics Workbook', workbook_name, 'Passed', time_taken, ''))
                            elif len(driver.find_elements_by_xpath("//div[@class='nodata']")) != 0:
                                test_case_id += 1
                                ws.append((test_case_id, 'Analytics Workbook', workbook_name, 'Failed', time_taken,
                                           'No data for the selected filters', driver.current_url))
                            # ASPY EDIT -------------------------------------------------------------------------------------------------------------------------
                            # '''
                            loader_element = 'sm_download_cssload_loader_wrap'
                            loader_element2 = 'toast sm_small_toast_message'
                            WebDriverWait(driver, 100).until(
                                EC.invisibility_of_element_located((By.CLASS_NAME, loader_element)))
                            WebDriverWait(driver, 100).until(
                                EC.invisibility_of_element_located((By.CLASS_NAME, loader_element2)))
                            # WebDriverWait(driver, 30).until(
                            # EC.element_to_be_clickable((By.XPATH, "// *[ @ id = 'sm_select_all'] / i")))
                            time.sleep(0.5)
                            select_all_present = True
                            try:
                                driver.find_element_by_xpath("// *[ @ id = 'sm_select_all'] / i").click()
                            except Exception as e:
                                print("No select all checkbox")
                                select_all_present = False
                            if select_all_present:
                                Drilldown_links = driver.find_element_by_class_name("breadcrumb_dropdown"). \
                                    find_elements_by_tag_name("a")
                            print("hello???")
                            # Next_drilldown_present = True
                            Drilldown_links.pop(0)
                            for link in Drilldown_links:
                                if select_all_present:
                                    link.click()
                                    start_time = time.perf_counter()
                                    WebDriverWait(driver, 500).until(
                                        EC.invisibility_of_element_located((By.CLASS_NAME, loader_element)))
                                    time_taken = time.perf_counter() - start_time
                                    WebDriverWait(driver, 100).until(
                                        EC.invisibility_of_element_located((By.CLASS_NAME, loader_element2)))
                                    Worksheet_name = link.text
                                    sf.captureScreenshot(driver, workbook_name + Worksheet_name, report_folder)
                                    if len(driver.find_elements_by_xpath("//div[@class='nodata']")) == 0:
                                        print(Worksheet_name + "Passed")
                                        test_case_id += 1
                                        ws.append(
                                            (test_case_id, 'Analytics Worksheet', Worksheet_name, 'Passed',
                                             (str)(round(time_taken, 3))))
                                    elif len(driver.find_elements_by_xpath("//div[@class='nodata']")) != 0:
                                        test_case_id += 1
                                        ws.append((test_case_id, 'Analytics Worksheet',
                                                   workbook_name + "-" + Worksheet_name, 'Failed',
                                                   (str)(round(time_taken, 3)),
                                                   'No data for the selected filters', driver.current_url))
                                    try:
                                        driver.find_element_by_xpath("// *[ @ id = 'sm_select_all'] / i").click()
                                        print("Found Select all")
                                    except Exception as e:
                                        print("No select all checkbox")
                                        print(e)
                                        select_all_present = False
                                        break
                                        # '''
                            # Aspyedit ends here ---------------------------------------------------------------------------------------------------------------------------------------
                            driver.find_element_by_xpath("//a[@id='sm_back']").click()

                        except Exception as e:
                            print(e)
                            traceback.print_exc()
                            print(workbook_name + "Failed!Exception occurred!")
                            test_case_id += 1
                            ws.append((test_case_id, 'Analytics Workbook', workbook_name, 'Failed', '', '', driver.current_url))
                            driver.get(current_url)

                        finally:
                            workbook_link += 1
                            all_workbooks = driver.find_elements_by_xpath(locator.xpath_total_Workbooks)

        except Exception as e:
            print(e)
            traceback.print_exc()
            test_case_id += 1
            ws.append((test_case_id, 'Analytics', 'Navigation to Analytics', 'Failed', '', 'Exception occurred!', driver.current_url))
        finally:
            driver.close()
            time.sleep(1)
            driver.switch_to.window(driver.window_handles[0])


    except Exception as e:
        print(e)
        traceback.print_exc()
        test_case_id += 1
        ws.append((test_case_id, 'Analytics', 'Navigation to Analytics', 'Failed', 'Exception occurred!', driver.current_url))
        driver.get(last_url)
        sf.ajax_preloader_wait(driver)
        WebDriverWait(driver, 30).until(
            EC.presence_of_element_located((By.XPATH, locator.xpath_app_Tray_Link)))

    rows = ws.max_row
    cols = ws.max_column
    for i in range(2, rows + 1):
        for j in range(3, cols + 1):
            if ws.cell(i, j).value == 'Passed':
                ws.cell(i, j).fill = PatternFill('solid', fgColor='0FC404')
            elif ws.cell(i, j).value == 'Failed':
                ws.cell(i, j).fill = PatternFill('solid', fgColor='FC0E03')
            elif ws.cell(i, j).value == 'Data table is empty':
                ws.cell(i, j).fill = PatternFill('solid', fgColor='FCC0BB')


def click_on_each_metric(customer, driver, workbook, path, report_folder):
    ws = workbook.create_sheet(customer)
    ws = workbook[customer]
    ws.append(['ID', 'List', 'Context', 'Time-Taken'])
    workbook.save(path + "\\Report.xlsx")

    metrics = driver.find_element_by_id("registry_body").find_elements_by_tag_name('li')
    tracker = 1
    i = 0
    while i < len(metrics):
        print(metrics[i].text)
        try:
            #driver.execute_script("arguments[0].scrollIntoView();", metrics[i])
            metrics[i].click()
            start = time.perf_counter()
            sf.ajax_preloader_wait(driver)
            WebDriverWait(driver, 100).until(
                EC.presence_of_element_located((By.CLASS_NAME, "tabs")))
            total_time = time.perf_counter() - start
            # print('Getting Here')
            context = driver.find_element_by_class_name("metric_specific_patient_list_title").text
            print("getting here")
            ws.append([tracker, 'Provider\'s Tab', context, total_time])
            clicky = driver.find_element_by_class_name('tabs').find_elements_by_tag_name('li')
            print(*clicky)
            clicky[0].click()
            start = time.perf_counter()
            sf.ajax_preloader_wait(driver)
            WebDriverWait(driver, 100).until(
                EC.presence_of_element_located((By.CLASS_NAME, "tabs")))
            total_time = time.perf_counter() - start
            ws.append([tracker, 'Practice Tab', context, total_time])
            print("Practice done")
            clicky = driver.find_element_by_class_name('tabs').find_elements_by_tag_name('li')
            print("found tabs from practice tab")
            clicky[2].click()
            print("Patients clicked")
            start = time.perf_counter()
            sf.ajax_preloader_wait(driver)
            WebDriverWait(driver, 100).until(
                EC.presence_of_element_located((By.CLASS_NAME, "tabs")))
            total_time = time.perf_counter() - start
            ws.append([tracker, 'Patients Tab', context, total_time])
            workbook.save(path + "\\Report.xlsx")
            driver.find_element_by_class_name("breadcrumb").click()
            print('Back clicked')
            WebDriverWait(driver, 100).until(
                EC.presence_of_element_located((By.XPATH, locator.side_nav_SlideOut)))
            print('Sidenav found')
            sf.ajax_preloader_wait(driver)
            print("Back to registries")
            i = i + 1
            metrics = driver.find_element_by_id("registry_body").find_elements_by_tag_name('li')


        except Exception as e:
            ws.append([tracker, 'Failed', 'Failed', driver.current_url])
            print(e)
            traceback.print_exc()
        finally:
            tracker = tracker + 1

    driver.find_element_by_xpath(locator.side_nav_SlideOut).click()
    driver.find_element_by_id("providers-list").click()
    start = time.perf_counter()
    sf.ajax_preloader_wait(driver)
    WebDriverWait(driver, 100).until(
        EC.presence_of_element_located((By.ID, "qt-mt-support-ls")))
    total_time = time.perf_counter() - start
    ws.append([tracker, 'Provider\'s List', total_time])
    tracker = tracker + 1
    driver.find_elements_by_class_name("handler")[0].click()
    start = time.perf_counter()
    sf.ajax_preloader_wait(driver)
    WebDriverWait(driver, 100).until(
        EC.presence_of_element_located((By.ID, "qt-mt-support-ls")))
    total_time = time.perf_counter() - start
    ws.append([tracker, 'Practice List', '', total_time])
    workbook.save(path + "\\Report.xlsx")


'''
def SupportpageAccordionValidationx(driver, workbook, logger, run_from):
    try:
        loader = WebDriverWait(driver, 300)
        loader.until(EC.invisibility_of_element_located((By.XPATH, "//div/div[contains(@class,'ajax_preloader')]")))
        LOBdropdownelement = driver.find_element_by_xpath("//*[@id='qt-filter-label']")
        LOBdropdownelement.click()
        default_quarter = driver.find_element_by_xpath(
            "//*[@id='filter-quarter']//following-sibling::li[@class='highlight']/span/a")
        print(default_quarter.text)
        logger.info("Default Quarter selected as  : " + default_quarter.text)
        logger.info("--------------------------------------------------------")
        Flag = config['SelectMeasurementYear']['Flag_Support']
        # Navigation with customize Measurement Year
        # LOBquarterlist = []
        print(Flag)
        LOBquarter = LOBdropdownelement.find_elements_by_xpath("//*[@id='filter-quarter']/li")
        if (Flag == "True"):
            for i in range(0, len(LOBquarter)):
                # LOBquarterlist.append(LOBquarter[i].text)
                if (LOBquarter[i].text == config['SelectMeasurementYear']['MeasurementYear_Support'] or LOBquarter[
                    i].text == config['SelectMeasurementYear']['MeasurementYearQuarter_Support']):
                    logger.info("Current Quarter selected as  : " + LOBquarter[i].text)
                    logger.info("--------------------------------------------------------")
                    LOBquarter[i].click()
                    break
                LOBquarter = LOBdropdownelement.find_elements_by_xpath("//*[@id='filter-quarter']/li")

        time.sleep(1)

        LOBname = LOBdropdownelement.find_element_by_xpath("//*[@id='filter-lob']")
        LOBnamelist = LOBname.find_elements_by_tag_name("li")
        Payername = LOBdropdownelement.find_elements_by_xpath("//*[@id='filter-payer']")
        for j in range(0, len(LOBnamelist)):

            print(LOBnamelist[j].text)
            print("--------------------------------")
            LOBnamelist[j].click()
            currentLOBName = LOBnamelist[j].text
            logger.info("LOB Selected  :: " + currentLOBName)
            logger.info("---------------------------------------")
            driver.find_element_by_xpath("//*[@id='reg-filter-apply']").click()
            time.sleep(2)
            loader = WebDriverWait(driver, 300)
            loader.until(
                EC.invisibility_of_element_located((By.XPATH, "//div/div[contains(@class,'ajax_preloader')]")))
            logger.captureScreenshot(driver, currentLOBName, screenshot_path)
            # Checking Patient count in Lob. Raise error if it is 0
            Lob_type = ["ALL", "Medicare", "Medicare ACO"]
            try:
                if currentLOBName in Lob_type:
                    lobpatientcount = driver.find_element_by_xpath(
                        "//*[@id='quality_registry']/div/div[1]/div[4]/div[2]").text
                    logger.info("--------------------------------------------------------")

                else:
                    lobpatientcount = driver.find_element_by_xpath(
                        "//*[@id='quality_registry']/div/div[1]/div[3]/div[2]").text
                    logger.info("--------------------------------------------------------")
            except Exception as e:
                lobpatientcount = driver.find_element_by_xpath(
                    "//*[@id='quality_registry']/div/div[1]/div[3]/div[2]").text
                logger.info("--------------------------------------------------------")

            if (lobpatientcount == "0"):
                logger.critical(
                    "Registry  -> " + str(currentLOBName) + " Patient count is 0.Please check.")
                logger.info(
                    "Registry  -> " + str(currentLOBName) + " Patient count is 0.")
                logger.info("--------------------------------------------------------")
            else:
                logger.info(
                    "Registry  ->  " + str(
                        currentLOBName) + " Patient count is : " + str(lobpatientcount))
                logger.info("--------------------------------------------------------")

            # Accordian metric validation started
            try:
                time.sleep(1)
                driver.find_element_by_xpath("//*[@id='metric_scorecard']/div/div[1]/div/span/a[3]").click()
                time.sleep(2)
                driver.find_element_by_xpath("//*[@id='qt-reg-nav-filters']/li[1]/label").click()
                time.sleep(2)
                driver.find_element_by_xpath("//*[@id='qt-apply-search']").click()
                time.sleep(2)
                total_accordion_metric = driver.find_elements_by_xpath("//*[@class='accordion active']")
                print("Total Accordion Metric(s) :  " + str(len(total_accordion_metric)))
                logger.info("Total Accordion Metric(s) :  " + str(len(total_accordion_metric)))

                accordion_metric_list = []
                for i in range(0, len(total_accordion_metric)):

                    print("Accordion Metric Id : " + total_accordion_metric[i].get_attribute('id'))
                    logger.info("Accordion Metric Id : " + total_accordion_metric[i].get_attribute('id'))
                    print(total_accordion_metric[i].get_attribute('id'))

                    # ["382","212","2053","2052","497","85"] -- Corresponding accordion metric id validation have been skipped
                    if (total_accordion_metric[i].get_attribute('id') in ["382", "212", "2053", "2052", "497",
                                                                          "85"]):
                        print("Accordion Metric id have been skipped")

                    else:
                        accordion_metric_list.append(total_accordion_metric[i].get_attribute('id'))
                    # print(accordion_metric_list)
                print("----------------------------------------------------------------------------")
                logger.info("--------------------------------------------------------")
                print("----------------------------------------------------------------------------")
                logger.info("--------------------------------------------------------")

                for i in range(0, len(accordion_metric_list)):
                    parent_num_den = driver.find_element_by_xpath("//*[@id='" + str(
                        accordion_metric_list[i]) + "']/div[1]/div/a/div/div[1]/div[2]/div[2]/span[2]")
                    print("Measure Num/Denom score of the Parent metric id  " + str(
                        accordion_metric_list[i]) + "  :  " + parent_num_den.text)
                    logger.info("Measure Num/Denom score of the Parent metric id  " + str(
                        accordion_metric_list[i]) + "  :  " + parent_num_den.text,
                                )
                    parent_score = parent_num_den.text
                    parent_num_den_extract = re.search('\(([^)]+)', parent_score).group(1)
                    # print(parent_num_den_extract)
                    parent_num_den_val = parent_num_den_extract.replace(",", "")
                    parent_num_den_split = parent_num_den_val.split("/", 1)
                    parent_num_value = parent_num_den_split[0]
                    print("Numerator value of the Parent metric id  " + str(
                        accordion_metric_list[i]) + "  :  " + parent_num_value)
                    logger.info("Numerator value of the Parent metric id  " + str(
                        accordion_metric_list[i]) + "  :  " + parent_num_value,
                                )
                    parent_den_value = parent_num_den_split[1]
                    print("Denominator value of the Parent metric id  " + str(
                        accordion_metric_list[i]) + "  :  " + parent_den_value)
                    logger.info("Denominator value of the Parent metric id  " + str(
                        accordion_metric_list[i]) + "  :  " + parent_den_value,
                                )

                    child_metric = driver.find_elements_by_xpath("//*[@id='" + str(
                        accordion_metric_list[i]) + "']/div[2]//ancestor::div[@class='qt-metric']")
                    print("Total Child Measures of the Parent metric id  " + str(
                        accordion_metric_list[i]) + " :  " + str(len(child_metric)))
                    logger.info("Total Child Measures of the Parent metric id  " + str(
                        accordion_metric_list[i]) + " :  " + str(len(child_metric)),
                                )
                    child_sum_num = 0
                    child_sum_den = 0
                    for j in range(0, len(child_metric)):
                        j = j + 1
                        child_num_den = driver.find_element_by_xpath(
                            "//*[@id='" + str(accordion_metric_list[i]) + "']/div[2]/div[" + str(
                                j) + "]/a/div/div[1]/div[2]/div[2]/span[2]")
                        # print("//*[@id='"+str(accordian_metric_list[i])+"']/div[2]/div["+str(j)+"]/a/div/div[1]/div[2]/div[2]/span[2]")
                        print("Child [" + str(j) + "] Num/Den score : " + child_num_den.text)
                        logger.info("Child [" + str(j) + "] Num/Den score : " + child_num_den.text,
                                    )
                        child_score = child_num_den.text
                        child_num_den_extract = re.search('\(([^)]+)', child_score).group(1)
                        child_num_den_val = child_num_den_extract.replace(",", "")
                        child_num_den_val = child_num_den_extract.replace(",", "")
                        child_num_den_split = child_num_den_val.split("/", 1)
                        child_num_value = child_num_den_split[0]
                        # print(child_num_value)
                        # print(int(child_num_value))
                        child_sum_num = int(child_num_value) + child_sum_num
                        # print(child_sum_num)
                        child_den_value = child_num_den_split[1]
                        # print(child_den_value)
                        child_sum_den = int(child_den_value) + child_sum_den
                        # print(child_sum_den)
                    print("Total Numerator score of the all child metric(s) :  " + str(child_sum_num))
                    logger.info("Total Numerator score of the all child metric(s) :  " + str(child_sum_num),
                                )
                    print("Total Denominator score of the all child metric(s) :  " + str(child_sum_den))
                    logger.info("Total Denominator score of the all child metric(s) :  " + str(child_sum_den),
                                )
                    if (int(parent_num_value) == child_sum_num and int(parent_den_value) == child_sum_den):
                        print("Sum of child score is matching with parent score")
                        logger.info("Sum of child score is matching with parent score",
                                    )
                    else:
                        print("Score didn't matched")
                        logger.info("###### Score didn't matched ######")
                        logger.critical("LOB Selected  :: " + currentLOBName)
                        logger.critical("---------------------------------------")
                        logger.critical("Parent metric id  :   " + str(accordion_metric_list[i]))
                        logger.critical("Score didn't matched for the corresponding Parent metric id",
                                        )
                        logger.critical("--------------------------------------------------------",
                                        )
                    print("----------------------------------------------------------------------------")
                    logger.info("--------------------------------------------------------")



            except Exception as e:
                print(e)

            time.sleep(1)
            LOBdropdownelement = driver.find_element_by_xpath("//*[@id='qt-filter-label']")
            LOBdropdownelement.click()
            time.sleep(1)
            LOBname = LOBdropdownelement.find_element_by_xpath("//*[@id='filter-lob']")
            LOBnamelist = LOBname.find_elements_by_tag_name("li")

     except Exception as e:
         logger.critical(
             "Registry  -> Accordion Measure validation have been suspended due to error!!Please check.")
'''


def SupportpageAccordionValidation(driver, workbook, logger, run_from, report_folder):
    logger.info("--------------------------------------------------------")
    logger.info("Skipping this validation since the environment is set to Stage")
    workbook.create_sheet('Patients, Rating and Scores')
    ws_counts = workbook['Patients, Rating and Scores']
    ws_counts.append(['Skipping this validation since the environment is set to Stage'])
    logger.info("--------------------------------------------------------")
    # try:
    #     workbook.create_sheet('Accordian Validation')
    #     ws = workbook['Accordian Validation']
    #
    #     ws.append(['ID', 'LoB Name', 'Metric ID', 'Status', 'Comments'])
    #     header_font = Font(color='FFFFFF', bold=False, size=12)
    #     header_cell_color = PatternFill('solid', fgColor='030303')
    #     ws['A1'].font = header_font
    #     ws['A1'].fill = header_cell_color
    #     ws['B1'].font = header_font
    #     ws['B1'].fill = header_cell_color
    #     ws['C1'].font = header_font
    #     ws['C1'].fill = header_cell_color
    #     ws['D1'].font = header_font
    #     ws['D1'].fill = header_cell_color
    #     ws['E1'].font = header_font
    #     ws['E1'].fill = header_cell_color
    #     ws.name = "Arial"
    #     test_case_id = 1
    #
    #     workbook.create_sheet('Patients, Rating and Scores')
    #     ws_counts = workbook['Patients, Rating and Scores']
    #
    #     ws_counts.append(['LoB Name', 'Context', 'Count/Rating', 'Status', 'Comments'])
    #     header_font = Font(color='FFFFFF', bold=False, size=12)
    #     header_cell_color = PatternFill('solid', fgColor='030303')
    #     ws_counts['A1'].font = header_font
    #     ws_counts['A1'].fill = header_cell_color
    #     ws_counts['B1'].font = header_font
    #     ws_counts['B1'].fill = header_cell_color
    #     ws_counts['C1'].font = header_font
    #     ws_counts['C1'].fill = header_cell_color
    #     ws_counts['D1'].font = header_font
    #     ws_counts['D1'].fill = header_cell_color
    #     ws_counts['E1'].font = header_font
    #     ws_counts['E1'].fill = header_cell_color
    #     ws_counts.name = "Arial"
    #     loader = WebDriverWait(driver, 300)
    #     loader.until(EC.invisibility_of_element_located((By.XPATH, "//div/div[contains(@class,'ajax_preloader')]")))
    #     LOBdropdownelement = driver.find_element_by_xpath("//*[@id='qt-filter-label']")
    #     LOBdropdownelement.click()
    #     default_quarter = driver.find_element_by_xpath(
    #         "//*[@id='filter-quarter']//following-sibling::li[@class=' highlight ']/span/a")
    #     print(default_quarter.text)
    #     logger.info("Default Quarter selected as  : " + default_quarter.text)
    #     logger.info("--------------------------------------------------------")
    #     Flag = locator.select_measurement_year_flag_support
    #     # Navigation with customize Measurement Year
    #     # LOBquarterlist = []
    #     print(Flag)
    #     LOBquarter = LOBdropdownelement.find_elements_by_xpath("//*[@id='filter-quarter']/li")
    #     if (Flag == "True"):
    #         for i in range(0, len(LOBquarter)):
    #             # LOBquarterlist.append(LOBquarter[i].text)
    #             if LOBquarter[i].text == locator.MeasurementYear_Support or LOBquarter[i].text == locator.MeasurementYearQuarter_Support:
    #                 logger.info("Current Quarter selected as  : " + LOBquarter[i].text)
    #                 logger.info("--------------------------------------------------------")
    #                 LOBquarter[i].click()
    #                 break
    #             LOBquarter = LOBdropdownelement.find_elements_by_xpath("//*[@id='filter-quarter']/li")
    #
    #     time.sleep(1)
    #     WebDriverWait(driver, 30).until(
    #         EC.presence_of_element_located((By.XPATH, "//*[@id='filter-lob']")))
    #     LOBname = LOBdropdownelement.find_element_by_xpath("//*[@id='filter-lob']")
    #     LOBnamelist = LOBname.find_elements_by_tag_name("li")
    #     print(*LOBnamelist)
    #     Payername = LOBdropdownelement.find_elements_by_xpath("//*[@id='filter-payer']")
    #     # LOBdropdownelement.click()
    #     time.sleep(1)
    #     for j in range(0, len(LOBnamelist)):
    #         # LOBdropdownelement.click()
    #         time.sleep(1)
    #         print(LOBnamelist[j].text)
    #         print("--------------------------------")
    #         try:
    #             LOBnamelist[j].click()
    #         except ElementNotInteractableException as e:
    #             continue
    #         currentLOBName = LOBnamelist[j].text
    #         logger.info("LOB Selected  :: " + currentLOBName)
    #         logger.info("---------------------------------------")
    #         time.sleep(4)
    #         driver.find_element_by_xpath("//*[@id='reg-filter-apply']").click()
    #         time.sleep(2)
    #         loader = WebDriverWait(driver, 300)
    #         loader.until(
    #             EC.invisibility_of_element_located((By.XPATH, "//div/div[contains(@class,'ajax_preloader')]")))
    #         # logger.captureScreenshot(driver, currentLOBName, screenshot_path)
    #         # Checking Patient count in Lob. Raise error if it is 0
    #         Lob_type = ["ALL", "Medicare", "Medicare ACO"]
    #         #Patient count and performance count
    #         sf.ajax_preloader_wait(driver)
    #         WebDriverWait(driver, 30).until(
    #             EC.presence_of_element_located((By.XPATH, locator.xpath_filter_measure_list)))
    #         try:
    #             summaryList = driver.find_element_by_class_name("registry_header_panel").find_elements_by_tag_name("div")
    #             overall_rating = "0"
    #             patient_count = "0"
    #             for div, next_div in zip(summaryList, summaryList[1:] + [summaryList[0]]):
    #                 if div.text == "Overall Rating":
    #                     overall_rating = next_div.text.replace("%","").strip()
    #                     if 'Stars' in overall_rating:
    #                         overall_rating = overall_rating.replace("Stars","").strip()
    #                 if div.text == "Patients":
    #                     patient_count = next_div.text.replace(",","").strip()
    #             print(overall_rating)
    #             if float(overall_rating)< 1:
    #                 ws_counts.append([currentLOBName, "Overall Rating", overall_rating, "Failed", "Rating is 0", driver.current_url])
    #             else:
    #                 ws_counts.append([currentLOBName, "Overall Rating", overall_rating, "Passed"])
    #             if float(patient_count)< 1:
    #                 ws_counts.append([currentLOBName, "Patient Count", patient_count, "Failed", "Patient count is 0", driver.current_url])
    #             else:
    #                 ws_counts.append([currentLOBName, "Patient Count", patient_count, "Passed"])
    #
    #
    #         except Exception as e:
    #             traceback.print_exc()
    #             ws_counts.append([currentLOBName, "x", "x", "Couldn't fetch patient counts/rating"])
    #
    #         try:
    #             if currentLOBName in Lob_type:
    #                 lobpatientcount = driver.find_element_by_xpath(
    #                     "//*[@id='quality_registry']/div/div[1]/div[4]/div[2]").text
    #                 logger.info("--------------------------------------------------------")
    #
    #             else:
    #                 lobpatientcount = driver.find_element_by_xpath(
    #                     "//*[@id='quality_registry']/div/div[1]/div[3]/div[2]").text
    #                 logger.info("--------------------------------------------------------")
    #         except Exception as e:
    #             lobpatientcount = driver.find_element_by_xpath(
    #                 "//*[@id='quality_registry']/div/div[1]/div[3]/div[2]").text
    #             logger.info("--------------------------------------------------------")
    #
    #         if (lobpatientcount == "0"):
    #             logger.critical(
    #                 "Registry  -> " + str(currentLOBName) + " Patient count is 0.Please check.")
    #             logger.info(
    #                 "Registry  -> " + str(currentLOBName) + " Patient count is 0.")
    #             logger.info("--------------------------------------------------------")
    #         else:
    #             logger.info(
    #                 "Registry  ->  " + str(
    #                     currentLOBName) + " Patient count is : " + str(lobpatientcount))
    #             logger.info("--------------------------------------------------------")
    #             try:
    #
    #                 time.sleep(1)
    #                 driver.find_element_by_xpath("//*[@id='metric_scorecard']/div/div[1]/div/span/a[3]").click()
    #                 time.sleep(2)
    #                 driver.find_element_by_xpath("//*[@id='qt-reg-nav-filters']/li[1]/label").click()
    #                 time.sleep(2)
    #                 driver.find_element_by_xpath("//*[@id='qt-apply-search']").click()
    #                 time.sleep(2)
    #                 # counting 0/0 measures
    #                 try:
    #                     zero_measures = 0
    #                     measure_scores_elements = driver.find_elements_by_class_name("num-den")
    #                     measure_scores = []
    #                     for score in measure_scores_elements:
    #                         measure_scores.append(score.text.strip())
    #
    #                     for score in measure_scores:
    #                         if score == "(0/0)":
    #                             zero_measures += 1
    #
    #                     if zero_measures > 10:
    #                         ws_counts.append([currentLOBName, "0/0 measures", zero_measures, "Failed",
    #                                           "More than 10 measures are 0/0", driver.current_url])
    #                     else:
    #                         ws_counts.append([currentLOBName, "0/0 Measures", zero_measures, "Passed"])
    #
    #                 except Exception as e:
    #                     traceback.print_exc()
    #                     ws_counts.append([currentLOBName, "0/0 Measures", "Null", "Failed",
    #                                       "Couldn't count 0/0 measures in this lob"])
    #
    #                 total_accordion_metric = driver.find_elements_by_xpath("//*[@class='accordion active']")
    #                 print("Total Accordion Metric(s) :  " + str(len(total_accordion_metric)))
    #                 logger.info("Total Accordion Metric(s) :  " + str(len(total_accordion_metric)))
    #
    #                 accordion_metric_list = []
    #                 for i in range(0, len(total_accordion_metric)):
    #
    #                     print("Accordion Metric Id : " + total_accordion_metric[i].get_attribute('id'))
    #                     logger.info("Accordion Metric Id : " + total_accordion_metric[i].get_attribute('id'))
    #                     print(total_accordion_metric[i].get_attribute('id'))
    #
    #                     # ["382","212","2053","2052","497","85"] -- Corresponding accordion metric id validation have been skipped
    #                     if total_accordion_metric[i].get_attribute('id') in ["382", "212", "2053", "2052", "497","85","60","508"]:
    #                         print("Accordion Metric id have been skipped")
    #
    #                     else:
    #                         accordion_metric_list.append(total_accordion_metric[i].get_attribute('id'))
    #                     # print(accordion_metric_list)
    #                 print("----------------------------------------------------------------------------")
    #                 logger.info("--------------------------------------------------------")
    #                 print("----------------------------------------------------------------------------")
    #                 logger.info("--------------------------------------------------------")
    #
    #                 for i in range(0, len(accordion_metric_list)):
    #                     parent_num_den = driver.find_element_by_xpath("//*[@id='" + str(
    #                         accordion_metric_list[i]) + "']/div[1]/div/a/div/div[1]/div[2]/div[2]/span[2]")
    #                     print("Measure Num/Denom score of the Parent metric id  " + str(
    #                         accordion_metric_list[i]) + "  :  " + parent_num_den.text)
    #                     logger.info("Measure Num/Denom score of the Parent metric id  " + str(
    #                         accordion_metric_list[i]) + "  :  " + parent_num_den.text, )
    #                     parent_score = parent_num_den.text
    #                     parent_num_den_extract = re.search('\(([^)]+)', parent_score).group(1)
    #                     # print(parent_num_den_extract)
    #                     parent_num_den_val = parent_num_den_extract.replace(",", "")
    #                     parent_num_den_split = parent_num_den_val.split("/", 1)
    #                     parent_num_value = parent_num_den_split[0]
    #                     print("Numerator value of the Parent metric id  " + str(
    #                         accordion_metric_list[i]) + "  :  " + parent_num_value)
    #                     logger.info("Numerator value of the Parent metric id  " + str(
    #                         accordion_metric_list[i]) + "  :  " + parent_num_value,
    #                                 )
    #                     parent_den_value = parent_num_den_split[1]
    #                     print("Denominator value of the Parent metric id  " + str(
    #                         accordion_metric_list[i]) + "  :  " + parent_den_value)
    #                     logger.info("Denominator value of the Parent metric id  " + str(
    #                         accordion_metric_list[i]) + "  :  " + parent_den_value,
    #                                 )
    #
    #                     child_metric = driver.find_elements_by_xpath("//*[@id='" + str(
    #                         accordion_metric_list[i]) + "']/div[2]//ancestor::div[@class='qt-metric']")
    #                     print("Total Child Measures of the Parent metric id  " + str(
    #                         accordion_metric_list[i]) + " :  " + str(len(child_metric)))
    #                     logger.info("Total Child Measures of the Parent metric id  " + str(
    #                         accordion_metric_list[i]) + " :  " + str(len(child_metric)),
    #                                 )
    #                     child_sum_num = 0
    #                     child_sum_den = 0
    #                     for j in range(0, len(child_metric)):
    #                         j = j + 1
    #                         child_num_den = driver.find_element_by_xpath(
    #                             "//*[@id='" + str(accordion_metric_list[i]) + "']/div[2]/div[" + str(
    #                                 j) + "]/a/div/div[1]/div[2]/div[2]/span[2]")
    #                         # print("//*[@id='"+str(accordian_metric_list[i])+"']/div[2]/div["+str(j)+"]/a/div/div[1]/div[2]/div[2]/span[2]")
    #                         print("Child [" + str(j) + "] Num/Den score : " + child_num_den.text)
    #                         logger.info("Child [" + str(j) + "] Num/Den score : " + child_num_den.text,
    #                                     )
    #                         child_score = child_num_den.text
    #                         child_num_den_extract = re.search('\(([^)]+)', child_score).group(1)
    #                         child_num_den_val = child_num_den_extract.replace(",", "")
    #                         child_num_den_val = child_num_den_extract.replace(",", "")
    #                         child_num_den_split = child_num_den_val.split("/", 1)
    #                         child_num_value = child_num_den_split[0]
    #                         # print(child_num_value)
    #                         # print(int(child_num_value))
    #                         child_sum_num = int(child_num_value) + child_sum_num
    #                         # print(child_sum_num)
    #                         child_den_value = child_num_den_split[1]
    #                         # print(child_den_value)
    #                         child_sum_den = int(child_den_value) + child_sum_den
    #                         # print(child_sum_den)
    #                     print("Total Numerator score of the all child metric(s) :  " + str(child_sum_num))
    #                     logger.info("Total Numerator score of the all child metric(s) :  " + str(child_sum_num),
    #                                 )
    #                     print("Total Denominator score of the all child metric(s) :  " + str(child_sum_den))
    #                     logger.info("Total Denominator score of the all child metric(s) :  " + str(child_sum_den),
    #                                 )
    #                     if (int(parent_num_value) == child_sum_num and int(parent_den_value) == child_sum_den):
    #                         print("Sum of child score is matching with parent score")
    #                         logger.info("Sum of child score is matching with parent score", )
    #                         ws.append([test_case_id, currentLOBName, accordion_metric_list[i], "Passed",
    #                                    "Sum of child score is matching with parent score. Parent Score: " + parent_num_den.text + " ,Child score sum: (" + str(
    #                                        child_sum_num) + "/" + str(child_sum_den) + ")"])
    #                         test_case_id += 1
    #                     else:
    #                         print("Score didn't matched")
    #                         logger.info("###### Score didn't matched ######")
    #                         logger.critical("LOB Selected  :: " + currentLOBName)
    #                         logger.critical("---------------------------------------")
    #                         logger.critical("Parent metric id  :   " + str(accordion_metric_list[i]))
    #                         logger.critical("Score didn't matched for the corresponding Parent metric id",
    #                                         )
    #                         logger.critical("--------------------------------------------------------",
    #                                         )
    #                         ws.append([test_case_id, currentLOBName, accordion_metric_list[i], "Failed",
    #                                    "Sum of child score is not matching with parent score. Parent Score: " + parent_num_den.text + " ,Child score sum: (" + str(
    #                                        child_sum_num) + "/" + str(child_sum_den) + ")", driver.current_url])
    #                         test_case_id += 1
    #
    #                     print("----------------------------------------------------------------------------")
    #                     logger.info("--------------------------------------------------------")
    #
    #             except Exception as e:
    #                 print(e)
    #
    #         time.sleep(2)
    #         LOBdropdownelement = driver.find_element_by_xpath("//*[@id='qt-filter-label']")
    #         LOBdropdownelement.click()
    #         time.sleep(3)
    #         LOBname = LOBdropdownelement.find_element_by_xpath("//*[@id='filter-lob']")
    #         LOBnamelist = LOBname.find_elements_by_tag_name("li")
    #
    # except Exception as e:
    #     traceback.print_exc()
    #     logger.critical(
    #         "Registry  -> Accordion Measure validation have been suspended due to error!!Please check.")
    #
    # rows = ws.max_row
    # cols = ws.max_column
    # for i in range(2, rows + 1):
    #     for j in range(3, cols + 1):
    #         if ws.cell(i, j).value == 'Passed':
    #             ws.cell(i, j).fill = PatternFill('solid', fgColor='0FC404')
    #         elif ws.cell(i, j).value == 'Failed':
    #             ws.cell(i, j).fill = PatternFill('solid', fgColor='FC0E03')
    #         elif ws.cell(i, j).value == 'Data table is empty':
    #             ws.cell(i, j).fill = PatternFill('solid', fgColor='FCC0BB')
    #
    # rows = ws_counts.max_row
    # cols = ws_counts.max_column
    # for i in range(2, rows + 1):
    #     for j in range(3, cols + 1):
    #         if ws_counts.cell(i, j).value == 'Passed':
    #             ws_counts.cell(i, j).fill = PatternFill('solid', fgColor='0FC404')
    #         elif ws_counts.cell(i, j).value == 'Failed':
    #             ws_counts.cell(i, j).fill = PatternFill('solid', fgColor='FC0E03')
    #         elif ws_counts.cell(i, j).value == 'Data table is empty':
    #             ws_counts.cell(i, j).fill = PatternFill('solid', fgColor='FCC0BB')
    #

def group_menubar(driver, workbook, logger, screenshot_path, run_from):
    workbook.create_sheet('Group Menubar')
    ws = workbook['Group Menubar']
    main_registry_url = driver.current_url
    try:
        context_dropdown_arrow = driver.find_element_by_xpath("//*[@id='context_dropdown_arrow']")
        # print("Arrow found")
        context_dropdown_arrow.click()
        time.sleep(3)
        group_list_container = driver.find_element_by_xpath("//*[@id='ul_1']")
        group_list_element = group_list_container.find_elements_by_tag_name("li")
        randomList = random.choices(group_list_element, k=1)  # randomly 1 group is checking
        for i in range(0, len(randomList)):
            print(randomList[i].text)
            group_name = randomList[i].text
            randomList[i].click()
            time.sleep(5)
            # print("Arrow found....2")
            try:
                group2_list_container = driver.find_element_by_xpath("//*[@id='ul_2']")
                group2_list_element = group2_list_container.find_elements_by_tag_name("li")
                group2_list_element_count = len(group2_list_element)
                randomList2 = random.choices(group2_list_element, k=1)  # randomly 1 group is checking
                for i in range(0, len(randomList2)):
                    print(randomList2[i].text)
                    group_name2 = randomList2[i].text
                    group_all = (group_name + "_" + group_name2)
                    print(group_all)
                    try:
                        randomList2[i].click()
                        time.sleep(5)
                        targetpath = driver.current_url
                        access_check = sf.URLAccessCheck(targetpath, driver)
                        if (access_check):
                            sf.captureScreenshot(driver, group_name2, screenshot_path)
                            logger.critical(
                                "Group List ->  Access Denied found on clicking " + group_all + " .Please check.",
                                targetpath)
                        else:
                            sf.captureScreenshot(driver, group_name2, screenshot_path)

                            support_menubar(driver, workbook, ws, logger, run_from, "Group level")
                    except Exception as e:
                        print(e)
                        logger.critical(
                            "Group Navigation -> Issue occurred while navigating to Group1-Group2 : " + group_all,
                            targetpath)

                    context_dropdown_arrow = driver.find_element_by_xpath("//*[@id='context_dropdown_arrow']")
                    context_dropdown_arrow.click()
                    group2_list_container = driver.find_element_by_xpath("//*[@id='ul_2']")
                    # print("Arrow found....3")
                    group2_list_element = group2_list_container.find_elements_by_tag_name("li")
                    randomList2 = random.choices(group2_list_element, k=1)

            except Exception as e:
                try:
                    targetpath = driver.current_url
                    access_check = sf.URLAccessCheck(targetpath, driver)
                    if access_check:
                        sf.captureScreenshot(driver, group_name, screenshot_path)
                        logger.critical(
                            "Group List ->  Access Denied found on clicking " + group_name + " .Please check.",
                            targetpath)
                    else:
                        sf.captureScreenshot(driver, group_name, screenshot_path)
                        print("Test.......1")
                        support_menubar(driver, workbook, ws, logger, run_from)
                except Exception as e:
                    logger.critical(
                        "Group Navigation -> Issue occurred while navigating to Group1 : " + group_name, targetpath)
            context_dropdown_arrow = driver.find_element_by_xpath("//*[@id='context_dropdown_arrow']")
            context_dropdown_arrow.click()
            time.sleep(5)
            group_list_container = driver.find_element_by_xpath("//*[@id='ul_1']")
            group_list_element = group_list_container.find_elements_by_tag_name("li")
            randomList = random.choices(group_list_element, k=1)

        driver.refresh()
        time.sleep(3)

    except Exception as e:
        print("Group1 is not available or having some issue while navigating.")
        logger.info("** Group1 is not available or having some issue while navigating.")

    driver.get(main_registry_url)
    sf.ajax_preloader_wait(driver)


def practice_tab_ss(driver, workbook, logger, screenshot_path, run_from):
    workbook.create_sheet('Support Level Tabs')
    ws = workbook['Support Level Tabs']
    
    ws.append(['ID', 'Context', 'Scenario', 'Status', 'Time Taken', 'Comments'])
    header_font = Font(color='FFFFFF', bold=False, size=12)
    header_cell_color = PatternFill('solid', fgColor='030303')
    ws['A1'].font = header_font
    ws['A1'].fill = header_cell_color
    ws['B1'].font = header_font
    ws['B1'].fill = header_cell_color
    ws['C1'].font = header_font
    ws['C1'].fill = header_cell_color
    ws['D1'].font = header_font
    ws['D1'].fill = header_cell_color
    ws['E1'].font = header_font
    ws['E1'].fill = header_cell_color
    ws['F1'].font = header_font
    ws['F1'].fill = header_cell_color
    ws.name = "Arial"
    test_case_id = 1

    registry_url = driver.current_url
    # Selecting tabs from Support MSPL
    context_name = "Couldn't Fetch"
    try:
        WebDriverWait(driver, 30).until(
            EC.presence_of_element_located((By.ID, "registry_body")))
        selected_metric_name = 'Couldnt fetch Metric Name'
        context_name = driver.find_element_by_xpath(locator.xpath_context_Name).text

        metrics = driver.find_element_by_id("registry_body").find_elements_by_tag_name('li')
        percent = '0.00'
        while percent == '0.00' or percent == '0.00%':
            selectedMetric = metrics[sf.RandomNumberGenerator(len(metrics), 1)[0]]
            percent = selectedMetric.find_element_by_class_name('percent').text
        selected_metric_name = selectedMetric.find_element_by_class_name('met-name').text
        selectedMetric.click()
        sf.ajax_preloader_wait(driver)
        WebDriverWait(driver, 30).until(
            EC.presence_of_element_located((By.CLASS_NAME, 'tab')))
        metric_url = driver.current_url
        # nav 1 : Practice Tab
        try:
            selectedPracticeName = 'Couldn\'t Fetch'
            driver.find_element_by_class_name('tabs').find_elements_by_class_name('tab')[0].click()
            start_time = time.perf_counter()
            sf.ajax_preloader_wait(driver)
            WebDriverWait(driver, 60).until(
                EC.presence_of_element_located((By.ID, "metric-support-prac-ls")))
            time_taken = time.perf_counter() - start_time
            print("Page Loaded")
            if len(driver.find_elements_by_id("metric-support-prac-ls")) != 0:
                sf.captureScreenshot(driver, selected_metric_name+'Practice_tab', screenshot_path)
                ws.append((test_case_id, selected_metric_name, "Navigation to practice tab", 'Passed', time_taken))
                print("Screenshot taken")

            practices = driver.find_element_by_id("metric-support-prac-ls").find_element_by_tag_name(
                'tbody').find_elements_by_tag_name('tr')
            global global_search_prac
            if len(practices) > 1:
                selectedPractice = \
                    practices[sf.RandomNumberGenerator(len(practices), 1)[0]].find_elements_by_tag_name('a')[1]
                selectedPracticeName = selectedPractice.text
                #global global_search_prac
                global_search_prac = selectedPracticeName
            else:
                try:
                    if practices[0].text == "No data available":
                        global_search_prac = "Couldn't fetch a practice"
                    else:
                        selectedPractice = practices[0].find_elements_by_tag_name('a')[1]
                        selectedPracticeName = selectedPractice.text
                        global_search_prac = selectedPracticeName
                except Exception as e:
                    ws.append([test_case_id, context_name,
                               'Navigation to a practice registry from the pratice tab of support MSPL :' + selected_metric_name,
                               'Failed', '',
                               'No Practices available',
                               driver.current_url])
                    test_case_id += 1
                    print(e)
                    traceback.print_exc()



        except Exception as e:
            ws.append([test_case_id, context_name,
                       'Navigation to a practice registry from the pratice tab of support MSPL :' + selected_metric_name,
                       'Failed', '',
                       'Couldnt click on practice tab or a random practice name: ' + selectedPracticeName, driver.current_url])
            test_case_id += 1
            print(e)
            traceback.print_exc()
        driver.get(metric_url)

        # Nav to provider registry
        try:
            sf.ajax_preloader_wait(driver)
            WebDriverWait(driver, 30).until(
                EC.presence_of_element_located((By.CLASS_NAME, 'tab')))
            selectedProviderName = 'Couldn\'t Fetch'
            driver.find_element_by_class_name('tabs').find_elements_by_class_name('tab')[1].click()
            start_time = time.perf_counter()
            sf.ajax_preloader_wait(driver)
            WebDriverWait(driver, 30).until(
                EC.presence_of_element_located((By.ID, "metric-support-prov-ls")))
            time_taken = time.perf_counter() - start_time
            if len(driver.find_elements_by_id("metric-support-prov-ls")) != 0:
                sf.captureScreenshot(driver, selected_metric_name+'Provider_tab', screenshot_path)
                ws.append((test_case_id, selected_metric_name, "Navigation to provider tab", 'Passed', time_taken))
                print("Screenshot taken")
            providers = driver.find_element_by_id("metric-support-prov-ls").find_element_by_tag_name(
                'tbody').find_elements_by_tag_name('tr')
            if len(providers) > 1:
                selectedProvider = \
                    providers[sf.RandomNumberGenerator(len(providers), 1)[0]].find_elements_by_tag_name('a')[2]
                selectedProviderName = selectedProvider.text
                global global_search_prov
                global_search_prov = selectedProviderName
            else:
                try:
                    if providers[0].text == "No Data Available":
                        global_search_prov = "Couldn't fetch provider name"
                    else:
                        selectedProvider = providers[0].find_elements_by_tag_name('a')[2]
                        selectedProviderNameName = selectedProvider.text
                        # global global_search_prov
                        global_search_prov = selectedProviderName
                except Exception as e:
                    print(e)
                    traceback.print_exc()
                    ws.append([test_case_id, context_name,
                               'Navigation to a provider registry from the provider tab of support MSPL :' + selected_metric_name,
                               'Failed', '',
                               'No providers available',
                               driver.current_url])
                    test_case_id += 1



        except Exception as e:
            print(e)
            traceback.print_exc()
            ws.append([test_case_id, context_name,
                       'Navigation to a provider registry from the provider tab of support MSPL :' + selected_metric_name,
                       'Failed', '',
                       'Couldnt click on provider tab or a random provider name: ' + selectedProviderName, driver.current_url])
            test_case_id += 1
        driver.get(metric_url)

        # nav 3 : Patient context
        try:
            sf.ajax_preloader_wait(driver)
            WebDriverWait(driver, 30).until(
                EC.presence_of_element_located((By.CLASS_NAME, 'tab')))
            patient_id = 'Couldn\'t Fetch'
            driver.find_element_by_class_name('tabs').find_elements_by_class_name('tab')[2].click()
            start_time = time.perf_counter()
            sf.ajax_preloader_wait(driver)
            WebDriverWait(driver, 30).until(
                EC.presence_of_element_located((By.ID, "metric-support-pat-ls")))
            time_taken = time.perf_counter() - start_time
            if len(driver.find_elements_by_id("metric-support-pat-ls")) != 0:
                sf.captureScreenshot(driver, selected_metric_name+'Patient_tab', screenshot_path)
                ws.append((test_case_id, selected_metric_name, "Navigation to patient tab", 'Passed', time_taken))
                print("Screenshot taken")
            patients = driver.find_element_by_id("metric-support-pat-ls").find_element_by_tag_name(
                'tbody').find_elements_by_tag_name('tr')
            global global_search_pat
            if len(patients) > 1:
                selectedPatient = \
                    patients[sf.RandomNumberGenerator(len(patients), 1)[0]].find_elements_by_class_name('pat_name')[
                        0].get_attribute("href")
                czid = sf.get_patient_id(selectedPatient)
                global_search_pat = czid
            else:
                try:
                    if patients[0].text == "No Data Available":
                        global_search_pat = "Couldn't fetch a patient ID"
                    else:
                        selectedPatient = patients[0].find_elements_by_class_name('pat_name')[0].get_attribute("href")
                        czid = sf.get_patient_id(selectedPatient)
                        global_search_pat = czid
                except Exception as e:
                    print(e)
                    traceback.print_exc()
                    ws.append([test_case_id, context_name,
                               'Navigation to patient context from the patients tab of support MSPL :' + selected_metric_name,
                               'Failed', '', 'No available patients',
                               driver.current_url])
                    test_case_id += 1



        except Exception as e:
            print(e)
            traceback.print_exc()
            ws.append([test_case_id, context_name,
                       'Navigation to patient context from the patients tab of support MSPL :' + selected_metric_name,
                       'Failed', '', 'Couldnt click on patient tab or a random patient : ' + patient_id, driver.current_url])
            test_case_id += 1
        driver.get(metric_url)

        # nav 4 : Performance Statistics
        try:
            sf.ajax_preloader_wait(driver)
            WebDriverWait(driver, 30).until(
                EC.presence_of_element_located((By.CLASS_NAME, 'tab')))
            driver.find_element_by_class_name('tabs').find_elements_by_class_name('tab')[3].click()
            start_time = time.perf_counter()
            sf.ajax_preloader_wait(driver)
            WebDriverWait(driver, 30).until(
                EC.presence_of_element_located((By.CLASS_NAME, 'tabs')))
            time_taken = time.perf_counter() - start_time
            if driver.find_elements_by_id('performance_details') != 0:
                ws.append([test_case_id, context_name,
                           'Navigation to Performance Stats from Support Metric : ' + selected_metric_name,
                           'Passed',
                           time_taken])
                test_case_id += 1
                sf.captureScreenshot(driver, selected_metric_name + 'Performance_tab', screenshot_path)
            else:
                ws.append([test_case_id, context_name,
                           'Navigation to Performance Stats from Support Metric : ' + selected_metric_name,
                           'Failed', driver.current_url])
                test_case_id += 1
                sf.captureScreenshot(driver, selected_metric_name + 'Performance_tab', screenshot_path)

        except Exception as e:
            print(e)
            traceback.print_exc()
            ws.append(
                [test_case_id, context_name, 'Navigation to Performance Stats from Practice MSPL', 'Failed', '',
                 'Couldnt click on the performance tab of metric :' + selected_metric_name, driver.current_url])
            test_case_id += 1
            sf.captureScreenshot(driver, selected_metric_name + 'Performance_tab', screenshot_path)


    except Exception as e:
        print(e)
        traceback.print_exc()
        ws.append([test_case_id, context_name, 'Navigation to Support MSPL', 'Failed', '',
                   'Unable to click on a random metric: ' + selected_metric_name, driver.current_url])
        test_case_id += 1
        sf.captureScreenshot(driver, 'Click on metric', screenshot_path)
        driver.get(registry_url)

    driver.get(registry_url)
    sf.ajax_preloader_wait(driver)

    rows = ws.max_row
    cols = ws.max_column
    for i in range(2, rows + 1):
        for j in range(3, cols + 1):
            if ws.cell(i, j).value == 'Passed':
                ws.cell(i, j).fill = PatternFill('solid', fgColor='0FC404')
            elif ws.cell(i, j).value == 'Failed':
                ws.cell(i, j).fill = PatternFill('solid', fgColor='FC0E03')
            elif ws.cell(i, j).value == 'Data table is empty':
                ws.cell(i, j).fill = PatternFill('solid', fgColor='FCC0BB')


def patient_medication(driver ,workbook, logger, screenshot_path, run_from):
    #pick a compliant patient from a PDC metric.
    sf.ajax_preloader_wait(driver)
    main_registry = driver.current_url
    WebDriverWait(driver, 60).until(EC.presence_of_element_located((By.XPATH, locator.xpath_filter_measure_list)))
    try:
        window_switched = 0
        driver.find_element_by_id("qt-filter-label").click()
        time.sleep(1)
        lobs = driver.find_element_by_id("filter-lob").find_elements_by_tag_name('li')
        for lob in lobs:
            if 'medicare' in lob.text or 'Medicare' in lob.text:
                lob.click()
                break
        driver.find_element_by_id("reg-filter-apply").click()
        sf.ajax_preloader_wait(driver)
        WebDriverWait(driver, 60).until(EC.presence_of_element_located((By.XPATH, locator.xpath_filter_measure_list)))
        driver.find_element_by_xpath(locator.xpath_filter_measure_list).click()
        WebDriverWait(driver, 60).until(EC.presence_of_element_located((By.ID, "qt-search-met")))
        driver.find_element_by_id("qt-search-met").clear()
        driver.find_element_by_id("qt-search-met").send_keys('pdc')
        time.sleep(2)
        driver.find_element_by_id("qt-apply-search").click()
        sf.ajax_preloader_wait(driver)
        driver.find_element_by_id("registry_body").find_elements_by_tag_name("li")[0].click()
        sf.ajax_preloader_wait(driver)
        WebDriverWait(driver, 30).until(
            EC.presence_of_element_located((By.CLASS_NAME, 'tab')))
        driver.find_element_by_class_name('tabs').find_elements_by_class_name('tab')[2].click()
        sf.ajax_preloader_wait(driver)
        WebDriverWait(driver, 30).until(
            EC.presence_of_element_located((By.ID, "metric-support-pat-ls")))
        driver.find_element_by_class_name("datatable_filter_dropdown").click()
        time.sleep(1)
        driver.find_element_by_id("table_dropdown_metric-support-pat-ls").find_element_by_xpath('//*[@id="table_dropdown_metric-support-pat-ls"]/div[5]').find_element_by_class_name("dropdown-trigger").click()
        time.sleep(1)
        driver.find_element_by_id("table_dropdown_metric-support-pat-ls").find_element_by_xpath('//*[@id="table_dropdown_metric-support-pat-ls"]/div[5]').find_element_by_class_name("select-wrapper").find_elements_by_tag_name('li')[1].click()
        time.sleep(1)
        driver.find_element_by_id("table_dropdown_metric-support-pat-ls").find_element_by_xpath('//*[@id="table_dropdown_metric-support-pat-ls"]/div[6]').find_element_by_class_name("datatable_apply").click()
        sf.ajax_preloader_wait(driver)
        WebDriverWait(driver, 30).until(
            EC.presence_of_element_located((By.ID, "metric-support-pat-ls")))
        driver.find_element_by_id("metric-support-pat-ls").find_element_by_tag_name(
            'tbody').find_elements_by_tag_name('tr')[0].find_elements_by_class_name('pat_name')[0].click()
        driver.switch_to.window(driver.window_handles[1])
        window_switched = 1
        sf.ajax_preloader_wait(driver)
        WebDriverWait(driver, 30).until(
            EC.presence_of_element_located((By.XPATH, locator.xpath_cozeva_Id)))
        chart_icon = driver.find_element_by_class_name("medical_adherence_chart_icon")
        driver.execute_script("arguments[0].scrollIntoView();", chart_icon)
        chart_icon.click()
        WebDriverWait(driver, 60).until(EC.presence_of_element_located((By.ID, "medication_chart_div_id")))
        sf.captureScreenshot(driver, "Medications_"+driver.find_element_by_xpath(locator.xpath_cozeva_Id).text, screenshot_path)
        time.sleep(1)
        driver.close()
        driver.switch_to.window(driver.window_handles[0])

    except Exception as e:
        traceback.print_exc()
        print(e)
        sf.captureScreenshot(driver, "Medication check failed", screenshot_path)
        if window_switched == 1:
            driver.switch_to.window(driver.window_handles[0])
    driver.get(main_registry)
    sf.ajax_preloader_wait(driver)
    WebDriverWait(driver, 60).until(EC.presence_of_element_located((By.XPATH, locator.xpath_filter_measure_list)))


def apptray_access_check(driver, workbook,logger,screenshot_path, run_from):
    workbook.create_sheet('Apptray Access Check')
    ws = workbook['Apptray Access Check']

    ws.append(['ID', 'Context', 'Scenario', 'Status', 'Time Taken', 'Comments'])
    header_font = Font(color='FFFFFF', bold=False, size=12)
    header_cell_color = PatternFill('solid', fgColor='030303')
    ws['A1'].font = header_font
    ws['A1'].fill = header_cell_color
    ws['B1'].font = header_font
    ws['B1'].fill = header_cell_color
    ws['C1'].font = header_font
    ws['C1'].fill = header_cell_color
    ws['D1'].font = header_font
    ws['D1'].fill = header_cell_color
    ws['E1'].font = header_font
    ws['E1'].fill = header_cell_color
    ws['F1'].font = header_font
    ws['F1'].fill = header_cell_color
    ws.name = "Arial"
    test_case_id = 1

    registry_url = driver.current_url
    try:
        last_url = driver.current_url
        window_switched = 0
        driver.find_element_by_xpath(locator.xpath_app_Tray_Link).click()
        driver.find_element_by_xpath(locator.xpath_app_Time_Capsule).click()
        driver.switch_to.window(driver.window_handles[1])
        window_switched = 1
        sf.ajax_preloader_wait(driver)
        try:
            sf.ajax_preloader_wait(driver)
            current_url = driver.current_url
            access_message = sf.CheckAccessDenied(current_url)

            if access_message == 1:
                print("Access Denied found!")
                # logger.critical("Access Denied found!")
                test_case_id+= 1
                ws.append(
                    (test_case_id, 'Time Capsule', 'Access check for Time Capsule', 'Failed', 'Access Denied', driver.current_url))
                sf.captureScreenshot(driver, 'Time Capsule Access denied', screenshot_path)

            else:
                print("Access Check done!")
                # logger.info("Access Check done!")
                error_message = sf.CheckErrorMessage(driver)

                if error_message == 1:
                    print("Error toast message is displayed")
                    # logger.critical("ERROR TOAST MESSAGE IS DISPLAYED!")
                    test_case_id += 1
                    ws.append((test_case_id, 'Time Capsule', 'Navigation to Time Capsule without error message',
                                'Failed', 'Error toast message is displayed', driver.current_url))
                    sf.captureScreenshot(driver, 'Time Capsule error message', screenshot_path)

                else:
                    ws.append((test_case_id, 'Time Capsule', 'Time Capsule page loading',
                                    'Passed'))
                    sf.captureScreenshot(driver, 'Time Capsule', screenshot_path)

        except Exception as e:
            print(e)
            test_case_id += 1
            ws.append(
                (test_case_id, 'Time Capsule', 'Navigation to Time Capsule', 'Failed', 'Exception occurred!', driver.current_url))
            sf.captureScreenshot(driver, 'Time Capsule access', screenshot_path)
        finally:
            driver.close()
            time.sleep(1)
            if window_switched == 1:
                driver.switch_to.window(driver.window_handles[0])

    except Exception as e:
        print(e)
        test_case_id += 1
        ws.append(
            (test_case_id, 'Time Capsule', 'Navigation to Time Capsule', 'Failed', 'Exception occurred!', driver.current_url))
        sf.captureScreenshot(driver, 'Time Capsule Access denied', screenshot_path)
        driver.get(last_url)
        sf.ajax_preloader_wait(driver)
        WebDriverWait(driver, 30).until(
            EC.presence_of_element_located((By.XPATH, locator.xpath_app_Tray_Link)))

    driver.get(registry_url)
    #secure messaging
    window_switched = 0
    try:
        last_url = driver.current_url
        WebDriverWait(driver, 30).until(
            EC.presence_of_element_located((By.XPATH, locator.xpath_app_Tray_Link)))
        driver.find_element_by_xpath(locator.xpath_app_Tray_Link).click()
        WebDriverWait(driver, 30).until(
            EC.presence_of_element_located((By.XPATH, locator.xpath_app_Secure_Messaging)))
        driver.find_element_by_xpath(locator.xpath_app_Secure_Messaging).click()
        driver.switch_to.window(driver.window_handles[1])
        window_switched = 1
        try:
            sf.ajax_preloader_wait(driver)
            current_url = driver.current_url
            access_message = sf.CheckAccessDenied(current_url)

            if access_message == 1:
                print("Access Denied found!")
                # logger.critical("Access Denied found!")
                test_case_id += 1
                ws.append(
                    (test_case_id, 'Secure Messaging', 'Access check for Secure Messaging', 'Failed',
                     'Access Denied', driver.current_url))
                sf.captureScreenshot(driver, 'Secure Messaging Access Denied', screenshot_path)

            else:
                print("Access Check done!")
                # logger.info("Access Check done!")
                error_message = sf.CheckErrorMessage(driver)

                if error_message == 1:
                    print("Error toast message is displayed")
                    # logger.critical("ERROR TOAST MESSAGE IS DISPLAYED!")
                    test_case_id += 1
                    ws.append(
                        (test_case_id, 'Secure Messaging',
                         'Navigation to Secure Messaging without error message',
                         'Failed', 'Error toast message is displayed', driver.current_url))
                    sf.captureScreenshot(driver, 'Secure Messaging Error toast', screenshot_path)

                else:
                    test_case_id += 1
                    ws.append((test_case_id, 'Secure Messaging', 'Navigation to Secure Messaging', 'Passed',))
                    sf.captureScreenshot(driver, 'Secure Messaging', screenshot_path)
        except Exception as e:
            print(e)
            test_case_id += 1
            ws.append((test_case_id, 'Secure Messaging', 'Navigation to Secure Messaging', 'Failed',
                       'Exception occurred!', driver.current_url))
            sf.captureScreenshot(driver, 'Secure Messaging Access Denied', screenshot_path)
        finally:
            driver.close()
            time.sleep(1)
            driver.switch_to.window(driver.window_handles[0])

    except Exception as e:
        print(e)
        test_case_id += 1
        ws.append((test_case_id, 'Secure Messaging', 'Navigation to Secure Messaging', 'Failed',
                   'Exception occurred!', driver.current_url))
        sf.captureScreenshot(driver, 'Secure Messaging Access Denied', screenshot_path)
        driver.get(last_url)
        sf.ajax_preloader_wait(driver)
        WebDriverWait(driver, 30).until(
            EC.presence_of_element_located((By.XPATH, locator.xpath_app_Tray_Link)))

    driver.get(registry_url)
    sf.ajax_preloader_wait(driver)
    #analytics
    try:
        WebDriverWait(driver, 30).until(EC.presence_of_element_located((By.XPATH, locator.xpath_app_Tray_Link)))
        driver.find_element_by_xpath(locator.xpath_app_Tray_Link).click()
        WebDriverWait(driver, 30).until(
            EC.presence_of_element_located((By.XPATH, locator.xpath_app_Analytics)))
        driver.find_element_by_xpath(locator.xpath_app_Analytics).click()
        driver.switch_to.window(driver.window_handles[1])
        window_switched = 1
        try:
            sf.ajax_preloader_wait(driver)
            current_url = driver.current_url
            access_message = sf.CheckAccessDenied(current_url)

            if access_message == 1:
                print("Access Denied found!")
                # logger.critical("Access Denied found!")
                test_case_id += 1
                ws.append(
                    (test_case_id, 'Analytics', 'Access check for Analytics', 'Failed', 'x', 'Access Denied', driver.current_url))
                sf.captureScreenshot(driver, 'Analytics Access Denied', screenshot_path)

            else:
                print("Access Check done!")
                # logger.info("Access Check done!")
                error_message = sf.CheckErrorMessage(driver)

                if error_message == 1:
                    print("Error toast message is displayed")
                    # logger.critical("ERROR TOAST MESSAGE IS DISPLAYED!")
                    test_case_id += 1
                    ws.append((test_case_id, 'Analytics', 'Navigation to Analytics without error message',
                               'Failed', 'x', 'Error toast message is displayed', driver.current_url))
                    sf.captureScreenshot(driver, 'Analytics error toast', screenshot_path)

                else:

                    total_workbooks = len(driver.find_elements_by_xpath(locator.xpath_total_Workbooks))
                    all_workbooks = driver.find_elements_by_xpath(locator.xpath_total_Workbooks)
                    test_case_id += 1
                    ws.append((test_case_id, 'Analytics', 'Navigation to Analytics', 'Passed', 'x',
                               'Number of Workbook links: ' + str(total_workbooks)))

                    sf.captureScreenshot(driver, 'Analytics', screenshot_path)
        except Exception as e:
            print(e)
            traceback.print_exc()
            test_case_id += 1
            ws.append((test_case_id, 'Analytics', 'Navigation to Analytics', 'Failed', '', 'Exception occurred!', driver.current_url))
        finally:
            driver.close()
            time.sleep(1)
            if window_switched == 1:
                driver.switch_to.window(driver.window_handles[0])
                window_switched == 0


    except Exception as e:
        print(e)
        traceback.print_exc()
        test_case_id += 1
        ws.append((test_case_id, 'Analytics', 'Navigation to Analytics', 'Failed', 'Exception occurred!', driver.current_url))
        sf.captureScreenshot(driver, 'Analytics Access Denied', screenshot_path)
        if window_switched == 1:
            driver.switch_to.window(driver.window_handles[0])
            window_switched == 0
        driver.get(last_url)
        sf.ajax_preloader_wait(driver)
        WebDriverWait(driver, 30).until(
            EC.presence_of_element_located((By.XPATH, locator.xpath_app_Tray_Link)))

    driver.get(registry_url)
    sf.ajax_preloader_wait(driver)
    WebDriverWait(driver, 60).until(EC.presence_of_element_located((By.XPATH, locator.xpath_filter_measure_list)))


    rows = ws.max_row
    cols = ws.max_column
    for i in range(2, rows + 1):
        for j in range(3, cols + 1):
            if ws.cell(i, j).value == 'Passed':
                ws.cell(i, j).fill = PatternFill('solid', fgColor='0FC404')
            elif ws.cell(i, j).value == 'Failed':
                ws.cell(i, j).fill = PatternFill('solid', fgColor='FC0E03')
            elif ws.cell(i, j).value == 'Data table is empty':
                ws.cell(i, j).fill = PatternFill('solid', fgColor='FCC0BB')


def training_resources(driver, workbook, logger, screenshot_path, run_from):
    sf.ajax_preloader_wait(driver)
    main_registry = driver.current_url
    WebDriverWait(driver, 60).until(EC.presence_of_element_located((By.XPATH, locator.xpath_filter_measure_list)))
    try:
        driver.find_element_by_xpath(locator.xpath_resources_link).click()
        time.sleep(1)
        driver.find_element_by_id("help_menu_options").find_elements_by_tag_name("li")[1].click()
        sf.ajax_preloader_wait(driver)
        WebDriverWait(driver, 60).until(EC.presence_of_element_located((By.XPATH, locator.xpath_resources_page_prac_sup)))
        sf.captureScreenshot(driver, 'Resources Page', screenshot_path)


    except Exception as e:
        traceback.print_exc()
        print(e)

    driver.get(main_registry)
    sf.ajax_preloader_wait(driver)
    WebDriverWait(driver, 60).until(EC.presence_of_element_located((By.XPATH, locator.xpath_filter_measure_list)))


def patient_timeline(driver, workbook, logger, screenshot_path, run_from):
    x=0


def sticket_validation(driver, workbook, logger, screenshot_path, run_from, customer_id):
    def validate_all_columns(list1):
        result = all(element == list1[0] for element in list1)
        status = ""
        if result:
            status = "PASS"
            return status
            # print("All the elements are Equal")
        else:
            status = "FAIL"
            return status
            # print("All Elements are not equal")

    def validate_date_time_format(s):
        format = "%m/%d/%Y %H:%M:%S"
        status = ""
        try:
            dt.datetime.strptime(s, format)
            print("This is the correct date string format.")
            status = "PASS"
            return status
        except ValueError:
            print("This is the incorrect date string format. It should be YYYY-MM-DD")
            status = "Unexpected Value"
            return status

    def validate_is_a_proper_string(s):
        status = ""
        if len(s) == 0:
            status = "Value is missing"
            logger.info(str(s) + " is empty ")
            return status
        if len(s) > 3:
            status = "PASS"
            logger.info(str(s) + "is a proper string")
            return status
        if len(s) < 3:
            status = "Unexpected Value "
            logger.info(str(s) + " has unexpected value")
            return status

    def validate_dob_original(s):
        format = "%m/%d/%Y"
        status = ""
        try:
            dt.datetime.strptime(s, format)
            logger.info("This is the correct date string format.")
            status = "PASS"
            return status
        except ValueError:
            status = "Unexpected Value"
            return status
            logger.info("This is the incorrect date string format. It should be YYYY-MM-DD")

    def validate_created(created):
        status = validate_date_time_format(created)
        return status

    def validate_last_updated(last_updated):
        status = validate_date_time_format(last_updated)
        return status

    def validate_created_by(created_by):
        status = validate_is_a_proper_string(created_by)
        return status

    def validate_last_updated_by(last_updated_by):
        status = validate_is_a_proper_string(last_updated_by)
        return status

    def validate_patient_status(patient):
        status = validate_is_a_proper_string(patient)
        return status

    def validate_dob(dob):
        status = validate_dob_original(dob)
        return status

    def validate_member_id_status(member_id):
        status = validate_is_a_proper_string(member_id)
        return status

    def validate_member_phone(member_phone):
        status = validate_is_a_proper_string(member_phone)
        return status

    def validate_pcp(pcp):
        status = validate_is_a_proper_string(pcp)
        return status

    def validate_latest_note(latest_note):
        status = validate_is_a_proper_string(latest_note)
        return status

    def verify_filter(driver, filter_name, value):
        # click on filter icon
        filter = driver.find_element_by_xpath(config.get("sticket-log-locator", "filter_list"))
        sf.action_click(filter, driver)
        if (filter_name == "created"):
            # extract date value
            original_format = "%m/%d/%Y %H:%M:%S"
            date_original_value = dt.datetime.strptime(value, original_format)
            format = "%m/%d/%Y"
            date_value_string = date_original_value.strftime(format)
            date_value_date = datetime.strptime(date_value_string, format)
            print(date_value_date)

            # send one day less to upper
            yesterday = date_value_date - timedelta(days=1)
            # convert yeaterday to proper format
            yesterday_date_value_string = yesterday.strftime(format)

            created_upper_input = driver.find_element_by_xpath(config.get("sticket-log-locator", "created_box_1"))
            created_upper_input.send_keys(yesterday_date_value_string)

            # send current date  to lower

            created_lower_input = driver.find_element_by_xpath(config.get("sticket-log-locator", "created_box_2"))
            created_lower_input.send_keys(date_value_string)

            # click on apply

            apply = driver.find_element_by_xpath(config.get("sticket-log-locator", "apply_xpath"))
            sf.action_click(apply, driver)
            # check number of returned records

            sf.ajax_preloader_wait(driver)
            num_of_entries = find_number_of_rows(driver)

            # if >1 pass else fail

            if num_of_entries >= 1:
                return "PASS"
            else:
                return "FAIL"

        if filter_name == "last_updated":

            created_upper_input = driver.find_element_by_xpath(config.get("sticket-log-locator", "created_box_1"))
            created_upper_input.clear()
            created_lower_input = driver.find_element_by_xpath(config.get("sticket-log-locator", "created_box_2"))
            created_lower_input.clear()
            original_format = "%m/%d/%Y %H:%M:%S"
            date_original_value = dt.datetime.strptime(value, original_format)
            format = "%m/%d/%Y"
            date_value_string = date_original_value.strftime(format)
            date_value_date = datetime.strptime(date_value_string, format)
            print(date_value_date)

            # send one day less to upper
            yesterday = date_value_date - timedelta(days=1)
            # convert yeaterday to proper format
            yesterday_date_value_string = yesterday.strftime(format)

            last_updated_upper_input = driver.find_element_by_xpath(
                config.get("sticket-log-locator", "last_updated_box_1"))
            last_updated_upper_input.send_keys(yesterday_date_value_string)

            # send current date  to lower

            last_updated__lower_input = driver.find_element_by_xpath(
                config.get("sticket-log-locator", "last_updated_box_2"))
            last_updated__lower_input.send_keys(date_value_string)

            # click on apply

            apply = driver.find_element_by_xpath(config.get("sticket-log-locator", "apply_xpath"))
            sf.action_click(apply, driver)
            # check number of returned records

            sf.ajax_preloader_wait(driver)
            num_of_entries = find_number_of_rows(driver)

            # if >1 pass else fail

            if num_of_entries >= 1:
                return "PASS"
            else:
                return "FAIL"

        if filter_name == "created_by":
            created_upper_input = driver.find_element_by_xpath(config.get("sticket-log-locator", "created_box_1"))
            created_upper_input.clear()
            created_lower_input = driver.find_element_by_xpath(config.get("sticket-log-locator", "created_box_2"))
            created_lower_input.clear()
            last_updated_upper_input = driver.find_element_by_xpath(
                config.get("sticket-log-locator", "last_updated_box_1"))
            last_updated_upper_input.clear()
            last_updated__lower_input = driver.find_element_by_xpath(
                config.get("sticket-log-locator", "last_updated_box_2"))
            last_updated__lower_input.clear()
            created_by = driver.find_element_by_xpath(config.get("sticket-log-locator", "created_by"))
            created_by.send_keys(value)
            # click on apply

            apply = driver.find_element_by_xpath(config.get("sticket-log-locator", "apply_xpath"))
            sf.action_click(apply, driver)
            # check number of returned records

            sf.ajax_preloader_wait(driver)

            num_of_entries = find_number_of_rows(driver)
            # if >1 pass else fail
            print("Number of entries in created by " + str(num_of_entries))
            if num_of_entries >= 1:
                return "PASS"
            else:
                return "FAIL"

        if filter_name == "last_updated_by":
            # print("In llast updatedby ")
            created_upper_input = driver.find_element_by_xpath(config.get("sticket-log-locator", "created_box_1"))
            created_upper_input.clear()
            created_lower_input = driver.find_element_by_xpath(config.get("sticket-log-locator", "created_box_2"))
            created_lower_input.clear()
            last_updated_upper_input = driver.find_element_by_xpath(
                config.get("sticket-log-locator", "last_updated_box_1"))
            last_updated_upper_input.clear()
            last_updated__lower_input = driver.find_element_by_xpath(
                config.get("sticket-log-locator", "last_updated_box_2"))
            last_updated__lower_input.clear()
            created_by = driver.find_element_by_xpath(config.get("sticket-log-locator", "created_by"))
            created_by.clear()

            last_updated_by = driver.find_element_by_xpath(config.get("sticket-log-locator", "last_updated_by"))
            last_updated_by.send_keys(value)

            # click on apply

            apply = driver.find_element_by_xpath(config.get("sticket-log-locator", "apply_xpath"))
            sf.action_click(apply, driver)
            # check number of returned records

            sf.ajax_preloader_wait(driver)
            num_of_entries = find_number_of_rows(driver)
            # if >1 pass else fail

            if (num_of_entries >= 1):
                return "PASS"
            else:
                return "FAIL"

        # if(filter_name=="patient"):

    def open_customer_messaging(cust_id):
        sm_customer_id = cust_id  # enter customer_id
        session_var = 'app_id=cozeva_messages&custId=' + str(sm_customer_id) + '&orgId=' + str(sm_customer_id)
        encoded_string = base64.b64encode(session_var.encode('utf-8'))
        driver.get(config.get("runner", "URL") + "cozeva_messages?session=" + encoded_string.decode(
            'utf-8') + "&tab=MessageList&label=Inbox&first_load=true")

    # def try_to_click(element):
    #     attempt=0
    #     while(attempt<3):

    def find_number_of_rows(driver):
        num_of_rows_total_xpath = config.get("sticket-log-locator", "num_of_rows_total_xpath")
        element = driver.find_elements_by_xpath(num_of_rows_total_xpath)
        num_of_entries = len(element) - 1
        return num_of_entries

    def find_number_of_columns(driver, column_xpath):
        element = driver.find_elements_by_xpath(column_xpath)
        num_of_columns = len(element)
        print("Number of columns is " + str(num_of_columns))
        return num_of_columns

    def extract_name_of_columns(driver, column_xpath):
        element = driver.find_elements_by_xpath(column_xpath)
        num_of_columns = len(element)
        header = []
        for i in range(1, num_of_columns):
            header = driver.find_element_by_xpath(column_xpath + "[" + str(i) + "]")
            header_text = header.get_attribute("innerHTML")
            header.append(header_text)
        return header

    def extract_patient_id(href):
        cozeva_id = re.search('/patient_detail/(.*)?session', href)
        return (cozeva_id.group(1).replace("?", ""))

    def apply_conditional_formatting(ws):
        red_text = Font(color="9C0006")
        red_fill = PatternFill(bgColor="FFC7CE")
        dxf = DifferentialStyle(font=red_text, fill=red_fill)
        rule = Rule(type="containsText", operator="containsText", text="FAIL", dxf=dxf)
        rule.formula = ['NOT(ISERROR(SEARCH("highlight",A1)))']

        green_text = Font(color="00FF00FF")
        green_fill = PatternFill(bgColor="0000FF00")
        dxf = DifferentialStyle(font=green_text, fill=green_fill)
        rule1 = Rule(type="containsText", operator="containsText", text="PASS", dxf=dxf)
        rule1.formula = ['NOT(ISERROR(SEARCH("highlight",A1)))']

        ws.conditional_formatting.add('B1:B10000', rule)
        ws.conditional_formatting.add('B1:B10000', rule1)

    def open_registry_page(customer_id):
        customer_list_url = []
        sm_customer_id = str(customer_id)
        sm_customer_id = sm_customer_id.split(".")[0]
        session_var = 'app_id=registries&custId=' + str(sm_customer_id) + '&payerId=' + str(
            sm_customer_id) + '&orgId=' + str(sm_customer_id)
        encoded_string = base64.b64encode(session_var.encode('utf-8'))
        customer_list_url.append(encoded_string)
        for idx, val in enumerate(customer_list_url):
            driver.get(config.get("runner", "URL") + "registries?session=" + val.decode('utf-8'))

    def retrieve_modal_attributes(logger, cozeva_id):
        number_of_logs_in_modal = driver.find_elements_by_xpath(config.get("locator", "no_of_logs_xpath"))
        logger.info("Number of sticket/contact for " + str(cozeva_id) + " is " + str(len(number_of_logs_in_modal) - 1))

    def validate_time_displayed(logger, time_displayed):
        time_displayed_list = time_displayed.split("<br>")
        print(time_displayed_list)
        timestamp = time_displayed_list
        date_string = time_displayed_list[0]
        format = "%m/%d/%Y"
        date_status = ""
        try:
            dt.datetime.strptime(date_string, format)
            date_status = "PASS"
            print("This is the correct date string format.")
            logger.info("This is the correct date string format.")
        except ValueError:
            date_status = "FAIL"
            print("This is the incorrect date string format. It should be MM-DD-YYYY")
            logger.error("This is the incorrect date string format. It should be MM-DD-YYYY")

        time_string = time_displayed_list[1]
        format2 = "%H:%M:%S"
        time_status = ""
        try:
            dt.datetime.strptime(date_string, format)
            time_status = "PASS"
            print("This is the correct time string format.")
            logger.info("This is the correct time string format.")
        except ValueError:
            time_status = "FAIL"
            print("This is the incorrect time string format. It should be %H:%M:%S")
            logger.error("This is the incorrect time string format. It should be %H:%M:%S")

        if (date_status == "PASS" and time_status == "PASS"):
            logger.info("Validate Date and Time Displayed -Done - PASS")
            return "PASS"
        else:
            logger.info("Validate Date and Time Displayed -Done - FAIL")
            return "FAIL"

    def validate_sender_displayed(logger, sender_displayed):
        num_of_words = len(sender_displayed.split())
        if (num_of_words > 1):
            logger.info("Sender Displayed is valid")
            return "PASS"
        else:
            logger.error("Please check Sender name reflection")
            return "FAIL"

    def validate_signature_displayed(logger, signature_displayed):
        if "Sent by" in signature_displayed:
            logger.info("Signature Sent By Displayed is valid")
            return "PASS"
        else:
            logger.error("Please check Signature Sent By reflection")
            return "FAIL"

    def assert_added(logger):
        time_displayed = driver.find_element_by_xpath(config.get("locator", "time_displayed_xpath")).get_attribute(
            "innerHTML")
        logger.info("Time Displayed in sticket is " + str(time_displayed))
        time_displayed_status = validate_time_displayed(logger, time_displayed)
        print("Time Displayed Status ", time_displayed_status)

        sender_displayed = driver.find_element_by_xpath(config.get("locator", "sender_displayed_xpath")).get_attribute(
            "innerHTML")
        logger.info("Sender Displayed in sticket is " + str(sender_displayed))
        sender_displayed_status = validate_sender_displayed(logger, sender_displayed)

        signature_displayed = driver.find_element_by_xpath(config.get("locator", "signature_xpath")).get_attribute(
            "innerHTML")
        logger.info("Signature Displayed in sticket is " + str(signature_displayed))
        signature_displayed_status = validate_signature_displayed(logger, signature_displayed)

        logger.info("Date and Time Test Case " + str(
            time_displayed_status) + " Sender Displayed Test Case in sticket is " + str(
            sender_displayed_status) + " Signature Displayed Test Case in sticket is  " + str(
            signature_displayed_status))
        if (time_displayed_status == sender_displayed_status == signature_displayed_status == "PASS"):
            return "PASS"
        else:
            return "FAIL"

    def add_sticket(logger, cozeva_id):
        assert_added_status = "N/A"
        delete_status = "N/A"
        sf.ajax_preloader_wait(driver)
        if len(driver.find_elements_by_xpath(
                config.get("locator", "xpath_patient_Header_Dropdown_Arrow"))) != 0:
            sf.action_click(driver.find_element_by_xpath(config.get("locator", "patient_drop_down")), driver)
            sf.action_click(driver.find_element_by_xpath(config.get("locator", "messages_arrow")), driver)
            sf.action_click(driver.find_element_by_xpath(config.get("locator", "new_sticket")), driver)
            logger.info("clicked on New Sticket for " + str(cozeva_id))
            sf.ajax_preloader_wait(driver)
            retrieve_modal_attributes(logger, cozeva_id)
            driver.find_element_by_xpath(config.get("locator", "sticket_modal")).send_keys(config.get("runner", "text"))
            logger.info("Entered text for sticket ")
            time.sleep(2)
            sf.action_click(driver.find_element_by_xpath(config.get("locator", "save_button")), driver)
            logger.info("Saved sticket")
            time.sleep(5)
            added = 1
            WebDriverWait(driver, 40).until(
                EC.text_to_be_present_in_element((By.XPATH, '(//*[text()="test!@#@##@ 123"])[1]'), "test!@#@##@ 123"))
            assert_added_status = assert_added(logger)
            logger.info("Assert Added status " + str(assert_added_status))
            return assert_added_status

    timestamp = []

    def assert_deleted():  # uses text
        sticket_by_text = driver.find_element_by_xpath(config.get("locator", "sticket_by_text_xpath"))
        # sticket_by_timestamp_xpath="//div[@class='col s2 message_time' and normalize-space(text()[1])="+"'"+timestamp[0]+"'"+"and normalize-space(text()[2])="+"'"+timestamp[1]+"'"+"]"
        try:
            WebDriverWait(driver, 30).until(
                EC.invisibility_of_element_located((By.XPATH, config.get("locator", "sticket_by_text_xpath"))))
            return "PASS"
        except TimeoutException:
            print("Failed in assert delete")
            return "FAIL"

    #
    # def verify_add_sticket(driver, workbook, logger, run_from, customer_id):
    #     ws1 = workbook.create_sheet("AddedSticket")
    #     ws = workbook["AddedSticket"]
    #
    #         return [cozeva_id,status]

    def verify_sticket(driver, workbook, logger, run_from, customer_id):
        try:
            workbook.create_sheet("Stickets")
            ws = workbook["Stickets"]
            if (run_from == "Cozeva Support"):
                # initialize report
                ws['A1'].value = "Test Case"
                ws['A1'].font = Font(bold=True, size=13)
                ws['B1'].value = "Status"
                ws['B1'].font = Font(bold=True, size=13)
                ws['C1'].value = "Comments"
                ws['C1'].font = Font(bold=True, size=13)
                ws['A2'] = "Sticket page loads in less than 60 sec"
                ws['A3'] = "All Columns appearing properly "
                ws['A4'] = "Column Data Display"
                ws['A5'] = "Created"
                ws['A6'] = " Last Updated"
                ws['A7'] = "Created By"
                ws['A8'] = "Last Updated By"
                ws['A9'] = "Patient"
                ws['A10'] = "DOB"
                ws['A11'] = "Member ID"
                ws['A12'] = "Member Phone #"
                ws['A13'] = "PCP"
                ws['A14'] = "Latest Note"
                ws['A15'] = "Filter Status"
                ws['A16'] = "Created"
                ws['A17'] = "Last Updated"
                ws['A18'] = "Created By"
                ws['A19'] = "Last Updated By"
                ws['A21'] = "Reflection of Added sticket"
                #open_registry_page(customer_id)
                sf.ajax_preloader_wait(driver)
                main_registry_url = driver.current_url
                logger.info("Opened customer registry" + str(config.get("runner", "customer")))
                # click on filter icon

                filter_icon = driver.find_element_by_xpath(config.get("locator", "filter_list"))
                sf.action_click(filter_icon, driver)

                # sort the registry

                sort_by = driver.find_element_by_xpath(config.get("locator", "sort_by_xpath"))
                sf.action_click(sort_by, driver)

                denominator_option = driver.find_element_by_xpath(config.get("locator", "denominator_option_xpath"))
                sf.action_click(denominator_option, driver)

                apply_button = driver.find_element_by_xpath(config.get("locator", "apply_button_xpath"))
                sf.action_click(apply_button, driver)

                # click on first metric

                first_metric = driver.find_element_by_xpath(config.get("locator", "first_metric_xpath"))
                sf.action_click(first_metric, driver)

                # wait for page to load

                sf.ajax_preloader_wait(driver)
                name_header_xpath = config.get("locator", "table_header_xpath")
                WebDriverWait(driver, 30).until(EC.visibility_of_element_located((By.XPATH, name_header_xpath)))
                # open patients tab

                patient_tab = driver.find_element_by_xpath(config.get("locator", "patient_xpath"))
                sf.action_click(patient_tab, driver)

                # wait for page to load

                sf.ajax_preloader_wait(driver)
                table_header_patient_xpath = config.get("locator", "table_header_patient_xpath")
                WebDriverWait(driver, 30).until(
                    EC.visibility_of_element_located((By.XPATH, table_header_patient_xpath)))
                # store the list of patients

                patient_num = 0
                patient_id_links = driver.find_elements_by_xpath(config.get("locator", "patient_name_list"))
                # click first patient and add sticket
                for patient_num in range(1, 2):
                    patient_xpath_final = "(" + config.get("locator", "patient_name_list") + ")" + "[" + str(
                        patient_num) + "]"
                    patient_link = driver.find_element_by_xpath(patient_xpath_final)
                    cozeva_id = extract_patient_id(patient_link.get_attribute("href"))
                    # click on patient
                    sf.action_click(patient_link, driver)

                driver.switch_to.window(driver.window_handles[1])
                window_switched = 1
                # returns Pass only if added data is reflected properly in the modal
                status = add_sticket(logger, cozeva_id)
                driver.switch_to.window(driver.window_handles[0])
                # open customer messaging
                open_customer_messaging(customer_id)
                logger.info("Navigating to customer messaging ")
                # ascertain time to load
                time_to_load_start = datetime.now()
                sf.ajax_preloader_wait(driver)
                time_to_load_end = datetime.now()
                time_to_load = time_to_load_end - time_to_load_start
                print("Time to load page", time_to_load)
                logger.info("Time to load messaging page " + str(time_to_load))
                # click on sticket drop down
                sticket_drop_down = driver.find_element_by_xpath(
                    config.get("sticket-log-locator", "sticket_dropdown_icon_xpath"))
                sf.action_click(sticket_drop_down, driver)
                logger.info("Clicked on collapsible drop down")
                # scroll down page
                sticket_log_link = driver.find_element_by_xpath(config.get("sticket-log-locator", "sticket_log_xpath"))
                driver.execute_script("arguments[0].scrollIntoView();", sticket_log_link)

                # click on sticket log

                sf.action_click(sticket_log_link, driver)
                time_to_load_sticket_page_start = datetime.now()

                # record time for page load
                created_column_xpath = config.get("sticket-log-locator", "created_column_xpath")
                try:
                    WebDriverWait(driver, 60).until(EC.visibility_of_element_located((By.XPATH, created_column_xpath)))
                    page_load_status = "PASS"
                except TimeoutException:
                    logger.error("Page is taking more than 60 seconds to load ")
                    page_load_status = "FAIL"

                time_to_load_sticket_page_end = datetime.now()

                time_to_load_sticket_page = time_to_load_sticket_page_end - time_to_load_sticket_page_start
                print("Time to load sticket ", time_to_load_sticket_page)
                logger.info("Time to load sticket page " + str(time_to_load_sticket_page))

                # count number of rows

                num_of_rows_total = find_number_of_rows(driver)
                try:
                    if (num_of_rows_total >= 1):
                        ws['A20'] = "Number of records"
                        ws['C20'] = num_of_rows_total
                except:
                    sticket_added_reflection = "FAIL"
                    raise Exception("Sticket records is empty")

                column_set_string = config.get("runner", "column_set")
                column_set = list(column_set_string.split(","))
                print("column set" + str(column_set))
                column_xpath = "(//tr[@role='row'])[1]//child::th"
                column_set_match = ""
                all_column_status = []
                all_columns_comment = ""
                if (find_number_of_columns(driver, column_xpath) - 1 == len(column_set)):
                    column_set_match = "PASS"
                    created = []
                    last_updated = []
                    created_by = []
                    last_updated_by = []
                    patient = []
                    dob = []
                    member_id = []
                    member_phone = []
                    pcp = []
                    latest_note = []

                    # will check from first four records
                    for i in range(2, 3):
                        row_xpath_2 = config.get("sticket-log-locator", "row_xpath_2") + str(i) + "]"

                        created_xpath = row_xpath_2 + "//child::td[1]"
                        last_updated_xpath = row_xpath_2 + "//child::td[2]"
                        created_by_xpath = row_xpath_2 + "//child::td[3]"
                        last_updated_by_xpath = row_xpath_2 + "//child::td[4]"
                        patient_xpath = row_xpath_2 + "//child::td[5]//child::a"
                        dob_xpath = row_xpath_2 + "//child::td[6]"
                        member_id_xpath = row_xpath_2 + "//child::td[7]//div"
                        member_phone_xpath = row_xpath_2 + "//child::td[8]"
                        pcp_xpath = row_xpath_2 + "//child::td[9]"
                        latest_note_xpath = row_xpath_2 + "//child::td[10]//span[@style='display: inline-block;word-break: break-all;']"

                        try:
                            data = "created_value"
                            created_text = driver.find_element_by_xpath(created_xpath).get_attribute("innerHTML")
                            created.append(created_text)

                            data = "last_updated_value"
                            last_updated_text = driver.find_element_by_xpath(last_updated_xpath).get_attribute(
                                "innerHTML")
                            last_updated.append(last_updated_text)

                            data = "last_updated_value"
                            created_by_text = driver.find_element_by_xpath(created_by_xpath).get_attribute("innerHTML")
                            created_by.append(created_by_text)

                            data = "last_updated_by_value"
                            last_updated_by_text = driver.find_element_by_xpath(last_updated_by_xpath).get_attribute(
                                "innerHTML")
                            last_updated_by.append(last_updated_by_text)

                            data = "patient_value"
                            patient_text = extract_patient_id(
                                driver.find_element_by_xpath(patient_xpath).get_attribute("href"))
                            patient.append(patient_text)

                            data = "dob_value"
                            dob_text = driver.find_element_by_xpath(dob_xpath).get_attribute("innerHTML")
                            dob.append(dob_text)

                            data = "member_id_value"
                            try:
                                member_id_text = driver.find_element_by_xpath(member_id_xpath).get_attribute(
                                    "innerHTML")
                                member_id.append(member_id_text)
                            except NoSuchElementException:
                                member_id.append(" ")
                                pass

                            data = "member_phone_value"
                            try:
                                member_phone_text = driver.find_element_by_xpath(member_phone_xpath).get_attribute(
                                    "innerHTML")
                                member_phone.append(member_phone_text)
                            except NoSuchElementException:
                                member_phone.append(" ")
                                pass

                            data = "pcp_value"
                            pcp_text = driver.find_element_by_xpath(pcp_xpath).get_attribute("innerHTML")
                            pcp.append(pcp_text)

                            data = "latest_note_value"
                            latest_note_text = driver.find_element_by_xpath(latest_note_xpath).get_attribute(
                                "innerHTML")
                            latest_note.append(latest_note_text)
                        except NoSuchElementException:
                            print(str(data) + "Not found")
                            pass

                        print(created)
                        print(last_updated)
                        print(created_by)
                        print(last_updated_by)
                        print(patient)
                        print(dob)
                        print(member_id)
                        print(member_phone)
                        print(pcp)
                        print(latest_note)

                        all_columns_status = []
                    created_status = ""
                    last_updated_status = ""
                    created_by_status = ""
                    last_updated_by_status = ""
                    patient_status = ""
                    dob_status = ""
                    member_id_status = ""
                    member_phone_status = ""
                    pcp_status = ""
                    latest_note_status = ""
                    for i in range(0, 1):
                        created_status = validate_created(created[i])
                        all_columns_status.append(created_status)
                        logger.info("Validated created_status " + str(created_status))

                        last_updated_status = validate_last_updated(last_updated[i])
                        all_columns_status.append(last_updated_status)
                        logger.info("Validated last_updated_status" + str(last_updated_status))

                        created_by_status = validate_created_by(created_by[i])
                        all_columns_status.append(created_by_status)
                        logger.info("Validated created_by_status" + str(created_by_status))

                        last_updated_by_status = validate_last_updated_by(last_updated_by[i])
                        all_columns_status.append(last_updated_by_status)
                        logger.info("Validated last_updated_by_status" + str(last_updated_by_status))

                        patient_status = validate_patient_status(patient[i])
                        all_columns_status.append(patient_status)
                        if (patient[i] == cozeva_id):
                            print("patient[i]" + str(patient[i]))
                            print("cozevaid" + str(cozeva_id))
                            sticket_added_reflection = "PASS"
                        logger.info("Validated patient_status" + str(patient_status))

                        dob_status = validate_dob(dob[i])
                        all_columns_status.append(dob_status)
                        logger.info("Validated dob_status" + str(dob_status))

                        member_id_status = validate_member_id_status(member_id[i])
                        all_columns_status.append(member_id_status)
                        logger.info("Validated member_id_status" + str(member_id_status))

                        member_phone_status = validate_member_phone(member_phone[i])
                        all_columns_status.append(member_phone_status)
                        logger.info("Validated member_phone_status" + str(member_phone_status))

                        pcp_status = validate_pcp(pcp[i])
                        all_columns_status.append(pcp_status)
                        logger.info("Validated pcp_status" + str(pcp_status))

                        latest_note_status = validate_latest_note(latest_note[i])
                        all_columns_status.append(latest_note_status)

                        logger.info("Validated latest_note_status " + str(latest_note_status))

                    all_column_status = validate_all_columns(all_columns_status)
                    logger.info("Result of all columns  " + str(all_columns_status))

                    # verifies from second record
                    random = 0
                    created_filter_status = ""
                    last_updated_filter_status = ""
                    created_by_filter_status = ""
                    last_updated_by_filter_status = ""
                    if (created_status == "PASS"):
                        created_filter_status = verify_filter(driver, "created", created[random])
                        print("Created Filter Status" + str(created_filter_status))
                    if (last_updated_status == "PASS"):
                        last_updated_filter_status = verify_filter(driver, "last_updated", last_updated[random])

                    if (created_by_status == "PASS"):
                        created_by_filter_status = verify_filter(driver, "created_by", created_by[random])
                        print("Createdby Filter Status" + str(created_by_filter_status))
                    #

                    if (last_updated_by_status == "PASS"):
                        last_updated_by_filter_status = verify_filter(driver, "last_updated_by",
                                                                      last_updated_by[random])
                        print(last_updated_by_filter_status)

                    # delete sticket

                    ws['B2'] = page_load_status
                    ws['C2'] = time_to_load_sticket_page
                    ws['B3'] = column_set_match
                    ws['C3'] = all_columns_comment
                    ws['C4'] = "Fail status expected for Onshore clients"
                    ws['C18'] = "Fail status expected for Onshore clients"
                    ws['B4'] = all_column_status
                    ws['B5'] = created_status
                    ws['B6'] = last_updated_status
                    ws['B7'] = created_by_status
                    ws['B8'] = last_updated_by_status
                    ws['B9'] = patient_status
                    ws['B10'] = dob_status
                    ws['B11'] = member_id_status
                    ws['B12'] = member_phone_status
                    ws['C11'] = "For onshore customers blank is expected"
                    ws['C12'] = "For onshore customers blank is expected"
                    ws['B13'] = pcp_status
                    ws['B14'] = latest_note_status
                    ws['B16'] = created_filter_status
                    ws['B17'] = last_updated_filter_status
                    ws['B18'] = created_by_filter_status
                    ws['B19'] = last_updated_by_filter_status
                    ws['B21'] = sticket_added_reflection
                    ws['C21'] = cozeva_id

                    try:
                        driver.switch_to.window(driver.window_handles[1])

                        sf.action_click(
                            driver.find_element_by_xpath(
                                config.get("locator", "sticket_delete_icon_for_concerned_text")), driver)
                        logger.info("Clicked on delete ")
                        print("Clicked on Deleted")
                        deleted = 1
                        time.sleep(2)
                        sf.action_click(driver.find_element_by_xpath(config.get("locator", "confirm_button")), driver)
                        logger.info("Confirmed deletion from confirmation modal")
                        time.sleep(4)
                        delete_status = assert_deleted()
                        driver.close()
                        logger.info("Closed patient's tab " + str(cozeva_id))
                        driver.switch_to.window(driver.window_handles[0])

                    except NoSuchElementException:
                        print("Delete from main")
                        delete_status = "FAIL"
                        pass

                    driver.refresh()
                    sf.ajax_preloader_wait(driver)
                    # record time for page load
                    created_column_xpath = config.get("sticket-log-locator", "created_column_xpath")
                    WebDriverWait(driver, 60).until(EC.visibility_of_element_located((By.XPATH, created_column_xpath)))

                    try:
                        patient_text2 = extract_patient_id(
                            driver.find_element_by_xpath(patient_xpath).get_attribute("href"))
                    except NoSuchElementException:
                        patient_text2 = ""
                    print(patient_text2)
                    if (patient_text2 != cozeva_id):
                        deleted_status = "PASS"
                    else:
                        deleted_status = "FAIL"

                    if (deleted_status == "PASS"):
                        ws.append(("Sticket Delete functionality", "PASS"))
                    else:
                        ws.append(("Sticket Delete functionality", "FAIL",
                                    "Manual intervention required for cozeva id given "))

                    apply_conditional_formatting(ws)
                    #workbook.save("Report_Sticket_Log.xlsx")

                    # ws['D4'] = "Number of Pages " + str(page)


                else:
                    column_set_match = False
                    list_of_columns = extract_name_of_columns(driver, column_xpath)
                    if (len(list_of_columns) < len([column_set])):
                        column_missing = list(set(column_set).difference(list_of_columns))
                        all_columns_comment = " ".join(column_missing) + " is missing"
                    else:
                        all_columns_comment = "extra columns found"
        except Exception as e:
            ws['A1'] = "Sticket page empty or failed to load "
            #workbook.save("Report_Sticket_Log.xlsx")
            print("Failed to continue sticket " + str(e))
            traceback.print_exc()

        rows = ws.max_row
        cols = ws.max_column
        for i in range(1, rows + 1):
            for j in range(1, cols + 1):
                if ws.cell(i, j).value == 'PASS':
                    ws.cell(i, j).fill = PatternFill('solid', fgColor='0FC404')
                elif ws.cell(i, j).value == 'FAIL':
                    ws.cell(i, j).fill = PatternFill('solid', fgColor='FC0E03')
                elif ws.cell(i, j).value == 'Data table is empty':
                    ws.cell(i, j).fill = PatternFill('solid', fgColor='FCC0BB')
        driver.get(main_registry_url)


        # try:
        #   count number of records
        #

    # create Folder or working directory
    # dateandtime = date_time()
    # master_directory = config.get("runner", "report_directory_input")
    # os.chdir(master_directory)
    # path = makedir(dateandtime)
    # LOG_FORMAT = "%(levelname)s %(asctime)s - %(message)s"
    # logging.basicConfig(filename=path + "\\" + "Sticket-Log.log", level=logging.INFO, format=LOG_FORMAT, filemode='w')
    # logger = logging.getLogger()
    # # logger.setLevel(logging.INFO)
    # os.chdir(path)

    # downloaddefault = config.get("runner", "downloaddefault")
    # makedir(downloaddefault)
    # driver = setup("Chrome", downloaddefault)
    # begin_time = datetime.now()
    # loc = config.get("runner", "login_file")

    # login
    # login(driver, loc)
    # logger.info("Login successful")

    # Initialize Worksheet

    # wb = openpyxl.Workbook()
    # def verify_sticket_functionality(driver, workbook, logger, run_from, customer_id):
    # added_status = [cozevaid, status] = verify_add_sticket(driver, workbook, logger, run_from, customer_id)  # worksheetappend
    # cozeva_id=
    # delete_status = verify_delete_sticket(added_status, driver, workbook, logger, run_from, customer_id)  # worksheetappend
    # if (delete_status == False):
    #     print("Failed to delete Sticket " + " for " + str(added_status[0]) + " Customer " + str(
    #         customer_id))  # add formatting if possible

    #
    # verify_sticket_functionality(driver, wb, logger, "CS", customer_id)

    # verify_sticket(driver, workbook, logger, run_from, customer_id)
    verify_sticket(driver, workbook, logger, run_from, customer_id)


def map_codingtool(driver, workbook, logger, run_from, customer_id):
    def apply_conditional_formatting(ws):
        red_text = Font(color="9C0006")
        red_fill = PatternFill(bgColor="FFC7CE")
        dxf = DifferentialStyle(font=red_text, fill=red_fill)
        rule = Rule(type="containsText", operator="containsText", text="FAIL", dxf=dxf)
        rule.formula = ['NOT(ISERROR(SEARCH("highlight",A1)))']

        green_text = Font(color="00FF00FF")
        green_fill = PatternFill(bgColor="0000FF00")
        dxf = DifferentialStyle(font=green_text, fill=green_fill)
        rule1 = Rule(type="containsText", operator="containsText", text="PASS", dxf=dxf)
        rule1.formula = ['NOT(ISERROR(SEARCH("highlight",A1)))']

        ws.conditional_formatting.add('B1:B10000', rule)
        ws.conditional_formatting.add('B1:B10000', rule1)

    def open_pendinglist(customer_id):
        customer_list_url = []
        sm_customer_id = str(customer_id)
        sm_customer_id = sm_customer_id.split(".")[0]
        session_var = 'app_id=registries&custId=' + str(sm_customer_id) + '&payerId=' + str(
            sm_customer_id) + '&orgId=' + str(sm_customer_id)
        encoded_string = base64.b64encode(session_var.encode('utf-8'))
        customer_list_url.append(encoded_string)
        for idx, val in enumerate(customer_list_url):
            driver.get(config.get("runner", "URL") + "registries/pending-list?session=" + val.decode('utf-8'))

    def extract_patient_id(href):
        cozeva_id = re.search('/patient_detail/(.*)?session', href)
        return (cozeva_id.group(1).replace("?", ""))

    def format_string(s):
        s1 = s.replace("-", '')
        s2 = s1.replace(" ", '')
        return s2

    def PatientDashboard(driver, sheet, quarter_name, lob_name, metric_name_4_patientdashboard,
                         add_supdata_flag_MSPL,
                         map_flag_MSPL, caregap_MSPL, mspl_url, provider_name):
        # driver.get("https://stage.cozeva.com/patient_detail/1R0ADY3?tab_type=CareOps&cozeva_id=1R0ADY3&patient_id=9290597&cozeva_id=1R0ADY3&session"
        #          "=YXBwX2lkPXJlZ2lzdHJpZXMmY3VzdElkPTE1MDAmZG9jdG9yc1BlcnNvbklkPTExODUxNTYzJmRvY3Rvcl91aWQ9MTE4MzE0ODkmcGF5ZXJJZD0xNTAwJnF1YXJ0ZXI9MjAyMC0xMi0zMSZob21lPVlYQndYMmxrUFhKbFoybHpkSEpwWlhNbVkzVnpkRWxrUFRFMU1EQW1jR0Y1WlhKSlpEMHhOVEF3Sm05eVowbGtQVEUxTURB&first_load=1")
        global add_supdata_flag_pt, map_flag_pt, cozeva_id, pcp_name

        try:
            sf.ajax_preloader_wait(driver)
            if len(driver.find_elements_by_xpath("(//div/span[@data-tooltip='Cozeva Id (Click to Copy)'])[1]")) != 0:
                cozeva_id = driver.find_element_by_xpath(
                    "(//div/span[@data-tooltip='Cozeva Id (Click to Copy)'])[1]").text
            elif len(driver.find_elements_by_xpath("(//div/span[@data-tooltip='Cozeva Id (Click to Copy)'])[1]")) == 0:
                cozeva_id = "Blank ; Please check "
                return
            sheet['B2'] = quarter_name + " | " + lob_name
            sheet['B3'] = metric_name_4_patientdashboard
            sheet['B4'] = cozeva_id
            # find metric pencil icon

            # Red dot count
            caregap_pt = len(driver.find_elements_by_xpath("//div[@class='non_compliant red_dot']"))

            print("Metric name for patient dashboard" + str(metric_name_4_patientdashboard))
            metric_name = metric_name_4_patientdashboard
            # list of metrics in patient_dashboard
            hcc_counter = 0
            if sf.check_exists_by_xpath(driver, "//table[@id='table_4']//div[@class='text-bold sub-title']"):
                hcc_metric_patientdashboard = driver.find_elements_by_xpath(
                    "//table[@id='table_4']//div[@class='text-bold sub-title']")
                hcc_counter = len(hcc_metric_patientdashboard)
                print("number of hcc metric " + str(hcc_counter))
            # value is taken in quality so index is of quality
            metrics_patientdashboard = driver.find_elements_by_xpath(
                "//table[@id='table_1']//div[@class='text-bold sub-title']")
            measure_display_flag = 0
            # hcc_collapse_xpath='//i[@class="material-icons hcc_toggle tooltipped"]'
            # if(check_exists_by_xpath(driver,hcc_collapse_xpath)):
            #     action_click(driver.find_element_by_xpath(hcc_collapse_xpath))
            total_metric_counter = hcc_counter + len(metrics_patientdashboard)
            print("Total metric " + str(total_metric_counter))
            print(range(total_metric_counter))
            for metric_counter in range(total_metric_counter):
                print("In Metric Counter block")
                print("Metric Counter" + str(metric_counter))
                metric_counter1 = metric_counter + 1
                print("Metric counter1 value " + str(metric_counter1))
                xpath1 = "(" + "//div[@class='text-bold sub-title']" + ")" + "[" + str(
                    hcc_counter + metric_counter1) + "]"
                xpath_metric_row = xpath1 + "/../../../../../.."
                xpath_pencil_patientdashboard = xpath1 + "/../../../../../../td/div/div[@class='dropdown']//child::a[@class='addSuppData-trigger pts']//child::i"
                metric_name_patientdashboard = metrics_patientdashboard[metric_counter].text
                print("Metric name in patient dashboard" + str(metric_name_patientdashboard))
                metric_row = driver.find_element_by_xpath(xpath_metric_row)
                # print(metric_name_patientdashboard)#Print all measures in Patient dashboard

                driver.execute_script("arguments[0].scrollIntoView(true);", metric_row)

                if metric_name_patientdashboard == metric_name:

                    ActionChains(driver).move_to_element(metric_row).perform()

                    if len(driver.find_elements_by_xpath(xpath_pencil_patientdashboard)) == 1:
                        sheet.append(("Pencil icon Present ?", "PASS"))
                        sf.action_click(driver.find_element_by_xpath(xpath_pencil_patientdashboard), driver)
                        print("Clicked on Pencil icon")
                        time.sleep(1)
                        xpath_pencil_options = xpath_pencil_patientdashboard + "//..//..//child::ul/li"
                        WebDriverWait(driver, 30).until(
                            EC.visibility_of_element_located((By.XPATH, xpath_pencil_options)))
                        pencil_options = driver.find_elements_by_xpath(xpath_pencil_options)
                        add_supdata_flag_pt = 0
                        map_flag_pt = 0

                        for option_counter in range(len(pencil_options)):

                            print((pencil_options[option_counter]).text)
                            pencil_options_pt_text = (pencil_options[option_counter]).text
                            if pencil_options_pt_text.strip() == "Add Supplemental Data":
                                add_supdata_flag_pt = 1
                                sheet.append(("Add Supplemental Data Present ?", "PASS"))
                                sf.action_click(pencil_options[option_counter], driver)
                                sf.ajax_preloader_wait(driver)
                                # verify submit button
                                submit_button_xpath = config.get("MAP", "submit_xpath")
                                try:
                                    driver.find_element_by_xpath(submit_button_xpath)
                                    sheet.append(("Submit button appearing in Supp data", "PASS"))
                                except NoSuchElementException:
                                    sheet.append(("Submit button appearing in Supp data", "FAIL"))

                                # verify delete button
                                delete_button_xpath = config.get("MAP", "delete_xpath")
                                try:
                                    driver.find_element_by_xpath(delete_button_xpath)
                                    sheet.append(("Delete button appearing in Supp data", "PASS"))
                                except NoSuchElementException:
                                    sheet.append(("Delete button appearing in Supp data", "FAIL"))

                                # verify Task id
                                task_text_xpath = config.get("MAP", "task_id_xpath")
                                try:
                                    task_text = driver.find_element_by_xpath(task_text_xpath).text
                                    sheet.append(("Task id appearing in Supp data", "PASS", str(task_text)))
                                except NoSuchElementException:
                                    sheet.append(("Task id appearing in Supp data", "FAIL"))

                                # verify attachment section
                                attachment_xpath = config.get("MAP", "attachment_xpath")
                                try:
                                    driver.find_element_by_xpath(attachment_xpath)
                                    sheet.append(("Attachment section appearing in Supp data", "PASS"))
                                except NoSuchElementException:
                                    sheet.append(("Attachment section appearing in Supp data", "FAIL"))

                                # Delete the task
                                delete_button_xpath = config.get("MAP", "delete_xpath")
                                try:
                                    sf.action_click(driver.find_element_by_xpath(delete_button_xpath), driver)

                                    # give reason
                                    reason_modal_xpath = config.get("MAP", "reason_input_modal")
                                    reason_modal = driver.find_element_by_xpath(reason_modal_xpath)
                                    reason_modal.send_keys("Cozeva QA")
                                    time.sleep(1)
                                    sf.action_click(driver.find_element_by_xpath(config.get("MAP", "confirm_modal_xpath")), driver)
                                    time.sleep(5)
                                    sf.ajax_preloader_wait(driver)

                                    sheet.append(("Task Deleted", "PASS"))
                                except NoSuchElementException:
                                    sheet.append(("Task Deleted", "FAIL", "Manual intervention required "))

                                sf.action_click(driver.find_element_by_xpath(xpath_pencil_patientdashboard), driver)

                            if pencil_options_pt_text.strip() == "Mark as Pending":
                                sheet.append(("Mark As Pending Present ?", "PASS"))
                                # click on MAP
                                map_flag_pt = 1
                                sf.action_click(pencil_options[option_counter], driver)

                                # click on confirm
                                # time.sleep(1)
                                # sf.action_click(driver.find_element_by_xpath(config.get("MAP", "confirm_modal_xpath")), driver)
                                # time.sleep(2)
                                # wait for page to load
                                sf.ajax_preloader_wait(driver)

                                # check for stale icon
                                restored = 0
                                stale_icon = 0
                                x = 1
                                start_time1 = timeit.default_timer()
                                while True:
                                    driver.refresh()
                                    sf.ajax_preloader_wait(driver)
                                    if (sf.check_exists_by_xpath(driver, config.get("MAP", "stale_icon_xpath"))):
                                        print("Stale icon found ")
                                        stale_icon = 1
                                        break
                                    if (x == 25):
                                        break
                                    x = x + 1
                                time_elapsed1_value = timeit.default_timer() - start_time1
                                time_elapsed1 = '{0:.2f}'.format(time_elapsed1_value)
                                if (stale_icon == 1):
                                    timestring = "Time taken(in s) " + str(time_elapsed1)
                                    sheet.append(("Mark As pending - Stale icon ", "PASS", str(timestring)))
                                else:
                                    timestring = "Time taken(in s) " + str(time_elapsed1)
                                    sheet.append(("Mark As pending -Stale icon", "FAIL", str(timestring)))

                                # Keep refreshing till you see the hollow dot
                                # Refresh 10 times to verify appearing of hollow dot
                                start_time = timeit.default_timer()
                                hollow_dot_found = 0
                                dot_status_xpath = "(//div[@class='text-bold sub-title'])" + "[" + str(hcc_counter +
                                    metric_counter1) + "]" + "//ancestor::tr//child::td[1]//child::div[contains(@style,'margin: 8px 0px 0px 4px;')]"
                                dot_status = driver.find_element_by_xpath(dot_status_xpath).get_attribute("class")
                                print("Dot status " + str(dot_status))
                                y = 1
                                while True:
                                    driver.refresh()
                                    sf.ajax_preloader_wait(driver)
                                    dot_status_xpath = "(//div[@class='text-bold sub-title'])" + "[" + str(
                                        hcc_counter+metric_counter1) + "]" + "//ancestor::tr//child::td[1]//child::div[contains(@style,'margin: 8px 0px 0px 4px;')]"
                                    dot_status = driver.find_element_by_xpath(dot_status_xpath).get_attribute("class")
                                    if (dot_status == "non_compliant hollow_dot"):
                                        print("Checking for hollow dot ")
                                        hollow_dot_found = 1
                                        break
                                    if (y == 25):
                                        break
                                    print("Dot Status while checking for hollow dot " + str(dot_status))
                                    y = y + 1

                                elapsed_value = timeit.default_timer() - start_time
                                elapsed = '{0:.2f}'.format(elapsed_value)
                                if (hollow_dot_found == 1):  # Click on pencil icon and unmark as Pending
                                    sf.ajax_preloader_wait(driver)
                                    timestring = "Time taken(in s) " + str(elapsed)
                                    sheet.append(("Hollow dot ", "PASS", str(timestring)))

                                    # check in pending list
                                    # open pending list
                                    open_pendinglist(customer_id)
                                    sf.ajax_preloader_wait(driver)
                                    provider_heading_xpath = '//*[text()="Provider"]'
                                    WebDriverWait(driver, 60).until(
                                        EC.visibility_of_element_located((By.XPATH, provider_heading_xpath)))
                                    print("Page loaded completely ")
                                    # extract patient cozeva id
                                    WebDriverWait(driver, 60).until(
                                        EC.visibility_of_element_located(
                                            (By.XPATH, config.get("MAP", "patient_link_xpath"))))
                                    patient_link = driver.find_element_by_xpath(config.get("MAP", "patient_link_xpath"))
                                    cozeva_id_pending_list = extract_patient_id(patient_link.get_attribute("href"))
                                    if (format_string(cozeva_id) == format(cozeva_id_pending_list)):
                                        sheet.append(("Displayed in Pending List", "PASS"))
                                    else:
                                        sheet.append(("Displayed in Pending List", "FAIL", "Please check"))
                                    # unmark as pending
                                    driver.back()
                                    sf.ajax_preloader_wait(driver)
                                    WebDriverWait(driver, 30).until(
                                        EC.element_to_be_clickable((By.XPATH, xpath_pencil_patientdashboard)))
                                    sf.action_click(driver.find_element_by_xpath(xpath_pencil_patientdashboard), driver)
                                    print("Clicked on Pencil icon")
                                    unmark_as_pending_xpath = '(//*[text()="Unmark as Pending"])[' + str(1) + ']'
                                    time.sleep(3)
                                    unmark_as_pending = driver.find_element_by_xpath(unmark_as_pending_xpath)
                                    sf.action_click(unmark_as_pending, driver)
                                    print("Clicked on unmark as pending icon")
                                    sf.ajax_preloader_wait(driver)
                                    dot_status_xpath = "(//div[@class='text-bold sub-title'])" + "[" + str(hcc_counter+
                                        metric_counter1) + "]" + "//ancestor::tr//child::td[1]//child::div[contains(@style,'margin: 8px 0px 0px 4px;')]"
                                    dot_status = driver.find_element_by_xpath(dot_status_xpath).get_attribute("class")
                                    z = 0
                                    while True:
                                        driver.refresh()
                                        sf.ajax_preloader_wait(driver)
                                        dot_status_xpath = "(//div[@class='text-bold sub-title'])" + "[" + str(hcc_counter+
                                            metric_counter1) + "]" + "//ancestor::tr//child::td[1]//child::div[contains(@style,'margin: 8px 0px 0px 4px;')]"
                                        dot_status = driver.find_element_by_xpath(dot_status_xpath).get_attribute(
                                            "class")
                                        if (dot_status == "non_compliant red_dot"):
                                            restored = 1
                                            break
                                        if (z == 25):
                                            break
                                        z = z + 1
                                    if (restored == 1):
                                        timestring = "Time waited for hollow dot " + str(elapsed)
                                        sheet.append(
                                            ("Unmark as Pending", "PASS", str(timestring)))
                                    else:
                                        timestring = "Manual intervention required , Time waited for hollow dot " + str(
                                            elapsed)
                                        sheet.append(
                                            ("Unmark as pending hasn't occurred as red dot has not re-appear", "FAIL",
                                             str(timestring)))
                                else:
                                    timestring = "Manual intervention required , Time waited for hollow dot " + str(
                                        elapsed)
                                    sheet.append(
                                        ("Unmark as pending hasn't occurred as hollow dot has not appear", "FAIL",
                                         str(timestring)))

                            pencil_options = driver.find_elements_by_xpath(xpath_pencil_options)
                        if (add_supdata_flag_pt != 1):
                            sheet.append(("Add Supplemental Data Present ?", "FAIL"))
                        if (map_flag_pt != 1):
                            sheet.append(("Mark As Pending Present ?", "FAIL"))



                    elif len(driver.find_elements_by_xpath(xpath_pencil_patientdashboard)) == 0:
                        print("NO PENCIL")
                        sheet.append(("Pencil icon Present ?", "FAIL"))
                        add_supdata_flag_pt = 0
                        map_flag_pt = 0
                        return False

                    print("Supdata flag(Pt): " + str(add_supdata_flag_pt))
                    print("Map flag(Pt): " + str(map_flag_pt))
                    measure_display_flag = 1
                    break
                else:
                    print("Metric name is not equal")

            if (add_supdata_flag_pt == map_flag_pt == 1):
                return True
            else:
                return False

        except Exception as e:
            print(e)
            return False

            # Click on MAP
            # confirm yes on the modal
            # check stale icon
            # 4 th test case pass
            # click on Add Supp data option
            # verify submit and delete button
            # 5th Test Case pass
            # navigate to pending list
            # check for patient cozeva id
            # refresh 5-6 times
            # 6h test case pass
            # return to Patient dashboard
            # check for hollow dot
            # unmark as pending
            # 7th test case pass
            # check for no hollow dot

    def verify_mark_as_pending(driver, workbook, logger, run_from):
        workbook.create_sheet("MAPCodingTool")
        ws = workbook["MAPCodingTool"]
        ws['A1'].value = "Test Data"
        ws['A1'].font = Font(bold=True, size=13)
        ws['A2'].value = "LOB"
        ws['A2'].font = Font(bold=True, size=13)
        ws['A3'].value = "Metric"
        ws['A3'].font = Font(bold=True, size=13)
        ws['A4'].value = "Cozeva ID"
        ws['A4'].font = Font(bold=True, size=13)
        ws['A5'].value = "Test Case"
        ws['A5'].font = Font(bold=True, size=13)
        ws['B5'].value = "Status"
        ws['B5'].font = Font(bold=True, size=13)
        ws['C5'].value = "Comments"
        ws['C5'].font = Font(bold=True, size=13)
        patient_verified = ""
        # navigate to registry
        sf.ajax_preloader_wait(driver)
        driver.refresh()
        sf.ajax_preloader_wait(driver)
        WebDriverWait(driver, 60).until(EC.presence_of_element_located((By.XPATH, locator.xpath_filter_measure_list)))
        # customer_name = driver.find_element_by_xpath(config['LOCATOR']['xpath_contextName']).text
        driver.find_element_by_xpath("//a[@id='qt-filter-label']").click()
        time.sleep(1)
        quarters = driver.find_elements_by_xpath("//ul[@id='filter-quarter']/li")
        lobs = driver.find_elements_by_xpath("//ul[@id='filter-lob']/li[@class!='hide']")
        driver.find_element_by_xpath("//a[@id='qt-filter-label']").click()
        patient_found = ""
        for quarter in range(2):
            quarter = quarter + 1
            if (patient_found == "Found"):
                break
            for lob in range(len(lobs)):
                if (patient_found == "Found"):
                    break
                # for lob in range(1):
                # lob = lob + 3
                time.sleep(0.5)
                WebDriverWait(driver, 30).until(
                    EC.presence_of_element_located((By.XPATH, "//a[@id='qt-filter-label']")))
                driver.find_element_by_xpath("//a[@id='qt-filter-label']").click()
                time.sleep(0.25)
                quarter_name = quarters[quarter].text
                print(colored(quarter_name, 'blue'))
                quarters[quarter].click()
                time.sleep(0.25)
                lobs[lob].click()
                lob_name = lobs[lob].text
                print(colored(lob_name, 'magenta'))
                driver.find_element_by_xpath("//a[@id='reg-filter-apply']").click()
                WebDriverWait(driver, 90).until(
                    EC.invisibility_of_element((By.XPATH, "//div[@class='ajax_preloader']")))
                WebDriverWait(driver, 30).until(
                    EC.presence_of_element_located((By.XPATH, "//a[@data-target='qt-reg-nav-filters']")))
                driver.find_element_by_xpath("//a[@data-target='qt-reg-nav-filters']").click()
                time.sleep(0.25)
                WebDriverWait(driver, 30).until(EC.presence_of_element_located((By.XPATH, "//label[@class='col s12']")))
                driver.find_element_by_xpath("//label[@class='col s12']").click()
                time.sleep(0.25)
                driver.find_element_by_xpath("//button[@id='qt-apply-search']").click()
                WebDriverWait(driver, 90).until(
                    EC.invisibility_of_element((By.XPATH, "//div[@class='ajax_preloader']")))

                """
                **** SUPPORT MEASURE REGISTRY NAVIGATION ****
                """
                measures_all = driver.find_elements_by_xpath("//div/span[@class='met-name']")
                scores = driver.find_elements_by_xpath("//span[@class='num-den']")
                measure_counter = 0
                score = 0

                while measure_counter < len(measures_all) and score < len(scores):
                    if (patient_found == "Found" or patient_verified == True):
                        break
                    WebDriverWait(driver, 60).until(
                        EC.presence_of_element_located((By.XPATH, "//a[@id='reg-faq-trigger']")))
                    time.sleep(0.5)
                    #driver.execute_script("arguments[0].scrollIntoView();", measures_all[measure_counter])
                    measure_name = (measures_all[measure_counter]).text
                    print("Measure name: ", measure_name)
                    numdeno = scores[score].text
                    numdeno = numdeno.lstrip("(")
                    numdeno = numdeno.rstrip(")")
                    numdeno = numdeno.split("/")
                    Numerator = numdeno[0].replace(',', '')
                    Denominator = numdeno[1].replace(',', '')
                    print("Numerator=", Numerator)
                    print("Denominator=", Denominator)
                    last_url = driver.current_url
                    measures_all[measure_counter].click()
                    try:
                        sf.ajax_preloader_wait(driver)
                        if len(driver.find_elements_by_xpath("//td/a[contains(@href,'/registries/')]")) == 0 and float(
                                Denominator) != 0 and float(Numerator) != 0:
                            met_name = driver.find_element_by_xpath(
                                "//div[@class='ch metric_specific_patient_list_title']").text
                            logger.info("Metric name: %s", measure_name)
                            logger.warning("Providers list is blank. Please check manually.")
                            ws.append((quarter_name + " | " + lob_name, measure_name, 'Blank Providers List'))

                        elif len(
                                driver.find_elements_by_xpath("//td/a[contains(@href,'/registries/')]")) == 0 and float(
                                Denominator) == 0:
                            met_name = driver.find_element_by_xpath(
                                "//div[@class='ch metric_specific_patient_list_title']").text
                            logger.info("Metric name: %s", measure_name)
                            logger.info("Providers list is blank since measure score is zero.")

                        elif len(driver.find_elements_by_xpath("//td/a[contains(@href,'/registries/')]")) == 0:
                            met_name = driver.find_element_by_xpath(
                                "//div[@class='ch metric_specific_patient_list_title']").text
                            logger.info("Metric name: %s", measure_name)
                            logger.info("Providers list is blank. Please check manually.")

                        else:
                            if len(driver.find_elements_by_xpath("(//td/a[contains(@href,'/registries/')])[2]")) != 0:
                                patientlist_link = driver.find_element_by_xpath(
                                    "(//td/a[contains(@href,'/registries/')])[2]")

                            else:
                                patientlist_link = driver.find_element_by_xpath(
                                    "(//td/a[contains(@href,'/registries/')])[1]")

                            ActionChains(driver).move_to_element(patientlist_link).perform()
                            ActionChains(driver).key_down(Keys.CONTROL).click(patientlist_link).key_up(
                                Keys.CONTROL).perform()

                            # """ *********** Termed column check **************** """
                            # ajax_preloader_wait()
                            # if len(driver.find_elements_by_xpath("//th[@aria-label='Termed: activate to sort column ascending']"))!=0:
                            #     print("Termed column exists!")
                            # else:
                            #     print("Please check")
                            # driver.find_element_by_xpath("//a[@class='datatable_filter_dropdown sidenav-trigger']").click()
                            # time.sleep(1)
                            #
                            # if len(driver.find_elements_by_xpath("//div[text()='Termed:']"))!=0:
                            #     print("Filter is present")
                            # else:
                            #     print("Check Filter")

                            # **** CALCULATE CARE GAP LIST ****
                            try:
                                driver.switch_to.window(driver.window_handles[1])
                                sf.ajax_preloader_wait(driver)

                                provider_name = driver.find_element_by_xpath("//a[@id='context_trigger']/div/span").text
                                logger.info("Provider Name: %s", provider_name)
                                metric_name = driver.find_element_by_xpath(
                                    "//div[@class='ch metric_specific_patient_list_title']").text
                                logger.info("Metric name: %s", metric_name)
                                my_lob_ce = driver.find_element_by_xpath(
                                    "//div[@class='metric_patient_list_filter left']").text
                                print(my_lob_ce)
                                x = my_lob_ce.split("\u2002·\u2002")
                                my_lob_ce_final = x[0] + " " + "|" + " " + x[1] + " " + "|" + " " + x[3]
                                logger.info("%s", my_lob_ce_final)

                                # Data for patient dashboard:
                                y = metric_name.split("|")

                                metric_name_4_patientdashboard1 = y[1].strip()
                                metric_name_4_patientdashboard = metric_name_4_patientdashboard1.replace('*', '')
                                print(metric_name_4_patientdashboard)

                                if len(driver.find_elements_by_xpath(
                                        "//td/div/a[contains(@href,'/patient_detail/')]")) == 0:
                                    print("Patient list is blank!")
                                    ws.append((quarter_name + " | " + lob_name, metric_name_4_patientdashboard,
                                                provider_name, 'No Non-compliant patient found'))
                                    measure_name4screenshot = ''.join(
                                        e for e in str(measure_name) if (e.isalnum() or e.isspace()))


                                # Pencil icon presence:
                                elif len(driver.find_elements_by_xpath(
                                        "//td/div/a[contains(@href,'/patient_detail/')]")) != 0:
                                    time.sleep(1)
                                    if len(driver.find_elements_by_xpath("//td[contains(@class,' pencil_icon')]")) != 0:
                                        driver.find_element_by_xpath("//td[contains(@class,' pencil_icon')]").click()
                                        time.sleep(0.5)
                                        pencil_options = driver.find_elements_by_xpath(
                                            "(//td[contains(@class,' pencil_icon')])[1]/div/ul[contains(@class,'dropdown-content patient-menu-list')]/li")
                                        # Available options in Pencil icon:
                                        add_supdata_flag_MSPL = 0
                                        map_flag_MSPL = 0
                                        option_counter = 0

                                        for option_counter in range(len(pencil_options)):
                                            print((pencil_options[option_counter]).text)
                                            pencil_options_text = (pencil_options[option_counter]).text
                                            if pencil_options_text.strip() == "Add Supplemental Data":
                                                add_supdata_flag_MSPL = 1
                                            elif pencil_options_text.strip() == "Mark as Pending":
                                                map_flag_MSPL = 1
                                            elif pencil_options_text.strip() == "Confirm/Disconfirm":
                                                add_supdata_flag_MSPL = "Confirm/Disconfirm"
                                            pencil_options = driver.find_elements_by_xpath(
                                                "(//td[contains(@class,' pencil_icon')])[1]/div/ul[contains(@class,'dropdown-content patient-menu-list')]/li")
                                            if (map_flag_MSPL == add_supdata_flag_MSPL == 1):
                                                patient_found = "Found"

                                    # Pencil icon is not present:
                                    elif len(driver.find_elements_by_xpath(
                                            "//td[contains(@class,' pencil_icon')]")) == 0:
                                        print("No Pencil in MSPL")
                                        add_supdata_flag_MSPL = 0
                                        map_flag_MSPL = 0
                                    print("Supdata flag(MSPL): " + str(add_supdata_flag_MSPL))
                                    print("Map flag(MSPL): " + str(map_flag_MSPL))

                                    # CareGap in MSPL:
                                    if len(driver.find_elements_by_xpath("//td[contains(@class,'care_ops')]")) != 0:
                                        caregap_MSPL = driver.find_element_by_xpath(
                                            "(//td[contains(@class,'care_ops')])[1]").text
                                        print("CareGap in MSPL:" + caregap_MSPL)
                                    elif len(driver.find_elements_by_xpath("//td[contains(@class,' care_ops')]")) == 0:
                                        caregap_MSPL = "Not present"
                                        print("MSPL: CareGap is Not present")

                                    # call PATIENT DASHBOARD:
                                    mspl_url = driver.current_url
                                    driver.find_element_by_xpath(
                                        "//td/div/a[contains(@href,'/patient_detail/')]").click()
                                    try:
                                        driver.switch_to.window(driver.window_handles[2])

                                        if (map_flag_MSPL == 1 and add_supdata_flag_MSPL == 1):
                                            patient_verified = PatientDashboard(driver, ws, quarter_name, lob_name,
                                                                                metric_name_4_patientdashboard,
                                                                                add_supdata_flag_MSPL,
                                                                                map_flag_MSPL, caregap_MSPL, mspl_url,
                                                                                provider_name)



                                    except Exception as e:
                                        print(e)
                                    finally:

                                        driver.close()
                                        driver.switch_to.window(driver.window_handles[1])





                            # Exception in MSPL block
                            except Exception as e:
                                print(e)
                                logger.critical(
                                    measure_name + '\n' + provider_name + '\n' + "Metric specific patients list is not opening!Exception occurred!!")
                                ws.append((quarter_name + " | " + lob_name, measure_name, provider_name, 'Error'))

                            finally:
                                driver.close()
                                driver.switch_to.window(driver.window_handles[0])
                                apply_conditional_formatting(ws)


                        WebDriverWait(driver, 30).until(
                            EC.presence_of_element_located((By.XPATH, "//a[@class='breadcrumb']")))
                        driver.find_element_by_xpath("//a[@class='breadcrumb']").click()

                    # Providers list open exception block
                    except Exception as e:

                        print(e)
                        driver.get(last_url)
                    finally:
                        # MEASURE COUNTER

                        measures_all = driver.find_elements_by_xpath("//div/span[@class='met-name']")
                        scores = driver.find_elements_by_xpath("//span[@class='num-den']")
                        measure_counter += 1
                        score += 1

                lobs = driver.find_elements_by_xpath("//ul[@id='filter-lob']/li[@class!='hide']")
                quarters = driver.find_elements_by_xpath("//ul[@id='filter-quarter']/li")

        rows = ws.max_row
        cols = ws.max_column
        for i in range(1, rows + 1):
            for j in range(1, cols + 1):
                if ws.cell(i, j).value == 'PASS':
                    ws.cell(i, j).fill = PatternFill('solid', fgColor='0FC404')
                elif ws.cell(i, j).value == 'FAIL':
                    ws.cell(i, j).fill = PatternFill('solid', fgColor='FC0E03')
                elif ws.cell(i, j).value == 'Data table is empty':
                    ws.cell(i, j).fill = PatternFill('solid', fgColor='FCC0BB')

    # store Cozeva ID
    # search the metric
    # Check options in Pencil icon
    # Click on Mark As Pending
    # Page will refresh ;Check for stale icon
    # Click on Add Supplemental Data
    # Check Task Id
    # Check reflection on Pending List  compare patient and metric name
    # Refresh twice or thrice
    # Come to patient dashboard
    # Unmark As pending

    # initialize Workbook

    verify_mark_as_pending(driver, workbook, logger, "Cozeva Support")


def market_sheet(driver, workbook, logger, run_from):
    workbook.create_sheet('Market Sheet')
    ws = workbook['Market Sheet']

    ws.append(['ID', 'Context', 'Scenario', 'Status', 'Comments'])
    header_font = Font(color='FFFFFF', bold=False, size=12)
    header_cell_color = PatternFill('solid', fgColor='030303')
    ws['A1'].font = header_font
    ws['A1'].fill = header_cell_color
    ws['B1'].font = header_font
    ws['B1'].fill = header_cell_color
    ws['C1'].font = header_font
    ws['C1'].fill = header_cell_color
    ws['D1'].font = header_font
    ws['D1'].fill = header_cell_color
    ws['E1'].font = header_font
    ws['E1'].fill = header_cell_color
    ws.name = "Arial"
    test_case_id = 1

    registry_url = driver.current_url
    sf.ajax_preloader_wait(driver)
    WebDriverWait(driver, 60).until(EC.presence_of_element_located((By.XPATH, "//*[@id='conti_enroll']")))
    CE_checkbox = driver.find_element_by_xpath("//*[@id='conti_enroll']")
    # print(CE_checkbox)

    if (CE_checkbox.is_selected()):
        CEstatus = 'ON'

    else:
        CEstatus = 'OFF'

    if CEstatus == config.get("market-sheet", "ce_status"):
        ws.append((test_case_id, 'Market Sheet Sync', 'CE Status', 'Passed', 'CE status in market sheet is '+config.get("market-sheet", "ce_status")+' and default CE for client is '+CEstatus))
    else:
        ws.append((test_case_id, 'Market Sheet Sync', 'CE Status', 'Failed',
                   'CE status in market sheet is ' + config.get("market-sheet",
                                                                "ce_status") + ' and default CE for client is ' + CEstatus, driver.current_url))
    test_case_id += 1
    displayname = driver.find_element_by_xpath(locator.xpath_context_Name).text
    if displayname == config.get("market-sheet", "display_name"):
        ws.append((test_case_id, 'Market Sheet Sync', 'Display Name', 'Passed',
                   'Display Name in market sheet is ' + config.get("market-sheet",
                                                                "display_name") + ' and display name on PROD is ' + displayname))
    else:
        ws.append((test_case_id, 'Market Sheet Sync', 'Display Name', 'Failed',
                   'Display name in market sheet is ' + config.get("market-sheet",
                                                                "display_name") + ' and display name on PROD is ' + displayname, driver.current_url))
    test_case_id += 1
    try:
        driver.find_element_by_id("qt-filter-label").click()
        time.sleep(1)
        def_lob = driver.find_element_by_id("filter-lob").find_element_by_class_name('highlight').text
    except:
        ws.append((test_case_id, 'Market Sheet Sync', 'Default Lob', 'Failed', 'Unable to click on LoB dropdown', driver.current_url))

    if def_lob == config.get("market-sheet", "def_lob"):
        ws.append((test_case_id, 'Market Sheet Sync', 'Default Lob', 'Passed',
                   'Default LoB in market sheet is ' + config.get("market-sheet",
                                                                   "def_lob") + ' and default Lob on PROD is ' + def_lob))
    else:
        ws.append((test_case_id, 'Market Sheet Sync', 'Display Name', 'Failed',
                   'Default Lob in market sheet is ' + config.get("market-sheet",
                                                                   "def_lob") + ' and default Lob on PROD is' + def_lob, driver.current_url))

    time.sleep(1)
    driver.get(registry_url)
    sf.ajax_preloader_wait(driver)
    rows = ws.max_row
    cols = ws.max_column
    for i in range(2, rows + 1):
        for j in range(3, cols + 1):
            if ws.cell(i, j).value == 'Passed':
                ws.cell(i, j).fill = PatternFill('solid', fgColor='0FC404')
            elif ws.cell(i, j).value == 'Failed':
                ws.cell(i, j).fill = PatternFill('solid', fgColor='FC0E03')
            elif ws.cell(i, j).value == 'Data table is empty':
                ws.cell(i, j).fill = PatternFill('solid', fgColor='FCC0BB')


def cetoggle(driver, workbook, logger, screenshot_path, run_from):
    registry_url = driver.current_url
    sf.ajax_preloader_wait(driver)
    WebDriverWait(driver, 60).until(EC.presence_of_element_located((By.XPATH, "//*[@id='conti_enroll']")))
    CE_checkbox = driver.find_element_by_xpath("//*[@id='conti_enroll']")
    # print(CE_checkbox)

    if CE_checkbox.is_selected():
        CEstatus = 'ON'
    else:
        CEstatus = 'OFF'

    sf.captureScreenshot(driver, "CE "+CEstatus, screenshot_path)
    time.sleep(1)
    driver.find_element_by_xpath("//*[@id='metric_scorecard']/div/div[1]/div/div/div/div[2]/label").click()
    time.sleep(4)

    CE_checkbox = driver.find_element_by_xpath("//*[@id='conti_enroll']")
    # print(CE_checkbox)

    if CE_checkbox.is_selected():
        CEstatus = 'ON'

    else:
        CEstatus = 'OFF'

    sf.captureScreenshot(driver, "CE " + CEstatus, screenshot_path)
    time.sleep(1)


def hccvalidation(driver, workbook, logger, screenshot_path, run_from):
    workbook.create_sheet('HCC Validation')
    ws = workbook['HCC Validation']
    #Push to HCc validation

    ws.append(
        ['LOB', 'HCC Measure Name', 'Domain Name Check', 'Performance Statistics Check', 'Network Comparison Check',
         'Risk Score Check', 'Clinical Score Check', 'Suspect Score Check', 'Comments', 'Provider Row URL'])
    header = NamedStyle(name="header")
    header.font = Font(bold=True)
    header.border = Border(bottom=Side(border_style="thin"))
    header.alignment = Alignment(horizontal="center", vertical="center")
    header_row = ws[1]
    for cell in header_row:
        cell.style = header
    red_background = PatternFill(patternType='solid', fgColor="00FF0000")
    green_background = PatternFill(patternType='solid', fgColor="50C878")
    gray_background = PatternFill(patternType='solid', fgColor="5F9EA0")
    # diff_style1 = DifferentialStyle(fill=red_background)
    # rule1 = Rule(type="text", dxf=diff_style1, text="Fail")
    # diff_style2 = DifferentialStyle(fill=green_background)
    # rule2 = Rule(type="text", dxf=diff_style2, text="Pass")
    # diff_style3 = DifferentialStyle(fill=gray_background)
    # rule3 = Rule(type="text", dxf=diff_style3, text="Unexecuted")

    ws.name = "Arial"
    test_case_id = 1

    def addition(total, row, switch, arr):
        try:
            for i in range(len(arr[1])):
                if switch == 1:
                    if str(arr[1][i]).endswith("Gaps"):
                        total = total + int(arr[row][i])
                if switch == 2:
                    if str(arr[1][i]).endswith("Conditions"):
                        total = total + int(arr[row][i])
                if switch == 3:
                    if str(arr[1][i]).endswith("Disconfirms"):
                        total = total + int(arr[row][i])
                if switch == 4:
                    if str(arr[1][i]).endswith("Clinical RAF"):
                        total = total + float(arr[row][i])
                if switch == 5:
                    if str(arr[1][i]).endswith("Potential RAF"):
                        total = total + float(arr[row][i])
                if switch == 6:
                    if str(arr[1][i]).endswith("Coded RAF"):
                        total = total + float(arr[row][i])

            return total
        except ValueError:
            print("No number detected in " + arr[1][i] + " .For row no = " + str(row + 1))
            return total

    def csvAddition(filepath):
        import csv
        with open(filepath, newline='') as csvfile:
            rows = csv.reader(csvfile, delimiter=',')
            rows = list(rows)
        Gaps = 0
        Conditions = 0
        Disconfirms = 0
        Clinical = 0
        Potential = 0
        Coded = 0
        for ind in range(2, len(rows)):
            Gaps = addition(Gaps, ind, 1, rows)
            Conditions = addition(Conditions, ind, 2, rows)
            Disconfirms = addition(Disconfirms, ind, 3, rows)
            Clinical = addition(Clinical, ind, 4, rows)
            Potential = addition(Potential, ind, 5, rows)
            Coded = addition(Coded, ind, 6, rows)
        return Gaps, Conditions, Disconfirms, Clinical, Potential, Coded, (len(rows) - 2)


    sf.ajax_preloader_wait(driver)
    WebDriverWait(driver, 120).until(EC.presence_of_element_located((By.XPATH, "//*[@id='qt-filter-label']")))
    print("Selected page= " + driver.title)
    Registry_URL = driver.current_url
    sf.ajax_preloader_wait(driver)
    driver.find_element(By.XPATH, "//*[@id='qt-filter-label']").click()
    Quarter_list = driver.find_element(By.XPATH, "//*[@id='filter-quarter']").find_elements(By.TAG_NAME, "li")
    for quarter in Quarter_list:
        if quarter.text == "2023":
            quarter.click()
            break
    #Quarter_list[0].click()
    LOB_list = driver.find_element(By.XPATH, "//*[@id='filter-lob']").find_elements(By.TAG_NAME, 'li')
    for i in range(0, len(LOB_list)-1):
        LOB_Name = LOB_list[i].text
        print("LOB Name for 2022: " + LOB_Name)
        try:
            LOB_list[i].click()
        except ElementNotInteractableException as e:
            continue
        #LOB_list[i].click()
        driver.find_element(By.ID, "reg-filter-apply").click()
        LOB_Specific_URL = driver.current_url
        sf.ajax_preloader_wait(driver)
        if driver.find_element(By.XPATH, "//*[@id='conti_enroll']").is_selected():
            driver.find_element(By.XPATH, "//*[@class='cont_disc_toggle']").click()
        print("LOB URL: " + LOB_Specific_URL)

        try:
            summaryList = driver.find_element_by_class_name("registry_header_panel").find_elements_by_tag_name("div")
            for div, next_div in zip(summaryList, summaryList[1:] + [summaryList[0]]):
                if div.text == "HCC":
                    hcc_summary = next_div.text.strip()

            if hcc_summary == "0(0)/0(0)" or hcc_summary == "0.000(0.000)/0.000":
                ws.append(
                    [currentLOBName, "HCC Summary Score", hcc_summary, "Failed"])


        except Exception as e:
            traceback.print_exc()
            ws_counts.append([currentLOBName, "x", "x", "Couldn't fetch patient counts/rating"])
        HCC_measure_checklist = [33, 551, 553, 554, 555, 556]
        for i in HCC_measure_checklist:
            flag = 0
            try:
                Measure_Specific_url = driver.find_element(By.XPATH, "//*[@id=" + str(i) + "]//a").get_attribute('href')
                Measure = driver.find_element(By.XPATH, "//*[@id=" + str(i) + "]//*[@class='met-name']").text
                Domain_name_UI = driver.find_element(By.XPATH, "//*[@id=" + str(i) + "]/div/div").text
                print("Measure Name: " + Measure)
                print("Measure URL: " + Measure_Specific_url)
                print("Domain Name registry page: " + Domain_name_UI)
                driver.get(Measure_Specific_url)
                sf.ajax_preloader_wait(driver)
                #driver.implicitly_wait(3)
                Domain_name_MSPL = driver.find_element(By.XPATH, "//*[@class='ch metric_specific_patient_list_title']").text
                print("Domain Name MSPL page: " + Domain_name_MSPL)
                if Domain_name_UI in Domain_name_MSPL:
                    Domain_comment = "Passed"
                else:
                    Domain_comment = "Failed"
                ListRow = driver.find_element(By.XPATH, "//*[@id='metric-support-prov-ls']").find_element(By.TAG_NAME, "tbody").find_elements(By.TAG_NAME, 'tr')
                if "No data available" in ListRow[0].text and len(ListRow) == 1:
                    Comments = "No provider data in HCC measure " + str(i)
                    print(Comments)
                    ws.append([cus_id, Measure, Domain_comment, 'Undetermined', 'Undetermined', 'Undetermined', 'Undetermined', 'Undetermined', Comments, URL])
                elif len(ListRow) == 1:
                    ListRow[0].find_elements(By.TAG_NAME, 'a')[1].click()
                    sf.ajax_preloader_wait(driver)
                    print(driver.current_url)
                else:
                    #k = random.randint(2, len(ListRow)-1)
                    #k = int(input("Enter the number of providers you want to check for the Measure "+Measure+" among the number of providers "+str(len(ListRow))+":\n"))
                    k = 2
                    Performance_num_UI = 0
                    Performance_denum_UI = 0
                    Performance_percentage_UI = 00.00
                    Performance_percentage_calculated = 00.00
                    while k != 0:
                        Row = ListRow[random.randint(0, len(ListRow)-1)]
                        #Row = ListRow[random.randint(0, len(ListRow)-k)]
                        Row.find_elements(By.TAG_NAME, 'a')[1].click()
                        sf.ajax_preloader_wait(driver)
                        DataToBeValidated = driver.find_element(By.XPATH, "//*[@class='tab']").find_elements(By.TAG_NAME,
                                                                                                             'span')
                        Provider_Specific_url = driver.current_url
                        print("Provider URL: " + Provider_Specific_url)
                        DataToBeValidated_num = DataToBeValidated[0].text
                        DataToBeValidated_num = DataToBeValidated_num.replace(',', '')
                        print("MSPL Numerator: " + DataToBeValidated_num)
                        DataToBeValidated_denum = DataToBeValidated[1].text
                        DataToBeValidated_denum = DataToBeValidated_denum.replace(',', '')
                        print("MSPL Denominator: " + DataToBeValidated_denum)

                        #Patient Dashboard checks
                        hcc_flag_list = ["Failed", "Failed", "Failed", "Failed"]
                        '''
                        0 = HCC count in care gap
                        1 = Total count of HCCs
                        2 = Clinical score present
                        3 = suspect score present
                        '''
                        try:
                            window_switched = 0
                            pat_dashboard_flag = "No scores on dashboard"
                            table = driver.find_element_by_id(
                                "quality_registry_list").find_element_by_tag_name(
                                "tbody").find_elements_by_tag_name('tr')
                            if len(table) == 1:
                                if "No data" in table[0].text:
                                    pat_dashboard_flag = "No patients for this provider"
                                else:
                                    table[0].find_element_by_class_name("pat_name").click()
                            else:
                                table[sf.RandomNumberGenerator(len(table), 1)[0]].find_element_by_class_name("pat_name").click()
                            time.sleep(1)
                            driver.switch_to.window(driver.window_handles[1])
                            sf.ajax_preloader_wait(driver)
                            WebDriverWait(driver, 120).until(
                                EC.presence_of_element_located((By.CLASS_NAME, "patient_header_wrapper")))
                            window_switched = 1
                            #switched to patient dashboard, Checking for HCCs and scores
                            '''
                            0 = HCC count in care gap
                            1 = Total count of HCCs
                            2 = Clinical score present
                            3 = suspect score present
                            '''
                            hcc_element = driver.find_element_by_id("table_4")
                            hcctable = hcc_element.find_element_by_tag_name('tbody').find_elements_by_class_name("row-group")
                            # print("Dashboard Hcc Count = "+str(len(hcctable)))
                            suspect_list = []
                            clinical_score_list = []
                            #Clinical score Check
                            for hcc_measures in hcctable:
                                clinical_text = hcc_measures.find_element_by_class_name("hcc_details").find_elements_by_tag_name("span")[1].text
                                try:
                                    clinical_score_list.append(float(clinical_text.replace('Clinical Factor', '').strip()))
                                except Exception as e:
                                    hcc_flag_list[2] = "Failed"
                                    break
                            if len(clinical_score_list) > 0:
                                hcc_flag_list[2] = "Passed"
                            #Clinical Score check
                            for hcc_measures in hcctable:
                                try:
                                    suspect_text = hcc_measures.find_element_by_xpath("//div[contains(text(),'Suspect')]").text
                                except Exception as e:
                                    print(e)
                                    hcc_flag_list[3] = "Failed"
                                    break
                                try:
                                    suspect_list.append(float(suspect_text.replace('Suspect', '').strip()))
                                except Exception as e:
                                    hcc_flag_list[3] = "Failed"
                                    break
                            if len(suspect_list) > 0:
                                hcc_flag_list[3] = "Passed"
                            driver.close()
                            driver.switch_to.window(driver.window_handles[0])
                            sf.ajax_preloader_wait(driver)
                            window_switched = 0

                        except Exception as e:
                            if window_switched == 1:
                                driver.close()
                                driver.switch_to.window(driver.window_handles[0])
                                sf.ajax_preloader_wait(driver)
                                window_switched = 0
                            print(e)
                            traceback.print_exc()
                        driver.find_element(By.XPATH,
                                            "//*[@data-target='datatable_bulk_filter_0_quality_registry_list']").click()
                        driver.find_element(By.XPATH, "//*[contains(text(),'Export all to CSV')]").click()
                        sf.ajax_preloader_wait(driver)
                        try:
                            driver.find_element(By.XPATH, "//*[@class='tabs']").find_elements(By.TAG_NAME, 'li')[1].click()
                            sf.ajax_preloader_wait(driver)
                            Performance_percentage_UI = driver.find_element(By.XPATH, "//*[@class='performance_value']").text
                            Performance_percentage_UI = Performance_percentage_UI.replace('%', '')
                            Performance_num_UI = driver.find_element(By.XPATH, "//*[@class='numerator']").text
                            Performance_num_UI = Performance_num_UI.replace('Numerator: ', '')
                            Performance_denum_UI = driver.find_element(By.XPATH, "//*[@class='denominator']").text
                            Performance_denum_UI = Performance_denum_UI.replace('Denominator: ', '')
                            Performance_percentage_calculated = round((float(Performance_num_UI)/float(Performance_denum_UI))*100, 4)
                            print("Performance Tab Numerator: " + Performance_num_UI)
                            print("Performance Tab Denominator: " + Performance_denum_UI)
                            print("Performance Tab Percentage: " + Performance_percentage_UI)
                            print("Performance calculated: " + str(Performance_percentage_calculated))
                            if (float(Performance_percentage_UI) - float(Performance_percentage_calculated))< 0.02:
                                Performance_comment = "Passed"
                            else:
                                Performance_comment = "Failed"
                        except ElementNotInteractableException:
                            Performance_comment = "The Performance tab is not clickable"
                        try:
                            driver.find_element(By.XPATH, "//*[@class='tabs']").find_elements(By.TAG_NAME, 'li')[2].click()
                            sf.ajax_preloader_wait(driver)
                            if EC.presence_of_element_located((By.XPATH, "//*[@id='network_comparison_chart']")):
                                Network_comment = "Passed"
                            else:
                                Network_comment = "Failed"
                        except ElementNotInteractableException:
                            Network_comment = "Failed"
                        onlyfiles = [f for f in listdir(locator.download_dir) if
                                     isfile(join(locator.download_dir, f))]
                        path = locator.download_dir + onlyfiles[0]
                        result = csvAddition(path)
                        print("Total Gap Count: " + str(result[0]))
                        print("Total Condition Count: " + str(result[1]))
                        print("Total Disconfirm Count: " + str(result[2]))
                        print("Total Clinical RAF Score: " + str(result[3]))
                        print("Total Potential Score: " + str(result[4]))
                        print("Total Coded Score: " + str(result[5]))
                        print("Total Patient Count: " + str(result[6]))
                        os.remove(path)
                        if i == 553 or i == 556:
                            DataToBeValidated_num = float(DataToBeValidated_denum) - float(DataToBeValidated_num)
                            DataToBeValidated_num = round(DataToBeValidated_num, 3)
                            DataToBeValidated_denum = round(DataToBeValidated_denum, 3)
                            num = float(result[3] / result[6])
                            num = round(num, 3)
                            temp = float((result[5]-result[3])/result[6])
                            denum = float(result[4] / result[6])
                            denum = denum - temp
                            denum = round(denum, 3)
                            if abs(float(DataToBeValidated_num) - num) < 0.015 and abs(float(DataToBeValidated_denum) - denum) < 0.015:
                                ws.append([LOB_Name, Measure, Domain_comment, Performance_comment, Network_comment, 'Passed',hcc_flag_list[2], hcc_flag_list[3], '-', Provider_Specific_url])
                            else:
                                ws.append([LOB_Name, Measure, Domain_comment, Performance_comment, Network_comment, 'Failed',hcc_flag_list[2], hcc_flag_list[3], '-', Provider_Specific_url])
                        elif i == 554 or i == 555:
                            if int(DataToBeValidated_num) == int(DataToBeValidated_denum)-int(result[1])-int(result[2]):
                                ws.append([LOB_Name, Measure, Domain_comment, Performance_comment, Network_comment, 'Passed',hcc_flag_list[2], hcc_flag_list[3], '-', Provider_Specific_url])
                            else:
                                ws.append([LOB_Name, Measure, Domain_comment, Performance_comment, Network_comment, 'Failed',hcc_flag_list[2], hcc_flag_list[3], '-', Provider_Specific_url])
                        else:
                            if int(DataToBeValidated_num) == int(result[0]) and int(DataToBeValidated_denum) == int(
                                    result[0] + result[1] + result[2]):
                                ws.append([LOB_Name, Measure, Domain_comment, Performance_comment, Network_comment, 'Passed',hcc_flag_list[2], hcc_flag_list[3], '-', Provider_Specific_url])
                            else:
                                ws.append([LOB_Name, Measure, Domain_comment, Performance_comment, Network_comment, 'Failed',hcc_flag_list[2], hcc_flag_list[3], '-', Provider_Specific_url])
                        k -= 1
                        driver.get(Measure_Specific_url)
                        sf.ajax_preloader_wait(driver)
                        ListRow = driver.find_element(By.XPATH, "//*[@id='metric-support-prov-ls']").find_element(
                        By.TAG_NAME,"tbody").find_elements(By.TAG_NAME, 'tr')
            except NoSuchElementException:
                if i == 33:
                    Comments = "For LOB "+ LOB_Name +",the Risk Measure- One Year Recapture Rate/Review of Chronic Conditions(ID:"+str(i)+") is not present"
                    Measure = "One Year Recapture Rate/Review of Chronic Conditions"
                elif i == 551:
                    Comments = "For LOB " + LOB_Name + ",the Risk Measure- Review of Suspect Conditions(ID:" + str(
                        i) + ") is not present"
                    Measure = "Review of Suspect Conditions"
                elif i == 553:
                    Comments = "For LOB " + LOB_Name + ",the Risk Measure- HCC Score(ID:" + str(
                        i) + ") is not present"
                    Measure = "HCC Score"
                elif i == 554:
                    Comments = "For LOB " + LOB_Name + ",the Risk Measure- Review of ACA Chronic Conditions(ID:" + str(
                        i) + ") is not present"
                    Measure= "Review of ACA Chronic Conditions"
                elif i == 555:
                    Comments = "For LOB " + LOB_Name + ",the Risk Measure- Review of ACA Suspect Conditions(ID:" + str(
                        i) + ") is not present"
                    Measure = "Review of ACA Suspect Conditions"
                else:
                    Comments = "For LOB " + LOB_Name + ",the Risk Measure- ACA HCC Score(ID:" + str(
                        i) + ") is not present"
                    Measure = "ACA HCC Score"
                ws.append([LOB_Name, Measure, '-', '-', '-', 'Not Present','-','-', Comments,'-'])
                # cellname = "N" + str(SheetRowName)
                # ws[''+cellname+''].fill = gray_background
                # # ws.conditional_formatting.add("J1:O100", rule3)
                print(Comments)
                workbook.save(screenshot_path + "\\report.xlsx")
                continue
                if flag == 5:
                    print("No more HCC measures switching to next LOB")
            except Exception as e:
                traceback.print_exc()
                print(e)
            #finally:
            #flag += 1
            driver.get(LOB_Specific_URL)
            sf.ajax_preloader_wait(driver)
            if driver.find_element(By.XPATH, "//*[@id='conti_enroll']").is_selected():
                driver.find_element(By.XPATH, "//*[@class='cont_disc_toggle']").click()
            #time.sleep(3)
            if driver.find_element(By.XPATH, "//*[@id='conti_enroll']").is_selected():
                driver.find_element(By.XPATH, "//*[@class='cont_disc_toggle']").click()
            workbook.save(screenshot_path + "\\Report.xlsx")
        sf.ajax_preloader_wait(driver)
        driver.find_element(By.XPATH, "//*[@id='qt-filter-label']").click()
        LOB_list = driver.find_element(By.XPATH, "//*[@id='filter-lob']").find_elements(By.TAG_NAME, 'li')

    rows = ws.max_row
    cols = ws.max_column
    for i in range(2, rows + 1):
        for j in range(3, cols + 1):
            if ws.cell(i, j).value == 'Passed':
                ws.cell(i, j).fill = PatternFill('solid', fgColor='0FC404')
            elif ws.cell(i, j).value == 'Failed':
                ws.cell(i, j).fill = PatternFill('solid', fgColor='FC0E03')
            elif ws.cell(i, j).value == 'Not Present':
                ws.cell(i, j).fill = PatternFill('solid', fgColor='808080')
            elif ws.cell(i, j).value == 'Data table is empty':
                ws.cell(i, j).fill = PatternFill('solid', fgColor='FCC0BB')


def hccvalidation_multi(driver, cus_id, year, workbook, provider_count, screenshot_path, run_from, workbook_title):
    print("Customer "+str(cus_id))
    workbook.create_sheet(str(cus_id))
    ws = workbook[str(cus_id)]

    #Push to HCc validation

    ws.append(
        ['Customer ID', 'LOB', 'HCC Measure Name', 'Domain Name Check', 'Performance Statistics Check', 'Network Comparison Check',
         'Risk Score Check','Clinical Score Check', 'Suspect score check', 'Comments','Provider Row URL'])
    # header = NamedStyle(name="header")
    # header.font = Font(bold=True)
    # header.border = Border(bottom=Side(border_style="thin"))
    # header.alignment = Alignment(horizontal="center", vertical="center")
    # header_row = ws[1]
    # for cell in header_row:
    #     cell.style = header
    red_background = PatternFill(patternType='solid', fgColor="00FF0000")
    green_background = PatternFill(patternType='solid', fgColor="50C878")
    gray_background = PatternFill(patternType='solid', fgColor="5F9EA0")
    # diff_style1 = DifferentialStyle(fill=red_background)
    # rule1 = Rule(type="text", dxf=diff_style1, text="Fail")
    # diff_style2 = DifferentialStyle(fill=green_background)
    # rule2 = Rule(type="text", dxf=diff_style2, text="Pass")
    # diff_style3 = DifferentialStyle(fill=gray_background)
    # rule3 = Rule(type="text", dxf=diff_style3, text="Unexecuted")

    ws.name = "Arial"
    test_case_id = 1

    def addition(total, row, switch, arr):
        try:
            for i in range(len(arr[1])):
                if switch == 1:
                    if str(arr[1][i]).endswith("Gaps"):
                        total = total + int(arr[row][i])
                if switch == 2:
                    if str(arr[1][i]).endswith("Conditions"):
                        total = total + int(arr[row][i])
                if switch == 3:
                    if str(arr[1][i]).endswith("Disconfirms"):
                        total = total + int(arr[row][i])
                if switch == 4:
                    if str(arr[1][i]).endswith("Clinical RAF"):
                        total = total + float(arr[row][i])
                if switch == 5:
                    if str(arr[1][i]).endswith("Potential RAF"):
                        total = total + float(arr[row][i])
                if switch == 6:
                    if str(arr[1][i]).endswith("Coded RAF"):
                        total = total + float(arr[row][i])

            return total
        except ValueError:
            print("No number detected in " + arr[1][i] + " .For row no = " + str(row + 1))
            return total

    def csvAddition(filepath):
        import csv
        with open(filepath, newline='') as csvfile:
            rows = csv.reader(csvfile, delimiter=',')
            rows = list(rows)
        Gaps = 0
        Conditions = 0
        Disconfirms = 0
        Clinical = 0
        Potential = 0
        Coded = 0
        for ind in range(2, len(rows)):
            Gaps = addition(Gaps, ind, 1, rows)
            Conditions = addition(Conditions, ind, 2, rows)
            Disconfirms = addition(Disconfirms, ind, 3, rows)
            Clinical = addition(Clinical, ind, 4, rows)
            Potential = addition(Potential, ind, 5, rows)
            Coded = addition(Coded, ind, 6, rows)
        return Gaps, Conditions, Disconfirms, Clinical, Potential, Coded, (len(rows) - 2)


    sf.ajax_preloader_wait(driver)
    WebDriverWait(driver, 120).until(EC.presence_of_element_located((By.XPATH, "//*[@id='qt-filter-label']")))
    print("Selected page= " + driver.title)
    Registry_URL = driver.current_url
    sf.ajax_preloader_wait(driver)
    driver.find_element(By.XPATH, "//*[@id='qt-filter-label']").click()
    Quarter_list = driver.find_element(By.XPATH, "//*[@id='filter-quarter']").find_elements(By.TAG_NAME, "li")
    for quarter in Quarter_list:
        if quarter.text == year:
            quarter.click()
            break
    #Quarter_list[0].click()
    LOB_list = driver.find_element(By.XPATH, "//*[@id='filter-lob']").find_elements(By.TAG_NAME, 'li')
    for i in range(0, len(LOB_list)):
        LOB_Name = LOB_list[i].text
        print("LOB Name for 2022: " + LOB_Name)
        try:
            LOB_list[i].click()
        except ElementNotInteractableException as e:
            continue
        #LOB_list[i].click()
        driver.find_element(By.ID, "reg-filter-apply").click()
        LOB_Specific_URL = driver.current_url
        sf.ajax_preloader_wait(driver)
        if driver.find_element(By.XPATH, "//*[@id='conti_enroll']").is_selected():
            driver.find_element(By.XPATH, "//*[@class='cont_disc_toggle']").click()
        print("LOB URL: " + LOB_Specific_URL)
        HCC_measure_checklist = [33, 551, 553, 554, 555, 556]
        for i in HCC_measure_checklist:
            flag = 0
            try:
                Measure_Specific_url = driver.find_element(By.XPATH, "//*[@id=" + str(i) + "]//a").get_attribute('href')
                Measure = driver.find_element(By.XPATH, "//*[@id=" + str(i) + "]//*[@class='met-name']").text
                Domain_name_UI = driver.find_element(By.XPATH, "//*[@id=" + str(i) + "]/div/div").text
                print("Measure Name: " + Measure)
                print("Measure URL: " + Measure_Specific_url)
                print("Domain Name registry page: " + Domain_name_UI)
                driver.get(Measure_Specific_url)
                sf.ajax_preloader_wait(driver)
                #driver.implicitly_wait(3)
                Domain_name_MSPL = driver.find_element(By.XPATH, "//*[@class='ch metric_specific_patient_list_title']").text
                print("Domain Name MSPL page: " + Domain_name_MSPL)
                if Domain_name_UI in Domain_name_MSPL:
                    Domain_comment = "Passed"
                else:
                    Domain_comment = "Failed"
                ListRow = driver.find_element(By.XPATH, "//*[@id='metric-support-prov-ls']").find_element(By.TAG_NAME, "tbody").find_elements(By.TAG_NAME, 'tr')
                if "No data available" in ListRow[0].text and len(ListRow) == 1:
                    Comments = "No provider data in HCC measure " + str(i)
                    print(Comments)
                    URL = driver.current_url
                    ws.append([cus_id, Measure, Domain_comment, 'Undetermined', 'Undetermined', 'Undetermined', 'Undetermined', 'Undetermined', Comments, URL])
                elif len(ListRow) == 1:
                    ListRow[0].find_elements(By.TAG_NAME, 'a')[1].click()
                    sf.ajax_preloader_wait(driver)
                    print(driver.current_url)
                else:
                    #k = random.randint(2, len(ListRow)-1)
                    #k = int(input("Enter the number of providers you want to check for the Measure "+Measure+" among the number of providers "+str(len(ListRow))+":\n"))
                    k = provider_count
                    Performance_num_UI = 0
                    Performance_denum_UI = 0
                    Performance_percentage_UI = 00.00
                    Performance_percentage_calculated = 00.00
                    while k != 0:
                        Row = ListRow[random.randint(0, len(ListRow)-1)]
                        #Row = ListRow[random.randint(0, len(ListRow)-k)]
                        Row.find_elements(By.TAG_NAME, 'a')[1].click()
                        sf.ajax_preloader_wait(driver)
                        DataToBeValidated = driver.find_element(By.XPATH, "//*[@class='tab']").find_elements(By.TAG_NAME,
                                                                                                             'span')
                        Provider_Specific_url = driver.current_url
                        print("Provider URL: " + Provider_Specific_url)
                        DataToBeValidated_num = DataToBeValidated[0].text
                        DataToBeValidated_num = DataToBeValidated_num.replace(',', '')
                        print("MSPL Numerator: " + DataToBeValidated_num)
                        DataToBeValidated_denum = DataToBeValidated[1].text
                        DataToBeValidated_denum = DataToBeValidated_denum.replace(',', '')
                        print("MSPL Denominator: " + DataToBeValidated_denum)

                        #Patient Dashboard checks
                        hcc_flag_list = ["Failed", "Failed", "Failed", "Failed"]
                        '''
                        0 = HCC count in care gap
                        1 = Total count of HCCs
                        2 = Clinical score present
                        3 = suspect score present
                        '''
                        try:
                            window_switched = 0
                            pat_dashboard_flag = "No scores on dashboard"
                            table = driver.find_element_by_id(
                                "quality_registry_list").find_element_by_tag_name(
                                "tbody").find_elements_by_tag_name('tr')
                            if len(table) == 1:
                                if "No data" in table[0].text:
                                    pat_dashboard_flag = "No patients for this provider"
                                else:
                                    table[0].find_element_by_class_name("pat_name").click()
                            else:
                                table[sf.RandomNumberGenerator(len(table), 1)[0]].find_element_by_class_name("pat_name").click()
                            time.sleep(1)
                            driver.switch_to.window(driver.window_handles[1])
                            sf.ajax_preloader_wait(driver)
                            WebDriverWait(driver, 120).until(
                                EC.presence_of_element_located((By.CLASS_NAME, "patient_header_wrapper")))
                            window_switched = 1
                            print("Patient Dashboard URL: "+driver.current_url)
                            #switched to patient dashboard, Checking for HCCs and scores
                            '''
                            0 = HCC count in care gap
                            1 = Total count of HCCs
                            2 = Clinical score present
                            3 = suspect score present
                            '''
                            hcc_element = driver.find_element_by_id("table_4")
                            hcctable = hcc_element.find_element_by_tag_name('tbody').find_elements_by_class_name("row-group")
                            # print("Dashboard Hcc Count = "+str(len(hcctable)))
                            suspect_list = []
                            clinical_score_list = []
                            tag_count_list = [0,0,0]
                            '''
                            tag_count_list
                            [0]= Recaptures
                            [1]= Suspects
                            [2]= New
                            '''
                            try:
                                tag_count_list[0] = len(hcc_element.find_elements(By.XPATH, "//span[@class='tag-block magenta-color tiny-title']"))
                                tag_count_list[1] = len(
                                    hcc_element.find_elements(By.XPATH, "//span[@class='tag-block yellow-ochre-color tiny-title']"))
                                tag_count_list[2] = len(hcc_element.find_elements(By.XPATH, "//span[@class='tag-block blue-color tiny-title']"))
                            except Exception as e:
                                traceback.print_exc()
                                print(e)
                                tag_count_list = ["NaN", "NaN", "NaN"]


                            #Clinical score Check
                            for hcc_measures in hcctable:
                                clinical_text = hcc_measures.find_element_by_class_name("hcc_details").find_elements_by_tag_name("span")[1].text
                                try:
                                    clinical_score_list.append(float(clinical_text.replace('Clinical Factor', '').strip()))
                                except Exception as e:
                                    hcc_flag_list[2] = "Failed"
                                    break
                            if len(clinical_score_list) > 0:
                                hcc_flag_list[2] = "Passed"
                            #Clinical Score check
                            for hcc_measures in hcctable:
                                try:
                                    suspect_text = hcc_measures.find_element_by_xpath("//div[contains(text(),'Suspect')]").text
                                except Exception as e:
                                    print(e)
                                    hcc_flag_list[3] = "Failed"
                                    break
                                try:
                                    suspect_list.append(float(suspect_text.replace('Suspect', '').strip()))
                                except Exception as e:
                                    hcc_flag_list[3] = "Failed"
                                    break
                            if len(suspect_list) > 0:
                                hcc_flag_list[3] = "Passed"
                            driver.close()
                            driver.switch_to.window(driver.window_handles[0])
                            sf.ajax_preloader_wait(driver)
                            window_switched = 0

                        except Exception as e:
                            if window_switched == 1:
                                driver.close()
                                driver.switch_to.window(driver.window_handles[0])
                                sf.ajax_preloader_wait(driver)
                                window_switched = 0
                            print(e)
                            traceback.print_exc()
                        driver.find_element(By.XPATH,
                                            "//*[@data-target='datatable_bulk_filter_0_quality_registry_list']").click()
                        driver.find_element(By.XPATH, "//*[contains(text(),'Export all to CSV')]").click()
                        sf.ajax_preloader_wait(driver)
                        try:
                            driver.find_element(By.XPATH, "//*[@class='tabs']").find_elements(By.TAG_NAME, 'li')[1].click()
                            sf.ajax_preloader_wait(driver)
                            Performance_percentage_UI = driver.find_element(By.XPATH, "//*[@class='performance_value']").text
                            Performance_percentage_UI = Performance_percentage_UI.replace('%', '')
                            Performance_num_UI = driver.find_element(By.XPATH, "//*[@class='numerator']").text
                            Performance_num_UI = Performance_num_UI.replace('Numerator: ', '')
                            Performance_denum_UI = driver.find_element(By.XPATH, "//*[@class='denominator']").text
                            Performance_denum_UI = Performance_denum_UI.replace('Denominator: ', '')
                            Performance_percentage_calculated = round((float(Performance_num_UI)/float(Performance_denum_UI))*100, 4)
                            print("Performance Tab Numerator: " + Performance_num_UI)
                            print("Performance Tab Denominator: " + Performance_denum_UI)
                            print("Performance Tab Percentage: " + Performance_percentage_UI)
                            print("Performance calculated: " + str(Performance_percentage_calculated))
                            if (float(Performance_percentage_UI) - float(Performance_percentage_calculated))< 0.02:
                                Performance_comment = "Passed"
                            else:
                                Performance_comment = "Failed"
                        except ElementNotInteractableException:
                            Performance_comment = "The Performance tab is not clickable"
                        try:
                            driver.find_element(By.XPATH, "//*[@class='tabs']").find_elements(By.TAG_NAME, 'li')[2].click()
                            sf.ajax_preloader_wait(driver)
                            if EC.presence_of_element_located((By.XPATH, "//*[@id='network_comparison_chart']")):
                                Network_comment = "Passed"
                            else:
                                Network_comment = "Failed"
                        except ElementNotInteractableException:
                            Network_comment = "Failed"
                        onlyfiles = [f for f in listdir(locator.download_dir) if
                                     isfile(join(locator.download_dir, f))]
                        path = locator.download_dir + onlyfiles[0]
                        result = csvAddition(path)
                        print("Total Gap Count: " + str(result[0]))
                        print("Total Condition Count: " + str(result[1]))
                        print("Total Disconfirm Count: " + str(result[2]))
                        print("Total Clinical RAF Score: " + str(result[3]))
                        print("Total Potential Score: " + str(result[4]))
                        print("Total Coded Score: " + str(result[5]))
                        print("Total Patient Count: " + str(result[6]))
                        os.remove(path)
                        comments = "Recapture HCCs: %s, Suspects HCCs: %s, New HCCs: %s" % (str(tag_count_list[0]), str(tag_count_list[1]), str(tag_count_list[2]))
                        if i == 553 or i == 556:
                            DataToBeValidated_num = float(DataToBeValidated_denum) - float(DataToBeValidated_num)
                            DataToBeValidated_num = round(DataToBeValidated_num, 3)
                            DataToBeValidated_denum = round(DataToBeValidated_denum, 3)
                            num = float(result[3] / result[6])
                            num = round(num, 3)
                            temp = float((result[5]-result[3])/result[6])
                            denum = float(result[4] / result[6])
                            denum = denum - temp
                            denum = round(denum, 3)
                            if abs(float(DataToBeValidated_num) - num) < 0.015 and abs(float(DataToBeValidated_denum) - denum) < 0.015:
                                ws.append([cus_id,LOB_Name, Measure, Domain_comment, Performance_comment, Network_comment, 'Passed', hcc_flag_list[2], hcc_flag_list[3], comments, Provider_Specific_url])
                            else:
                                ws.append([cus_id,LOB_Name, Measure, Domain_comment, Performance_comment, Network_comment, 'Failed', hcc_flag_list[2], hcc_flag_list[3], comments, Provider_Specific_url])
                        elif i == 554 or i == 555:
                            if int(DataToBeValidated_num) == int(DataToBeValidated_denum)-int(result[1])-int(result[2]):
                                ws.append([cus_id,LOB_Name, Measure, Domain_comment, Performance_comment, Network_comment, 'Passed', hcc_flag_list[2], hcc_flag_list[3], comments, Provider_Specific_url])
                            else:
                                ws.append([cus_id,LOB_Name, Measure, Domain_comment, Performance_comment, Network_comment, 'Failed', hcc_flag_list[2], hcc_flag_list[3], comments, Provider_Specific_url])
                        else:
                            if int(DataToBeValidated_num) == int(result[0]) and int(DataToBeValidated_denum) == int(
                                    result[0] + result[1] + result[2]):
                                ws.append([cus_id,LOB_Name, Measure, Domain_comment, Performance_comment, Network_comment, 'Passed', hcc_flag_list[2], hcc_flag_list[3], comments, Provider_Specific_url])
                            else:
                                ws.append([cus_id,LOB_Name, Measure, Domain_comment, Performance_comment, Network_comment, 'Failed', hcc_flag_list[2], hcc_flag_list[3], comments, Provider_Specific_url])
                        k -= 1
                        driver.get(Measure_Specific_url)
                        sf.ajax_preloader_wait(driver)
                        ListRow = driver.find_element(By.XPATH, "//*[@id='metric-support-prov-ls']").find_element(
                        By.TAG_NAME,"tbody").find_elements(By.TAG_NAME, 'tr')
            except NoSuchElementException:
                if i == 33:
                    Comments = "For LOB "+ LOB_Name +",the Risk Measure- One Year Recapture Rate/Review of Chronic Conditions(ID:"+str(i)+") is not present"
                    Measure = "One Year Recapture Rate/Review of Chronic Conditions"
                elif i == 551:
                    Comments = "For LOB " + LOB_Name + ",the Risk Measure- Review of Suspect Conditions(ID:" + str(
                        i) + ") is not present"
                    Measure = "Review of Suspect Conditions"
                elif i == 553:
                    Comments = "For LOB " + LOB_Name + ",the Risk Measure- HCC Score(ID:" + str(
                        i) + ") is not present"
                    Measure = "HCC Score"
                elif i == 554:
                    Comments = "For LOB " + LOB_Name + ",the Risk Measure- Review of ACA Chronic Conditions(ID:" + str(
                        i) + ") is not present"
                    Measure= "Review of ACA Chronic Conditions"
                elif i == 555:
                    Comments = "For LOB " + LOB_Name + ",the Risk Measure- Review of ACA Suspect Conditions(ID:" + str(
                        i) + ") is not present"
                    Measure = "Review of ACA Suspect Conditions"
                else:
                    Comments = "For LOB " + LOB_Name + ",the Risk Measure- ACA HCC Score(ID:" + str(
                        i) + ") is not present"
                    Measure = "ACA HCC Score"
                ws.append([cus_id,LOB_Name, Measure, '-', '-', '-', 'Not Present','-','-', Comments,'-'])
                # cellname = "N" + str(SheetRowName)
                # ws[''+cellname+''].fill = gray_background
                # # ws.conditional_formatting.add("J1:O100", rule3)
                print(Comments)
                workbook.save(screenshot_path + "\\" + workbook_title)
                continue
            except Exception as e:
                traceback.print_exc()
                print(e)
            #finally:
            #flag += 1
            driver.get(LOB_Specific_URL)
            sf.ajax_preloader_wait(driver)
            if driver.find_element(By.XPATH, "//*[@id='conti_enroll']").is_selected():
                driver.find_element(By.XPATH, "//*[@class='cont_disc_toggle']").click()
            #time.sleep(3)
            if driver.find_element(By.XPATH, "//*[@id='conti_enroll']").is_selected():
                driver.find_element(By.XPATH, "//*[@class='cont_disc_toggle']").click()
            workbook.save(screenshot_path + "\\" + workbook_title)
        sf.ajax_preloader_wait(driver)
        driver.find_element(By.XPATH, "//*[@id='qt-filter-label']").click()
        LOB_list = driver.find_element(By.XPATH, "//*[@id='filter-lob']").find_elements(By.TAG_NAME, 'li')

    rows = ws.max_row
    cols = ws.max_column
    for i in range(2, rows + 1):
        for j in range(3, cols + 1):
            if ws.cell(i, j).value == 'Passed':
                ws.cell(i, j).fill = PatternFill('solid', fgColor='0FC404')
            elif ws.cell(i, j).value == 'Failed':
                ws.cell(i, j).fill = PatternFill('solid', fgColor='FC0E03')
            elif ws.cell(i, j).value == 'Undetermined':
                ws.cell(i, j).fill = PatternFill('solid', fgColor='808080')
            elif ws.cell(i, j).value == 'Data table is empty':
                ws.cell(i, j).fill = PatternFill('solid', fgColor='FCC0BB')
    workbook.save(screenshot_path + "\\" + workbook_title)


def hccvalidation1(driver, workbook, logger, screenshot_path, run_from):
    workbook.create_sheet('HCC Validation')
    ws = workbook['HCC Validation']

    ws.append(
        ['LOB', 'HCC Measure', 'Patient Non Compliant count(UI)', 'Patient Total count(UI)', 'HCC Score(UI)', 'Gaps',
         'Conditions',
         'Disconfirms', 'Clinical RAF', 'Potential RAF', 'Non Compliant Count(Export)', 'Total Count(Export)',
         'HCC Score Calculated(Export)',
         'Status', 'Comments'])
    header = NamedStyle(name="header")
    header.font = Font(bold=True)
    header.border = Border(bottom=Side(border_style="thin"))
    header.alignment = Alignment(horizontal="center", vertical="center")
    header_row = ws[1]
    for cell in header_row:
        cell.style = header
    red_background = PatternFill(patternType='solid', fgColor="00FF0000")
    green_background = PatternFill(patternType='solid', fgColor="50C878")
    gray_background = PatternFill(patternType='solid', fgColor="5F9EA0")
    # diff_style1 = DifferentialStyle(fill=red_background)
    # rule1 = Rule(type="text", dxf=diff_style1, text="Fail")
    # diff_style2 = DifferentialStyle(fill=green_background)
    # rule2 = Rule(type="text", dxf=diff_style2, text="Pass")
    # diff_style3 = DifferentialStyle(fill=gray_background)
    # rule3 = Rule(type="text", dxf=diff_style3, text="Unexecuted")

    ws.name = "Arial"
    test_case_id = 1

    def addition(total, row, switch, arr):
        try:
            for i in range(len(arr[1])):
                if switch == 1:
                    if str(arr[1][i]).endswith("Gaps"):
                        total = total + int(arr[row][i])
                if switch == 2:
                    if str(arr[1][i]).endswith("Conditions"):
                        total = total + int(arr[row][i])
                if switch == 3:
                    if str(arr[1][i]).endswith("Disconfirms"):
                        total = total + int(arr[row][i])
                if switch == 4:
                    if str(arr[1][i]).endswith("Clinical RAF"):
                        total = total + float(arr[row][i])
                if switch == 5:
                    if str(arr[1][i]).endswith("Potential RAF"):
                        total = total + float(arr[row][i])
                if switch == 6:
                    if str(arr[1][i]).endswith("Coded RAF"):
                        total = total + float(arr[row][i])

            return total
        except ValueError:
            print("No number detected in " + arr[1][i] + " .For row no = " + str(row + 1))
            return total

    def csvAddition(filepath):
        import csv
        with open(filepath, newline='') as csvfile:
            rows = csv.reader(csvfile, delimiter=',')
            rows = list(rows)
        Gaps = 0
        Conditions = 0
        Disconfirms = 0
        Clinical = 0
        Potential = 0
        Coded = 0
        for ind in range(2, len(rows)):
            Gaps = addition(Gaps, ind, 1, rows)
            Conditions = addition(Conditions, ind, 2, rows)
            Disconfirms = addition(Disconfirms, ind, 3, rows)
            Clinical = addition(Clinical, ind, 4, rows)
            Potential = addition(Potential, ind, 5, rows)
            Coded = addition(Coded, ind, 6, rows)
        print(Gaps)
        print(Conditions)
        print(Disconfirms)
        print(Clinical)
        print(Potential)
        print(Coded)
        print(len(rows) - 2)
        return Gaps, Conditions, Disconfirms, Clinical, Potential, Coded, (len(rows) - 2)


    sf.ajax_preloader_wait(driver)
    WebDriverWait(driver, 120).until(EC.presence_of_element_located((By.XPATH, "//*[@id='qt-filter-label']")))
    print("Selected page= " + driver.title)
    Registry_URL = driver.current_url
    sf.ajax_preloader_wait(driver)
    driver.find_element(By.XPATH, "//*[@id='qt-filter-label']").click()
    Quarter_list = driver.find_element(By.XPATH, "//*[@id='filter-quarter']").find_elements(By.TAG_NAME, "li")
    # for quarter in Quarter_list:
    #     if quarter.text == "The year from GUI/Console":
    #         quarter.click()
    #         break
    Quarter_list[0].click()
    LOB_list = driver.find_element(By.XPATH, "//*[@id='filter-lob']").find_elements(By.TAG_NAME, 'li')
    for i in range(0, len(LOB_list)):
        LOB_Name = LOB_list[i].text
        print(LOB_Name)
        try:
            LOB_list[i].click()
        except ElementNotInteractableException as e:
            continue
        #LOB_list[i].click()
        driver.find_element(By.ID, "reg-filter-apply").click()
        LOB_Specific_URL = driver.current_url
        sf.ajax_preloader_wait(driver)
        if driver.find_element(By.XPATH, "//*[@id='conti_enroll']").is_selected():
            driver.find_element(By.XPATH, "//*[@class='cont_disc_toggle']").click()
        print(LOB_Specific_URL)
        HCC_measure_checklist = [33, 551, 553, 554, 555, 556]
        Provider_Specific_url = ''
        for i in HCC_measure_checklist:
            flag = 0
            try:
                Measure_Specific_url = driver.find_element(By.XPATH, "//*[@id=" + str(i) + "]//a").get_attribute('href')
                Measure = driver.find_element(By.XPATH, "//*[@id=" + str(i) + "]//*[@class='met-name']").text
                print(Measure)
                print(Measure_Specific_url)
                driver.get(Measure_Specific_url)
                sf.ajax_preloader_wait(driver)
                #driver.implicitly_wait(3)
                ListRow = driver.find_element(By.XPATH, "//*[@id='metric-support-prov-ls']").find_element(By.TAG_NAME,
                                                                                                          "tbody").find_elements(
                    By.TAG_NAME, 'tr')
                if "No data available" in ListRow[0].text and len(ListRow) == 1:
                    Comments = "No provider data in HCC measure " + str(i)
                    print(Comments)
                    ws.append([LOB_Name, Measure, '0', '0', '0', '0', '0', '0', '0', '0', '0', '0', '0', 'Undetermined',
                               Comments])
                elif len(ListRow) == 1:
                    ListRow[0].find_elements(By.TAG_NAME, 'a')[1].click()
                    sf.ajax_preloader_wait(driver)
                    print(driver.current_url)
                else:
                    #k = random.randint(2, len(ListRow)-1)
                    #k = int(input("Enter the number of providers you want to check for the Measure "+Measure+" among the number of providers "+str(len(ListRow))+":\n"))
                    k = 1
                    while k != 0:
                        Row = ListRow[random.randint(0, len(ListRow)-1)]
                        #Row = ListRow[random.randint(0, len(ListRow)-k)]
                        Row.find_elements(By.TAG_NAME, 'a')[1].click()
                        sf.ajax_preloader_wait(driver)
                        DataToBeValidated = driver.find_element(By.XPATH, "//*[@class='tab']").find_elements(By.TAG_NAME,
                                                                                                             'span')
                        Provider_Specific_url = driver.current_url
                        print(Provider_Specific_url)
                        DataToBeValidated_num = DataToBeValidated[0].text
                        DataToBeValidated_num = DataToBeValidated_num.replace(',', '')
                        print(DataToBeValidated_num)
                        DataToBeValidated_denum = DataToBeValidated[1].text
                        DataToBeValidated_denum = DataToBeValidated_denum.replace(',', '')
                        print(DataToBeValidated_denum)
                        driver.find_element(By.XPATH,
                                            "//*[@data-target='datatable_bulk_filter_0_quality_registry_list']").click()
                        driver.find_element(By.XPATH, "//*[contains(text(),'Export all to CSV')]").click()
                        sf.ajax_preloader_wait(driver)
                        onlyfiles = [f for f in listdir(locator.download_dir) if
                                     isfile(join(locator.download_dir, f))]
                        path = locator.download_dir + onlyfiles[0]
                        result = csvAddition(path)
                        os.remove(path)
                        if i == 553 or i == 556:
                            DataToBeValidated_num = float(DataToBeValidated_denum) - float(DataToBeValidated_num)
                            DataToBeValidated_num = round(DataToBeValidated_num, 3)
                            DataToBeValidated_denum = round(DataToBeValidated_denum, 3)
                            num = float(result[3] / result[6])
                            num = round(num, 3)
                            temp = float((result[5]-result[3])/result[6])
                            denum = float(result[4] / result[6])
                            denum = denum - temp
                            denum = round(denum, 3)
                            if abs(float(DataToBeValidated_num) - num) < 0.015 and abs(float(DataToBeValidated_denum) - denum) < 0.015:
                                ws.append([LOB_Name, Measure, "NA", "NA",
                                           str(DataToBeValidated_num) + "/" + str(DataToBeValidated_denum), "NA", "NA",
                                           "NA", result[3], result[4], "NA", "NA", str(num) + "/" + str(denum), 'Passed',
                                           "The HCC score for this measure is matching with UI and Export"])
                            else:
                                ws.append([LOB_Name, Measure, "NA", "NA",
                                           str(DataToBeValidated_num) + "/" + str(DataToBeValidated_denum), "NA", "NA",
                                           "NA", result[3], result[4], "NA", "NA", str(num) + "/" + str(denum), 'Failed',
                                           "The HCC score for this measure is not matching with UI and Export"])
                        elif i == 554 or i == 555:
                            if int(DataToBeValidated_num) == int(DataToBeValidated_denum)-int(result[1])-int(result[2]):
                                ws.append(
                                    [LOB_Name, Measure, DataToBeValidated_num, DataToBeValidated_denum, "NA",
                                     int(DataToBeValidated_denum)-int(result[1])-int(result[2]), result[1],
                                     result[2], "NA", "NA", int(DataToBeValidated_denum)-int(result[1])-int(result[2]),
                                     result[0] + result[1] + result[2], "NA", 'Passed',
                                     "The Compliant and total patient count of UI is matching with export"])
                            else:
                                ws.append(
                                    [LOB_Name, Measure, DataToBeValidated_num, DataToBeValidated_denum, "NA",
                                     int(DataToBeValidated_denum)-int(result[1])-int(result[2]), result[1],
                                     result[2], "NA", "NA", result[0], result[0] + result[1] + result[2], "NA", 'Failed',
                                     'The Compliant and total patient count of UI is not matching with export'])
                        else:
                            if int(DataToBeValidated_num) == int(result[0]) and int(DataToBeValidated_denum) == int(
                                    result[0] + result[1] + result[2]):
                                ws.append(
                                    [LOB_Name, Measure, DataToBeValidated_num, DataToBeValidated_denum, "NA", result[0],
                                     result[1],
                                     result[2], "NA", "NA", result[0], result[0] + result[1] + result[2], "NA", 'Passed',
                                     "The Compliant and total patient count of UI is matching with export"])

                                # ws.conditional_formatting.add("J1:O100", rule2)
                            else:
                                ws.append(
                                    [LOB_Name, Measure, DataToBeValidated_num, DataToBeValidated_denum, "NA", result[0],
                                     result[1],
                                     result[2], "NA", "NA", result[0], result[0] + result[1] + result[2], "NA", 'Failed',
                                     'The Compliant and total patient count of UI is not matching with export'])
                                # ws.conditional_formatting.add("J1:O100", rule1)
                    k -= 1
                    driver.get(Provider_Specific_url)
            except NoSuchElementException:
                Comments = "HCC measure with id " + str(i) + " not found in LOB " + LOB_Name
                # ws.append([LOB_Name, str(i), '0', '0', '0', '0', '0', '0', '0', '0', '0', '0', '0', 'Unexecuted', Comments])
                # cellname = "N" + str(SheetRowName)
                # ws[''+cellname+''].fill = gray_background
                # # ws.conditional_formatting.add("J1:O100", rule3)
                print(Comments)
                if flag == 5:
                    print("No more HCC measures switching to next LOB")
            except Exception as e:
                traceback.print_exc()
                print(e)
            finally:
                flag += 1
                driver.get(LOB_Specific_URL)
                sf.ajax_preloader_wait(driver)
                #time.sleep(3)
                workbook.save(screenshot_path + "\\Report.xlsx")
        sf.ajax_preloader_wait(driver)
        driver.find_element(By.XPATH, "//*[@id='qt-filter-label']").click()
        LOB_list = driver.find_element(By.XPATH, "//*[@id='filter-lob']").find_elements(By.TAG_NAME, 'li')

    rows = ws.max_row
    cols = ws.max_column
    for i in range(2, rows + 1):
        for j in range(3, cols + 1):
            if ws.cell(i, j).value == 'Passed':
                ws.cell(i, j).fill = PatternFill('solid', fgColor='0FC404')
            elif ws.cell(i, j).value == 'Failed':
                ws.cell(i, j).fill = PatternFill('solid', fgColor='FC0E03')
            elif ws.cell(i, j).value == 'Not Present':
                ws.cell(i, j).fill = PatternFill('solid', fgColor='808080')
            elif ws.cell(i, j).value == 'Data table is empty':
                ws.cell(i, j).fill = PatternFill('solid', fgColor='FCC0BB')












