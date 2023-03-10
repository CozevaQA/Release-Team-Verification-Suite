import os
import time
import traceback

from openpyxl.styles import Font, PatternFill
from selenium.webdriver.common.by import By
from selenium.webdriver.support.wait import WebDriverWait
from tkinter import *
import setups
import logging
import ExcelProcessor as db
import context_functions as cf
import support_functions as sf
import setups as st
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
import variablestorage as locator
from openpyxl import Workbook, load_workbook
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import NoSuchElementException, ElementNotInteractableException, \
    ElementClickInterceptedException, TimeoutException

ENV = 'PROD'
client_id = str(1500)
MY = "2022"

def fetch_client_name():
    root = Tk()


    def on_start():
        global client_id, MY
        client_id = db.fetchCustomerID(selected_cust.get())
        MY = my_entry.get()
        root.destroy()


    customer_label = Label(root, text="Select customer", font=("Nunito Sans", 10))
    selected_cust = StringVar()
    selected_cust.set("Customer")
    customer_list = db.getCustomerList()  # vs.customer_list
    customer_drop = OptionMenu(root, selected_cust, *customer_list)


    customer_label.grid(row=0, column=0)
    customer_drop.grid(row=1, column=0)
    Label(root, text="Measurement Year",  font=("Nunito Sans", 10)).grid(row=0, column=3)
    my_entry = Entry(root)
    my_entry.grid(row=1, column=3)

    Button(root, text="Start Test", command=on_start, font=("Nunito Sans", 10)).grid(row=2, column=0, columnspan=5)





    root.title("Special column Validation")
    root.iconbitmap("assets/icon.ico")
    # root.geometry("400x400")
    root.mainloop()




fetch_client_name()
print(client_id, MY)
column_master_list = [["RCC", 33, 554],                         #Exclamation icon (!) beside last visit column which denotes that the patient has not visited since last 9 months
                      ["RSC", 551, 555],                        #Exclamation icon (!) beside last visit column which denotes that the patient has not visited since last 9 months
                      ["BCS", 1, 321, 608, 900, 946, 947],      #Last Test & Last test type should not be blank (atleast few patient should have values)
                      ["COL", 3, 607, 907, 954, 955],           #Last Test and Test Date
                      ["CHL", 165, 355],                        #Denom Eligibility column
                      ["PPC", 183, 184],                        #Den Events
                      ["TRC", 153, 315, 154, 314],              #Den Events
                      ["CDC", 6, 84, 20, 39],                   #Test Result
                      ["CIS", 212],                             #Immunization
                      ["W30", 36, 37],                          #Remaining
                      ["AWV", 219],                             #Last Test Type
                      ["PDC", 27, 28, 29],                      #PDC col, Fall in/out
                      ["OMW", 51, 351],                         #Denom Eligibility column
                      ["AMR", 180, 380, 480]]                   # -- ?? --


#column_master_list = [["BCS", 1, 321, 608, 900, 946, 947]]



print("Hello World")

#make reporting folder

report_folder = os.path.join(locator.parent_dir,"Special Column Reports")
isdir = os.path.isdir(report_folder)
if not isdir:
    os.mkdir(report_folder)

wb = Workbook()
ws = wb.active
workbook_title = str(db.fetchCustomerName(client_id))+"_Special Columns_"+sf.date_time()+"_"+ENV+".xlsx"
ws.title = str(client_id)+"_"+ENV
ws.append(["ID", "LoB", "Metric Name/ID", "Provider", "Scenario", "Status", "Comments", "URL"])
wb.save(report_folder+ "\\"+workbook_title)
test_case_id = 1
driver = setups.driver_setup()
if ENV == 'CERT':
    setups.login_to_cozeva_cert(client_id)
elif ENV == 'STAGE':
    setups.login_to_cozeva_stage()
elif ENV == "PROD":
    setups.login_to_cozeva(client_id)
else:
    print("ENV INVALID")
    exit(3)

sf.ajax_preloader_wait(driver)
WebDriverWait(driver, 120).until(EC.presence_of_element_located((By.XPATH, "//*[@id='qt-filter-label']")))
print("Selected page= " + driver.title)
Registry_URL = driver.current_url
sf.ajax_preloader_wait(driver)
driver.find_element(By.XPATH, "//*[@id='qt-filter-label']").click()
Quarter_list = driver.find_element(By.XPATH, "//*[@id='filter-quarter']").find_elements(By.TAG_NAME, "li")
for quarter in Quarter_list:
    if quarter.text == MY:
        quarter.click()
        break
#Quarter_list[0].click()
LOB_list = driver.find_element(By.XPATH, "//*[@id='filter-lob']").find_elements(By.TAG_NAME, 'li')
for i in range(0, len(LOB_list)):
    LOB_Name = LOB_list[i].text
    print("LOB Name for 2022: " + LOB_Name)
    try:
        LOB_list[i].click()
    except ElementNotInteractableException as e:
        continue
    #LOB_list[i].click()
    driver.find_element(By.ID, "reg-filter-apply").click()
    LOB_Specific_URL = driver.current_url
    sf.ajax_preloader_wait(driver)
    if driver.find_element(By.XPATH, "//*[@id='conti_enroll']").is_selected():
        driver.find_element(By.XPATH, "//*[@class='cont_disc_toggle']").click()
    print("LOB URL: " + LOB_Specific_URL)

    #Search the measure abbreviation
    for measure_index, measure in enumerate(column_master_list):
        driver.find_element_by_xpath(locator.xpath_filter_measure_list).click()
        time.sleep(2)
        driver.find_element_by_id("qt-search-met").clear()
        driver.find_element_by_id("qt-search-met").send_keys(measure[0])
        driver.execute_script("arguments[0].scrollIntoView();", driver.find_element_by_id("qt-apply-search"))
        driver.find_element_by_id("qt-apply-search").click()
        time.sleep(1)
        registry_element = driver.find_element_by_id("registry_body").find_element_by_tag_name("ul")
        #measure_elements = driver.find_element_by_id("registry_body").find_element_by_tag_name("ul").find_elements_by_tag_name("li")
        for k, metric_id in enumerate(measure):
            print(LOB_Name + "-" + str(metric_id) + "-" + measure[0])
            if k == 0:
                continue
            if registry_element.find_elements_by_id(metric_id):
                registry_element.find_element_by_id(metric_id).click()
                print('validating')
                sf.ajax_preloader_wait(driver)
                #click on first man icon
                first_provider_element = driver.find_element(By.XPATH, "//*[@id='metric-support-prov-ls']").find_element(By.TAG_NAME,"tbody").find_elements(By.TAG_NAME, 'tr')[0]
                if first_provider_element.text == "No Data Available":
                    driver.get(LOB_Specific_URL)
                    sf.ajax_preloader_wait(driver)
                    WebDriverWait(driver, 60).until(
                        EC.presence_of_element_located((By.XPATH, locator.xpath_filter_measure_list)))
                    driver.find_element_by_xpath(locator.xpath_filter_measure_list).click()
                    time.sleep(2)
                    driver.find_element_by_id("qt-search-met").clear()
                    driver.find_element_by_id("qt-search-met").send_keys(measure[0])
                    driver.execute_script("arguments[0].scrollIntoView();",
                                          driver.find_element_by_id("qt-apply-search"))
                    driver.find_element_by_id("qt-apply-search").click()
                    time.sleep(1)
                    registry_element = driver.find_element_by_id("registry_body").find_element_by_tag_name("ul")
                    continue
                else:
                    first_provider_element.find_elements(By.TAG_NAME, 'a')[1].click()
                    sf.ajax_preloader_wait(driver)
                    provider_name = driver.find_element(By.XPATH, locator.xpath_context_Name).text
                    if measure[0] == 'RCC': #RCC -
                        RCC_counter = driver.find_elements(By.XPATH, "//span[@title='Patient has not been seen for over nine months.']")
                        if len(RCC_counter) > 0:
                            ws.append([test_case_id, LOB_Name, measure[0] + "_" + str(metric_id), provider_name, "Last Visit - Presence of '!'", "Passed", "Number of '!' :"+str(len(RCC_counter)), driver.current_url])
                        else:
                            ws.append(
                                [test_case_id, LOB_Name, measure[0] + "_" + str(metric_id), provider_name, "Last Visit - Presence of '!'",
                                 "Failed", "'!' not present", driver.current_url])
                        test_case_id+=1
                    elif measure[0] == 'RSC': #RSC
                        RSC_counter = driver.find_elements(By.XPATH,
                                                           "//span[@title='Patient has not been seen for over nine months.']")
                        if len(RSC_counter) > 0:
                            ws.append([test_case_id, LOB_Name, measure[0] + "_" + str(metric_id), provider_name,
                                       "Last Visit - Presence of '!'", "Passed",
                                       "Number of '!' :" + str(len(RSC_counter)), driver.current_url])
                        else:
                            ws.append(
                                [test_case_id, LOB_Name, measure[0] + "_" + str(metric_id), provider_name,
                                 "Last Visit - Presence of '!'",
                                 "Failed", "'!' not present", driver.current_url])
                        test_case_id += 1
                    elif measure[0] == 'BCS': #BCS
                        last_test_elements = driver.find_elements(By.XPATH, "//td[@class=' last_test_date last_test_date_pt']")
                        last_test_type_elements = driver.find_elements(By.XPATH, "//td[@class=' last_test_name last_test_name_pt']")
                        last_test_counter = 0
                        last_test_type_counter = 0
                        print(len(last_test_elements))
                        for element in last_test_elements:
                            if len(element.text)>0:
                                last_test_counter+=1
                        for element in last_test_type_elements:
                            if len(element.text)>0:
                                last_test_type_counter+=1

                        if last_test_counter > 0:
                            ws.append([test_case_id, LOB_Name, measure[0] + "_" + str(metric_id), provider_name,
                                       "Last Test", "Passed",
                                       "Number of Last Test entries :" + str(last_test_counter), driver.current_url])
                        else:
                            ws.append([test_case_id, LOB_Name, measure[0] + "_" + str(metric_id), provider_name,
                                       "Last Test", "Failed",
                                       "Number of Last Test entries : 0", driver.current_url])

                        if last_test_type_counter > 0:
                            ws.append([test_case_id, LOB_Name, measure[0] + "_" + str(metric_id), provider_name,
                                       "Last Test Type", "Passed",
                                       "Number of Last Test entries :" + str(last_test_type_counter), driver.current_url])
                        else:
                            ws.append([test_case_id, LOB_Name, measure[0] + "_" + str(metric_id), provider_name,
                                       "Last Test Type", "Failed",
                                       "Number of Last Test entries : 0", driver.current_url])
                        last_test_counter = 0
                        last_test_type_counter = 0
                    elif measure[0] == 'COL': #COL
                        last_test_elements = driver.find_elements(By.XPATH,
                                                                  "//td[@class=' last_test_date last_test_date_pt']")
                        last_test_type_elements =  driver.find_elements(By.XPATH, "//td[@class=' last_test_name last_test_name_pt']")
                        last_test_counter = 0
                        last_test_type_counter = 0
                        print(len(last_test_elements))
                        for element in last_test_elements:
                            if len(element.text) > 0:
                                last_test_counter += 1
                        for element in last_test_type_elements:
                            if len(element.text) > 0:
                                last_test_type_counter += 1

                        if last_test_counter > 0:
                            ws.append([test_case_id, LOB_Name, measure[0] + "_" + str(metric_id), provider_name,
                                       "Last Test", "Passed",
                                       "Number of Last Test entries :" + str(last_test_counter), driver.current_url])
                        else:
                            ws.append([test_case_id, LOB_Name, measure[0] + "_" + str(metric_id), provider_name,
                                       "Last Test", "Failed",
                                       "Number of Last Test entries : 0", driver.current_url])

                        if last_test_type_counter > 0:
                            ws.append([test_case_id, LOB_Name, measure[0] + "_" + str(metric_id), provider_name,
                                       "Last Test Type", "Passed",
                                       "Number of Last Test entries :" + str(last_test_type_counter),
                                       driver.current_url])
                        else:
                            ws.append([test_case_id, LOB_Name, measure[0] + "_" + str(metric_id), provider_name,
                                       "Last Test Type", "Failed",
                                       "Number of Last Test entries : 0", driver.current_url])
                        last_test_counter = 0
                        last_test_type_counter = 0


                    elif measure[0] == 'CHL': #CHL
                        denom_eli_elements = driver.find_elements(By.XPATH,
                                                                  "//td[@class=' denom_date denom_date_pt']")
                        denom_eli_counter = 0

                        for element in denom_eli_elements:
                            if len(element.text) > 0:
                                denom_eli_counter+=1
                        if denom_eli_counter > 0:
                            ws.append([test_case_id, LOB_Name, measure[0] + "_" + str(metric_id), provider_name,
                                       "Denominator Eligiblity column", "Passed",
                                       "Number of Denominator Eligiblity entries :" + str(denom_eli_counter), driver.current_url])
                        else:
                            ws.append([test_case_id, LOB_Name, measure[0] + "_" + str(metric_id), provider_name,
                                       "Denominator Eligiblity column", "Failed",
                                       "Number of Denominator Eligiblity entries : 0", driver.current_url])

                    elif measure[0] == 'PPC': #PPC
                        den_events_elements =  driver.find_elements(By.XPATH,
                                                                  "//td[@class=' util_count2 util_count2_pt']")

                        den_event_counter = 0

                        for element in den_events_elements:
                            if len(element.text) > 0:
                                den_event_counter += 1
                        if den_event_counter > 0:
                            ws.append([test_case_id, LOB_Name, measure[0] + "_" + str(metric_id), provider_name,
                                       "Denominator Events Column", "Passed",
                                       "Number of Denominator Event entries :" + str(den_event_counter),
                                       driver.current_url])
                        else:
                            ws.append([test_case_id, LOB_Name, measure[0] + "_" + str(metric_id), provider_name,
                                       "Denominator Events Column", "Failed",
                                       "Number of Denominator Event entries : 0", driver.current_url])


                    elif measure[0] == 'TRC': #TRC
                        den_events_elements = driver.find_elements(By.XPATH,
                                                                   "//td[@class=' util_count2 util_count2_pt']")

                        den_event_counter = 0

                        for element in den_events_elements:
                            if len(element.text) > 0:
                                den_event_counter += 1
                        if den_event_counter > 0:
                            ws.append([test_case_id, LOB_Name, measure[0] + "_" + str(metric_id), provider_name,
                                       "Denominator Events Column", "Passed",
                                       "Number of Denominator Event entries :" + str(den_event_counter),
                                       driver.current_url])
                        else:
                            ws.append([test_case_id, LOB_Name, measure[0] + "_" + str(metric_id), provider_name,
                                       "Denominator Events Column", "Failed",
                                       "Number of Denominator Event entries : 0", driver.current_url])

                    elif measure[0] == 'CDC': #CDC
                        test_result_elements = []
                    elif measure[0] == 'CIS': #CIS
                        immunization_elements = driver.find_elements(By.XPATH,
                                                                   "//td[@class=' vaccine_gap vaccine_gap_pt']")

                        immunization_counter = 0

                        for element in immunization_elements:
                            if len(element.text) > 0:
                                immunization_counter += 1
                        if immunization_counter > 0:
                            ws.append([test_case_id, LOB_Name, measure[0] + "_" + str(metric_id), provider_name,
                                       "Immunization Column", "Passed",
                                       "Number of Immunization Column entries :" + str(immunization_counter),
                                       driver.current_url])
                        else:
                            ws.append([test_case_id, LOB_Name, measure[0] + "_" + str(metric_id), provider_name,
                                       "Immunization Column", "Failed",
                                       "Number of Immunization Column entries : 0", driver.current_url])


                    elif measure[0] == 'W30': #W30
                        remaining_column_elements = driver.find_elements(By.XPATH,
                                                                   "//td[@class=' gap_count gap_count_pt']")

                        remaining_counter = 0

                        for element in remaining_column_elements:
                            if len(element.text) > 0:
                                remaining_counter += 1
                        if remaining_counter > 0:
                            ws.append([test_case_id, LOB_Name, measure[0] + "_" + str(metric_id), provider_name,
                                       "Remaining Column", "Passed",
                                       "Number of Remaining Column entries :" + str(remaining_counter),
                                       driver.current_url])
                        else:
                            ws.append([test_case_id, LOB_Name, measure[0] + "_" + str(metric_id), provider_name,
                                       "Remaining Column", "Failed",
                                       "Number of Remaining Column entries : 0", driver.current_url])

                    elif measure[0] == 'AWV': #AWV
                        last_test_type_elements = driver.find_elements(By.XPATH,
                                                                   "//td[@class=' last_test_name last_test_name_pt']")

                        last_test_type_counter = 0

                        for element in last_test_type_elements:
                            if len(element.text) > 0:
                                last_test_type_counter += 1
                        if last_test_type_counter > 0:
                            ws.append([test_case_id, LOB_Name, measure[0] + "_" + str(metric_id), provider_name,
                                       "Last Test Type Column", "Passed",
                                       "Number of Last Test Type column entries :" + str(last_test_type_counter),
                                       driver.current_url])
                        else:
                            ws.append([test_case_id, LOB_Name, measure[0] + "_" + str(metric_id), provider_name,
                                       "Last Test Type Column", "Failed",
                                       "Number of Last Test Type column entries : 0", driver.current_url])


                    elif measure[0] == 'PDC': #PDC
                        fall_in_out_elements = driver.find_elements(By.XPATH,
                                                                   "//td[@class=' pdc_due_date pdc_due_date_pt']")

                        fall_in_out_counter = 0

                        for element in fall_in_out_elements:
                            if len(element.text) > 0:
                                fall_in_out_counter += 1
                        if fall_in_out_counter > 0:
                            ws.append([test_case_id, LOB_Name, measure[0] + "_" + str(metric_id), provider_name,
                                       "Fall in/out Column", "Passed",
                                       "Number of Fall in/out Column entries :" + str(fall_in_out_counter),
                                       driver.current_url])
                        else:
                            ws.append([test_case_id, LOB_Name, measure[0] + "_" + str(metric_id), provider_name,
                                       "Fall in/out Column", "Failed",
                                       "Number of Fall in/out Column entries : 0", driver.current_url])


                    elif measure[0] == 'OMW': #OMW
                        denom_eli_elements = driver.find_elements(By.XPATH,
                                                                  "//td[@class=' denom_date denom_date_pt']")
                        denom_eli_counter = 0

                        for element in denom_eli_elements:
                            if len(element.text) > 0:
                                denom_eli_counter += 1
                        if denom_eli_counter > 0:
                            ws.append([test_case_id, LOB_Name, measure[0] + "_" + str(metric_id), provider_name,
                                       "Denominator Eligiblity column", "Passed",
                                       "Number of Denominator Eligiblity entries :" + str(denom_eli_counter),
                                       driver.current_url])
                        else:
                            ws.append([test_case_id, LOB_Name, measure[0] + "_" + str(metric_id), provider_name,
                                       "Denominator Eligiblity column", "Failed",
                                       "Number of Denominator Eligiblity entries : 0", driver.current_url])
                    elif measure[0] == 'AMR': #AMR
                        x=0
                    wb.save(report_folder + "\\" + workbook_title)
                    driver.get(LOB_Specific_URL)
                    sf.ajax_preloader_wait(driver)
                    WebDriverWait(driver, 60).until(
                        EC.presence_of_element_located((By.XPATH, locator.xpath_filter_measure_list)))
                    driver.find_element_by_xpath(locator.xpath_filter_measure_list).click()
                    time.sleep(2)
                    driver.find_element_by_id("qt-search-met").clear()
                    driver.find_element_by_id("qt-search-met").send_keys(measure[0])
                    driver.execute_script("arguments[0].scrollIntoView();",
                                          driver.find_element_by_id("qt-apply-search"))
                    driver.find_element_by_id("qt-apply-search").click()
                    time.sleep(1)
                    registry_element = driver.find_element_by_id("registry_body").find_element_by_tag_name("ul")
                    #this should open the first MSPL
            else:
                continue

    driver.get(Registry_URL)
    sf.ajax_preloader_wait(driver)
    WebDriverWait(driver, 60).until(
        EC.presence_of_element_located((By.XPATH, locator.xpath_filter_measure_list)))
    driver.find_element(By.XPATH, "//*[@id='qt-filter-label']").click()
    LOB_list = driver.find_element(By.XPATH, "//*[@id='filter-lob']").find_elements(By.TAG_NAME, 'li')

rows = ws.max_row
cols = ws.max_column
for i in range(2, rows + 1):
    for j in range(3, cols + 1):
        if ws.cell(i, j).value == 'Passed':
            ws.cell(i, j).fill = PatternFill('solid', fgColor='0FC404')
        elif ws.cell(i, j).value == 'Failed':
            ws.cell(i, j).fill = PatternFill('solid', fgColor='FC0E03')
        elif ws.cell(i, j).value == 'Data table is empty':
            ws.cell(i, j).fill = PatternFill('solid', fgColor='FCC0BB')

wb.save(report_folder + "\\" + workbook_title)
driver.quit()




